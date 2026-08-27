"""
WhatsApp Bot State Machine (Twilio or Meta Cloud API compatible)
Handles the full conversation flow for lorry assignment.

State flow:
  IDLE -> AWAIT_USER_ID -> AWAIT_MASTER_UPLOAD -> AWAIT_TRIP_DAY -> AWAIT_EXCEL
       -> CONFIRMING -> DONE
  After login the user uploads the daily master lorry file; the bot reads which
  lorries are Available for that user (cross-use supported), rejecting the file
  if any plate is Available under two users. Then Today/Tomorrow, then the DOs.
  (Auto-assigns best lorry for all DOs silently, shows summary for confirmation)
"""

import io
import json
import os
import re
import threading
import time
from datetime import date, datetime, time as dtime
import pandas as pd
from lorry_engine import LorryEngine
from assignment_config import (
    # Capacity / utilisation
    CAPACITY_TARGET, MIN_UTIL_TO_ASSIGN, MAX_STOPS_PER_LORRY,
    NAIK_FACTOR, SAME_ROUTE_NAIK, REBAL_THRESHOLD, FILL_TARGET,
    MAX_DOS_PER_LORRY,
    # Lorry size
    LORRY_LARGE_MIN_TON, LORRY_SMALL_MAX_TON, LORRY_TINY_EXCL_TON,
    OUTSTATION_MIN_TON as _OUTSTATION_MIN_TON,
    TINY_OUTSTATION_PREFIXES as _TINY_OUTSTATION_PREFIXES,
    TINY_OUTSTATION_MAX_TON as _TINY_OUTSTATION_MAX_TON,
    URBAN_MAX_TON as _URBAN_MAX_TON,
    # Geographic
    DEPOT_LAT, DEPOT_LON,
    CROSS_BEARING_LIMIT as _CROSS_BEARING_LIMIT,
    MAX_CITY_MERGE_KM_OUTSTATION as _MAX_CITY_MERGE_KM_OUTSTATION,
    MAX_GEO_GAP_DEG        as _MAX_GEO_GAP_DEG,
    URBAN_MERGE_SPREAD_DEG as _URBAN_MERGE_SPREAD,
    # Destination classification
    DEST_MIN_TON         as _DEST_MIN_TON,
    DEST_LARGE_LONG_CLUSTERS  as _DEST_LARGE_LONG_CLUSTERS,
    DEST_MEDIUM_LONG_CLUSTERS as _DEST_MEDIUM_LONG_CLUSTERS,
    DEST_MEDIUM_LONG_KV_CODES as _DEST_MEDIUM_LONG_KV_CODES,
    DEST_URBAN_GROUPS    as _DEST_URBAN_GROUPS,
    DEST_SORT_PRI        as _DEST_SORT_PRI,
    URBAN_COMPATIBLE_STATES as _URBAN_COMPATIBLE_STATES,
    # Geography / normalisation
    STATE_NAME_NORM      as _STATE_NAME_NORM,
    POSTCODE_STATE_RANGES as _POSTCODE_STATE_RANGES,
    # Schedule / remarks parsing
    SCHD_DAY_MAP         as _SCHD_DAY_MAP,
    REMARKS_KEYWORD_DAY  as _REMARKS_KEYWORD_DAY,
    REMARKS_FIELD3_TON_CAP as _REMARKS_FIELD3_TON_CAP,
    REMARKS_SIZE_ALIASES   as _REMARKS_SIZE_ALIASES,
    # Route rules
    ROUTE_CORRIDOR_GROUPS as _ROUTE_CORRIDOR_GROUPS,
    LORRY_STRICT_ROUTE   as _LORRY_STRICT_ROUTE,
    ROUTE_PREFERRED_LORRY as _ROUTE_PREFERRED_LORRY,
    # Tiny-item guard
    TINY_ITEM_AVG_WEIGHT_T as _TINY_ITEM_AVG_WEIGHT_T,
)

_HERE = os.path.dirname(os.path.abspath(__file__))

# ── Load assignment rules file ────────────────────────────────────────────────
# ASSIGNMENT_RULES.md is the authoritative rules reference.  It is loaded once
# at startup so the contents are available for logging/display.  All assignment
# logic in this file must conform to the rules defined there.
_RULES_PATH = os.path.join(_HERE, "ASSIGNMENT_RULES.md")
try:
    with open(_RULES_PATH, encoding="utf-8") as _rf:
        ASSIGNMENT_RULES_TEXT = _rf.read()
    print(f"[RULES] Loaded {_RULES_PATH} ({len(ASSIGNMENT_RULES_TEXT)} chars)")
except FileNotFoundError:
    ASSIGNMENT_RULES_TEXT = ""
    print(f"[RULES] WARNING: {_RULES_PATH} not found — rules file missing!")
# Keep data files in a separate subfolder so Flask's watchdog reloader
# never sees Excel writes and restarts the server mid-request.
_DATA_DIR = os.path.join(_HERE, "data")
os.makedirs(_DATA_DIR, exist_ok=True)

PLANNING_PATH  = os.path.join(_DATA_DIR, "LORRY DAILY PLANNING.xlsx")         # lorry naik + route codes
MASTER_PATH    = PLANNING_PATH   # alias kept for backwards compat inside engine calls
MASTER_LORRY_PATH = os.path.join(_DATA_DIR, "master_lorry.xlsx")              # capacity lookup for simple lorry lists
# History path — single source of truth (manual-assignment history, new format
# with LONGITUD GPS column).  Use the .xlsx everywhere; the old .xls duplicate is
# no longer referenced so it can be deleted.
HISTORY_PATH     = os.path.join(_DATA_DIR, "ZSDOROUTEWRH.xlsx")               # primary (new format, manual assignments)
HISTORY_PATH_ALT = os.path.join(_DATA_DIR, "ZSDOROUTEWRH-bot.xlsx")          # bot-exported (new format)
HISTORY_PATH_OLD = os.path.join(_DATA_DIR, "126-A BI(ES) TRIP ROUTE CODE.xlsx")  # legacy reference

# ── Malaysia States & Cities reference data ───────────────────────────────────
# Loaded once at startup from the "Malaysia States & Cities" sheet in
# LORRY DAILY PLANNING.xlsx.  Used to:
#   1. Normalise / fill missing STATE values from CITY in DO items.
#   2. Validate that two cities belong to the same state before merging.
_CITY_TO_STATE: dict[str, str] = {}   # city.upper() → canonical state key
_STATE_TO_CITIES: dict[str, set] = {} # canonical state key → set of city names (upper)
_POSTCODE_TO_STATE: dict[int, str] = {}   # exact postcode (int) → canonical state key

# _STATE_NAME_NORM imported from assignment_config

def _norm_state(st: str) -> str:
    """Normalise a state name to the canonical short form used in DO items."""
    return _STATE_NAME_NORM.get(st, st)

def _load_malaysia_geo() -> None:
    """Populate _CITY_TO_STATE, _STATE_TO_CITIES and _POSTCODE_TO_STATE from the
    "Malaysia States & Cities" sheet (columns: STATE, CITY / TOWN, POSTCODE)."""
    global _CITY_TO_STATE, _STATE_TO_CITIES, _POSTCODE_TO_STATE
    if not os.path.exists(PLANNING_PATH):
        return
    try:
        df = pd.read_excel(PLANNING_PATH,
                           sheet_name="Malaysia States & Cities",
                           usecols=[0, 1, 2], header=0)
        df.columns = ["STATE", "CITY", "POSTCODE"]
        df = df.dropna(subset=["STATE", "CITY"])
        for _, row in df.iterrows():
            st_raw = str(row["STATE"]).strip().upper()
            st  = _norm_state(st_raw)
            cty = str(row["CITY"]).strip().upper()
            if st and cty:
                _CITY_TO_STATE[cty] = st
                _STATE_TO_CITIES.setdefault(st, set()).add(cty)
            # Exact postcode → state (from the sheet's POSTCODE column)
            _pc_raw = row.get("POSTCODE")
            if st and _pc_raw is not None and str(_pc_raw).strip() not in ("", "nan", "NaN"):
                try:
                    _POSTCODE_TO_STATE[int(float(str(_pc_raw).strip()))] = st
                except (ValueError, TypeError):
                    pass
    except Exception:
        pass   # sheet missing or malformed — fall back to item STATE column

_load_malaysia_geo()


def _resolve_state(state_raw: str, city_raw: str) -> str:
    """Return the canonical STATE for an item.
    Priority: explicit STATE column → lookup city in _CITY_TO_STATE.
    """
    st = state_raw.strip().upper()
    if st and st not in ("NAN", "NONE", "-", ""):
        return st
    cty = city_raw.strip().upper()
    return _CITY_TO_STATE.get(cty, "")


def _load_user_route_prefixes(user: str) -> set | None:
    """Return the set of route-code prefixes assigned to *user* by reading
    the '{USER} ROUTE' sheet from LORRY DAILY PLANNING.xlsx.
    Scans all columns for route-like strings (robust to column layout changes).
    Returns None if the file/sheet doesn't exist (no filtering applied).
    """
    if not os.path.exists(PLANNING_PATH):
        return None
    try:
        u = user.strip().upper()
        sheet_name = f"{u} ROUTE"   # e.g. "ABI ROUTE" or "VIVIAN ROUTE"
        df = pd.read_excel(PLANNING_PATH, sheet_name=sheet_name, header=None)
        prefixes: set[str] = set()
        # Scan every cell — route strings start with a code like KV01A, PH09, NS04
        _route_pat = re.compile(r'^([A-Za-z]{2,4}\d{1,2}[A-Za-z]?)')
        for col_idx in range(df.shape[1]):
            for val in df.iloc[:, col_idx].dropna().astype(str):
                m = _route_pat.match(val.strip())
                if m:
                    prefixes.add(m.group(1).upper())
        return prefixes if prefixes else None
    except Exception:
        return None

# _SCHD_DAY_MAP and _REMARKS_KEYWORD_DAY imported from assignment_config

def _parse_remarks_days(remarks: str) -> set[int] | None:
    """Parse a REMARKS string into a set of delivery weekday integers.

    Returns:
        None  — no day information found (no restriction applies)
        set   — explicit day set; caller should check trip weekday is in set
    """
    if not remarks or str(remarks).strip().lower() in ("nan", "none", ""):
        return None
    txt = str(remarks).strip().upper()

    # "SETIAP HARI" / "DAILY" / "EVERY DAY" → no restriction
    if re.search(r'\bSETIAP\s+HARI\b|\bDAILY\b|\bEVERY\s+DAY\b', txt):
        return None

    # Timing/logistics notes that are not day restrictions — always assign.
    #   "NEXT DAY MUST DELIVER" — instruction to deliver next day, not a day filter
    #   "SAME DAY DELIVERY"     — urgency note
    #   "LUNCH TIME"            — time window note
    #   "MORNING TRIP"          — trip timing note
    #   "SMALL LORRY"           — lorry size suggestion (handled elsewhere)
    if re.search(
        r'\bNEXT\s+DAY\b|\bSAME\s+DAY\b|\bLUNCH\s+TIME\b'
        r'|\bMORNING\s+TRIP\b|\bAFTERNOON\s+TRIP\b'
        r'|\bSMALL\s+LORRY\b|\bBIG\s+LORRY\b|\bLARGE\s+LORRY\b'
        r'|\bAM\s+FIRST\s+TRIP\b|\bPM\s+TRIP\b',
        txt
    ):
        return None

    # Operational / informational remarks that mention days as context (not
    # delivery-day restrictions).  Examples:
    #   "LORRY OPERASI ISNIN-SABTU 4PM"  — lorry hours, not customer restriction
    #   "WAKTU OPERASI ..."               — operation hours note
    if re.search(r'\b(?:OPERASI|WAKTU\s+OPERASI|OPERATION\s+HOURS?|OPERATING\s+HOURS?)\b', txt):
        return None

    # Negative patterns: a day mentioned as CLOSED / OFF / not accepted.
    # Patterns: "<DAY> OFF", "<DAY> TUTUP", "<DAY> TAK TERIMA", "TUTUP <DAY>",
    #           "CLOSED <DAY>", "TAK TERIMA <DAY>", "<DAY> TIDAK HANTAR",
    #           "KEDAI TUTUP" (shop closed on those days),
    #           "JANGAN HANTAR <DAY>" (do not deliver on those days)
    _NEG_SUFFIX = (
        r'(?:[\s:]+\w+)*[\s:]+(?:OFF|TUTUP|TAK\s+TERIMA(?:\s+BARANG)?'
        r'|TIDAK\s+(?:TERIMA|HANTAR|ACCEPT|BOLEH)'
        r'|CLOSED|BLOCKED|SKIP|TOLAK)'
    )
    _NEG_PREFIX = (
        r'(?:TUTUP|CLOSED|OFF|TAK\s+TERIMA(?:\s+BARANG)?'
        r'|TIDAK\s+(?:TERIMA|HANTAR)'
        r'|JANGAN\s+HANTAR(?:\s+BARANG)?)'
        r'(?:\s+(?:DALAM|PADA|HARI|BARANG))*\s+'
    )
    # Sentence-level negation: if the whole clause contains a block keyword,
    # all day keywords found in that same clause are treated as negated.
    # Split on common clause separators (-->  /  ;  ,  &-free boundary).
    _CLAUSE_NEG = re.compile(
        r'\b(?:JANGAN\s+HANTAR(?:\s+BARANG)?'
        r'|TIDAK\s+(?:TERIMA|HANTAR)'
        r'|TAK\s+TERIMA(?:\s+BARANG)?'
        r'|KEDAI\s+TUTUP|TUTUP|CLOSED|OFF)\b'
    )
    # Split remark into clauses on --> or ; separators
    clauses = re.split(r'-->', txt)
    clause_neg_days: set[int] = set()
    for clause in clauses:
        if _CLAUSE_NEG.search(clause):
            for kw, wd in _REMARKS_KEYWORD_DAY:
                if re.search(r'\b' + re.escape(kw) + r'\b', clause):
                    clause_neg_days.add(wd)

    neg_days: set[int] = set(clause_neg_days)
    for kw, wd in _REMARKS_KEYWORD_DAY:
        kw_re = r'\b' + re.escape(kw) + r'\b'
        if re.search(kw_re + _NEG_SUFFIX, txt):
            neg_days.add(wd)
        if re.search(_NEG_PREFIX + kw_re, txt):
            neg_days.add(wd)

    # Positive day keywords (not in the negated set)
    days: set[int] = set()
    for kw, wd in _REMARKS_KEYWORD_DAY:
        if wd not in neg_days and re.search(r'\b' + re.escape(kw) + r'\b', txt):
            days.add(wd)

    # If only negation was found (no positive days), invert: allow all weekdays
    # except the blocked ones (Mon-Sat = 0-5, skip Sun=6 as no deliveries).
    if not days and neg_days:
        return {d for d in range(6) if d not in neg_days}
    return days if days else None


def _row_trip_session(remarks: str) -> str | None:
    """Detect an explicit AM/PM trip-timing note in a REMARKS string.

    Returns "MORNING", "AFTERNOON", or None (no trip-timing note — this DO
    isn't restricted to either half of the day, so it's fine on both).
    """
    if not remarks or str(remarks).strip().lower() in ("nan", "none", ""):
        return None
    txt = str(remarks).strip().upper()
    if re.search(r'\bMORNING\s+TRIP\b|\bAM\s+FIRST\s+TRIP\b|\bAM\s+TRIP\b', txt):
        return "MORNING"
    if re.search(r'\bAFTERNOON\s+TRIP\b|\bPM\s+TRIP\b', txt):
        return "AFTERNOON"
    return None


# ── LLM-assisted REMARKS parsing (Claude API) ────────────────────────────────
# Free-text remarks vary a lot ("HANTAR SELASA SAHAJA", "x hantar hari jumaat",
# "deliver only on Mon & Thu", typos, mixed BM/English). The keyword regex
# above catches the common patterns; for anything else we ask Claude Haiku to
# normalize the remark into a weekday set. Results are cached on disk so each
# unique remark is only ever sent to the API once.
REMARKS_LLM_CACHE_PATH = os.path.join(_DATA_DIR, "remarks_day_cache.json")
_remarks_llm_cache: dict[str, list | None] | None = None
_remarks_llm_lock = threading.Lock()

# If the local LLM (LM Studio etc.) times out or errors, don't pay that same
# timeout penalty again on every upload/re-assignment for a while — a stuck
# or model-less local server previously meant every request in a session
# (upload, then "assign off-schedule DOs too") each separately waited out
# the full timeout. Skip local calls for this long after a failure.
_LOCAL_LLM_COOLDOWN_SEC = 300
_local_llm_unavailable_until = 0.0

def _load_remarks_llm_cache() -> dict:
    global _remarks_llm_cache
    if _remarks_llm_cache is None:
        try:
            with open(REMARKS_LLM_CACHE_PATH, encoding="utf-8") as f:
                _remarks_llm_cache = json.load(f)
        except Exception:
            _remarks_llm_cache = {}
    return _remarks_llm_cache

def _save_remarks_llm_cache() -> None:
    try:
        with open(REMARKS_LLM_CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(_remarks_llm_cache, f, ensure_ascii=False, indent=1)
    except Exception:
        pass

_REMARKS_LLM_SYSTEM = (
    "You normalize Malaysian delivery remarks (mixed Malay/English, "
    "often with typos) into delivery-day constraints.\n"
    "Weekday numbers: Monday=0 Tuesday=1 Wednesday=2 Thursday=3 "
    "Friday=4 Saturday=5 Sunday=6.\n"
    "Malay days: ISNIN=0 SELASA=1 RABU=2 KHAMIS=3 JUMAAT=4 SABTU=5 AHAD=6.\n"
    "IMPORTANT rules:\n"
    "- If the remark says a day is OFF/TUTUP/TAK TERIMA/CLOSED (negative), "
    "that day is EXCLUDED. If no positive delivery days remain, return null.\n"
    "- If the remark is about lorry/shop OPERATION HOURS (e.g. 'OPERASI "
    "ISNIN-SABTU 4PM', 'WAKTU OPERASI', 'JUMAAT HANYA 3PM'), return null — "
    "it is NOT a customer delivery-day restriction.\n"
    "- Return null for packaging, payment, location, 'setiap hari'/daily, "
    "or unintelligible remarks.\n"
    "For each numbered remark output the list of weekdays the customer "
    "ACCEPTS delivery, or null.\n"
    "Respond ONLY with valid JSON: "
    "{\"results\":[{\"index\":0,\"days\":[1,4]},{\"index\":1,\"days\":null}]}"
)

def _llm_call_local(todo: list[str]) -> dict | None:
    """Call LM Studio (OpenAI-compatible) local LLM. Returns parsed dict or None.

    Skips the attempt entirely while a prior failure's cooldown is active, so
    a stuck/model-less local server doesn't make every upload and every
    off-schedule re-assignment in the same session separately wait out the
    full timeout."""
    global _local_llm_unavailable_until
    now = time.time()
    if now < _local_llm_unavailable_until:
        return None
    import requests as _req
    base_url = os.environ.get("LOCAL_LLM_URL", "http://localhost:1234/v1")
    model = os.environ.get("LOCAL_LLM_MODEL", "local-model")
    numbered = "\n".join(f"{i}: {r}" for i, r in enumerate(todo))
    try:
        resp = _req.post(
            f"{base_url}/chat/completions",
            json={
                "model": model,
                "messages": [
                    {"role": "system", "content": _REMARKS_LLM_SYSTEM},
                    {"role": "user", "content": numbered},
                ],
                "temperature": 0,
                "max_tokens": 2048,
            },
            timeout=10,
        )
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        # Extract JSON from possible <think>...</think> wrapper (deepseek-r1)
        content = re.sub(r"<think>.*?</think>", "", content, flags=re.DOTALL).strip()
        # Extract first {...} block
        m = re.search(r"\{.*\}", content, re.DOTALL)
        if m:
            return json.loads(m.group())
    except Exception as e:
        print(f"⚠️ Local LLM call failed: {e}")
        _local_llm_unavailable_until = now + _LOCAL_LLM_COOLDOWN_SEC
    return None


def _llm_call_anthropic(todo: list[str]) -> dict | None:
    """Call Anthropic Claude Haiku. Returns parsed dict or None."""
    if not os.environ.get("ANTHROPIC_API_KEY"):
        return None
    try:
        import anthropic
        client = anthropic.Anthropic()
        numbered = "\n".join(f"{i}: {r}" for i, r in enumerate(todo))
        response = client.messages.create(
            model="claude-haiku-4-5",
            max_tokens=2048,
            system=_REMARKS_LLM_SYSTEM,
            messages=[{"role": "user", "content": numbered}],
            output_config={
                "format": {
                    "type": "json_schema",
                    "schema": {
                        "type": "object",
                        "properties": {
                            "results": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "index": {"type": "integer"},
                                        "days": {
                                            "anyOf": [
                                                {"type": "array",
                                                 "items": {"type": "integer",
                                                           "enum": [0, 1, 2, 3, 4, 5, 6]}},
                                                {"type": "null"},
                                            ]
                                        },
                                    },
                                    "required": ["index", "days"],
                                    "additionalProperties": False,
                                },
                            }
                        },
                        "required": ["results"],
                        "additionalProperties": False,
                    },
                }
            },
        )
        text = next(b.text for b in response.content if b.type == "text")
        return json.loads(text)
    except Exception as e:
        print(f"⚠️ Anthropic LLM call failed: {e}")
    return None


def _llm_parse_remarks_batch(remarks_list: list[str]) -> None:
    """Send un-cached remarks to LLM (local first, Anthropic fallback) in one batch.

    Priority: LM Studio local (LOCAL_LLM_URL env) → Anthropic API (ANTHROPIC_API_KEY env).
    Results cached to disk. Silently falls back to regex when both unavailable.
    """
    cache = _load_remarks_llm_cache()
    todo = sorted({
        r.strip() for r in remarks_list
        if r and r.strip() and r.strip().lower() not in ("nan", "none")
        and r.strip() not in cache
    })
    if not todo:
        return

    local_url = os.environ.get("LOCAL_LLM_URL", "http://localhost:1234/v1")
    use_local = bool(local_url)

    parsed = None
    if use_local:
        parsed = _llm_call_local(todo)
    if parsed is None and os.environ.get("ANTHROPIC_API_KEY"):
        parsed = _llm_call_anthropic(todo)
    if parsed is None:
        return

    with _remarks_llm_lock:
        for entry in parsed.get("results", []):
            idx = entry.get("index")
            if isinstance(idx, int) and 0 <= idx < len(todo):
                cache[todo[idx]] = entry.get("days")
        _save_remarks_llm_cache()

def _remarks_days_for(remarks: str) -> set[int] | None:
    """Delivery-day set for a remark: LLM cache first, regex fallback."""
    key = str(remarks).strip() if remarks else ""
    if key and key.lower() not in ("nan", "none"):
        cache = _load_remarks_llm_cache()
        if key in cache:
            val = cache[key]
            return set(val) if val is not None else None
    return _parse_remarks_days(remarks)


# ── REMARKS FIELD sheet (FIELD 3 — lorry tonnage requirement) ─────────────────
# The optional "REMARKS FIELD" sheet in LORRY DAILY PLANNING.xlsx lists the
# canonical remark phrases.  Column 3 (FIELD 3) holds the lorry-size phrases
# (VAN, BELOW 5 TON, …, ANY SIZE).  We load them so the size-cap detection stays
# in sync with the planners' spreadsheet; if the sheet is absent we fall back to
# the defaults in assignment_config.REMARKS_FIELD3_TON_CAP.
_remarks_field3_cache: dict[str, float | None] | None = None

def _ton_cap_from_field3_phrase(phrase: str) -> float | None:
    """Map a FIELD 3 phrase (e.g. 'BELOW 10 TON') to a tonnage cap."""
    p = phrase.strip().upper()
    if p in _REMARKS_FIELD3_TON_CAP:
        return _REMARKS_FIELD3_TON_CAP[p]
    if p == "VAN":
        return 2.0
    m = re.search(r'BELOW\s+(\d+(?:\.\d+)?)\s*TON', p)
    if m:
        return float(m.group(1))
    if "ANY SIZE" in p:
        return None
    return None

def _load_remarks_field3() -> dict[str, float | None]:
    """Return {phrase_upper: ton_cap} from the REMARKS FIELD sheet (FIELD 3),
    merged over the assignment_config defaults.  Cached after first read."""
    global _remarks_field3_cache
    if _remarks_field3_cache is not None:
        return _remarks_field3_cache
    table: dict[str, float | None] = dict(_REMARKS_FIELD3_TON_CAP)
    if os.path.exists(PLANNING_PATH):
        try:
            df = pd.read_excel(PLANNING_PATH, sheet_name="REMARKS FIELD", header=0)
            # FIELD 3 is the third labelled column (the lorry-size column).
            f3_col = None
            for c in df.columns:
                if str(c).strip().upper() in ("FIELD 3", "FIELD3"):
                    f3_col = c
                    break
            if f3_col is None and df.shape[1] >= 3:
                f3_col = df.columns[2]
            if f3_col is not None:
                for v in df[f3_col].dropna().astype(str):
                    phrase = v.strip().upper()
                    if phrase:
                        table[phrase] = _ton_cap_from_field3_phrase(phrase)
        except Exception:
            pass   # sheet missing or malformed → config defaults only
    _remarks_field3_cache = table
    return table

def _remarks_lorry_cap(remarks: str) -> float | None:
    """Return the maximum lorry tonnage allowed for a DO based on its REMARKS,
    or None if the remark imposes no size requirement.

    Detection order:
      1. Canonical FIELD 3 phrases (from the REMARKS FIELD sheet / config).
      2. Free-text size aliases (LORRY KECIL, VAN, LORI BESAR TIDAK BOLEH …).
    The tightest (smallest) matching cap wins.
    """
    if not remarks:
        return None
    txt = str(remarks).strip().upper()
    if not txt or txt in ("NAN", "NONE"):
        return None
    caps: list[float] = []

    def _phrase_in(phrase: str, text: str) -> bool:
        """Whole-word keyword match (case-insensitive, already upper).
        Word-boundary so short keywords like VAN don't match CARAVAN, and
        LORI/LORRY spelling variants are both accepted for the same phrase."""
        if not phrase:
            return False
        # Accept LORI/LORRY interchangeably in either the phrase or the text.
        _p = re.escape(phrase).replace(r"LORRY", r"LOR(?:RY|I)").replace(r"LORI", r"LOR(?:RY|I)")
        return re.search(rf"\b{_p}\b", text) is not None

    # 1. Canonical FIELD 3 phrases
    for phrase, cap in _load_remarks_field3().items():
        if cap is not None and _phrase_in(phrase, txt):
            caps.append(cap)
    # 2. Free-text aliases
    for phrase, cap in _REMARKS_SIZE_ALIASES.items():
        if cap is not None and _phrase_in(phrase, txt):
            caps.append(cap)
    # 3. Free-text "N TON" size caps, tolerant of filler words & ordering, e.g.
    #      "BELOW 5 TON"            "BELOW OR 5 TON LORRY"
    #      "5 TON OR BELOW"         "5 TON LORRY AND BELOW"
    #      "5 TON KE BAWAH"         "MAX 5 TON"
    #    Any of these caps the lorry at N tonnes.
    _below_words = r'(?:BELOW|MAX(?:IMUM)?|UNDER|KE\s*BAWAH|AND\s+BELOW|OR\s+BELOW)'
    _ton_patterns = [
        rf'{_below_words}\s+(?:OR\s+)?(\d+(?:\.\d+)?)\s*TON',   # BELOW [OR] N TON
        rf'(\d+(?:\.\d+)?)\s*TON(?:\s+LORRY)?\s+(?:OR\s+|AND\s+)?{_below_words}',  # N TON [LORRY] [OR/AND] BELOW
        rf'(\d+(?:\.\d+)?)\s*TON\s+KE\s*BAWAH',                  # N TON KE BAWAH
        rf'(?:MAX(?:IMUM)?|MAKS)\s+(\d+(?:\.\d+)?)\s*TON',      # MAX N TON
    ]
    for _pat in _ton_patterns:
        for _m in re.finditer(_pat, txt):
            try:
                caps.append(float(_m.group(1)))
            except (ValueError, IndexError):
                pass
    return min(caps) if caps else None


def _remarks_forbidden_plates(remarks: str, all_plates) -> set:
    """Return eligible plates a DO's REMARKS forbid by NUMBER.

    Handles "lorry 3875 tak boleh masuk" (BQU3875 forbidden), "3875 tidak boleh
    masuk", "no 3875 cannot enter", etc.  A plate FRAGMENT is a token containing
    a digit (so size words like BESAR / KECIL are never treated as a plate — those
    are handled by the size-cap rules).  Any eligible lorry whose plate contains
    the fragment is excluded for this DO.
    """
    if not remarks:
        return set()
    txt = str(remarks).upper()
    if "TAK BOLEH" not in txt and "TIDAK BOLEH" not in txt \
            and "TDK BOLEH" not in txt and "CANNOT" not in txt \
            and "CAN NOT" not in txt:
        return set()
    _forbid_kw = r'(?:TAK|TIDAK|TDK)\s+BOLEH|CAN\s?NOT'
    _frags = set()
    # plate fragment (has a digit) appearing shortly BEFORE the forbid keyword,
    # e.g. "3875 TAK BOLEH", "W3826C TIDAK BOLEH", "LORRY 3875 TAK BOLEH MASUK"
    for _m in re.finditer(rf'([A-Z]*\d[A-Z0-9]*)\s+(?:{_forbid_kw})', txt):
        _frags.add(_m.group(1))
    # ...or shortly AFTER ("TAK BOLEH MASUK LORI 3875")
    for _m in re.finditer(rf'(?:{_forbid_kw})[^0-9A-Z]*(?:MASUK\s+)?(?:LOR(?:RY|I)\s+)?([A-Z]*\d[A-Z0-9]*)', txt):
        _frags.add(_m.group(1))
    if not _frags:
        return set()
    _excl = set()
    for _p in all_plates:
        _pu = str(_p).strip().upper()
        for _f in _frags:
            if len(_f) >= 3 and _f in _pu:
                _excl.add(_pu)
                break
    return _excl


def _load_schedule(user: str) -> dict[int, set[str]]:
    """Return {weekday_int: set_of_route_prefixes} from the SCHD sheet.
    weekday_int follows Python's datetime.weekday(): Mon=0 … Sun=6.
    Returns empty dict if the sheet doesn't exist (no schedule filtering).
    """
    if not os.path.exists(PLANNING_PATH):
        return {}
    try:
        u = user.strip().upper()
        target = f"SCHD({u.lower()})"   # e.g. SCHD(abi)
        # Sheet names may have trailing spaces — find by stripped comparison
        xl = pd.ExcelFile(PLANNING_PATH)
        sheet = next((s for s in xl.sheet_names if s.strip() == target), None)
        if sheet is None:
            return {}
        df = pd.read_excel(PLANNING_PATH, sheet_name=sheet, header=None)
    except Exception:
        return {}

    schedule: dict[int, set[str]] = {}
    # Find the header row (contains day names like MON, TUES, WED …)
    header_row_idx = None
    day_cols: dict[int, int] = {}   # weekday_int → column index
    for ri in range(min(5, len(df))):
        for ci in range(len(df.columns)):
            val = str(df.iloc[ri, ci]).strip().upper().rstrip(".")
            if val in _SCHD_DAY_MAP:
                header_row_idx = ri
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        return {}

    for ci in range(len(df.columns)):
        val = str(df.iloc[header_row_idx, ci]).strip().upper().rstrip(".")
        if val in _SCHD_DAY_MAP:
            day_cols[_SCHD_DAY_MAP[val]] = ci

    # Rows below header contain route strings per day column
    for ri in range(header_row_idx + 1, len(df)):
        for wd, ci in day_cols.items():
            cell = str(df.iloc[ri, ci]).strip()
            if not cell or cell.lower() in ("nan", "none", ""):
                continue
            m = re.match(r'^([A-Za-z]{2,4}\d{1,2}[A-Za-z]?)', cell)
            if m:
                schedule.setdefault(wd, set()).add(m.group(1).upper())

    return schedule


def _scheduled_prefixes_for_upload(user: str, trip_day: str = "today") -> set[str] | None:
    """Return route prefixes scheduled for the target day (today or tomorrow).

    These come from the SCHD(abi/vivian) sheet's day column — only routes
    explicitly scheduled for that weekday are returned.  Routes that belong
    to the user but are on a different day are left for the YES/NO prompt
    (_not_today_count in the upload handler).

    Returns None if no schedule found (skip day-filtering entirely).
    """
    schedule = _load_schedule(user)
    if not schedule:
        return None

    from datetime import timedelta as _timedelta
    target_date = datetime.now().date()
    if trip_day == "tomorrow":
        target_date += _timedelta(days=1)
        while target_date.weekday() == 6:
            target_date += _timedelta(days=1)

    target_wd = target_date.weekday()   # 0=Mon … 6=Sun
    return schedule.get(target_wd, set())


def _extract_route_prefix(route: str) -> str:
    """Extract the leading route code token (e.g. 'KV19A', 'PH09', 'JH09')."""
    m = re.match(r'^([A-Za-z]{2,4}\d{1,2}[A-Za-z]?)', route.strip())
    return m.group(1).upper() if m else ""

# ── Destination state classification ─────────────────────────────────────────
# All constants imported from assignment_config — see ASSIGNMENT_RULES.md §3.

def _states_compatible(s1: str, s2: str) -> bool:
    """Return True if two destination states are allowed to share the same lorry."""
    if not s1 or not s2:
        return True
    if s1.upper() == s2.upper():
        return True
    # Urban states (KL / Selangor variants) share lorries freely
    if s1.upper() in _URBAN_COMPATIBLE_STATES and s2.upper() in _URBAN_COMPATIBLE_STATES:
        return True
    return False

def _atomic_components(item_list, rcode_fn, gps_fn, code_fn):
    """Split-atomicity grouping, in operator priority order. When a load must be
    split across lorries, two DOs must NEVER be separated if they share:
      1. the same GPS longitude/point            → "same longitud, same lorry"
      2. the same route code AND customer CODE    → "same route + same customer,
                                                     same lorry"
    (Same route but DIFFERENT customer — criterion 3 — is only a *preference*:
    those may ride together but may be split when capacity forces it, so they
    are NOT unioned here; callers keep same-route components together on a
    best-effort basis. Criterion 4, full weight utilisation, is the packer's
    job once these atomic units are fixed.)

    Union-find over the two hard keys returns a list of inseparable component
    lists; any split (heavy-group half-split, urban de-concentration) moves
    whole components only, so criteria 1 and 2 are never broken."""
    parent = list(range(len(item_list)))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    by_routecode: dict = {}   # (route code, customer CODE) → same lorry
    by_gps: dict = {}         # exact GPS point → same lorry
    for i, it in enumerate(item_list):
        by_routecode.setdefault((rcode_fn(it), str(code_fn(it)).strip().upper()),
                                []).append(i)
        g = gps_fn(it)
        if g and g[0] is not None and g[1] is not None:
            by_gps.setdefault((round(g[0], 4), round(g[1], 4)), []).append(i)
    for grp in list(by_routecode.values()) + list(by_gps.values()):
        for k in grp[1:]:
            union(grp[0], k)
    comps: dict = {}
    for i, it in enumerate(item_list):
        comps.setdefault(find(i), []).append(it)
    return list(comps.values())


def _route_code_of(it) -> str:
    """Canonical route code ('KV12A', 'PH03', …) — module-level twin of the
    nested _rcode helper, usable before that closure is defined."""
    r = str(it.get("ROUTE", "")).strip().upper()
    m = re.match(r"([A-Z]+\d+[A-Z]?)", r)
    return m.group(1) if m else r[:6]


def _eff_dest_min_ton(route: str, dest_grp: str, total_weight: float) -> float:
    """Minimum lorry tonnage for an outstation group, with the tiny-NS
    relaxation. A very small load (≤ TINY_OUTSTATION_MAX_TON) on a Negeri
    Sembilan / Seremban route (prefix in TINY_OUTSTATION_PREFIXES) waives the
    outstation minimum so it can ride a small lorry. Far outstation (Kuantan,
    Pahang, Johor, Perak, Terengganu) is never relaxed."""
    base = _DEST_MIN_TON.get(dest_grp, 0.0)
    if base <= 0:
        return 0.0
    prefix = _route_code_of({"ROUTE": route})[:2]
    if (prefix in _TINY_OUTSTATION_PREFIXES
            and total_weight <= _TINY_OUTSTATION_MAX_TON):
        return 0.0
    return base


def _classify_dest_group(route: str, state: str = "") -> str:
    """Return destination group for a route + optional explicit state.

    Groups:
      LARGE_LONG  — Pahang, Kuantan, Terengganu, Kelantan, Johor, Perak, …  (≥14T)
      MEDIUM_LONG — NS/Seremban, Rawang/T.Malim via KV01A/KV02A            (≥11T)
      KL          — Kuala Lumpur urban (STATE=KUALA LUMPUR or postcode 50-60k) (<11T)
      SELANGOR    — Selangor (STATE=SELANGOR or postcode 40-48k)            (<11T)
      KL_SELANGOR — fallback when state cannot be determined                 (<11T)
    """
    r = route.strip().upper()
    # KV routes heading outstation (Rawang / T.Malim direction)
    pfx5 = r[:5]
    if pfx5 in _DEST_MEDIUM_LONG_KV_CODES:
        return "MEDIUM_LONG"
    cluster = r[:2]
    if cluster in _DEST_LARGE_LONG_CLUSTERS:
        return "LARGE_LONG"
    if cluster in _DEST_MEDIUM_LONG_CLUSTERS:
        # NS cluster — Seremban/Negeri Sembilan direction
        return "MEDIUM_LONG"
    # KV or other KL/Selangor routes — refine using actual state
    st = state.strip().upper()
    if st == "KUALA LUMPUR":
        return "KL"
    if st == "SELANGOR":
        return "SELANGOR"
    return "KL_SELANGOR"   # unknown — treat as generic urban

# _DEST_SORT_PRI, _LORRY_STRICT_ROUTE, _ROUTE_CORRIDOR_GROUPS imported from assignment_config

# Actual destination STATE → outstation corridor. Used to group routes by their
# REAL geography instead of a possibly-misleading route code. e.g. some
# "NS04-->Port Dickson" DOs are actually delivered to Kuala Lipis, PAHANG — their
# state (PAHANG) puts them with the Pahang run (PH03/PH04), not the Seremban run.
_STATE_TO_CORRIDOR = {
    "PAHANG":           "PH_INT",
    "NEGERI SEMBILAN":  "NS",
    "PERAK":            "PERAK",
}


def _norm_cust(c: str) -> str:
    return re.sub(r'[^A-Z0-9 ]', '', str(c).upper()).strip()

# Exception to Kuantan-independence: these specific customers' Kuantan (PH09)
# DOs may ride with the TR02 / Kemaman run (Kuantan & Kemaman are adjacent on
# the east coast). Treated as the Kemaman corridor, not independent Kuantan.
_KUANTAN_TR02_MIX_CUSTOMERS = {
    _norm_cust("GOLDEN HP AGENCY SDN BHD"),
    _norm_cust("HASIL LAUTAN JUN KEE SDN. BHD."),
}


def _kemaman_ph09(route: str, customer: str) -> bool:
    """True if this is a PH09/Kuantan DO for one of the special customers that
    may mix with TR02/Kemaman."""
    r = str(route).strip().upper()
    return (r.startswith("PH09") or "KUANTAN" in r) \
        and _norm_cust(customer) in _KUANTAN_TR02_MIX_CUSTOMERS


def _is_kuantan(route: str, customer: str = "") -> bool:
    """Kuantan (PH09) is a far east-coast run that must ALWAYS be independent —
    it never shares a lorry with any non-Kuantan route. EXCEPTION: the two
    special customers ride with TR02/Kemaman, so they are NOT flagged Kuantan."""
    r = str(route).strip().upper()
    if not (r.startswith("PH09") or "KUANTAN" in r):
        return False
    return not _kemaman_ph09(route, customer)


def _state_corridor(state: str, route: str = "") -> str:
    # KV (Klang Valley) routes are handled by city/urban bucketing — never let a
    # border state (e.g. KV01A touching Perak at Tanjung Malim) reclassify them
    # into a deep-outstation corridor. Only genuine outstation route codes use
    # the state override.
    if str(route).strip().upper().startswith("KV"):
        return ""
    return _STATE_TO_CORRIDOR.get(str(state).strip().upper(), "")


def _prefix_corridor(route: str) -> str:
    r = str(route).strip().upper()
    for _name, _pfxs in _ROUTE_CORRIDOR_GROUPS.items():
        if any(r.startswith(p) for p in _pfxs):
            return _name
    return ""


def _same_corridor_group(route1: str, route2: str,
                         state1: str = "", state2: str = "",
                         cust1: str = "", cust2: str = "") -> bool:
    """Return True when both routes belong to the same delivery corridor group.

    When BOTH actual states are known outstation states, group purely by state
    (so a mislabeled route rides the truck for its real geography). Otherwise
    fall back to route-code corridor groups.
    """
    # Kuantan only ever groups with Kuantan — never with any other route
    # (except the two special customers, handled as the KEMAMAN corridor below).
    k1, k2 = _is_kuantan(route1, cust1), _is_kuantan(route2, cust2)
    if k1 or k2:
        return k1 and k2
    # Special-customer PH09 rides with TR02 (both map to the KEMAMAN direction).
    d1 = _direction_key(route1, state1, cust1)
    d2 = _direction_key(route2, state2, cust2)
    if d1 == "KEMAMAN" and d2 == "KEMAMAN":
        return True
    c1 = _state_corridor(state1, route1)
    c2 = _state_corridor(state2, route2)
    if c1 and c2:
        return c1 == c2
    r1 = route1.strip().upper()
    r2 = route2.strip().upper()
    for pfxs in _ROUTE_CORRIDOR_GROUPS.values():
        if (any(r1.startswith(p) for p in pfxs)
                and any(r2.startswith(p) for p in pfxs)):
            return True
    return False

def _direction_key(route: str, state: str = "", customer: str = "") -> str:
    """Coarse outstation 'direction' for a route: the corridor-group name if the
    route belongs to one (PH_INT, KV_NORTH, NS, …), else the route code. Used to
    reserve one lorry per distinct outstation direction so the biggest direction
    (e.g. Pahang/Kuantan) does not consume every lorry and starve a smaller one
    (e.g. the KV01A/KV02A northern run).

    If the DO's actual state maps to a known outstation corridor, that wins over
    the route-code prefix — so NS04-in-Pahang is directed with the Pahang run."""
    r = str(route).strip().upper()
    # Special-customer PH09 and any TR02 share the KEMAMAN direction (they may
    # ride one east-coast lorry together).
    if r.startswith("TR02") or _kemaman_ph09(route, customer):
        return "KEMAMAN"
    if _is_kuantan(route, customer):
        return "KUANTAN"                 # its own reserved direction, always
    sc = _state_corridor(state, route)
    if sc:
        return sc
    r = str(route).strip().upper()
    cg = _prefix_corridor(r)
    if cg:
        return cg
    _m = re.match(r"([A-Z]+\d+[A-Z]?)", r)
    return _m.group(1) if _m else r[:6]

# _ROUTE_PREFERRED_LORRY, _TINY_ITEM_AVG_WEIGHT_T, _CROSS_BEARING_LIMIT,
# _MAX_CITY_MERGE_KM_OUTSTATION imported from assignment_config


def _preferred_lorries_for_route(route_text: str, engine=None) -> list[str]:
    """Return the ordered list of preferred plates for this route, or [].

    Source priority:
      1. The FIT IN LORRY sheet (data-driven, per-owner — RULE 9A) via
         engine.fit_in_lorry_preferred(), when `engine` is passed and the route
         is listed there for the engine's owner.
      2. ROUTE_PREFERRED_LORRY in assignment_config.py (kept empty by design).
    Both match the LONGEST route-code prefix. HINT only — callers fall back to
    the full eligible fleet when none of these plates are available/fit.
    """
    r = route_text.strip().upper()
    if engine is not None:
        _fil = engine.fit_in_lorry_preferred(r)
        if _fil:
            return _fil
    best_pfx = ""
    best_plates: list[str] = []
    for pfx, plates in _ROUTE_PREFERRED_LORRY.items():
        if r.startswith(pfx) and len(pfx) > len(best_pfx):
            best_pfx = pfx
            best_plates = plates
    return best_plates

def _strict_route_excl(route_text: str) -> set:
    """Return plates that must NOT serve this route due to strict reservations.
    route_text: space-joined ROUTE strings for the group.
    """
    r = route_text.strip().upper()
    excl: set[str] = set()
    for plate, allowed_pfxs in _LORRY_STRICT_ROUTE.items():
        route_allowed = any(
            r.startswith(pfx) or re.search(r'\b' + re.escape(pfx), r)
            for pfx in allowed_pfxs
        )
        if not route_allowed:
            excl.add(plate)
    return excl

def _check_manual_placement(item: dict, plate: str, engine, sess: dict) -> list[str]:
    """Spot-check a manual drag-and-drop placement (board UI) against the hard
    rules the auto-assignment engine enforces. Never blocks the move — the
    caller always applies it — this only returns human-readable warnings for
    the UI to flag on the card, so a human can still make a judgment call the
    engine can't (e.g. a phone call from the customer) while seeing what
    they're overriding.

    Not an exhaustive re-run of every rule in the full assignment algorithm —
    covers the checks most likely to matter for a single manual move:
    eligibility, strict-route reservations, size/forbidden-plate caps,
    capacity, outstation minimum tonnage, and route compatibility with
    whatever else is already on that lorry today.
    """
    warnings: list[str] = []
    from lorry_engine import _routes_on_same_way

    route = str(item.get("ROUTE", ""))
    weight = float(item.get("WEIGHT", 0) or 0)

    cap = None
    in_eligible = False
    if engine is not None and getattr(engine, "eligible_lorries", None) is not None:
        _row = engine.eligible_lorries[engine.eligible_lorries["LORRY"] == plate]
        if not _row.empty:
            cap = float(_row.iloc[0]["TON"])
            in_eligible = True
    if cap is None and engine is not None and getattr(engine, "all_lorries", None) is not None:
        _row = engine.all_lorries[engine.all_lorries["LORRY"] == plate]
        if not _row.empty:
            cap = float(_row.iloc[0]["TON"])
    if cap is None:
        # Caller (board_move) is expected to have already hard-rejected a
        # plate unknown to the whole fleet before calling this — reaching
        # here with cap still None means the fleet lookup itself is broken,
        # not a real placement choice, so bail rather than compute nonsense.
        return warnings
    if not in_eligible:
        warnings.append(f"{plate} isn't Available in today's master lorry list (cross-owner or blocked override).")

    if plate in _strict_route_excl(route):
        warnings.append(f"{plate} is reserved for other routes only — not this one.")

    max_ton = item.get("MAX_TON")
    if max_ton is not None and cap > float(max_ton) + 1e-9:
        warnings.append(f"Exceeds this DO's size cap (max {max_ton}T, {plate} is {cap}T).")

    forbid = item.get("FORBID_PLATES")
    if forbid and plate in forbid:
        warnings.append(f"{plate} is explicitly forbidden for this DO.")

    others = [it for it in (sess.get("items") or [])
              if it is not item and it.get("LORRY") == plate]
    other_load = sum(float(o.get("WEIGHT", 0) or 0) for o in others)
    if other_load + weight > cap * NAIK_FACTOR + 1e-6:
        warnings.append(
            f"Over capacity: {round(other_load + weight, 3)}T on a {cap}T lorry.")

    incompatible = [o.get("ROUTE", "") for o in others
                    if not (_same_corridor_group(route, o.get("ROUTE", ""))
                            or _routes_on_same_way(route, o.get("ROUTE", "")))]
    if incompatible:
        warnings.append(
            f"Route doesn't match what's already on {plate} ({incompatible[0]}).")

    dest_grp = _classify_dest_group(route, item.get("STATE", ""))
    min_t = _eff_dest_min_ton(route, dest_grp, weight)
    if min_t > 0 and cap < min_t:
        warnings.append(f"Too small for this outstation route (needs ≥{min_t}T).")

    return warnings


def _resolve_history_path() -> str:
    """Return the best available history file.
    Single source of truth is the .xlsx; bot-exported and legacy files are
    fallbacks only.
    """
    for p in [HISTORY_PATH, HISTORY_PATH_ALT, HISTORY_PATH_OLD]:
        if os.path.exists(p):
            return p
    return HISTORY_PATH_OLD  # fallback even if missing — engine will warn
DAILY_LOG_PATH = os.path.join(_DATA_DIR, "daily_assignments.json")

# _POSTCODE_STATE_RANGES imported from assignment_config

def _postcode_to_state(postcode) -> str:
    """Return Malaysian state name from postcode, or '' if unknown.
    Priority: exact postcode from the "Malaysia States & Cities" sheet
              (_POSTCODE_TO_STATE) → hardcoded POSTCODE_STATE_RANGES fallback."""
    try:
        pc = int(str(postcode).strip().split()[0])
    except (ValueError, TypeError):
        return ""
    # 1. Exact postcode from the planning sheet (operator-maintained)
    _st = _POSTCODE_TO_STATE.get(pc)
    if _st:
        return _st
    # 2. Hardcoded range table fallback
    for lo, hi, state in _POSTCODE_STATE_RANGES:
        if lo <= pc <= hi:
            return state
    return ""

def _state_from_row(row) -> str:
    """Derive destination state from a DataFrame row.
    Priority: STATE column > CITY lookup (Malaysia States & Cities sheet)
             > POSCODE range table.
    """
    # 1. Explicit STATE column
    raw_state = str(row.get("STATE", "")).strip().upper()
    if raw_state and raw_state not in ("NAN", "NONE", "", "-"):
        return raw_state
    # 2. City lookup against Malaysia States & Cities reference
    raw_city = str(row.get("CITY", "")).strip().upper()
    if raw_city and raw_city not in ("NAN", "NONE", "", "-"):
        st = _CITY_TO_STATE.get(raw_city, "")
        if st:
            return st
    # 3. POSCODE column fallback
    pc_val = row.get("POSCODE", "") or row.get("POSTCODE", "")
    if pc_val:
        st = _postcode_to_state(pc_val)
        if st:
            return st.upper()
    return ""

# ── Shared UI constants ───────────────────────────────────────────────────────
_HI_BTN = {"_type": "buttons", "body": "Tap below to start a new session.",
            "buttons": [{"id": "hi", "title": "👋 Hi"}]}

# ── Daily assignment log (persists across conversations) ─────────────────────

def _today() -> str:
    return date.today().isoformat()   # e.g. "2026-05-11"

def _load_daily_log() -> dict:
    """
    Returns { "date": "YYYY-MM-DD", "assigned": ["PLATE1", ...] }
    Resets automatically if the stored date is not today.
    """
    if os.path.exists(DAILY_LOG_PATH):
        try:
            with open(DAILY_LOG_PATH, "r") as f:
                data = json.load(f)
            if data.get("date") == _today():
                return data
        except Exception:
            pass
    return {"date": _today(), "assigned": []}

def _save_daily_log(log: dict):
    # Always sanitise before saving — remove empty/blank plate strings
    if "assigned" in log:
        log["assigned"] = sorted({p for p in log["assigned"] if p and p.strip()})
    if "broken" in log:
        log["broken"] = {k: v for k, v in log["broken"].items() if k and k.strip()}
    with open(DAILY_LOG_PATH, "w") as f:
        json.dump(log, f, indent=2)

def get_assigned_today() -> set:
    """Return set of ALL lorry plates assigned today (never includes empty strings)."""
    return {p for p in _load_daily_log()["assigned"] if p and p.strip()}

def record_assignments_today(plates: list[str], user: str | None = None):
    """Add newly confirmed plates to today's log, recording WHO assigned each
    (so a shared SPARE lorry can be attributed to the user who took it)."""
    log = _load_daily_log()
    existing = set(log["assigned"])
    assigned_by = dict(log.get("assigned_by", {}))
    for p in plates:
        if p and p != "SKIPPED":
            pu = p.upper()
            existing.add(pu)
            if user:
                assigned_by[pu] = user.upper()
    log["assigned"] = sorted(existing)
    log["assigned_by"] = assigned_by
    _save_daily_log(log)


def get_assigned_by() -> dict:
    """Return { PLATE: USER } — who assigned each plate today (best-effort;
    plates recorded before attribution existed have no entry)."""
    return {str(k).upper(): str(v).upper()
            for k, v in _load_daily_log().get("assigned_by", {}).items()}


def release_specific_plates(plates: list[str]) -> bool:
    """Helper to remove plates from today's assignment and broken log."""
    log = _load_daily_log()
    assigned = set(log.get("assigned", []))
    broken = log.get("broken", {})
    changed = False
    for p in plates:
        p_up = p.upper()
        if p_up in assigned:
            assigned.discard(p_up)
            broken.pop(p_up, None)
            changed = True
    if changed:
        log["assigned"] = sorted(assigned)
        log["broken"] = broken
        _save_daily_log(log)
    return changed




def get_broken_lorries() -> dict:
    """Return dict of { broken_plate: replacement_plate } for today."""
    return _load_daily_log().get("broken", {})

def record_broken_lorry(broken: str, replacement: str):
    """Mark a lorry as broken and record its replacement for today."""
    log = _load_daily_log()
    if "broken" not in log:
        log["broken"] = {}
    log["broken"][broken.upper()] = replacement.upper()
    # Also block broken lorry from being assigned
    existing = set(log["assigned"])
    existing.add(broken.upper())
    log["assigned"] = sorted(existing)
    _save_daily_log(log)

def remove_broken_lorry(broken: str):
    """Mark a previously broken lorry as fixed — removes it from broken list."""
    log = _load_daily_log()
    broken_map = log.get("broken", {})
    plate = broken.upper()
    if plate in broken_map:
        del broken_map[plate]
        log["broken"] = broken_map
        # Also unblock from assigned list
        existing = set(log["assigned"])
        existing.discard(plate)
        log["assigned"] = sorted(existing)
        _save_daily_log(log)
        return True
    return False

def clear_daily_log_for_user(engine) -> list[str]:
    """
    Remove only the plates the user OWNS EXCLUSIVELY (USER != SPARE) from
    today's log, so re-uploading recomputes their own lorries. SPARE lorries
    are shared: one committed by another user's earlier run (e.g. ABI) must
    stay blocked when this user (e.g. VIVIAN) uploads — never freed here.
    Returns the list of plates actually removed.
    """
    el = engine.eligible_lorries
    if "USER" in el.columns:
        own_lorries = set(el[el["USER"].astype(str).str.upper() != "SPARE"]["LORRY"].str.upper())
    else:
        own_lorries = set(el["LORRY"].str.upper())
    log = _load_daily_log()
    all_plates  = set(log["assigned"])
    my_plates   = all_plates & own_lorries         # this user's OWN (non-SPARE) plates
    remaining   = sorted(all_plates - my_plates)   # keep other users' + all SPARE plates
    log["assigned"] = remaining
    log["assigned_by"] = {p: u for p, u in log.get("assigned_by", {}).items()
                          if p.upper() in set(remaining)}
    _save_daily_log(log)
    return sorted(my_plates)


LORRY_TOGGLE_PATH = os.path.join(_DATA_DIR, "lorry_toggle.json")


def _muatan_spare_plates() -> set:
    """Plates tagged SPARE in LORRY DAILY PLANNING.xlsx's MUATAN sheet — the
    ones shared between ABI and VIVIAN, subject to the board's on/off switch
    enforcing that only one of them has a given SPARE plate available at a
    time. Reads the sheet directly rather than a session's engine (unlike
    the older _spare_plates(engine) below) since the toggle state must be
    resolvable before any particular session's eligible_lorries exists."""
    try:
        df = pd.read_excel(PLANNING_PATH, sheet_name="MUATAN", header=None)
        if len(df) == 0 or str(df.iloc[0, 0]).strip().upper() != "NAME":
            return set()
        out = set()
        for _, r in df.iloc[1:].iterrows():
            user_val  = str(r.iloc[0]).strip().upper() if pd.notna(r.iloc[0]) else ""
            plate_val = str(r.iloc[1]).strip().upper() if pd.notna(r.iloc[1]) else ""
            if user_val == "SPARE" and plate_val:
                out.add(plate_val)
        return out
    except Exception:
        return set()


def _load_lorry_toggle() -> dict:
    """Returns {"date": "YYYY-MM-DD", "off_for": {"PLATE": ["ABI", ...]}} —
    which planners have switched each plate OFF (not available) today.
    Resets automatically at the start of a new day, same as the daily log."""
    if os.path.exists(LORRY_TOGGLE_PATH):
        try:
            with open(LORRY_TOGGLE_PATH, "r") as f:
                data = json.load(f)
            if data.get("date") == _today():
                return data
        except Exception:
            pass
    return {"date": _today(), "off_for": {}}


def _save_lorry_toggle(state: dict):
    state["off_for"] = {p: sorted(set(us)) for p, us in state.get("off_for", {}).items() if us}
    with open(LORRY_TOGGLE_PATH, "w") as f:
        json.dump(state, f, indent=2)


def get_unavailable_plates_for(user: str) -> set:
    """Plates this planner has switched OFF today — excluded from their
    eligible fleet when it's (re)built at login."""
    user = user.strip().upper()
    state = _load_lorry_toggle()
    return {p for p, us in state.get("off_for", {}).items() if user in us}


def set_plate_toggle(plate: str, user: str, on: bool) -> None:
    """Turn a plate ON/OFF for this planner today. Turning a SPARE plate ON
    for one planner forces it OFF for the other(s) that share it — only one
    of ABI/VIVIAN can actually take a shared lorry out on a given day."""
    plate = plate.strip().upper()
    user = user.strip().upper()
    state = _load_lorry_toggle()
    off_for = state.setdefault("off_for", {})
    plate_off = set(off_for.get(plate, []))
    if on:
        plate_off.discard(user)
        if plate in _muatan_spare_plates():
            for other in _MASTER_FILE_USERS:
                if other != user:
                    plate_off.add(other)
    else:
        plate_off.add(user)
    off_for[plate] = sorted(plate_off)
    _save_lorry_toggle(state)


def toggle_lorry_availability(sess, plate: str, on: bool) -> dict:
    """Board on/off switch: mark a plate available/unavailable for today for
    the logged-in planner, update this session's live eligible fleet.

    Switching OFF sends any DOs currently on that plate back to the
    unassigned pool — an unavailable lorry can't keep carrying a load — and
    remembers which DOs those were. Switching back ON tries to put those
    SAME DOs back on the plate (not a fresh assignment run, and not forced —
    only DOs that still fit cleanly, per the same rule checks a manual drag
    is spot-checked against, go back; anything that no longer fits stays in
    the pool for a human to place)."""
    user = sess.get("user_id")
    engine = sess.get("engine")
    if not user or engine is None:
        return {"error": "not_logged_in", "message": "Please pick a planner first."}
    plate = plate.strip().upper()
    set_plate_toggle(plate, user, on)

    el = engine.eligible_lorries
    unassigned_count = 0
    refilled_count = 0
    if on:
        if "LORRY" in el.columns and plate not in set(el["LORRY"].str.upper()):
            # Only re-add a plate that actually belongs to this planner's own
            # fleet (own lorries + shared SPARE) — captured in _full_fleet
            # before any toggle-filtering. A plate outside that fleet (e.g.
            # BIG/SELAYANG, or the other planner's own lorry) is never a valid
            # drop target for this user, toggle or not.
            _fleet_tons = {str(p).strip().upper(): t for p, t in sess.get("_full_fleet") or []}
            _ton = _fleet_tons.get(plate)
            if _ton is not None:
                engine.eligible_lorries = pd.concat([
                    el, pd.DataFrame([{"LORRY": plate, "TON": _ton, "USER": user, "Status": "Available"}])
                ], ignore_index=True)

        # Try to restore whatever this plate was carrying when it got turned
        # off — only DOs that are still sitting unassigned AND still fit
        # capacity-wise. These DOs were already validated as compatible
        # together (route, size cap, etc.) by the assignment engine that put
        # them on this exact plate in the first place, so the only thing
        # that could genuinely have changed since eviction is how much room
        # is left — that's the "not force" check we re-run here, not the
        # full manual-drag rule set (which flags same-route DOs against each
        # other on plate-code technicalities that don't apply when restoring
        # a group the engine already grouped).
        _evicted = sess.get("_toggle_evicted", {}).pop(plate, [])
        if _evicted:
            _by_do = {str(it.get("DO NUMBER")): it for it in sess.get("items", []) or []}
            _cap = None
            _row = engine.eligible_lorries[engine.eligible_lorries["LORRY"] == plate]
            if not _row.empty:
                _cap = float(_row.iloc[0]["TON"])
            if _cap is not None:
                _load = sum(float(it.get("WEIGHT", 0) or 0) for it in sess.get("items", []) or []
                            if str(it.get("LORRY", "")).strip().upper() == plate)
                for _do_num in _evicted:
                    it = _by_do.get(_do_num)
                    if it is None or it.get("LORRY") not in (None, "NO_LORRY"):
                        continue   # already re-placed elsewhere — leave it
                    _w = float(it.get("WEIGHT", 0) or 0)
                    if _load + _w > _cap * NAIK_FACTOR + 1e-6:
                        continue  # doesn't cleanly fit any more — don't force it
                    it["LORRY"] = plate
                    sess.setdefault("assigned", {})[it["DO NUMBER"]] = plate
                    sess.setdefault("unassigned_reasons", {}).pop(it["DO NUMBER"], None)
                    _load += _w
                    refilled_count += 1
            if refilled_count:
                for do in sess.get("pending_dos", []):
                    do["TOTAL_TON"] = round(sum(x["WEIGHT"] for x in do["ITEMS"]), 3)
                sess.pop("export_bytes", None)
    else:
        if "LORRY" in el.columns:
            engine.eligible_lorries = el[el["LORRY"].str.upper() != plate].reset_index(drop=True)
        _evicted_dos = []
        for it in sess.get("items", []) or []:
            if str(it.get("LORRY", "")).strip().upper() == plate:
                it["LORRY"] = None
                sess.setdefault("assigned", {}).pop(it["DO NUMBER"], None)
                _evicted_dos.append(str(it.get("DO NUMBER")))
                unassigned_count += 1
        if unassigned_count:
            sess.setdefault("_toggle_evicted", {})[plate] = _evicted_dos
            for do in sess.get("pending_dos", []):
                do["TOTAL_TON"] = round(sum(x["WEIGHT"] for x in do["ITEMS"]), 3)
            sess.pop("export_bytes", None)
    return {"ok": True, "plate": plate, "on": on,
            "unassigned_count": unassigned_count, "refilled_count": refilled_count}


def refresh_eligible_from_toggle(sess) -> None:
    """Rebuild engine.eligible_lorries from this session's full fleet, applying
    today's LIVE on/off toggle state fresh.

    A toggle mutates eligible_lorries directly for whichever session actually
    clicked it, but a shared SPARE plate's other planner may already have a
    session open with its own (now-stale) eligible_lorries snapshot — e.g.
    ABI turns WA6899M on right as VIVIAN is mid-setup; VIVIAN's session never
    saw that write. Calling this right before an assignment run makes the
    eligible fleet always reflect the current toggle state (including
    another planner's claim on a shared SPARE) regardless of when either
    session last touched it, so two planners can't both get handed the same
    SPARE lorry by assigning at the same time."""
    engine = sess.get("engine")
    user = sess.get("user_id")
    full_fleet = sess.get("_full_fleet")
    if engine is None or not user or full_fleet is None:
        return
    _off = get_unavailable_plates_for(user)
    _assignable = [(p, t) for p, t in full_fleet if str(p).strip().upper() not in _off]
    engine.eligible_lorries = pd.DataFrame(
        [{"LORRY": p, "TON": t, "USER": user, "Status": "Available"}
         for p, t in sorted(_assignable)]
    )


def _spare_plates(engine) -> set:
    """Plates in this user's fleet that are SPARE (shared)."""
    el = engine.eligible_lorries
    if "USER" not in el.columns:
        return set()
    return set(el[el["USER"].astype(str).str.upper() == "SPARE"]["LORRY"].str.upper())


def categorize_clear_plates(engine, user: str):
    """Split today's log plates (that fall in this user's fleet) into:
      own       — plates the user owns exclusively (clear freely)
      my_spare  — SPARE plates THIS user assigned (clear freely)
      others_spare — { plate: owner } SPARE plates assigned by ANOTHER user
                     (or unknown) → clearing needs confirmation.
    """
    user_up = user.upper()
    assigned = get_assigned_today()
    by = get_assigned_by()
    spare = _spare_plates(engine)
    fleet = set(engine.eligible_lorries["LORRY"].str.upper())
    own, my_spare, others_spare = [], [], {}
    for p in sorted(assigned & fleet):
        if p in spare:
            owner = by.get(p)
            if owner is None or owner == user_up:
                my_spare.append(p)      # this user's own (or legacy/unknown-as-mine)
            else:
                others_spare[p] = owner
        else:
            own.append(p)               # owner-exclusive lorry
    return own, my_spare, others_spare


def clear_specific_plates_from_log(plates) -> None:
    """Remove the given plates from today's log (assigned + assigned_by)."""
    up = {str(p).upper() for p in plates}
    log = _load_daily_log()
    log["assigned"] = sorted(set(log.get("assigned", [])) - up)
    log["assigned_by"] = {p: u for p, u in log.get("assigned_by", {}).items()
                          if p.upper() not in up}
    _save_daily_log(log)


def clear_daily_log():
    """Wipe entire log (legacy/midnight reset)."""
    _save_daily_log({"date": _today(), "assigned": []})


# ── Midnight auto-reset thread ────────────────────────────────────────────────

def _seconds_until_midnight() -> float:
    now = datetime.now()
    midnight = datetime.combine(now.date(), dtime(0, 0, 0))
    from datetime import timedelta
    next_midnight = midnight + timedelta(days=1)
    return (next_midnight - now).total_seconds()

def _midnight_reset_loop():
    """Background thread: waits until 00:00, clears the daily log, repeats."""
    while True:
        wait = _seconds_until_midnight()
        threading.Event().wait(wait)          # sleep until midnight
        clear_daily_log()
        # Also clear all in-memory sessions so lorries are fresh for the new day
        sessions.clear()

# Start the background thread once when bot.py is imported
_reset_thread = threading.Thread(target=_midnight_reset_loop, daemon=True)
_reset_thread.start()


# ── Conversation session store ────────────────────────────────────────────────
sessions: dict[str, dict] = {}

def get_session(phone: str) -> dict:
    if phone not in sessions:
        sessions[phone] = {
            "state": "IDLE",
            "user_id": None,
            "engine": None,
            "pending_dos": [],
            "current_do_index": 0,
            "suggestions": [],
            "unavailable": set(),   # marked unavailable this session
            "assigned": {},         # DO_NUMBER -> LORRY
        }
    return sessions[phone]

def reset_session(phone: str):
    sessions.pop(phone, None)


# ── Message handlers ──────────────────────────────────────────────────────────

def handle_message(phone: str, text: str = None,
                   file_bytes: bytes = None, file_mime: str = None) -> list[str]:
    sess = get_session(phone)
    state = sess["state"]
    text = (text or "").strip()
    cmd_lower = text.lower().strip()

    # ── Global commands ───────────────────────────────────────────────────────
    if text.lower() in ("reset", "restart", "start over"):
        reset_session(phone)
        return ["Session reset. Send *hi* to start again."]

    if text.lower() == "clear daily log":
        sess = get_session(phone)
        # Guard: if we are mid-flow waiting for user input (e.g. broken replacement),
        # ignore the button tap and remind the user what we are waiting for.
        if sess.get("state") == "AWAIT_BROKEN_REPLACEMENT":
            broken = sess.get("pending_broken_plate", "?")
            return [f"⚠️ Still waiting for a replacement lorry for *{broken}*.\n"
                    "Reply with the replacement plate or type *none* to skip."]
        engine = sess.get("engine")
        user   = sess.get("user_id")
        if not engine or not user:
            return ["❌ Please log in first (send *hi*) before clearing the log."]
        own, my_spare, others_spare = categorize_clear_plates(engine, user)
        mine = sorted(own + my_spare)
        sess["_clear_others_spare"] = others_spare   # remember for the follow-up
        # Nothing of the user's own to clear → skip the pointless "release 0"
        # step and go straight to the shared-lorry question (if any).
        if not mine:
            if others_spare:
                _shared = "\n".join(f"  • *{p}* — assigned to *{o}*"
                                    for p, o in sorted(others_spare.items()))
                return [{
                    "_type": "buttons",
                    "body": (
                        "\U0001f536 You have no assignments of your own to clear.\n\n"
                        "These SPARE lorry(s) are assigned to another user:\n"
                        f"{_shared}\n\n"
                        "Clear them so *you* can use them? This takes a shared lorry "
                        "already assigned to someone else — make sure it won't be "
                        "double-booked.\n\n"
                        "• *Yes* → you may use these lorries.\n"
                        "• *No* → you'll use only your other available lorries."
                    ),
                    "buttons": [
                        {"id": "confirm clear shared", "title": "✅ Yes, clear shared"},
                        {"id": "keep shared",          "title": "❌ No, keep them"},
                    ],
                }]
            return [
                "✅ Nothing to clear — you have no lorry assignments today.",
                {"_type": "buttons", "body": "Type *hi* to start.",
                 "buttons": [{"id": "hi", "title": "👋 Hi"}]},
            ]
        _body = (
            f"⚠️ *Confirm Clear Your Log ({user})?*\n\n"
            f"This will release *{len(mine)}* of your own assignment(s) today.\n"
            f"Plates: {', '.join(mine)}\n"
        )
        if others_spare:
            _shared = ", ".join(f"{p} ({o})" for p, o in sorted(others_spare.items()))
            _body += (f"\n🔶 *{len(others_spare)}* SPARE lorry(s) are assigned to "
                      f"other user(s): {_shared}.\nI'll ask about those next.\n")
        _body += "\nThis cannot be undone. Are you sure?"
        return [
            {
                "_type": "buttons",
                "body": _body,
                "buttons": [
                    {"id": "confirm clear daily log", "title": "✅ Yes, Clear"},
                    {"id": "cancel clear",            "title": "❌ Cancel"},
                ],
            }
        ]

    if text.lower() == "confirm clear daily log":
        sess = get_session(phone)
        engine = sess.get("engine")
        user   = sess.get("user_id")
        if not engine or not user:
            return ["❌ Please log in first (send *hi*) before clearing the log."]
        # Clear the user's OWN plates + their OWN SPARE assignments right away.
        own, my_spare, others_spare = categorize_clear_plates(engine, user)
        clear_specific_plates_from_log(own + my_spare)
        sess["unavailable"] = set(get_assigned_today()) | set(get_broken_lorries())
        removed = sorted(own + my_spare)
        _msgs = [
            f"\U0001f5d1\ufe0f *{user}*'s log cleared.\n"
            f"\U0001f4cb Plates released: {', '.join(removed) or 'none'}\n"
            "Your lorries are now available again."
        ]
        others_spare = sess.get("_clear_others_spare") or others_spare
        if others_spare:
            _shared = "\n".join(f"  • *{p}* — assigned to *{o}*"
                                for p, o in sorted(others_spare.items()))
            _msgs.append({
                "_type": "buttons",
                "body": (
                    "\U0001f536 These SPARE lorry(s) are already assigned to another user:\n"
                    f"{_shared}\n\n"
                    "Clear them too so *you* can use them? This lets you take a "
                    "shared lorry already assigned to someone else — make sure "
                    "it won't be double-booked.\n\n"
                    "• *Yes* → you may use these lorries.\n"
                    "• *No* → you'll use only your other available lorries."
                ),
                "buttons": [
                    {"id": "confirm clear shared", "title": "✅ Yes, clear shared"},
                    {"id": "keep shared",          "title": "❌ No, keep them"},
                ],
            })
        else:
            _msgs.append({
                "_type": "buttons",
                "body": "Tap below to start a new session, or type *hi* anytime.",
                "buttons": [{"id": "hi", "title": "👋 Hi"}],
            })
        return _msgs

    if text.lower() == "confirm clear shared":
        sess = get_session(phone)
        user = sess.get("user_id") or "?"
        others_spare = sess.get("_clear_others_spare") or {}
        if not others_spare:
            return ["Nothing shared to clear.",
                    {"_type": "buttons", "body": "Type *hi* to start.",
                     "buttons": [{"id": "hi", "title": "👋 Hi"}]}]
        clear_specific_plates_from_log(list(others_spare))
        sess["unavailable"] = set(get_assigned_today()) | set(get_broken_lorries())
        sess["_clear_others_spare"] = {}
        _plates = ", ".join(sorted(others_spare))
        return [
            f"✅ Shared lorry(s) released: {_plates}.\n"
            f"*{user}* may now use them. ⚠️ Make sure they are not "
            "double-booked with the other user's trip.",
            {"_type": "buttons", "body": "Type *hi* to start a new session.",
             "buttons": [{"id": "hi", "title": "👋 Hi"}]},
        ]

    if text.lower() == "keep shared":
        sess = get_session(phone)
        others_spare = sess.get("_clear_others_spare") or {}
        sess["_clear_others_spare"] = {}
        _plates = ", ".join(sorted(others_spare)) or "none"
        return [
            f"\U0001f44d Kept as-is. These shared lorry(s) stay with the other "
            f"user: {_plates}.\nYou'll be assigned only your other available lorries.",
            {"_type": "buttons", "body": "Type *hi* to start a new session.",
             "buttons": [{"id": "hi", "title": "👋 Hi"}]},
        ]

    if text.lower() == "cancel clear":
        return [
            "❌ Clear cancelled. Daily log unchanged.",
            {"_type": "buttons", "body": "What would you like to do next?",
             "buttons": [{"id": "hi", "title": "👋 Hi"}]},
        ]

    # ── Step 1: User tapped a DO# → show available lorry list ──────────────
    # ── Step 1: User tapped a DO# → show lorry picker (button pages) ────────
    if text.lower().startswith("select_do "):
        parts  = text.strip().split(" ")
        do_num = parts[1].strip().upper() if len(parts) > 1 else ""
        page   = int(parts[2]) if len(parts) > 2 and parts[2].isdigit() else 0
        return _lorry_picker_buttons(sess, do_num, page)

    if text.lower().startswith("select_lorry "):
        parts  = text.strip().split(" ", 2)
        if len(parts) < 3:
            return ["❌ Invalid lorry selection."]
        do_num = parts[1].strip().upper()
        plate  = parts[2].strip().upper()

        target_do = next((d for d in sess.get("pending_dos", [])
                          if d["DO NUMBER"] == do_num), None)
        if not target_do:
            return [f"❌ DO# *{do_num}* not found."]

        if plate == "__AUTO__":
            # Auto-pick: exclude current lorry so user always gets something different
            engine: LorryEngine = sess.get("engine")
            current_lorry = sess["assigned"].get(do_num, "")
            other_assigned = set()
            for d in sess.get("pending_dos", []):
                if d["DO NUMBER"] == do_num:
                    continue
                for it in d.get("ITEMS", []):
                    lv = it.get("LORRY", "")
                    if lv not in ("NO_LORRY", "SPLIT", "", None):
                        other_assigned.add(lv)
            excluded = (sess.get("unavailable", set()) | get_assigned_today() | other_assigned | {current_lorry})
            excluded.discard("")
            suggestions = engine.suggest(
                route=target_do["ROUTE"],
                total_ton=target_do["TOTAL_TON"],
                unavailable=excluded,
                top_n=1,
            )
            if not suggestions:
                return [f"⚠️ No eligible lorry found for *{do_num}*. All lorries may be assigned."]
            plate = suggestions[0]["LORRY"]

        # Validate — not already used by another DO
        already_used = set()
        for d in sess.get("pending_dos", []):
            if d["DO NUMBER"] == do_num:
                continue
            for it in d.get("ITEMS", []):
                lv = it.get("LORRY", "")
                if lv not in ("NO_LORRY", "SPLIT", "", None):
                    already_used.add(lv)
                for bin_ in (it.get("SPLIT_LORRIES") or []):
                    already_used.add(bin_["lorry"])

        blocked_today = get_assigned_today()
        old_lorry = sess["assigned"].get(do_num, "")

        if plate in already_used:
            other = next((d["DO NUMBER"] for d in sess.get("pending_dos", [])
                          for it in d.get("ITEMS", []) if it.get("LORRY") == plate), "another DO")
            return [f"❌ *{plate}* is already assigned to {other}.",
                    {"_type": "buttons", "body": "Pick another option:",
                     "buttons": [{"id": f"select_do {do_num}", "title": "🔄 Pick again"}]}]

        if plate in blocked_today and plate != old_lorry:
            return [f"❌ *{plate}* is blocked/assigned today.",
                    {"_type": "buttons", "body": "Pick another option:",
                     "buttons": [{"id": f"select_do {do_num}", "title": "🔄 Pick again"}]}]

        # Release old lorry, assign new
        if old_lorry and old_lorry not in ("NO_LORRY", "SPLIT", ""):
            sess.get("unavailable", set()).discard(old_lorry)
        for item in target_do.get("ITEMS", []):
            item["LORRY"] = plate
            item.pop("SPLIT_LORRIES", None)
        target_do["SPLIT"] = False
        target_do.pop("SPLIT_LORRIES", None)
        sess["assigned"][do_num] = plate
        sess.setdefault("unavailable", set()).add(plate)

        _sr = _build_summary(sess)
        return [f"✅ *{do_num}* → *{plate}*"] + (_sr if isinstance(_sr, list) else [_sr])

    # ══════════════════════════════════════════════════════════════════════════
    # LORRY MANAGEMENT — single consolidated implementation
    # ══════════════════════════════════════════════════════════════════════════

    def _get_engine_safe():
        """Return engine from session or reload from master."""
        _e = sess.get("engine")
        if _e is None and sess.get("user_id"):
            try:
                _e = LorryEngine(MASTER_PATH, _resolve_history_path(), owner_user=sess["user_id"])
            except Exception:
                pass
        return _e

    # ── Main menu ─────────────────────────────────────────────────────────────
    if cmd_lower == "manage lorry":
        return [{
            "_type": "buttons",
            "body":  "🚛 *Lorry Management*\nChoose an action:",
            "buttons": [
                {"id": "lorry_maint", "title": "🔧 Maintenance"},
                {"id": "hi",          "title": "🏠 Main Menu"},
            ]
        }]

    # ── Maintenance sub-menu ──────────────────────────────────────────────────
    if cmd_lower == "lorry_maint":
        # Use a list message — supports 4 actions (WhatsApp buttons cap at 3)
        return [{
            "_type":  "do_list",
            "header": "Lorry Management",
            "body":   "Select the action to apply:",
            "button": "Choose Action",
            "items":  [
                {"id": "maint_block",   "title": "Block",   "description": "Mark lorry unavailable for today"},
                {"id": "maint_broken",  "title": "Broken",  "description": "Log breakdown and find replacement"},
                {"id": "maint_release", "title": "Release", "description": "Unblock lorry from today log"},
                {"id": "maint_fixed",   "title": "Fixed",   "description": "Mark broken lorry as repaired"},
            ],
        }]

    # ── Show plate picker for each action ────────────────────────────────────
    def _maint_list(action: str) -> list:
        """
        Build the lorry picker list for the given action.
        For RELEASE and FIXED: prepend a 'Done — apply X' row when batch non-empty.
        BLOCK and BROKEN: single-select, no batch needed.
        """
        engine2 = _get_engine_safe()
        if not engine2:
            return ["Please log in first. Send hi to start."]
        taken_today2 = get_assigned_today()
        broken_map2  = get_broken_lorries()
        batch        = sess.get(f"_maint_batch_{action}", [])

        all_items2   = []   # unselected lorries
        selected_rows = []  # already-batched lorries (shown at top with toggle)

        for _, r in engine2.eligible_lorries.iterrows():
            p2        = str(r["LORRY"]).upper()
            cap2      = float(r["TON"])
            is_broken2  = p2 in broken_map2
            is_blocked2 = p2 in taken_today2
            if action == "BLOCK"   and is_blocked2: continue
            if action == "RELEASE" and not is_blocked2: continue
            if action == "FIXED"   and not is_broken2:  continue
            if action == "BROKEN"  and is_broken2:  continue
            if is_broken2:
                rep2  = broken_map2[p2]
                desc2 = f"{cap2}T | Broken->{rep2}" if rep2 != "NONE" else f"{cap2}T | Broken"
            elif is_blocked2:
                desc2 = f"{cap2}T | Blocked"
            else:
                desc2 = f"{cap2}T | Available"

            if p2 in batch:
                # Already selected — show with checkmark and allow tap-to-deselect
                selected_rows.append({
                    "id":          f"maint_toggle {action} {p2}",
                    "title":       f"[X] {p2}",
                    "description": f"Tap to deselect | {desc2}"[:72],
                })
            else:
                all_items2.append({
                    "id":          f"maint_exec {action} {p2}",
                    "title":       p2,
                    "description": desc2[:72],
                })

        header_map2 = {
            "BLOCK":   "Block a Lorry",
            "BROKEN":  "Log Breakdown",
            "RELEASE": "Release Lorries",
            "FIXED":   "Mark Lorries Fixed",
        }

        if not all_items2 and not batch:
            msg_map2 = {
                "BLOCK":   "All lorries already blocked today.",
                "RELEASE": "No lorries currently blocked.",
                "FIXED":   "No lorries marked as broken.",
                "BROKEN":  "All lorries already marked as broken.",
            }
            return [
                msg_map2.get(action, "No lorries to show."),
                {"_type": "buttons", "body": "What would you like to do?",
                 "buttons": [{"id": "lorry_maint", "title": "Back"},
                             {"id": "hi",          "title": "Main Menu"}]},
            ]

        list_items2 = []

        # For RELEASE / FIXED: show Done row + selected rows at the top,
        # then unselected lorries below (paginated)
        if action in ("RELEASE", "FIXED"):
            if batch:
                selected_str = ", ".join(batch)
                list_items2.append({
                    "id":          f"maint_batch_done {action}",
                    "title":       f"Done ({len(batch)} selected)",
                    "description": f"Confirm: {selected_str}"[:72],
                })
            list_items2 += selected_rows   # checked items always visible, no pagination

        # Paginate unselected items
        # Reserve rows for: Done (1) + selected rows + Next page (1)
        reserved    = (1 if batch else 0) + len(selected_rows) + 1   # +1 for possible Next
        PER_PAGE2   = max(1, 9 - reserved)
        total2      = len(all_items2)
        page2       = int(sess.get("maint_picker_page", {}).get(action, 0))
        total_pages2 = max(1, -(-total2 // PER_PAGE2)) if total2 else 1
        page2        = max(0, min(page2, total_pages2 - 1))
        start2       = page2 * PER_PAGE2
        chunk2       = all_items2[start2:start2 + PER_PAGE2]
        list_items2 += chunk2

        if total_pages2 > 1:
            next_p = (page2 + 1) % total_pages2
            list_items2.append({
                "id":          f"maint_page {action} {next_p}",
                "title":       "Next page...",
                "description": f"Showing {start2+1}-{start2+len(chunk2)} of {total2}",
            })

        body_map2 = {
            "BLOCK":   "Select lorry to block for today:",
            "BROKEN":  "Select lorry to log as broken:",
            "RELEASE": "Tap lorries to release. [X] = selected. Tap Done when ready:",
            "FIXED":   "Tap lorries to mark as fixed. [X] = selected. Tap Done when ready:",
        }

        return [{
            "_type":  "do_list",
            "header": header_map2.get(action, action),
            "body":   body_map2.get(action, "Select lorry:"),
            "button": "Pick Lorry",
            "items":  list_items2,
        }]

    if cmd_lower in ("maint_block", "maint_broken", "maint_release", "maint_fixed"):
        action_map = {
            "maint_block":   "BLOCK",
            "maint_broken":  "BROKEN",
            "maint_release": "RELEASE",
            "maint_fixed":   "FIXED",
        }
        action = action_map[cmd_lower]
        # Clear batch when user re-enters the action from scratch
        if action in ("RELEASE", "FIXED"):
            sess.pop(f"_maint_batch_{action}", None)
        sess.setdefault("maint_picker_page", {})[action] = 0
        return _maint_list(action)

    # ── Page turn for lorry picker ───────────────────────────────────────────
    if cmd_lower.startswith("maint_page "):
        parts = text.strip().split()
        if len(parts) >= 3:
            action_p = parts[1].upper()
            page_p   = int(parts[2])
            sess.setdefault("maint_picker_page", {})[action_p] = page_p
            return _maint_list(action_p)
        return ["Invalid page selection."]

    # ── Commit the multi-select batch ────────────────────────────────────────
    if cmd_lower.startswith("maint_batch_done "):
        action    = text.strip().split(" ", 1)[1].strip().upper()
        batch_key = f"_maint_batch_{action}"
        batch     = sess.pop(batch_key, [])
        if not batch:
            return ["Nothing selected. Tap lorries first, then Done."]
        for p3 in batch:
            if action == "RELEASE":
                release_specific_plates([p3])
                sess.setdefault("unavailable", set()).discard(p3)
            elif action == "FIXED":
                remove_broken_lorry(p3)
                sess.setdefault("unavailable", set()).discard(p3)
        action_label = "released" if action == "RELEASE" else "marked as fixed"
        plates_str   = ", ".join(batch)
        in_active = sess.get("state") in ("CONFIRMING", "REVIEWING", "AWAIT_OTHER_USER_REPLY") and sess.get("pending_dos")
        follow_up = _build_summary(sess) if in_active else [{
            "_type": "buttons", "body": "Need anything else?",
            "buttons": [{"id": "lorry_maint", "title": "More Actions"},
                        {"id": "hi",          "title": "Main Menu"}]
        }]
        noun = "lorry" if len(batch) == 1 else "lorries"
        return [f"{len(batch)} {noun} {action_label}: {plates_str}"] + follow_up

    # ── Toggle (deselect) a plate that was already queued ────────────────────
    if cmd_lower.startswith("maint_toggle "):
        parts = text.strip().split()
        if len(parts) < 3:
            return ["Invalid selection."]
        action    = parts[1].upper()
        plate     = parts[2].upper()
        batch_key = f"_maint_batch_{action}"
        batch     = sess.setdefault(batch_key, [])
        if plate in batch:
            batch.remove(plate)
        return _maint_list(action)

    # ── Execute action after tapping a plate ─────────────────────────────────
    if cmd_lower.startswith("maint_exec "):
        parts = text.strip().split()
        if len(parts) < 3:
            return ["Invalid selection."]
        action = parts[1].upper()
        plate  = parts[2].upper()

        if action == "BLOCK":
            record_assignments_today([plate], user=sess.get("user_id"))
            sess.setdefault("unavailable", set()).add(plate)
            in_active = sess.get("state") in ("CONFIRMING", "REVIEWING", "AWAIT_OTHER_USER_REPLY") and sess.get("pending_dos")
            follow_up = _build_summary(sess) if in_active else [{
                "_type": "buttons", "body": "Need anything else?",
                "buttons": [{"id": "lorry_maint", "title": "More Actions"},
                            {"id": "hi",          "title": "Main Menu"}]
            }]
            return [f"{plate} blocked."] + follow_up

        elif action == "BROKEN":
            return handle_message(phone, text=f"select_broken_lorry {plate}")

        elif action in ("RELEASE", "FIXED"):
            # Add to batch and show updated list
            batch_key = f"_maint_batch_{action}"
            batch     = sess.setdefault(batch_key, [])
            if plate not in batch:
                batch.append(plate)
            return _maint_list(action)

    # ── manage [PLATE]: shortcut ──────────────────────────────────────────────
    if cmd_lower.startswith("manage ") and not cmd_lower.startswith("manage_lorry_pick"):
        shortcut_plate = text.split(" ", 1)[1].strip().upper()
        return handle_message(phone, text=f"manage_lorry_pick {shortcut_plate}")

    # ── manage_lorry_pick: show action buttons for a chosen plate ────────────
    if cmd_lower.startswith("manage_lorry_pick "):
        plate       = text.split(" ", 1)[1].strip().upper()
        taken_today = get_assigned_today()
        broken_map  = get_broken_lorries()
        is_broken   = plate in broken_map
        is_blocked  = plate in taken_today

        buttons = []
        if not is_blocked:
            buttons.append({"id": f"select_block_lorry {plate}",   "title": "🚫 Block"})
        if not is_broken:
            buttons.append({"id": f"select_broken_lorry {plate}",  "title": "🔧 Broken"})
        if is_blocked:
            buttons.append({"id": f"select_release_lorry {plate}", "title": "🔓 Release"})
        if is_broken:
            buttons.append({"id": f"select_fixed_lorry {plate}",   "title": "✅ Fixed"})
        if len(buttons) < 3:
            buttons.append({"id": "lorry_maint", "title": "↩️ Pick Another"})

        engine = _get_engine_safe()
        cap    = ""
        if engine is not None:
            row = engine.eligible_lorries[engine.eligible_lorries["LORRY"] == plate]
            if not row.empty:
                cap = f"{float(row.iloc[0]['TON'])}T"

        status_parts = []
        if is_broken:
            rep = broken_map[plate]
            status_parts.append(f"🔧 Broken — replacement: {rep}")
        if is_blocked:
            status_parts.append("⛔ Assigned/blocked today")
        if not is_broken and not is_blocked:
            status_parts.append("✅ Available")

        body = f"*{plate}*  {cap}\n" + "\n".join(status_parts) + "\n\nWhat would you like to do?"
        return [{"_type": "buttons", "body": body, "buttons": buttons[:3]}]


    # ── Tap-only lorry action handlers ──────────────────────────────────────
    # These are triggered from manage_lorry_pick buttons

    if text.lower().startswith("select_block_lorry "):
        plate  = text.split(" ", 1)[1].strip().upper()
        engine = sess.get("engine")
        all_plates = set(engine.all_lorries["LORRY"].str.upper()) if engine else set()
        if all_plates and plate not in all_plates:
            return [f"⚠️ *{plate}* not found in master list."]
        if plate in sess.get("unavailable", set()) or plate in get_assigned_today():
            return [f"⚠️ *{plate}* is already blocked today."]
        sess.setdefault("unavailable", set()).add(plate)
        record_assignments_today([plate], user=sess.get("user_id"))
        in_active = sess.get("state") in ("CONFIRMING", "REVIEWING", "AWAIT_OTHER_USER_REPLY") and sess.get("pending_dos")
        follow_up = _build_summary(sess) if in_active else [_HI_BTN]
        return [f"🚫 *{plate}* blocked for today."] + follow_up

    if text.lower().startswith("select_release_lorry "):
        plate  = text.split(" ", 1)[1].strip().upper()
        engine = sess.get("engine")
        all_plates = set(engine.all_lorries["LORRY"].str.upper()) if engine else set()
        if all_plates and plate not in all_plates:
            close = _find_close_plate(plate, all_plates)
            hint  = f"\nDid you mean *{close}*?" if close else ""
            return [f"⚠️ *{plate}* not found in master list.{hint}"]
        released = release_specific_plates([plate])
        in_active = sess.get("state") in ("CONFIRMING", "REVIEWING", "AWAIT_OTHER_USER_REPLY") and sess.get("pending_dos")
        follow_up = _build_summary(sess) if in_active else [_HI_BTN]
        if released:
            sess.setdefault("unavailable", set()).discard(plate)
            return [f"✅ *{plate}* released and available again."] + follow_up
        return [f"⚠️ *{plate}* was not in today's log (already available)."] + follow_up

    if text.lower().startswith("select_fixed_lorry "):
        plate = text.split(" ", 1)[1].strip().upper()
        in_active = sess.get("state") in ("CONFIRMING", "REVIEWING", "AWAIT_OTHER_USER_REPLY") and sess.get("pending_dos")
        follow_up = _build_summary(sess) if in_active else [_HI_BTN]
        if remove_broken_lorry(plate):
            sess.setdefault("unavailable", set()).discard(plate)
            return [f"✅ *{plate}* marked as fixed and unblocked."] + follow_up
        return [f"⚠️ *{plate}* was not in today's broken list."] + follow_up

    if text.lower().startswith("select_replacement "):
        parts         = text.strip().split(" ", 2)
        broken_plate  = parts[1].strip().upper() if len(parts) > 1 else ""
        replace_plate = parts[2].strip().upper() if len(parts) > 2 else "NONE"
        if not broken_plate:
            return ["❌ Invalid selection."]
        record_broken_lorry(broken_plate, replace_plate)
        sess.setdefault("unavailable", set()).add(broken_plate)
        sess.pop("pending_broken_plate", None)
        rep_str = f"replaced by *{replace_plate}*" if replace_plate != "NONE" else "no replacement"
        in_active = sess.get("state") in ("CONFIRMING", "REVIEWING", "AWAIT_OTHER_USER_REPLY") and sess.get("pending_dos")
        follow_up = _build_summary(sess) if in_active else [_HI_BTN]
        return [
            f"🔧 *Breakdown logged:*\n"
            f"  ❌ Broken:      *{broken_plate}*\n"
            f"  ✅ Replacement: {rep_str}\n\n"
            f"*{broken_plate}* is blocked for today.",
        ] + follow_up

    if text.lower().startswith("select_broken_lorry "):
        broken_plate = text.split(" ", 1)[1].strip().upper()
        engine = sess.get("engine")
        all_plates = set(engine.all_lorries["LORRY"].str.upper()) if engine else set()
        if all_plates and broken_plate not in all_plates:
            return [f"⚠️ *{broken_plate}* not found in master list."]
        excl = sess.get("unavailable", set()) | get_assigned_today() | {broken_plate}
        avail = []
        if engine:
            for _, r in engine.eligible_lorries.iterrows():
                if r["LORRY"] not in excl:
                    cap    = float(r["TON"])
                    status = "Blocked" if r["LORRY"] in get_assigned_today() else "Available"
                    avail.append({
                        "id":          f"select_replacement {broken_plate} {r['LORRY']}",
                        "title":       str(r["LORRY"])[:24],
                        "description": f"{cap}T | {status}",
                    })
        # WhatsApp list: max 10 rows per section, 1 section = 10 total max
        # Keep best 9 lorries + "No replacement" to always stay in one section
        avail = avail[:9]
        avail.append({
            "id":          f"select_replacement {broken_plate} NONE",
            "title":       "No replacement",
            "description": "Block only, no replacement needed",
        })
        sess["pending_broken_plate"] = broken_plate
        return [{
            "_type":  "do_list",
            "header": f"{broken_plate} broken - pick replacement",
            "body":   f"Select a replacement lorry for {broken_plate}:",
            "button": "Pick Lorry",
            "items":  avail,
        }]

    # ── Broken lorry commands ─────────────────────────────────────────────────
    # broken [PLATE]            — mark lorry as broken, bot asks for replacement
    # broken [PLATE] [REPLACE]  — mark broken + set replacement in one go
    # fixed [PLATE]             — mark lorry as repaired, remove from broken list
    # broken list               — show all broken lorries and replacements today

    if cmd_lower == "broken list":
        return _handle_broken_list(sess)

    if cmd_lower.startswith("fixed "):
        plate = text.split(" ", 1)[1].strip().upper()
        if remove_broken_lorry(plate):
            sess["unavailable"].discard(plate)
            return [
                f"✅ *{plate}* marked as *fixed* and unblocked for today.",
                {"_type": "buttons", "body": "What would you like to do next?",
                 "buttons": [{"id": "hi", "title": "👋 Hi"}]},
            ]
        return [f"⚠️ *{plate}* was not in today's broken list."]

    # ── release [PLATE1] [PLATE2...] ─────────────────────────────────────────
    # Remove specific plates from today's assigned log (lorry repaired/available again)
    # Does NOT require the plate to be in the broken list — handles any reason
    # Usage:
    #   release VJN9910            — release one plate
    #   release VJN9910 BQU3875    — release multiple plates at once
    if cmd_lower.startswith("release "):
        plates_to_release = [p.upper() for p in text.strip().split()[1:] if p]
        if not plates_to_release:
            return ["Usage: *release [PLATE1] [PLATE2...]*\ne.g. *release VJN9910* or *release VJN9910 BQU3875*"]

        engine       = sess.get("engine")
        all_plates   = set(engine.all_lorries["LORRY"].str.upper()) if engine else set()
        log          = _load_daily_log()
        assigned_set = set(log["assigned"])
        broken_map   = log.get("broken", {})

        released   = []
        not_in_log = []
        typos      = []   # (typed, suggested_correct)

        for plate in plates_to_release:
            # Step 1: check if plate exists in master list at all
            if all_plates and plate not in all_plates:
                close = _find_close_plate(plate, all_plates)
                typos.append((plate, close))
                continue
            # Step 2: check if it's in today's log
            if plate in assigned_set:
                assigned_set.discard(plate)
                broken_map.pop(plate, None)
                sess["unavailable"].discard(plate)
                released.append(plate)
            else:
                not_in_log.append(plate)

        log["assigned"] = sorted(assigned_set)
        log["broken"]   = broken_map
        _save_daily_log(log)

        lines = []
        if released:
            lines.append(f"✅ Released & available again: *{', '.join(released)}*")
        if not_in_log:
            lines.append(f"⚠️ Not in today's log (already available): *{', '.join(not_in_log)}*")
        if typos:
            for typed, close in typos:
                hint = f" → Did you mean *{close}*?" if close else ""
                lines.append(f"❌ *{typed}* not found in master list.{hint}")
        if not released and not lines:
            lines.append("No changes made.")

        # ── Re-evaluate active DOs against the newly released lorry(s) ──────────
        reassigned = []
        active_states = ("CONFIRMING", "REVIEWING", "AWAIT_OTHER_USER_REPLY")
        if released and engine and sess.get("state") in active_states and sess.get("pending_dos"):
            taken = (sess["unavailable"] | get_assigned_today()) - set(released)
            for do in sess["pending_dos"]:
                do_num = do["DO NUMBER"]
                for item in do.get("ITEMS", []):
                    current_lorry = item.get("LORRY", "")
                    if current_lorry in ("NO_LORRY", "SPLIT", "SKIPPED", ""):
                        continue
                    weight = item["WEIGHT"]
                    route  = item["ROUTE"]
                    cur_row = engine.eligible_lorries[
                        engine.eligible_lorries["LORRY"] == current_lorry
                    ]
                    if cur_row.empty:
                        continue
                    cur_surplus = float(cur_row.iloc[0]["TON"]) - weight
                    excl = (taken | sess["unavailable"]) - set(released) - {current_lorry}
                    best = engine.suggest(route=route, total_ton=weight,
                                         unavailable=excl, top_n=1)
                    if not best:
                        continue
                    best_lorry   = best[0]["LORRY"]
                    best_surplus = best[0]["SURPLUS"]
                    if best_lorry != current_lorry and best_surplus < cur_surplus - 0.001:
                        old = current_lorry
                        item["LORRY"] = best_lorry
                        sess["assigned"][do_num] = best_lorry
                        sess["unavailable"].discard(old)
                        sess["unavailable"].add(best_lorry)
                        taken.add(best_lorry)
                        taken.discard(old)
                        reassigned.append(f"  • {do_num}: {old} → *{best_lorry}* "
                                          f"({round(best_surplus,2)}T spare vs {round(cur_surplus,2)}T)")

        if reassigned:
            lines.append("\n🔄 *Better fits found after release:*\n" + "\n".join(reassigned))
        else:
            lines.append("\nThese lorries will now appear in suggestions for new DOs.")

        if sess.get("state") in active_states and sess.get("pending_dos"):
            return ["\n".join(lines), _build_summary(sess)]
        return ["\n".join(lines), {"_type": "buttons",
                                    "body": "What would you like to do next?",
                                    "buttons": [{"id": "hi", "title": "👋 Hi"}]}]

    if cmd_lower.startswith("broken "):
        parts = text.strip().split()
        broken_plate  = parts[1].upper() if len(parts) > 1 else None
        replace_plate = parts[2].upper() if len(parts) > 2 else None
        if not broken_plate:
            return ["Usage: *broken [PLATE]* or *broken [PLATE] [REPLACEMENT]*"]
        if replace_plate and broken_plate == replace_plate:
            return [f"⚠️ Replacement cannot be the same as the broken lorry (*{broken_plate}*)."]
        if replace_plate:
            record_broken_lorry(broken_plate, replace_plate)
            sess["unavailable"].add(broken_plate)
            return _broken_confirmed_reply(broken_plate, replace_plate, sess)
        else:
            sess["state_before_broken"] = sess["state"]
            sess["pending_broken_plate"] = broken_plate
            sess["state"] = "AWAIT_BROKEN_REPLACEMENT"
            engine = sess.get("engine")

            # ── Find what this broken lorry was assigned to ──────────────────
            # Look through current session items to get route + weight context
            broken_items = []
            for do in sess.get("pending_dos", []):
                for it in do.get("ITEMS", []):
                    if it.get("LORRY") == broken_plate:
                        broken_items.append((it, do))
                    elif it.get("LORRY") == "SPLIT":
                        for b in (it.get("SPLIT_LORRIES") or []):
                            if b.get("lorry") == broken_plate:
                                broken_items.append((it, do))

            # ── Auto-recommend best replacement ─────────────────────────────
            best_recs = []   # list of (plate, cap, util_pct, reason)
            excl = sess["unavailable"] | get_assigned_today() | {broken_plate}

            if engine is not None and broken_items:
                # Use first broken item's route + weight for recommendation
                ref_item, ref_do = broken_items[0]
                weight = ref_item.get("WEIGHT", 0)
                route  = ref_item.get("ROUTE", "")
                sug = engine.suggest(route=route, total_ton=weight,
                                     unavailable=excl, top_n=3)
                for s in sug:
                    cap  = s["TON_CAPACITY"]
                    util = round(weight / cap * 100, 1) if cap > 0 else 0
                    best_recs.append((s["LORRY"], cap, util, s["REASON"]))

            elif engine is not None:
                # No current assignment context — suggest by capacity only
                avail = engine.eligible_lorries[
                    ~engine.eligible_lorries["LORRY"].isin(excl)
                ].copy()
                avail = avail.sort_values("TON", ascending=False)
                for _, r in avail.head(3).iterrows():
                    best_recs.append((r["LORRY"], r["TON"], None, "Available lorry"))

            # ── Build message + buttons ──────────────────────────────────────
            lines = [f"🔧 *{broken_plate}* marked as broken."]

            if broken_items:
                items_str = ", ".join(
                    f"{do.get('CUSTOMER NAME','')[:18]} {round(it.get('WEIGHT',0),1)}T ({do.get('DO NUMBER','')})"
                    for it, do in broken_items[:3]
                )
                lines.append(f"Was assigned to: {items_str}")

            lines.append("")
            if best_recs:
                lines.append("🚛 *Recommended replacements:*")
                for plate, cap, util, reason in best_recs:
                    util_str = f" — {util}% util" if util is not None else ""
                    lines.append(f"  • *{plate}* ({cap}T){util_str}")
                lines.append("")
                lines.append("Tap a button to assign, or type any plate manually:")
            else:
                lines.append("No suitable replacement found automatically.")
                lines.append("Type a plate manually or *none* to skip.")

            # Buttons: top 2 recommendations + none option (max 3 buttons)
            btns = []
            for plate, cap, util, _ in best_recs[:2]:
                util_tag = f" {util}%" if util is not None else ""
                btns.append({"id": plate, "title": f"🚛 {plate}{util_tag}"[:20]})
            btns.append({"id": "none", "title": "⏭️ No replacement"})

            return [
                "\n".join(lines),
                {"_type": "buttons", "body": "Choose replacement:", "buttons": btns}
            ]


    if text.lower() == "show assigned today":
        today_plates = get_assigned_today()
        if not today_plates:
            return ["No lorries assigned yet today.", _HI_BTN]
        sess = get_session(phone)
        engine = sess.get("engine")
        if engine is not None:
            user_lorries = set(engine.eligible_lorries["LORRY"].str.upper())
            my_plates    = sorted(today_plates & user_lorries)
            other_plates = sorted(today_plates - user_lorries)
            lines = [f"🚛 *Lorries assigned today ({sess['user_id']}):*"]
            if my_plates:
                lines += [f"  • {p}" for p in my_plates]
            else:
                lines.append("  (none of your lorries assigned yet)")
            if other_plates:
                lines.append(f"\n_Other users: {', '.join(other_plates)}_")
            return ["\n".join(lines), _HI_BTN]
        return ["🚛 *Lorries already assigned today:*\n" +
                "\n".join(f"  • {p}" for p in sorted(today_plates)), _HI_BTN]

    if text.lower() in ("show blocked", "block list", "blocked list"):
        sess   = get_session(phone)
        engine = sess.get("engine")
        today  = get_assigned_today()
        broken_map = get_broken_lorries()   # {broken: replacement}

        # Only show this user's eligible lorries that are currently blocked
        user_lorries = set()
        if engine is not None:
            user_lorries = set(engine.eligible_lorries["LORRY"].str.upper())

        blocked_plates = sorted(today & user_lorries) if user_lorries else sorted(today)
        broken_plates  = set(broken_map.keys())

        if not blocked_plates:
            return ["✅ No blocked lorries for you today. All available.", _HI_BTN]

        lines = ["🚫 *Blocked lorries today:*\n"]
        for p in blocked_plates:
            if p in broken_plates:
                rep = broken_map[p]
                rep_str = f" → replaced by *{rep}*" if rep != "NONE" else " (no replacement)"
                lines.append(f"  🔧 *{p}* — broken{rep_str}")
            else:
                lines.append(f"  ⛔ *{p}* — assigned/blocked")

        lines.append("\nTap a plate to release it, or type *release [PLATE]*:")

        # Buttons: up to 3 plates as tappable release buttons
        # Button ID = "release PLATE" so it's handled by existing release handler
        btns = [
            {"id": f"release {p}", "title": f"🔓 {p}"[:20]}
            for p in blocked_plates[:3]
        ]

        return [
            "\n".join(lines),
            {"_type": "buttons", "body": "Tap to release:", "buttons": btns}
        ]

    if text.lower() in ("hi", "hello", "start"):
        return _start(phone, sess)

    # ── Bare plate release shortcut ───────────────────────────────────────────
    # If user types just a plate (e.g. "Wld8738") and it's in their blocked list,
    # treat it as "release [PLATE]" — works in any state
    if re.match(r'^[A-Za-z0-9]{4,10}$', text.strip()):
        candidate = text.strip().upper()
        _eng = sess.get("engine")
        if _eng is None and sess.get("user_id"):
            try:
                _hist = _resolve_history_path()
                _eng = LorryEngine(MASTER_PATH, _hist, owner_user=sess["user_id"])
            except Exception:
                pass
        _ul = set(_eng.eligible_lorries["LORRY"].str.upper()) if _eng else set()
        if candidate in (get_assigned_today() & _ul):
            return handle_message(phone, text=f"release {candidate}")

    if state == "IDLE":
        return _start(phone, sess)
    elif state == "AWAIT_USER_ID":
        return _handle_user_id(phone, sess, text)
    elif state == "AWAIT_MASTER_UPLOAD":
        if file_bytes:
            return _handle_master_upload(phone, sess, file_bytes)
        return ["📄 Please upload today's *master lorry file* (.xlsx) to continue."]
    elif state == "AWAIT_TRIP_DAY":
        return _handle_trip_day(phone, sess, text)
    elif state == "AWAIT_EXCEL":
        if file_bytes:
            return _handle_excel_upload(phone, sess, file_bytes)
        return ["Please upload the DO Excel file (.xlsx) to continue."]
    elif state == "AWAIT_OTHER_USER_REPLY":
        return _handle_other_user_reply(phone, sess, text)
    elif state in ("REVIEWING", "CONFIRMING"):
        # Allow lorry-status file upload at any point during an active session
        if file_bytes:
            try:
                _df_up = pd.read_excel(io.BytesIO(file_bytes))
                _df_up.columns = [c.strip().upper() for c in _df_up.columns]
                _status_result = _handle_lorry_status_upload(phone, sess, _df_up)
                if _status_result is not None:
                    # After updating statuses, re-show summary so user sees changes
                    return _status_result + [_build_summary(sess)]
            except Exception:
                pass   # not an Excel — fall through to text handler
        if state == "REVIEWING":
            return _handle_reviewing(phone, sess, text)
        return _handle_confirming(phone, sess, text)
    elif state == "DONE":
        # After export, user may still block/change a lorry — revert to CONFIRMING,
        # process the command, then auto re-export and send the updated file
        cmd = text.lower().strip()
        if (cmd.startswith("block ") or cmd.startswith("change ") or
                cmd.startswith("release ")):
            sess["state"] = "CONFIRMING"
            msgs = _handle_confirming(phone, sess, text)
            # If the command changed assignments, re-export automatically
            if sess.get("state") == "CONFIRMING" and sess.get("pending_dos"):
                try:
                    export_msgs = _export_result(sess)
                    msgs += export_msgs
                except Exception as _e:
                    msgs.append(f"⚠️ Could not regenerate export: {_e}")
            return msgs
        return ["Your assignments have been exported. Send *hi* to start a new session."]
    elif state == "AWAIT_BROKEN_REPLACEMENT":
        return _handle_broken_replacement(phone, sess, text)

    return ["Sorry, I didn't understand that. Send *hi* to start."]


# ── State handlers ────────────────────────────────────────────────────────────

def _get_valid_users() -> list[str]:
    """Read valid users from LORRY DAILY PLANNING.xlsx MUATAN sheet.
    Users are section headers (rows where col 0 has a non-lorry string like ABI/VIVIAN).
    """
    try:
        df = pd.read_excel(PLANNING_PATH, sheet_name="MUATAN", header=None)
        users = []
        for val in df.iloc[:, 0].dropna().astype(str):
            v = val.strip().upper()
            # Section headers are short non-numeric strings that aren't "LORRY"
            if v and v != "LORRY" and not re.match(r'^[A-Z]{1,3}\d', v):
                users.append(v)
        return sorted(set(users)) if users else ["ABI", "VIVIAN"]
    except Exception:
        return ["ABI", "VIVIAN", "SELAYANG", "BIG"]  # fallback


def _start(phone, sess):
    sess["state"] = "AWAIT_USER_ID"
    users = _get_valid_users()
    # ABI, VIVIAN, BIG get clickable buttons; everything else (SELAYANG, SPARE, …) is type-only
    BUTTON_ORDER = ["ABI", "VIVIAN", "BIG"]
    btn_users  = [u for u in BUTTON_ORDER if u in users]
    type_users = [u for u in users if u not in btn_users]
    body = "👋 *Lorry Assignment Bot*\n\nPlease tap your name below or type it to continue."
    if type_users:
        body += f"\nOr type: {', '.join(u.title() for u in type_users)}"
    return [{
        "_type": "buttons",
        "body": body,
        "buttons": [{"id": u.lower(), "title": u.title()} for u in btn_users],
    }]


# Users who share / cross-use lorries via the daily master file. Only these are
# asked to upload it at login; others use their default owner+SPARE fleet.
_MASTER_FILE_USERS = {"ABI", "VIVIAN"}


def _convert_simple_available_list(file_bytes, user: str):
    """Convert a simple 'today's available lorries' file — just a LORRY
    column (e.g. the daily delivery-performance report format: NO, LORRY,
    DRIVER, MUATAN(KG), ..., AREA, ...) — into the LORRY/USER/DESCRIPTION/
    Status shape _parse_master_lorry expects.

    Every plate listed becomes Available for the current user; every plate
    NOT listed is simply absent, so it won't be in that user's eligible
    fleet today (equivalent to Blocked). Capacity always comes from the
    current master_lorry.xlsx, not from this file — even if the upload has
    its own capacity-looking column (e.g. MUATAN(KG)), using it would let a
    stale figure silently override the maintained default.

    Returns (file_bytes, skipped_plates). file_bytes is None if this doesn't
    look like that format at all (caller falls back to the standard
    master_lorry.xlsx-shaped parse) or if every listed plate is unresolvable.
    skipped_plates lists any listed plate with no known capacity (typo, or a
    genuinely new plate not yet in master_lorry.xlsx) — these are dropped
    rather than passed through with an unknown/NaN capacity, which would
    otherwise silently corrupt the fleet and break the result JSON.
    """
    try:
        df = pd.read_excel(io.BytesIO(file_bytes))
    except Exception:
        return None, []
    df.columns = [str(c).strip().upper() for c in df.columns]
    if "LORRY" not in df.columns or "USER" in df.columns or "STATUS" in df.columns:
        return None, []

    plates = []
    seen = set()
    for v in df["LORRY"]:
        plate = str(v).strip().upper()
        if plate and plate not in ("NAN", "NONE") and plate not in seen:
            seen.add(plate)
            plates.append(plate)
    if not plates:
        return None, []

    cap_by_plate = {}
    try:
        _m = pd.read_excel(MASTER_LORRY_PATH)
        _m.columns = [str(c).strip().upper() for c in _m.columns]
        for _, r in _m.iterrows():
            _plate = str(r.get("LORRY", "")).strip().upper()
            _desc = r.get("DESCRIPTION")
            if _plate and _desc is not None and pd.notna(_desc):
                cap_by_plate[_plate] = _desc
    except Exception:
        pass

    resolved = [p for p in plates if p in cap_by_plate]
    skipped = [p for p in plates if p not in cap_by_plate]
    if not resolved:
        return None, skipped

    out = pd.DataFrame([{
        "LORRY": p, "USER": user, "Status": "Available",
        "DESCRIPTION": cap_by_plate[p],
    } for p in resolved])
    buf = io.BytesIO()
    out.to_excel(buf, index=False)
    return buf.getvalue(), skipped


def _parse_master_lorry(file_bytes):
    """Parse the daily master lorry file.

    Layout: columns LORRY, TON, DESCRIPTION (naik value in kg), USER, Status.
    The same plate appears once per user, carrying that user's Status for the
    day. A user's fleet = the plates marked 'Available' for them — so cross-use
    works: a lorry Available for ABI but Block for VIVIAN is ABI's to use.
    TON is computed from DESCRIPTION / 1000 (the TON column is a spreadsheet
    formula, so we read the kg value in DESCRIPTION).

    A row with USER == 'SPARE' is a shared lorry, not tied to one planner —
    when Available it's added to BOTH ABI's and VIVIAN's fleet (whichever of
    them logs in that day can use it), and never counts toward the
    'Available for >1 user' conflict check below (that check is for a real
    plate mistakenly double-booked between two named planners; a SPARE row
    being usable by both is the intended design, not a conflict).

    Returns (per_user, conflicts, err):
      per_user  = {USER: [(plate, ton), ...]}  Available plates per user
      conflicts = [(plate, [users]), ...]      plates Available under >1 user
      err       = str | None                   parse/format error
    """
    try:
        df = pd.read_excel(io.BytesIO(file_bytes))
    except Exception as e:
        return {}, [], f"could not read the file ({e})"
    df.columns = [str(c).strip().upper() for c in df.columns]
    for _need in ("LORRY", "USER", "STATUS"):
        if _need not in df.columns:
            return {}, [], "missing required columns (need LORRY, USER, Status)"
    plate_users: dict = {}   # plate -> {user: (ton, status)}
    for _, r in df.iterrows():
        plate = str(r.get("LORRY", "")).strip().upper()
        if not plate or plate in ("NAN", "NONE", ""):
            continue
        user   = str(r.get("USER", "")).strip().upper()
        status = str(r.get("STATUS", "")).strip().upper()
        ton = None
        try:
            ton = round(float(r.get("DESCRIPTION")) / 1000.0, 4)
        except Exception:
            try:
                ton = float(r.get("TON"))
            except Exception:
                ton = None
        if not user or ton is None:
            continue
        plate_users.setdefault(plate, {})[user] = (ton, status)
    per_user: dict = {}
    conflicts: list = []
    _SPARE_SHARED_WITH = ("ABI", "VIVIAN")
    for plate, um in plate_users.items():
        avail = [u for u, (t, s) in um.items() if s.startswith("AVAIL") and u != "SPARE"]
        if len(avail) > 1:
            conflicts.append((plate, sorted(avail)))
        for u, (t, s) in um.items():
            if not s.startswith("AVAIL"):
                continue
            if u == "SPARE":
                for shared_u in _SPARE_SHARED_WITH:
                    per_user.setdefault(shared_u, []).append((plate, t))
            else:
                per_user.setdefault(u, []).append((plate, t))
    return per_user, conflicts, None


def _fleet_and_trip_prompt(sess, user):
    """Build the 'your lorries' list + the Today/Tomorrow trip-day buttons from
    the engine's current eligible fleet (already set from the master file)."""
    lorries = sess["engine"].get_eligible_lorry_list()
    taken_today  = get_assigned_today()
    broken_today = get_broken_lorries()
    lines = []
    for _, r in lorries.iterrows():
        plate, ton, lorry_user = r["LORRY"], r["TON"], r["USER"]
        if plate in broken_today:
            rep = broken_today[plate]
            tag = f" 🔴 Broken→{rep}" if rep != "NONE" else " 🔴 Broken"
        elif plate in taken_today:
            tag = " ⛔ Assigned today"
        else:
            tag = " ✅ Available"
        lines.append(f"  • {plate} — {ton}T ({lorry_user}){tag}")

    from datetime import timedelta as _td
    _today_name    = datetime.now().strftime("%A")
    _tomorrow_date = datetime.now().date() + _td(days=1)
    while _tomorrow_date.weekday() == 6:
        _tomorrow_date += _td(days=1)
    _tomorrow_name = _tomorrow_date.strftime("%A")

    lorry_text = (f"✅ *{user}*'s available lorries today ({len(lines)}):\n"
                  + "\n".join(lines))
    trip_day_msg = {
        "_type": "buttons",
        "body": (
            f"📅 Which day's DOs are you planning now?\n\n"
            f"• *Today* = {_today_name}'s routes\n"
            f"• *Tomorrow* = {_tomorrow_name}'s routes"
        ),
        "buttons": [
            {"id": "trip_day_today",    "title": f"Today ({_today_name[:3]})"},
            {"id": "trip_day_tomorrow", "title": f"Tomorrow ({_tomorrow_name[:3]})"},
            {"id": "clear daily log",   "title": "🗑️ Clear Today Log"},
        ],
    }
    return [lorry_text, trip_day_msg]


def _handle_master_upload(phone, sess, file_bytes):
    """Step 2 of the flow: the logged-in user uploads the daily master lorry
    file. Validate it (no plate Available under two users) and build this
    user's available fleet from it, then move on to the trip-day question."""
    _simple, _simple_skipped = _convert_simple_available_list(file_bytes, sess.get("user_id", ""))
    if _simple is None and _simple_skipped:
        _lines = "\n".join(f"  • {p}" for p in _simple_skipped[:25])
        return [f"❌ None of the plate(s) in this file are recognised:\n{_lines}\n\n"
                "Check for typos against the master lorry list, then upload again."]
    if _simple is not None:
        file_bytes = _simple
    per_user, conflicts, err = _parse_master_lorry(file_bytes)
    if err:
        return [f"❌ Master lorry file: {err}.\nPlease upload the correct master "
                "lorry file (columns: LORRY, TON, DESCRIPTION, USER, Status)."]
    if conflicts:
        _lines = "\n".join(f"  • *{p}* — Available for {', '.join(us)}"
                           for p, us in conflicts[:25])
        return [
            "❌ *Duplicate Available lorries.*\n"
            "These plate(s) are marked *Available* for more than one user, so I "
            "can't tell who should use them — that would double-book a lorry:\n"
            f"{_lines}\n\n"
            "Please *Block* the duplicates so each lorry is Available for only "
            "ONE user, then upload the master lorry file again.\n"
            "_(Duplicate *Block* is fine — that just means the lorry is off today.)_"
        ]
    user = sess.get("user_id")
    my_fleet = per_user.get(user, [])
    if not my_fleet:
        return [f"⚠️ No *Available* lorries found for *{user}* in the master file.\n"
                "Check the Status column (should say 'Available' on your rows) and "
                "upload the master lorry file again."]
    # Fresh start: uploading the master file clears THIS user's previous
    # assignments from today's log, so a re-run isn't blocked by their own
    # earlier lorries. Other users' log entries are left intact.
    _by = get_assigned_by()
    _mine_in_log = [p for p, u in _by.items() if u == user]
    if _mine_in_log:
        clear_specific_plates_from_log(_mine_in_log)
    sess["unavailable"] = set(get_assigned_today()) | set(get_broken_lorries())

    # Keep the FULL per-user fleet (before toggle-filtering) so the board can
    # still render a lane — with a working switch — for a plate the user has
    # toggled OFF. Only `eligible_lorries` (used for actual AI assignment)
    # gets toggle-filtered below.
    sess["_full_fleet"] = list(my_fleet)

    # A lorry this planner has toggled OFF today (board's on/off switch) is
    # excluded from the ASSIGNABLE fleet — for a SPARE plate, turning it ON
    # for the OTHER planner already forces it into this set (see
    # set_plate_toggle), so it drops out here too.
    _toggled_off = get_unavailable_plates_for(user)
    _assignable_fleet = [(p, t) for p, t in my_fleet if p.upper() not in _toggled_off] if _toggled_off else my_fleet

    # Replace the engine's eligible fleet with TODAY's availability for this user.
    sess["engine"].eligible_lorries = pd.DataFrame(
        [{"LORRY": p, "TON": t, "USER": user, "Status": "Available"}
         for p, t in sorted(_assignable_fleet)]
    )
    sess["_master_uploaded"] = True
    sess["state"] = "AWAIT_TRIP_DAY"
    _cleared_note = (f"\n🗑️ Cleared *{user}*'s previous log ({len(_mine_in_log)} "
                     f"lorry) for a fresh run." if _mine_in_log else "")
    _msgs = _fleet_and_trip_prompt(sess, user)
    if _cleared_note:
        _msgs[0] = _msgs[0] + _cleared_note
    if _simple_skipped:
        _skip_lines = ", ".join(_simple_skipped[:25])
        _msgs.insert(0, f"⚠️ Not recognised, skipped: {_skip_lines} "
                         "(check for typos against the master lorry list).")
    return _msgs


def _handle_user_id(phone, sess, text):
    valid_users = _get_valid_users()
    user = text.upper().strip()
    if user not in valid_users:
        return [f"❌ User not recognised. Please reply with one of: {', '.join(valid_users)}"]

    sess["user_id"] = user
    # Build the engine (history/routes/GPS + default owner+SPARE fleet).
    _hist = _resolve_history_path()
    sess["engine"] = LorryEngine(MASTER_PATH, _hist, owner_user=user)
    sess["_master_uploaded"] = False

    # Only ABI and VIVIAN share/cross-use lorries via the daily master file, so
    # only they are asked to upload it. Other users (SELAYANG, SPARE, BIG, …)
    # skip straight to the trip-day question using their default fleet.
    if user in _MASTER_FILE_USERS:
        sess["state"] = "AWAIT_MASTER_UPLOAD"
        return [
            f"✅ Logged in as *{user}*\n\n"
            "📄 Please upload today's *master lorry file* (.xlsx).\n"
            "I'll read which lorries are *Available* for you today (including any "
            "cross-use lorries you've been given).\n\n"
            "_Rule: each lorry must be *Available* for only ONE user. If the same "
            "plate is Available for two users, I'll ask you to fix and re-upload._"
        ]

    sess["state"] = "AWAIT_TRIP_DAY"
    return _fleet_and_trip_prompt(sess, user)


def _handle_trip_day(phone, sess, text):
    """Handle Today / Tomorrow selection after login."""
    cmd = text.lower().strip()
    if cmd in ("trip_day_today", "today"):
        trip_day = "today"
    elif cmd in ("trip_day_tomorrow", "tomorrow"):
        trip_day = "tomorrow"
    else:
        from datetime import timedelta as _td
        _today_name    = datetime.now().strftime("%A")
        _tomorrow_date = datetime.now().date() + _td(days=1)
        while _tomorrow_date.weekday() == 6:
            _tomorrow_date += _td(days=1)
        _tomorrow_name = _tomorrow_date.strftime("%A")
        return [{
            "_type": "buttons",
            "body": "Please select which day's DOs to plan:",
            "buttons": [
                {"id": "trip_day_today",    "title": f"Today ({_today_name[:3]})"},
                {"id": "trip_day_tomorrow", "title": f"Tomorrow ({_tomorrow_name[:3]})"},
                {"id": "clear daily log",   "title": "🗑️ Clear Today Log"},
            ],
        }]

    sess["trip_day"] = trip_day
    sess["state"]   = "AWAIT_EXCEL"

    from datetime import timedelta as _td
    if trip_day == "today":
        _day_label = datetime.now().strftime("%A, %d %b")
    else:
        _d = datetime.now().date() + _td(days=1)
        while _d.weekday() == 6:
            _d += _td(days=1)
        _day_label = _d.strftime("%A, %d %b")

    return [
        f"✅ Planning for *{_day_label}*.\n\n"
        "📎 Please upload your DO Excel file (.xlsx) now.\n\n"
        "_Tip: you can also upload the master lorry file (with a_ *Status* _column) "
        "to block/release lorries in bulk._",
    ]



def _handle_prefilled_excel(phone, sess, raw: "pd.DataFrame", prefilled: "pd.DataFrame") -> list[str]:
    """
    Called when the uploaded Excel already has LICENSE plates filled in.
    Reads every plate (including comma-separated split plates), registers them
    as assigned today in the daily log, and returns a clear summary message.
    """
    SENTINELS = {"SKIPPED", "NO_LORRY", "SPLIT", "", "nan", "none", "n/a", "-", None}

    # Collect all plates from LICENSE column (handles "BMN3682, WUD4927" too)
    plates_found = []
    rows_summary = []

    for _, row in prefilled.iterrows():
        lic_raw   = str(row.get("LICENSE", "")).strip()
        customer  = str(row.get("CUSTOMER NAME", "")).strip()
        do_num    = str(row.get("DO NUMBER", "")).strip()
        # Support both formats: WEIGHT(T) (old) or GROSS WEIGHT converted (new)
        weight    = row.get("WEIGHT(T)", row.get("GROSS WEIGHT", ""))
        if sess.get("is_new_format") and weight != "":
            try:
                weight = round(float(weight) / 1000, 3)
            except Exception:
                pass
        itmref    = str(row.get("ITMREF_0", "")).strip()

        # Split on comma to handle multi-lorry cells
        plates_in_cell = [p.strip().upper() for p in lic_raw.split(",")
                          if p.strip().upper() not in {s.upper() for s in SENTINELS if s}]
        if not plates_in_cell:
            continue

        plates_found.extend(plates_in_cell)
        plate_display = ", ".join(f"*{p}*" for p in plates_in_cell)
        itmref_str    = f" ({itmref})" if itmref and itmref.lower() not in ("nan","") else ""
        rows_summary.append(
            f"  🚛 {plate_display}  ←  {customer}{itmref_str}  {do_num}  {weight}T"
        )

    if not plates_found:
        # All LICENSE cells were empty despite column existing — fall through to auto-assign
        return None  # caller must handle None → proceed with normal auto-assign

    # Register in daily log
    record_assignments_today(plates_found, user=sess.get("user_id"))

    unique_plates = sorted(set(plates_found))
    lines = []
    lines.append(f"📋 *Pre-filled assignment detected!*")
    lines.append(f"Found *{len(unique_plates)} lorry plate(s)* in the uploaded file.")
    lines.append("Registered as assigned today:")
    for p in unique_plates:
        lines.append(f"  ⛔ *{p}*")
    lines.append("")
    lines.append("─────────────────────")
    lines.append("*Row details:*")
    lines.extend(rows_summary)
    lines.append("─────────────────────")
    lines.append(f"✅ These lorries are now marked *unavailable* for today's auto-assignment.")

    # Reset session — user may want to do a fresh assignment next
    reset_session(phone)

    return [
        "\n".join(lines),
        {
            "_type": "buttons",
            "body": "Start a new session or check what's assigned today.",
            "buttons": [
                {"id": "hi",                  "title": "👋 Hi"},
                {"id": "show assigned today",  "title": "📋 Show Assigned"},
            {"id": "show blocked",         "title": "🚫 Show Blocked"},
            ],
        }
    ]

def _handle_lorry_status_upload(phone, sess, df: "pd.DataFrame") -> list:
    """Handle a lorry-status update file.

    Accepted column names (case-insensitive):
      LORRY / PLATE / LICENSE  — plate number
      STATUS                   — "Available" or "Blocked" (or "block"/"avail")

    Reads each row and blocks or releases the lorry in today's log.
    Returns a reply message list.
    """
    # Normalise column names
    col_map = {c.upper(): c for c in df.columns}

    # Find LORRY column
    lorry_col = next((col_map[k] for k in ("LORRY", "PLATE", "LICENSE") if k in col_map), None)
    status_col = col_map.get("STATUS")

    if not lorry_col or not status_col:
        return None   # not a lorry-status file — caller should fall through

    engine = sess.get("engine")
    all_plates = set(engine.all_lorries["LORRY"].str.upper()) if engine else set()

    blocked_now  = []
    released_now = []
    unknown      = []

    for _, row in df.iterrows():
        plate  = str(row[lorry_col]).strip().upper()
        status = str(row[status_col]).strip().lower()
        if not plate or plate in ("NAN", "NONE", ""):
            continue
        if all_plates and plate not in all_plates:
            unknown.append(plate)
            continue

        if status.startswith("block"):
            record_assignments_today([plate], user=sess.get("user_id"))
            sess.setdefault("unavailable", set()).add(plate)
            blocked_now.append(plate)
        elif status.startswith("avail") or status in ("ok", "free", "available"):
            release_specific_plates([plate])
            sess.setdefault("unavailable", set()).discard(plate)
            released_now.append(plate)

    if not blocked_now and not released_now and not unknown:
        return None   # nothing actionable — fall through to DO-file handler

    lines = ["📋 *Lorry Status Updated from File*\n"]
    if blocked_now:
        lines.append(f"⛔ *Blocked ({len(blocked_now)}):* {', '.join(blocked_now)}")
    if released_now:
        lines.append(f"✅ *Released ({len(released_now)}):* {', '.join(released_now)}")
    if unknown:
        lines.append(f"⚠️ *Not found in master:* {', '.join(unknown)}")
    lines.append("\nLorry availability updated for today.")

    return [
        "\n".join(lines),
        {
            "_type": "buttons",
            "body": "What would you like to do next?",
            "buttons": [
                {"id": "hi",                  "title": "👋 Upload DOs"},
                {"id": "show assigned today",  "title": "📋 Show Blocked"},
                {"id": "manage lorry",         "title": "🔧 Manage Lorry"},
            ],
        }
    ]


# Slight-overload allowance for the rescue pass: a lorry may be filled up to
# this fraction of its rated TON when the only alternative is leaving a DO
# unassigned. 1.15 = up to 15% over. Mirrors the manual planner, who slightly
# overloads small lorries rather than dropping urban DOs.
SLIGHT_OVERLOAD = 1.20


def _is_urban_do(it) -> bool:
    """True if a DO is an urban delivery (Kuala Lumpur / Selangor).

    Uses the engine's own destination classifier so it matches the rest of the
    assignment logic: KV Klang-Valley routes count as urban (KL/SELANGOR),
    while semi-outstation KV codes (e.g. KV01A Tanjung Malim → MEDIUM_LONG) and
    all PH/NS/JH/PK routes count as outstation — regardless of a border row's
    raw STATE value.
    """
    try:
        grp = _classify_dest_group(str(it.get("ROUTE", "")),
                                   str(it.get("STATE", "")))
        return grp in _DEST_URBAN_GROUPS
    except Exception:
        st = str(it.get("STATE", "")).strip().upper()
        return ("KUALA LUMPUR" in st) or ("SELANGOR" in st)


def _overload_rescue(sess, max_over: float = SLIGHT_OVERLOAD):
    """Final safety net for URBAN DOs only.

    Slight overload is allowed strictly for Kuala Lumpur / Selangor routes, and
    only onto lorries that carry urban DOs (or are empty) — an urban DO is
    NEVER added to a lorry that already carries an outstation route, so urban
    and outstation are never mixed. Outstation DOs are left exactly as the main
    assignment placed them (no overload, no reshuffling). Per-DO size caps
    (MAX_TON) and forbidden plates are always respected. Only ever REDUCES the
    unassigned count.
    """
    import re as _re
    eng = sess.get("engine")
    if eng is None:
        return
    try:
        caps = {str(r["LORRY"]).strip().upper(): float(r["TON"])
                for _, r in eng.eligible_lorries.iterrows()}
    except Exception:
        return
    if not caps:
        return
    items = sess.get("items", []) or []

    def _pfx(it):
        m = _re.match(r'\s*([A-Z]{1,3}\d+[A-Z]?)', str(it.get("ROUTE", "")))
        return m.group(1) if m else ""

    load = {l: 0.0 for l in caps}
    has_outstation = {l: False for l in caps}   # lorry carries a non-urban DO
    lpts = {l: [] for l in caps}                 # urban (lat, lon, code) per lorry
    for it in items:
        l = it.get("LORRY")
        if l in caps:
            load[l] += it.get("WEIGHT", 0.0) or 0.0
            if not _is_urban_do(it):
                has_outstation[l] = True
            _la, _lo = it.get("GPS_LAT"), it.get("GPS_LON")
            if _la is not None and _lo is not None:
                lpts[l].append((_la, _lo, _pfx(it)))

    def _allowed(it, l):
        mt = it.get("MAX_TON")
        if mt is not None and caps[l] > mt + 1e-9:
            return False
        fp = it.get("FORBID_PLATES")
        if fp and l in fp:
            return False
        if l in _strict_route_excl(it.get("ROUTE", "")):
            return False
        return True

    def _near_ok(it, l):
        # Don't add an urban DO to a lorry that already carries a DIFFERENT
        # urban route code more than _URBAN_MERGE_SPREAD away — no far mixing.
        _la, _lo = it.get("GPS_LAT"), it.get("GPS_LON")
        if _la is None or _lo is None:
            return True
        _c = _pfx(it)
        for _pla, _plo, _pc in lpts[l]:
            if _pc != _c and ((_la - _pla) ** 2 + (_lo - _plo) ** 2) ** 0.5 > _URBAN_MERGE_SPREAD:
                return False
        return True

    def _place(it, l):
        it["LORRY"] = l
        sess["assigned"][it["DO NUMBER"]] = l
        load[l] += it.get("WEIGHT", 0.0) or 0.0
        _la, _lo = it.get("GPS_LAT"), it.get("GPS_LON")
        if _la is not None and _lo is not None:
            lpts[l].append((_la, _lo, _pfx(it)))

    # Only urban DOs are eligible for the slight-overload rescue.
    unplaced = [it for it in items
                if it.get("LORRY") in ("NO_LORRY", "NO_ELIGIBLE_LORRY")
                and _is_urban_do(it)]
    unplaced.sort(key=lambda i: -(i.get("WEIGHT", 0.0) or 0.0))

    rescued = 0
    for it in unplaced:
        w = it.get("WEIGHT", 0.0) or 0.0
        # Target only lorries that carry NO outstation route (urban-only or
        # empty) AND are geographically near this DO (no far urban mixing).
        cands = [l for l in caps
                 if _allowed(it, l) and not has_outstation[l] and _near_ok(it, l)]
        # Tier 1 — fits within capacity (tightest fit first).
        fit = [l for l in cands if load[l] + w <= caps[l] + 1e-9]
        if fit:
            _place(it, max(fit, key=lambda l: (load[l] + w) / caps[l]))
            rescued += 1
            continue
        # Tier 2 — slight overload of an urban lorry (smallest resulting ratio).
        ov = [l for l in cands if load[l] + w <= caps[l] * max_over + 1e-9]
        if ov:
            _place(it, min(ov, key=lambda l: (load[l] + w) / caps[l]))
            rescued += 1
            continue
        # No urban lorry can take it (even slightly overloaded). Leave it
        # unassigned rather than mix it onto an outstation lorry.

    if rescued:
        import logging as _rlog
        _rlog.info("[OVERLOAD-RESCUE] placed %d urban DO(s) via slight overload.", rescued)


def _urban_rebalance(sess, max_over: float = SLIGHT_OVERLOAD):
    """Free a big lorry stuck with a lone small URBAN load by repacking all the
    small-urban loads (plus that orphan) tightly onto the smaller lorries — like
    a human planner putting KV10A+KV11A+KV12A on one small van instead of
    leaving KV12A on a 15T truck.

    Fully atomic: the repack is applied ONLY if EVERY pooled DO places on the
    smaller lorries (within slight overload, size caps, forbidden plates, and
    the urban distance rule, keeping each route code whole). So it can never
    unassign a DO or create a far mix. One big lorry freed per call.
    """
    import math as _math
    import re as _re
    eng = sess.get("engine")
    if eng is None:
        return
    try:
        caps = {str(r["LORRY"]).strip().upper(): float(r["TON"])
                for _, r in eng.eligible_lorries.iterrows()}
    except Exception:
        return
    if not caps:
        return
    items = sess.get("items", []) or []

    def _pfx(it):
        m = _re.match(r'\s*([A-Z]{1,3}\d+[A-Z]?)', str(it.get("ROUTE", "")))
        return m.group(1) if m else ""

    from collections import defaultdict as _dd

    def _state():
        on = _dd(list); load = _dd(float); has_out = _dd(bool)
        for it in items:
            l = it.get("LORRY")
            if l in caps:
                on[l].append(it)
                load[l] += it.get("WEIGHT", 0.0) or 0.0
                if not _is_urban_do(it):
                    has_out[l] = True
        return on, load, has_out

    def _centroid(_its):
        la = [x.get("GPS_LAT") for x in _its if x.get("GPS_LAT") is not None]
        lo = [x.get("GPS_LON") for x in _its if x.get("GPS_LON") is not None]
        return (sum(la) / len(la), sum(lo) / len(lo)) if la else None

    on, load, has_out = _state()
    for L in sorted(caps, key=lambda x: -caps[x]):
        if load.get(L, 0.0) <= 1e-9 or has_out[L]:
            continue                                   # empty or outstation
        if load[L] > caps[L] * 0.5:
            continue                                   # not under-used
        L_items = list(on[L])
        L_w = load[L]
        L_cent = _centroid(L_items)
        L_codes = {_pfx(x) for x in L_items}
        L_mt = min((x["MAX_TON"] for x in L_items if x.get("MAX_TON") is not None), default=None)
        L_fb = set()
        for x in L_items:
            if x.get("FORBID_PLATES"):
                L_fb |= x["FORBID_PLATES"]
        # Move the whole small load onto a NEAR urban lorry that has room, so the
        # big lorry is freed. Prefer the tightest such target.
        L_strict_excl: set = set()
        for x in L_items:
            L_strict_excl |= _strict_route_excl(x.get("ROUTE", ""))
        cand = []
        for m in caps:
            if m == L or has_out[m] or load.get(m, 0.0) <= 1e-9:
                continue                               # only onto a used urban lorry
            if m in L_fb:
                continue
            if m in L_strict_excl:
                continue
            if L_mt is not None and caps[m] > L_mt + 1e-9:
                continue
            if load[m] + L_w > caps[m] * max_over + 1e-9:
                continue                               # no room even overloaded
            # No far mix: the MAX pairwise distance between any two different
            # urban route codes on the COMBINED lorry must stay within the
            # guard (centroid-to-centroid is too loose for a wide cluster).
            comb = [(x.get("GPS_LAT"), x.get("GPS_LON"), _pfx(x))
                    for x in (on[m] + L_items)
                    if x.get("GPS_LAT") is not None and _is_urban_do(x)]
            _far_mix = False
            for _a in range(len(comb)):
                for _b in range(_a + 1, len(comb)):
                    if comb[_a][2] != comb[_b][2] and _math.hypot(
                            comb[_a][0] - comb[_b][0], comb[_a][1] - comb[_b][1]) > _URBAN_MERGE_SPREAD:
                        _far_mix = True
                        break
                if _far_mix:
                    break
            if _far_mix:
                continue
            cand.append(m)
        if not cand:
            continue
        tgt = min(cand, key=lambda m: caps[m])         # tightest fit
        for x in L_items:
            x["LORRY"] = tgt
            sess["assigned"][x["DO NUMBER"]] = tgt
        import logging as _rlog
        _rlog.info("[URBAN-REBALANCE] moved %s's load to %s (freed the big lorry).", L, tgt)
        return                                         # one per call


def _downsize_lorries(sess):
    """Reduce wasted capacity: if a used lorry's whole load would fit on a
    SMALLER idle lorry, move it there and free the bigger one — so e.g. a lone
    10T load stops rolling out on a 22T truck when an ~11T truck is idle.

    Moves a lorry's entire DO group intact to an empty lorry, so directions are
    never mixed and nothing becomes unassigned. Per-DO size caps and forbidden
    plates are respected. Purely cosmetic w.r.t. what gets delivered — it only
    right-sizes the truck used.
    """
    eng = sess.get("engine")
    if eng is None:
        return
    try:
        caps = {str(r["LORRY"]).strip().upper(): float(r["TON"])
                for _, r in eng.eligible_lorries.iterrows()}
    except Exception:
        return
    if not caps:
        return
    items = sess.get("items", []) or []

    from collections import defaultdict as _dd
    on = _dd(list)
    for it in items:
        l = it.get("LORRY")
        if l in caps:
            on[l].append(it)
    load = {l: sum((x.get("WEIGHT", 0.0) or 0.0) for x in v) for l, v in on.items()}
    used = set(on.keys())

    def _fits_all(dos, l):
        for x in dos:
            mt = x.get("MAX_TON")
            if mt is not None and caps[l] > mt + 1e-9:
                return False
            fp = x.get("FORBID_PLATES")
            if fp and l in fp:
                return False
        return True

    moved = 0
    # Tackle the biggest (most wasteful) lorries first.
    for l in sorted(list(used), key=lambda x: -caps[x]):
        dos = on.get(l)
        if not dos:
            continue
        # BQY7823 priority (explicit request): never downsize it away from a
        # route it's the FIT IN LORRY-preferred pick for — it should keep
        # serving that outstation route rather than get bumped to a smaller
        # idle lorry for a cosmetic capacity-fit gain.
        if l == "BQY7823" and any(
                "BQY7823" in _preferred_lorries_for_route(x.get("ROUTE", ""), eng)
                for x in dos):
            continue
        # Kuantan priority pool (explicit request): never downsize VJN9910 or
        # BQU3875 away from a Kuantan (PH09) load — Kuantan is restricted to
        # VJN9910/BQY7823/BQU3875 first (see the preferred-lorry pool logic
        # above), so bumping one of them to a smaller idle lorry outside that
        # pool would violate it. BQY7823 is already covered by the check
        # above (Kuantan is one of its listed FIT IN LORRY routes).
        if l in ("VJN9910", "BQU3875") and any(
                _is_kuantan(x.get("ROUTE", ""), x.get("CUSTOMER NAME", ""))
                for x in dos):
            continue
        L = load[l]
        # Smallest idle lorry that (a) still holds the load within capacity and
        # (b) is genuinely smaller than the current lorry.
        idle = [m for m in caps
                if m not in used
                and caps[m] + 1e-9 >= L
                and caps[m] < caps[l] - 1e-9
                and _fits_all(dos, m)]
        if not idle:
            continue
        tgt = min(idle, key=lambda m: caps[m])
        for x in dos:
            x["LORRY"] = tgt
            sess["assigned"][x["DO NUMBER"]] = tgt
        used.discard(l)
        used.add(tgt)
        on[tgt] = dos
        on[l] = []
        load[tgt] = L
        load[l] = 0.0
        moved += 1

    if moved:
        import logging as _rlog
        _rlog.info("[DOWNSIZE] right-sized %d lorry load(s) onto smaller trucks.", moved)


def reassign_unassigned(sess, plates: list) -> dict:
    """Assign still-unassigned (NO_LORRY) DOs onto the given available lorry
    plates — best-fit-decreasing (heaviest DO first, tightest-fitting plate
    wins) — WITHOUT re-uploading, since the items are already in sess["items"].
    Enforces the same hard rules as normal assignment: eligible fleet, strict
    route exclusions, forbidden plates, capacity (+naik), outstation minimum
    tonnage, REMARKS/SHIP size cap, and destination-state compatibility.

    Returns a summary dict for the caller (web/whatsapp) to display.
    """
    engine = sess.get("engine")
    if engine is None:
        return {"error": "no_engine", "assigned": 0}
    items = sess.get("items", []) or []
    no_lorry = [it for it in items
                if it.get("LORRY") in ("NO_LORRY", "NO_ELIGIBLE_LORRY")]
    total = len(no_lorry)
    if not no_lorry:
        return {"assigned": 0, "total": 0, "still": [], "used": [], "skipped": []}

    # The user is telling us which lorries are NOW free — so accept ANY of their
    # own lorries (owner + SPARE), including ones that were Blocked at upload
    # time (they're overriding that). Only reject a lorry genuinely taken by
    # ANOTHER user today.
    _me = str(sess.get("user_id", "")).strip().upper()
    _fleet_df = getattr(engine, "all_lorries", engine.eligible_lorries)
    cap_map = {}
    for _, r in _fleet_df.iterrows():
        _u = str(r.get("USER", "")).strip().upper()
        if _u in (_me, "SPARE"):
            cap_map[str(r["LORRY"]).strip().upper()] = float(r["TON"])
    known = set(cap_map)
    given = [str(p).strip().upper() for p in (plates or []) if str(p).strip()]
    _taken_by_other = {p for p, u in get_assigned_by().items() if u != _me}
    avail = [p for p in given if p in known and p not in _taken_by_other]
    skipped = [p for p in given if p not in known or p in _taken_by_other]
    if not avail:
        return {"assigned": 0, "total": total, "still": [], "used": [],
                "skipped": skipped, "error": "no_available_plates"}

    load = {p: 0.0 for p in avail}
    states = {p: set() for p in avail}
    reasons = sess.setdefault("unassigned_reasons", {})
    done, still = [], []
    for it in sorted(no_lorry, key=lambda x: -(x.get("WEIGHT", 0.0) or 0.0)):
        w = it.get("WEIGHT", 0.0) or 0.0
        route = it.get("ROUTE", "")
        st = str(it.get("STATE", "")).strip().upper()
        dest = _classify_dest_group(route, st)
        strict = _strict_route_excl(route)
        forbid = it.get("FORBID_PLATES") or set()
        cands = []
        for p in avail:
            if p in strict or p in forbid:
                continue
            cap = cap_map.get(p, 0.0)
            if cap * NAIK_FACTOR - load[p] < w:
                continue
            if cap < _eff_dest_min_ton(route, dest, load[p] + w):
                continue
            if it.get("MAX_TON") is not None and cap > it["MAX_TON"]:
                continue
            if st and states[p] and not any(_states_compatible(st, s) for s in states[p]):
                continue
            cands.append((cap * NAIK_FACTOR - load[p], p))
        if not cands:
            still.append(it)
            reasons[it["DO NUMBER"]] = "NO_FIT_ON_GIVEN_PLATES"
            continue
        cands.sort()                      # tightest remaining fit first
        chosen = cands[0][1]
        it["LORRY"] = chosen
        load[chosen] += w
        if st:
            states[chosen].add(st)
        sess["assigned"][it["DO NUMBER"]] = chosen
        reasons.pop(it["DO NUMBER"], None)
        done.append(it)

    for do in sess.get("pending_dos", []):
        do["TOTAL_TON"] = round(sum(x["WEIGHT"] for x in do["ITEMS"]), 3)

    return {
        "assigned": len(done), "total": total,
        "used": sorted({it["LORRY"] for it in done}),
        "skipped": skipped,
        "still": [{"do": str(x["DO NUMBER"]), "route": x.get("ROUTE", ""),
                   "weight": round(x.get("WEIGHT", 0.0) or 0.0, 3)} for x in still],
    }


def assign_specific_dos(sess, plate: str, do_numbers: list) -> dict:
    """Manually assign a hand-picked list of still-unassigned DOs onto ONE
    user-named plate — the counterpart to reassign_unassigned() (which
    auto-bin-packs across multiple lorries, and stays owner-scoped). This
    manual box is a deliberate override: it searches the FULL fleet across
    every owner (not just this user's + SPARE) and ignores the master
    file's Blocked/Available status, on the theory that a human is
    consciously picking one plate and a handful of DOs, not an algorithm —
    so a lorry belonging to another user, or one flagged Blocked, is still
    offered as long as it isn't already full. "Already full" is judged
    against this session's own knowledge only (existing load already on
    that plate within sess["items"]); a plate another user is loading up
    concurrently in a separate session isn't visible here.
    An unrecognised plate (not in ANY owner's fleet at all) is always
    rejected — never accepted with a manually-typed capacity. Every
    requested DO must still be unassigned, the combined weight (existing
    load + new picks) must not exceed the plate's rated tonnage, and each
    DO must still pass the same hard rules as normal assignment (route
    reservations, forbidden plates, outstation minimum, size cap, state
    compatibility). If anything fails, NOTHING is assigned and the caller
    gets a clear reason so the user can adjust their selection and retry.
    """
    engine = sess.get("engine")
    if engine is None:
        return {"error": "no_engine", "message": "Session expired — please upload the DO file again."}

    plate = str(plate).strip().upper()
    _fleet_df = getattr(engine, "all_lorries", engine.eligible_lorries)
    cap = None
    for _, r in _fleet_df.iterrows():
        if str(r["LORRY"]).strip().upper() == plate:
            cap = float(r["TON"])
            break
    if cap is None:
        return {"error": "unknown_plate",
                "message": f"{plate} is not a known lorry. "
                           f"Check the spelling, or add it to the master lorry file first."}

    items = sess.get("items", []) or []
    by_do = {str(it.get("DO NUMBER")): it for it in items}
    wanted = [str(d).strip() for d in (do_numbers or []) if str(d).strip()]
    if not wanted:
        return {"error": "no_dos", "message": "No DOs selected."}

    selected, missing = [], []
    for dn in wanted:
        it = by_do.get(dn)
        if it is None or it.get("LORRY") not in ("NO_LORRY", "NO_ELIGIBLE_LORRY"):
            missing.append(dn)
        else:
            selected.append(it)
    if missing:
        return {"error": "not_unassigned",
                "message": f"These DOs are no longer unassigned: {', '.join(missing)}. Refresh and try again."}

    existing_items = [it for it in items if it.get("LORRY") == plate]
    existing_load = sum((it.get("WEIGHT", 0.0) or 0.0) for it in existing_items)
    existing_states = {str(it.get("STATE", "")).strip().upper()
                        for it in existing_items if it.get("STATE")}

    new_w = sum((it.get("WEIGHT", 0.0) or 0.0) for it in selected)
    total_w = existing_load + new_w
    if total_w > cap * NAIK_FACTOR + 1e-6:
        return {"error": "over_capacity",
                "message": f"Selected DOs total {new_w:.3f}T"
                           + (f" (plus {existing_load:.3f}T already on {plate})" if existing_load else "")
                           + f" — exceeds {plate}'s {cap:.3f}T capacity by {total_w - cap:.3f}T. "
                             f"Deselect some and try again."}

    states = set(existing_states)
    violations = []
    for it in selected:
        route = it.get("ROUTE", "")
        st = str(it.get("STATE", "")).strip().upper()
        dest = _classify_dest_group(route, st)
        if plate in _strict_route_excl(route):
            violations.append((it["DO NUMBER"], "PLATE_FORBIDDEN_FOR_ROUTE"))
            continue
        if it.get("FORBID_PLATES") and plate in it["FORBID_PLATES"]:
            violations.append((it["DO NUMBER"], "PLATE_FORBIDDEN"))
            continue
        if cap < _eff_dest_min_ton(route, dest, total_w):
            violations.append((it["DO NUMBER"], "OUTSTATION_NEEDS_LARGER_LORRY"))
            continue
        if it.get("MAX_TON") is not None and cap > it["MAX_TON"]:
            violations.append((it["DO NUMBER"], "SIZE_CAP_EXCEEDED"))
            continue
        if st and states and not any(_states_compatible(st, s) for s in states):
            violations.append((it["DO NUMBER"], "STATE_MISMATCH"))
            continue
        if st:
            states.add(st)
    if violations:
        return {"error": "rule_violation",
                "message": "Some selected DOs can't go on " + plate + ": " +
                           "; ".join(f"{d}({r})" for d, r in violations)}

    reasons = sess.setdefault("unassigned_reasons", {})
    for it in selected:
        it["LORRY"] = plate
        sess["assigned"][it["DO NUMBER"]] = plate
        reasons.pop(it["DO NUMBER"], None)

    for do in sess.get("pending_dos", []):
        do["TOTAL_TON"] = round(sum(x["WEIGHT"] for x in do["ITEMS"]), 3)

    if sess.get("raw_df") is not None:
        try:
            _export_result(sess)
        except Exception:
            pass

    return {"ok": True, "assigned": len(selected), "plate": plate, "weight": round(new_w, 3)}


def board_move(sess, do_number: str, plate) -> dict:
    """Board UI: drag-and-drop one DO onto a lorry (plate), or back to the
    unassigned pool (plate falsy). Never blocks on a rule violation — the
    human stays in control — but returns human-readable warnings from
    _check_manual_placement() for the UI to flag on the card. Keeps
    pending_dos totals in sync; the export file itself is rebuilt lazily
    at download time (see below) rather than on every move — _export_result
    re-reads and rewrites the entire multi-MB history file, which made every
    single drag or cancel on the board take 20-30 seconds for no benefit to
    the board itself (nothing in this session reads the history file again
    until the NEXT login)."""
    engine = sess.get("engine")
    items = sess.get("items", []) or []
    by_do = {str(it.get("DO NUMBER")): it for it in items}
    it = by_do.get(str(do_number).strip())
    if it is None:
        return {"error": "unknown_do", "message": f"DO {do_number} not found in this session."}

    plate = str(plate).strip().upper() if plate else ""
    warnings: list[str] = []
    if not plate:
        it["LORRY"] = "NO_LORRY"
        sess.setdefault("assigned", {}).pop(it["DO NUMBER"], None)
    else:
        _fleet_df = getattr(engine, "all_lorries", None) if engine is not None else None
        if _fleet_df is None or _fleet_df[_fleet_df["LORRY"] == plate].empty:
            return {"error": "unknown_plate",
                    "message": f"{plate} is not a known lorry. Check the spelling."}
        warnings = _check_manual_placement(it, plate, engine, sess)
        _user = sess.get("user_id")
        if _user and plate in get_unavailable_plates_for(_user):
            warnings.append(f"{plate} is toggled OFF today — placed anyway, but it's not available.")
        it["LORRY"] = plate
        sess.setdefault("assigned", {})[it["DO NUMBER"]] = plate
        sess.setdefault("unassigned_reasons", {}).pop(it["DO NUMBER"], None)

    for do in sess.get("pending_dos", []):
        do["TOTAL_TON"] = round(sum(x["WEIGHT"] for x in do["ITEMS"]), 3)
    # Mark the export stale rather than rebuilding it now — /api/download
    # already rebuilds on demand when export_bytes is missing.
    sess.pop("export_bytes", None)

    return {"ok": True, "do": it["DO NUMBER"], "lorry": it["LORRY"], "warnings": warnings}


def board_move_route(sess, route: str, plate) -> dict:
    """Board UI: drag an entire route group (every one of its currently
    UNASSIGNED DOs — same set the pool panel shows under that route) onto
    one lorry in a single action, instead of dragging each card one at a
    time. Reuses board_move() per item so the same validation/warning
    behaviour applies to each — never blocks, just flags."""
    engine = sess.get("engine")
    items = sess.get("items", []) or []
    plate = str(plate).strip().upper() if plate else ""
    if not plate:
        return {"error": "missing_plate", "message": "Pick a lorry to drop the route onto."}

    _fleet_df = getattr(engine, "all_lorries", None) if engine is not None else None
    if _fleet_df is None or _fleet_df[_fleet_df["LORRY"] == plate].empty:
        return {"error": "unknown_plate",
                "message": f"{plate} is not a known lorry. Check the spelling."}

    _known_plates = set()
    if _fleet_df is not None:
        _known_plates = {str(p).strip().upper() for p in _fleet_df["LORRY"]}
    _do_numbers = [
        it["DO NUMBER"] for it in items
        if str(it.get("ROUTE", "")) == route
        and str(it.get("LORRY") or "").strip().upper() not in _known_plates
    ]
    if not _do_numbers:
        return {"error": "nothing_to_move", "message": f"No unassigned DOs found for route {route}."}

    moved = 0
    all_warnings: list[str] = []
    for do_num in _do_numbers:
        outcome = board_move(sess, do_num, plate)
        if outcome.get("ok"):
            moved += 1
            all_warnings.extend(outcome.get("warnings", []))
    return {"ok": True, "moved": moved, "total": len(_do_numbers), "warnings": all_warnings}


def _handle_excel_upload(phone, sess, file_bytes):
    try:
        # Always assign against the LATEST toggle state — not just whatever
        # this session's eligible_lorries happened to hold last. Covers a
        # shared SPARE plate the other planner claimed (or released) after
        # this session was already set up.
        refresh_eligible_from_toggle(sess)
        # Keep the raw upload so "assign off-schedule DOs" (YES) can re-run the
        # full assignment with the schedule filter off.
        if file_bytes and not sess.get("_ignore_schedule"):
            sess["_upload_bytes"] = file_bytes
        df = pd.read_excel(io.BytesIO(file_bytes))
        # Some exports (e.g. the ZSDOROUTEWRH .xls) put a title/date row above the
        # real header, so row 0 is not the column names. Detect the header row
        # (the one containing DO NUMBER / ROUTE) and re-read from there.
        _cols0 = [str(c).strip().upper() for c in df.columns]
        if not any(_k in _cols0 for _k in
                   ("DO NUMBER", "ROUTE", "GROSS WEIGHT", "WEIGHT(T)")):
            _raw = pd.read_excel(io.BytesIO(file_bytes), header=None)
            _hdr = None
            for _i in range(min(10, len(_raw))):
                _rv = [str(x).strip().upper() for x in _raw.iloc[_i].tolist()]
                if "DO NUMBER" in _rv or ("ROUTE" in _rv and "CUSTOMER NAME" in _rv):
                    _hdr = _i
                    break
            if _hdr is not None:
                df = pd.read_excel(io.BytesIO(file_bytes), header=_hdr)
        df.columns = [str(c).strip().upper() for c in df.columns]

        # Remember the EXACT columns (and order) of the uploaded file so the
        # export can be returned with the same layout — no added TRIP/DEST_STATE
        # or internal WEIGHT(T), and INVOICE DATE etc. keep their position.
        sess["_orig_cols"] = list(df.columns)

        # ── Detect lorry-status file (LORRY + STATUS columns) ───────────────
        # Must be checked BEFORE DO-file detection so a master-lorry upload
        # with Status column doesn't accidentally trigger DO assignment flow.
        _lorry_status_result = _handle_lorry_status_upload(phone, sess, df)
        if _lorry_status_result is not None:
            return _lorry_status_result

        # ── Detect format: new (ZSDOROUTEWRH) vs old ────────────────────────
        # New format: GROSS WEIGHT (kg), no WEIGHT(T) or ITMREF_0
        # Old format: WEIGHT(T), ITMREF_0
        IS_NEW_FORMAT = "GROSS WEIGHT" in df.columns and "WEIGHT(T)" not in df.columns

        if IS_NEW_FORMAT:
            # New format required columns
            required = {"DO NUMBER", "ROUTE", "CUSTOMER NAME", "GROSS WEIGHT"}
            missing = required - set(df.columns)
            if missing:
                return [f"❌ Missing columns: {', '.join(missing)}\nPlease check and re-upload."]
            # Convert GROSS WEIGHT kg → tonnes, store as WEIGHT(T) internally
            df["WEIGHT(T)"] = pd.to_numeric(df["GROSS WEIGHT"], errors="coerce").fillna(0) / 1000.0
            if "CODE" not in df.columns:
                df["CODE"] = ""
            # DATE arrives as "2026-05-11 00:00:00" (string of a datetime serial).
            # Strip the time part so it exports as "2026-05-11".
            if "DATE" in df.columns:
                df["DATE"] = (
                    df["DATE"]
                    .astype(str)
                    .str.strip()
                    .str.split(" ").str[0]   # keep only "2026-05-11"
                    .replace({"nan": "", "NaT": "", "None": ""})
                )
            # Convert DATE to formatted string NOW so raw_df never stores datetime64.
            # pandas to_excel will re-serialise datetime64 as a date cell regardless
            # of any later string assignment — converting at source is the only safe fix.
            if "DATE" in df.columns:
                def _fmt_date_on_load(v):
                    if not isinstance(v, str) and pd.isna(v):
                        return ""
                    try:
                        ts = pd.to_datetime(v, errors="coerce")
                        if pd.isna(ts):
                            return str(v)
                        return ts.strftime("%-d/%-m/%y")
                    except Exception:
                        return str(v)
                df["DATE"] = df["DATE"].apply(_fmt_date_on_load).astype(str)
        else:
            # Old format required columns
            required = {"CODE", "CUSTOMER NAME", "ROUTE", "DO NUMBER", "WEIGHT(T)"}
            missing = required - set(df.columns)
            if missing:
                return [f"❌ Missing columns: {', '.join(missing)}\nPlease check and re-upload."]

        if "LICENSE" not in df.columns:
            df["LICENSE"] = ""

        raw = df.dropna(subset=["ROUTE", "DO NUMBER"]).copy()
        raw = raw.reset_index(drop=True)
        # Store format flag so export knows not to touch DATE
        sess["is_new_format"] = IS_NEW_FORMAT

        # ── Pre-filled detection ─────────────────────────────────────────────
        # If the uploaded Excel already has LICENSE plates in it:
        #   Case A (fully prefilled): register plates as assigned today, show summary, stop.
        #   Case B (partial — some rows blank): track capacity in session but do NOT
        #     register plates as "assigned today" so they remain eligible for TRIP 2
        #     on the blank rows. Continue to auto-assign the blank rows.
        SENTINELS_STR = {"", "nan", "none", "n/a", "-"}
        _lic_lower = raw["LICENSE"].fillna("").astype(str).str.strip().str.lower()
        prefilled_rows = raw[_lic_lower.isin(SENTINELS_STR) == False].copy()
        _blank_rows = raw[_lic_lower.isin(SENTINELS_STR)].copy()

        if not prefilled_rows.empty:
            if _blank_rows.empty:
                # Case A: fully prefilled — register, summarise, and stop.
                result = _handle_prefilled_excel(phone, sess, raw, prefilled_rows)
                if result is not None:
                    return result
            else:
                # Case B: partial file — skip record_assignments_today so those
                # lorries stay available for the blank rows (as TRIP 2 if needed).
                # Just note how many were pre-filled so the bot can mention it.
                sess["_prefilled_count"] = len(prefilled_rows)
            # result is None → all plates were sentinels, fall through to auto-assign

        # ── Auto-free this user's lorries before assigning ────────────────────
        # Every DO-file upload is a fresh planning run: release the user's own
        # plates from today's "assigned today" log so ALL their lorries are
        # available for assignment.  Re-uploading a result file (or running
        # twice in a day) must not leave lorries wrongly blocked — capacity is
        # still tracked accurately from the file's pre-filled loads (Case B),
        # so this cannot double-book a lorry beyond its tonnage.  Other users'
        # plates are preserved.
        _eng_for_clear = sess.get("engine")
        if _eng_for_clear is not None:
            try:
                clear_daily_log_for_user(_eng_for_clear)
                sess["unavailable"] = set()
            except Exception:
                pass

        # ── Build item list: one item per Excel row ─────────────────────────
        # Each row is an independent item that needs its own lorry.
        # Items with the same DO NUMBER belong to the same customer/route
        # but may end up on different lorries (e.g. 17.5T row vs 3.5T row).

        # Route-code filtering: only assign rows whose route prefix belongs to
        # the logged-in user.  Rows for other users are kept in items (so they
        # appear in the export) but pre-marked as OTHER_USER so they get a
        # blank LICENSE in the exported file.
        _user_prefixes = _load_user_route_prefixes(sess.get("user_id", ""))

        # ── Schedule filter ──────────────────────────────────────────────────
        # User explicitly chose "Today" or "Tomorrow" at login.
        # Routes not on the chosen day's schedule are left blank (NOT_TODAY).
        # When the user has already said YES to "assign off-schedule DOs too",
        # the whole assignment is re-run with the schedule filter OFF so every
        # rule (reservation, geo, same-destination, corridor) applies uniformly
        # to ALL routes — not just a simplified per-route pass for one route.
        _trip_day       = sess.get("trip_day", "today")
        if sess.get("_ignore_schedule"):
            _sched_prefixes = None
        else:
            _sched_prefixes = _scheduled_prefixes_for_upload(sess.get("user_id", ""), trip_day=_trip_day)

        # Batch-parse all unique REMARKS via Claude (one API call, disk-cached).
        # Falls back silently to the keyword regex when no API key is set.
        if "REMARKS" in raw.columns:
            _llm_parse_remarks_batch(
                [str(v) for v in raw["REMARKS"].dropna().unique()]
            )

        items = []
        _other_user_count  = 0
        _not_today_count   = 0
        _past_date_count   = 0
        _wrong_trip_count  = 0
        _today_date = datetime.now().date()
        _trip_session = sess.get("trip_session")   # "MORNING" / "AFTERNOON" / None (any)
        # Lorry tonnage lookup for enforcing REMARKS size caps on pre-filled rows.
        _prefill_cap_map: dict[str, float] = {}
        _eng_pf = sess.get("engine")
        if _eng_pf is not None:
            try:
                _prefill_cap_map = {
                    str(r["LORRY"]).strip().upper(): float(r["TON"])
                    for _, r in _eng_pf.eligible_lorries.iterrows()
                }
            except Exception:
                _prefill_cap_map = {}
        for idx, row in raw.iterrows():
            route_str = str(row["ROUTE"]).strip()
            pfx = _extract_route_prefix(route_str)
            _is_mine     = True
            _is_today    = True

            if _user_prefixes and pfx and pfx not in _user_prefixes:
                _is_mine = False
                _other_user_count += 1

            # Schedule check: route must be on today's SCHD sheet.
            if _is_mine and _sched_prefixes is not None:
                if pfx not in _sched_prefixes:
                    _is_today = False
                    _not_today_count += 1

            # Date check: a DO whose own DATE has already passed is stale —
            # never auto-assign it (no opt-in prompt either, unlike the
            # day-of-week schedule check above, which the user can still say
            # yes to). Independent of _sched_prefixes/_ignore_schedule so it
            # still applies on the "assign off-schedule DOs too?" re-run.
            _is_past_date = False
            if _is_mine:
                _row_dt = pd.to_datetime(str(row.get("DATE", "")).strip(),
                                          dayfirst=True, errors="coerce")
                if pd.notna(_row_dt) and _row_dt.date() < _today_date:
                    _is_past_date = True

            _remarks_cell = row.get("REMARKS", "")
            _remarks_raw = "" if pd.isna(_remarks_cell) else str(_remarks_cell).strip()

            # Trip-session filter: only excludes a DO when the user picked a
            # specific half-day (Morning/Afternoon) AND the REMARKS explicitly
            # say the OTHER half. No trip-timing note in REMARKS → assign
            # either way (not restricted).
            _is_wrong_trip = False
            if _is_mine and _trip_session:
                _row_session = _row_trip_session(_remarks_raw)
                if _row_session and _row_session != _trip_session:
                    _is_wrong_trip = True

            # ── SHIP_DETAIL column (new file format) ──────────────────────────
            # Format: "<days>, [AM|PM], MAX <N> TON".  The SHIP_DETAIL "MAX N TON"
            # is AUTHORITATIVE and OVERRIDES the REMARKS size phrase: e.g. a DO
            # whose REMARKS say "SMALL LORRY" (≤5T) but whose SHIP_DETAIL says
            # "MAX 15 TON" may ride up to a 15T lorry (so it can merge with other
            # 15T-capable DOs). If SHIP_DETAIL has no MAX, the REMARKS cap applies.
            # (Day / AM-PM parts are ignored, per request.)
            # Rules:
            #  • SHIP_DETAIL "MAX 2 TON" is always enforced (van), regardless.
            #  • If REMARKS impose a size cap (e.g. SMALL LORRY ≤5T) AND
            #    SHIP_DETAIL says "MAX N TON", the SHIP value OVERRIDES/LIFTS the
            #    REMARKS cap (SMALL LORRY + MAX 15 TON → may ride up to 15T).
            #  • A SHIP_DETAIL MAX with NO REMARKS restriction is ignored (a
            #    plain "MAX 15 TON" does not newly cap an otherwise-free DO — so
            #    Kuantan can still use a 21T lorry).
            _ship_raw = str(row.get("SHIP_DETAIL", "")).strip()
            _cap_remarks = _remarks_lorry_cap(_remarks_raw)
            _cap_ship    = _remarks_lorry_cap(_ship_raw)   # MAX N TON from SHIP_DETAIL
            if _cap_ship == 2.0:
                _size_cap = 2.0                          # MAX 2 TON → van (always)
            elif _cap_remarks is not None and _cap_ship is not None:
                _size_cap = _cap_ship                    # SHIP MAX lifts REMARKS cap
            else:
                _size_cap = _cap_remarks                 # REMARKS cap, or no cap

            # Remarks-day filter DISABLED by request: assignment now depends only
            # on the trip day the user selected at login (today/tomorrow) plus the
            # SCHD schedule check above. Delivery-day hints in REMARKS (e.g.
            # "SABTU", "JANGAN HANTAR SELASA", "KEDAI TUTUP RABU") no longer block
            # assignment — the user decides the day.

            # SHIP_DETAIL "OUT SOURCE" → this DO is handled by a third party and
            # must NOT be assigned a lorry (LICENSE left blank, not counted as
            # unassigned). Takes priority over everything else.
            _is_outsource = "OUT SOURCE" in _ship_raw.upper() \
                or "OUTSOURCE" in _ship_raw.upper()

            _lorry_init = None
            if _is_outsource:
                _lorry_init = "OUT_SOURCE"
            elif not _is_mine:
                _lorry_init = "OTHER_USER"
            elif _is_past_date:
                _lorry_init = "PAST_DATE"
                _past_date_count += 1
            elif not _is_today:
                _lorry_init = "NOT_TODAY"
            elif _is_wrong_trip:
                _lorry_init = "WRONG_TRIP"
                _wrong_trip_count += 1
            elif sess.get("_prefilled_count"):
                # Case B: pre-filled file — rows that already have a valid lorry
                # plate in LICENSE keep that assignment so session_loads is accurate.
                _existing_lic = str(row.get("LICENSE", "")).strip()
                _lic_key = _existing_lic.lower()
                if _lic_key and _lic_key not in {"", "nan", "none", "n/a", "-"}:
                    _plate_up = _existing_lic.upper()
                    _cap_pf = _size_cap
                    _plate_ton_pf = _prefill_cap_map.get(_plate_up)
                    if _plate_up in get_assigned_today() or _plate_up in get_broken_lorries():
                        # The pre-filled lorry was already used on an earlier run
                        # today (e.g. a SPARE lorry taken by the ABI run before
                        # this VIVIAN run) — it is BLOCKED. Drop the pre-fill so
                        # this DO is reassigned to a still-available lorry.
                        _lorry_init = None
                    elif _plate_up not in _prefill_cap_map:
                        # Pre-filled plate is not in THIS user's eligible fleet
                        # (owner isolation) — drop so it is reassigned.
                        _lorry_init = None
                    # Enforce the REMARKS size cap even on pre-filled rows: a DO
                    # whose remark demands a small lorry (e.g. "small lorry" → ≤2T)
                    # must NOT keep a pre-set oversized lorry.
                    elif (_cap_pf is not None and _plate_ton_pf is not None
                            and _plate_ton_pf > _cap_pf):
                        _lorry_init = None
                    else:
                        _lorry_init = _plate_up

            # REMARKS forbidding a specific plate by number ("lorry 3875 tak boleh
            # masuk" → BQU3875 excluded for this DO). Size words ("besar") are not
            # plates and are handled by the size-cap rules instead.
            _forbid_plates = _remarks_forbidden_plates(_remarks_raw, _prefill_cap_map.keys())
            # If the pre-filled plate is itself forbidden, drop the pre-fill so it
            # gets reassigned to a lorry that IS allowed to enter.
            if _lorry_init and _lorry_init in _forbid_plates:
                _lorry_init = None

            # Parse GPS coordinates from LONGITUD column (format: "lat lon")
            _gps_lat, _gps_lon = None, None
            _loc_raw = str(row.get("LONGITUD", "")).strip()
            if _loc_raw and _loc_raw.lower() not in ("nan", "none", ""):
                _loc_parts = _loc_raw.replace(",", " ").split()
                if len(_loc_parts) >= 2:
                    try:
                        _gps_lat = float(_loc_parts[0])
                        _gps_lon = float(_loc_parts[1])
                    except ValueError:
                        pass

            items.append({
                "ROW_IDX":       idx,
                "DO NUMBER":     str(row["DO NUMBER"]).strip(),
                "CUSTOMER NAME": str(row["CUSTOMER NAME"]).strip(),
                "ROUTE":         route_str,
                "CODE":          str(row["CODE"]).strip(),
                "WEIGHT":        float(row["WEIGHT(T)"]),
                "ITMREF":        str(row.get("ITMREF_0", "")).strip(),
                "DATE":          str(row.get("DATE", "")).strip(),
                "ETD":           str(row.get("ETD", "")).strip(),
                "STATE":         _state_from_row(row),   # destination state from file
                "CITY":          str(row.get("CITY", "")).strip(),
                "REMARKS":       _remarks_raw,
                "GPS_LAT":       _gps_lat,
                "GPS_LON":       _gps_lon,
                "LORRY":         _lorry_init,
                "SPLIT_LORRIES": None,
                # Max lorry tonnage allowed by this DO's REMARKS and/or SHIP_DETAIL
                # ("MAX N TON") — tightest of the two (None = any size).
                "MAX_TON":       _size_cap,
                # Specific plates this DO forbids ("lorry 3875 tak boleh masuk").
                "FORBID_PLATES": _forbid_plates,
            })

        # ── Infer missing GPS from same-city (then same-postcode) neighbours ──
        # Some DOs arrive with a blank LONGITUD. Without coordinates a DO can't
        # be geo-separated and may wrongly ride a far lorry (e.g. a Semenyih DO
        # with no GPS landing on a Rawang lorry). Fill each missing coordinate
        # from the mean GPS of other DOs in this SAME upload that share its CITY
        # (fallback: its POSCODE) and do have GPS — data-driven, no hardcoding.
        def _norm_key(_it, _field):
            return str(_it.get(_field, "")).strip().upper()
        _city_gps: dict = {}
        _pc_gps: dict = {}
        for _it in items:
            if _it.get("GPS_LAT") is not None and _it.get("GPS_LON") is not None:
                _city_gps.setdefault(_norm_key(_it, "CITY"), []).append(
                    (_it["GPS_LAT"], _it["GPS_LON"]))
                _pc_gps.setdefault(_norm_key(_it, "POSCODE"), []).append(
                    (_it["GPS_LAT"], _it["GPS_LON"]))
        for _it in items:
            if _it.get("GPS_LAT") is None or _it.get("GPS_LON") is None:
                _pool = (_city_gps.get(_norm_key(_it, "CITY"))
                         or _pc_gps.get(_norm_key(_it, "POSCODE")))
                if _pool:
                    _it["GPS_LAT"] = sum(p[0] for p in _pool) / len(_pool)
                    _it["GPS_LON"] = sum(p[1] for p in _pool) / len(_pool)
                    _it["GPS_INFERRED"] = True

        # Sort eligible items by DATE ascending so oldest pending DOs are
        # assigned lorries first. Items already settled (NOT_TODAY/OTHER_USER/
        # pre-filled plate) sort last so they don't displace pending DOs.
        def _early_date_key(it):
            lorry = it.get("LORRY")
            if lorry not in (None, "NO_LORRY"):
                return "9999-99-99"
            raw_d = str(it.get("DATE", "")).strip()
            try:
                _ts = pd.to_datetime(raw_d, dayfirst=True, errors="coerce", format="mixed")
                if pd.notna(_ts):
                    return _ts.strftime("%Y-%m-%d")
            except Exception:
                pass
            return "9999-12-31"
        items.sort(key=_early_date_key)

        # Build schedule notice for the user
        _sched_notice = []
        if _sched_prefixes is not None:
            from datetime import timedelta as _td
            _day_names = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
            if _trip_day == "today":
                _tgt_date = datetime.now().date()
            else:
                _tgt_date = datetime.now().date() + _td(days=1)
                while _tgt_date.weekday() == 6:
                    _tgt_date += _td(days=1)
            _tgt_wd = _tgt_date.weekday()
            _sched_notice.append(
                f"📅 Schedule filter — assigning *{_trip_day}* ({_day_names[_tgt_wd]}). "
                f"Routes: {', '.join(sorted(_sched_prefixes)) if _sched_prefixes else 'none scheduled'}."
            )
            if _not_today_count:
                _sched_notice.append(
                    f"⏭ *{_not_today_count} DO(s)* not on {_day_names[_tgt_wd]}'s route list — left unassigned."
                )

        sess["items"]      = items          # row-level item list
        sess["raw_df"]     = raw
        # Seed unavailable with lorries already committed TODAY (e.g. a SPARE
        # lorry used on an earlier ABI run before this VIVIAN run) plus broken
        # lorries, so the core assignment never reassigns a blocked lorry.
        sess["unavailable"] = set(get_assigned_today()) | set(get_broken_lorries())
        sess["assigned"]   = {}             # kept for change/block compat (item ROW_IDX → lorry)
        sess["state"]      = "CONFIRMING"

        engine: LorryEngine = sess["engine"]

        # ── Auto-assign: two-pass global optimiser ─────────────────────────
        # PROBLEM with naive heaviest-first:
        #   Small lorries (e.g. VEA2818 1.07T) get consumed by slightly-heavier
        #   tiny DOs before even-lighter ones are processed, causing the lightest
        #   DO to get a much-too-large lorry.
        #
        # SOLUTION — two passes:
        #   Pass 1: loads ABOVE the smallest lorry capacity → heaviest first
        #           (these need large lorries; process early to claim them)
        #   Pass 2: loads AT OR BELOW the smallest lorry capacity → LIGHTEST first
        #           (tiny loads get the smallest available lorry — no waste)
        #
        # Within each pass, route frequency still acts as a tiebreaker.

        broken_map = get_broken_lorries()
        sess["unavailable"].update(broken_map.keys())

        eligible_caps   = sorted(engine.eligible_lorries["TON"].tolist())
        smallest_cap    = eligible_caps[0] if eligible_caps else 1.0

        # ── Route grouping with geographic corridor merging ────────────────────
        # Step 1: one item-list per exact route code.
        # Step 2: build "corridor super-groups" — all routes that travel in the
        #   same cluster+corridor direction AND pass through a shared geographic
        #   waypoint (or have no parseable waypoints, in which case corridor
        #   match alone is sufficient).
        # Step 3: if a super-group is too heavy for the largest lorry, bin-pack
        #   it into sub-groups (heaviest routes first).
        # No "heaviest stays alone" rule — every route that goes the same way
        # should share a lorry; capacity is the only hard limit.
        from collections import defaultdict
        from lorry_engine import (
            _extract_route_intelligence,
            _routes_on_same_way,
            _route_centroid,
            _haversine_km,
            _bearing_deg,
            _bearing_diff,
            _DEPOT,
            can_share_cross_cluster,
            MAX_STOPS_PER_LORRY as _MAX_STOPS,
            MIN_UTIL_TO_ASSIGN  as _MIN_UTIL,
        )

        # Step 1 — bucket by (ROUTE, STATE, CITY) for urban routes so items
        # in the same route code but different states or distant cities don't
        # share a lorry unnecessarily.
        # e.g. KV11A-KL-KUALA LUMPUR vs KV11A-SELANGOR-AMPANG are separate.
        # Outstation routes (LARGE_LONG / MEDIUM_LONG) are NOT split by city —
        # a single lorry handles all stops on a Kuantan or Seremban run.
        route_buckets: dict[str, list] = defaultdict(list)
        # Session loads to be populated — pre-filled items (Case B) seed it now
        # so that blank-row assignment knows lorry capacities are partially used.
        _prefill_loads: dict[str, float] = {}
        _prefill_routes: dict[str, str]  = {}
        _prefill_states: dict[str, set]  = {}   # plate → set of destination states
        SKIP_SENTINELS = {"OTHER_USER", "NOT_TODAY", "REMARKS_SKIP", "OUT_SOURCE", "PAST_DATE", "WRONG_TRIP", "NO_LORRY",
                          "SPLIT", "SKIPPED", None, ""}

        for it in items:
            lorry = it.get("LORRY")
            if lorry and lorry not in SKIP_SENTINELS:
                # Pre-filled item — record it as already assigned
                if it.get("LORRY") not in (None,):
                    plate = lorry.strip().upper()
                    if re.match(r'^[A-Z]{1,3}\d', plate):  # looks like a lorry plate
                        _prefill_loads[plate] = _prefill_loads.get(plate, 0) + it["WEIGHT"]
                        if plate not in _prefill_routes:
                            _prefill_routes[plate] = it["ROUTE"]
                        _st = it.get("STATE", "").strip().upper()
                        if _st:
                            _prefill_states.setdefault(plate, set()).add(_st)
                        sess["assigned"][it["DO NUMBER"]] = plate

        def _bearing_octant(lat: float, lon: float) -> str:
            """Map a GPS coordinate to a coarse 8-direction sector from depot.
            Used to sub-bucket DOs geographically so items going in very
            different directions (e.g. north vs south) never share a lorry.
            """
            from lorry_engine import _bearing_deg as _bd, _DEPOT as _DP
            dist_km = ((lat - _DP[0])**2 + (lon - _DP[1])**2) ** 0.5 * 111
            if dist_km < 8.0:
                return "LOCAL"   # too close to depot to have a meaningful direction
            b = _bd(_DP[0], _DP[1], lat, lon)
            # Divide 360° into 8 sectors, each 45° wide, named by compass point
            dirs = ["N", "NE", "E", "SE", "S", "SW", "W", "NW"]
            idx = int((b + 22.5) / 45) % 8
            return dirs[idx]

        for it in items:
            lorry = it.get("LORRY")
            if lorry in ("OTHER_USER", "NOT_TODAY", "REMARKS_SKIP", "OUT_SOURCE", "PAST_DATE", "WRONG_TRIP"):
                continue
            # Skip pre-filled items — they're already in sess["assigned"]
            if lorry and lorry not in (None,) and re.match(r'^[A-Z]{1,3}\d', lorry.strip().upper()):
                continue
            _rt_key = it["ROUTE"].strip().upper()
            # Special-customer PH09 rides the Kemaman run — bucket it apart from
            # ordinary Kuantan so it can group with TR02 instead.
            if _kemaman_ph09(it.get("ROUTE", ""), it.get("CUSTOMER NAME", "")):
                _rt_key = _rt_key + "@KMN"
            _st_key = it.get("STATE", "")
            _ct_key = it.get("CITY", "").strip().upper()
            dest_grp = _classify_dest_group(it["ROUTE"].strip().upper(), _st_key)

            # Geographic sub-bucketing — applies to ALL routes.
            # Primary key: state (identifies destination region).
            # Secondary key:
            #   • Urban routes  → city (tight street-level grouping)
            #   • Outstation    → GPS bearing octant from depot (N/NE/E/…/NW)
            #     This prevents two DOs with the same route code but pointing in
            #     genuinely opposite directions (e.g. north vs south) from sharing
            #     a lorry, which is the root cause of wrong mixed-direction runs.
            _sub_key = _st_key
            if dest_grp in _DEST_URBAN_GROUPS:
                # Urban: split by state + city
                if _ct_key and _ct_key not in ("NAN", "NONE", ""):
                    _sub_key = f"{_st_key}|{_ct_key}"
            else:
                # Outstation: split by state + GPS bearing octant when available
                _lat = it.get("GPS_LAT")
                _lon = it.get("GPS_LON")
                if _lat is not None and _lon is not None:
                    _octant = _bearing_octant(_lat, _lon)
                    if _octant != "LOCAL":
                        _sub_key = f"{_st_key}|{_octant}"
                    else:
                        # Very close to depot — use city as fallback sub-key
                        if _ct_key and _ct_key not in ("NAN", "NONE", ""):
                            _sub_key = f"{_st_key}|{_ct_key}"
                else:
                    # No GPS — fall back to state + city
                    if _ct_key and _ct_key not in ("NAN", "NONE", ""):
                        _sub_key = f"{_st_key}|{_ct_key}"

            if _sub_key:
                _rt_key = f"{_rt_key}||{_sub_key}"
            route_buckets[_rt_key].append(it)

        # Step 1.5 — Collapse same-route-prefix sub-buckets into one bucket.
        # The geographic sub-bucketing above (state|city or state|octant) can split
        # a single route like KV11A into several keys:
        #   KV11A - PUDU…||SELANGOR|CHERAS
        #   KV11A - PUDU…||WP|KUALA LUMPUR
        # These must ride the same lorry (same route = same physical road).
        # Pre-merge them here so the cross-route merge step (Step 2) sees one
        # unified bucket per route, not scattered fragments.
        _pfx_to_keys: dict[str, list] = {}
        for _bk in list(route_buckets.keys()):
            _bare_rt = _bk.split("||")[0]
            _pfx = _extract_route_prefix(_bare_rt)
            # Keep special-customer PH09 (@KMN) in its own prefix group so it is
            # never re-merged with ordinary Kuantan.
            if "@KMN" in _bare_rt:
                _pfx = (_pfx or "") + "@KMN"
            if _pfx:
                _pfx_to_keys.setdefault(_pfx, []).append(_bk)
        for _pfx, _bkeys in _pfx_to_keys.items():
            if len(_bkeys) <= 1:
                continue
            # Merge all sub-buckets for this prefix into the first key
            _primary = _bkeys[0]
            for _other in _bkeys[1:]:
                route_buckets[_primary].extend(route_buckets.pop(_other))

        # Step 2 — cluster same-way buckets into corridor super-groups
        # Each super-group is a list of route-bucket lists.
        max_lorry_cap = float(engine.eligible_lorries["TON"].max()) \
            if not engine.eligible_lorries.empty else 99.0

        # Build route→state map for state-mismatch guards in cross-cluster merge
        _route_state: dict[str, str] = {}
        for _rt_key, _bucket in route_buckets.items():
            _st = _bucket[0].get("STATE", "") if _bucket else ""
            if _st:
                _route_state[_rt_key.strip().upper()] = str(_st).strip().upper()

        # Build live GPS centroids from per-item GPS_LAT/GPS_LON coordinates.
        # Compute the TRUE average centroid for each bucket key (route||state|octant)
        # so that bearing checks use actual delivery coordinates, not a single sample.
        _live_centroids: dict[str, tuple] = {}
        _cent_acc: dict[str, list] = {}   # key → list of (lat, lon)
        for _bk, _bkt in route_buckets.items():
            for _it in _bkt:
                _la = _it.get("GPS_LAT")
                _lo = _it.get("GPS_LON")
                if _la is not None and _lo is not None:
                    _cent_acc.setdefault(_bk, []).append((_la, _lo))
        for _bk, _pts in _cent_acc.items():
            _avg_lat = sum(p[0] for p in _pts) / len(_pts)
            _avg_lon = sum(p[1] for p in _pts) / len(_pts)
            _live_centroids[_bk] = (_avg_lat, _avg_lon)
            # Also index by the bare route key (without ||suffix) for fallback
            _bare = _bk.split("||")[0]
            if _bare not in _live_centroids:
                _live_centroids[_bare] = (_avg_lat, _avg_lon)

        def _best_centroid(route_str: str) -> tuple | None:
            full_key = route_str.strip().upper()
            # Try full bucket key first (route||state|octant) for per-sub-bucket accuracy
            live = _live_centroids.get(full_key)
            if live:
                return live
            # Fallback to bare route key
            bare_key = full_key.split("||")[0]
            live = _live_centroids.get(bare_key)
            if live:
                return live
            return _route_centroid(route_str)

        # Keep (bucket_key, items) pairs so centroid lookups use the full key
        # (route||state|octant) for geographic accuracy.
        bucket_pairs = list(route_buckets.items())   # [(key, [item, …]), …]
        bucket_list  = [items for _, items in bucket_pairs]
        in_group     = [False] * len(bucket_list)
        super_groups: list[list] = []                # each entry = flat item list

        for i, base_bucket in enumerate(bucket_list):
            if in_group[i]:
                continue
            base_bkey  = bucket_pairs[i][0]          # full bucket key incl. ||suffix
            base_route = base_bucket[0]["ROUTE"]
            merged_items = list(base_bucket)
            in_group[i]  = True

            for j in range(i + 1, len(bucket_list)):
                if in_group[j]:
                    continue
                cand_bucket = bucket_list[j]
                cand_bkey   = bucket_pairs[j][0]
                cand_route  = cand_bucket[0]["ROUTE"]

                # All routes already absorbed into merged_items
                combined_w = sum(it["WEIGHT"] for it in merged_items) + \
                             sum(it["WEIGHT"] for it in cand_bucket)
                n_distinct = len({it["ROUTE"] for it in merged_items}) + 1

                # Don't merge routes whose preferred-lorry sets are disjoint —
                # unless BOTH routes are urban (KL/Selangor) and their GPS centroids
                # are within _MAX_CITY_MERGE_KM_OUTSTATION of each other, in which
                # case proximity overrides the preferred-lorry constraint so nearby
                # city routes (e.g. KV05A + KV19A, both around KL) can share one lorry.
                _base_pref = set(_preferred_lorries_for_route(base_route, engine))
                _cand_pref = set(_preferred_lorries_for_route(cand_route, engine))
                _base_dest_g_pref = _classify_dest_group(base_route, "")
                _cand_dest_g_pref = _classify_dest_group(cand_route, "")
                _both_urban = (_base_dest_g_pref in _DEST_URBAN_GROUPS
                               and _cand_dest_g_pref in _DEST_URBAN_GROUPS)
                # Check GPS proximity between the two route centroids.
                # Use live GPS from uploaded items first; fall back to static centroid DB.
                _bc_pref = (_live_centroids.get(base_bkey)
                            or _live_centroids.get(base_route)
                            or _route_centroid(base_route))
                _cc_pref = (_live_centroids.get(cand_bkey)
                            or _live_centroids.get(cand_route)
                            or _route_centroid(cand_route))
                _geo_close = (
                    _both_urban
                    and _bc_pref and _cc_pref
                    and _haversine_km(_bc_pref[0], _bc_pref[1], _cc_pref[0], _cc_pref[1])
                        <= _MAX_CITY_MERGE_KM_OUTSTATION
                )
                _pref_overlap = (
                    not _base_pref
                    or not _cand_pref
                    or bool(_base_pref & _cand_pref)
                    or _geo_close   # urban routes close together may share a lorry
                )

                # Don't merge buckets whose GPS bearing octants are incompatible,
                # or whose destination STATES differ.
                # State boundary rule: even if two sub-buckets share the same
                # route code and bearing direction (e.g. KV01A-SELANGOR-N vs
                # KV01A-PERAK-N), they are in different states → different lorry.
                # A route that "passes through" multiple states (e.g. KV01A going
                # through Batang Kali/Selangor before Tanjung Malim/Perak) should
                # split into per-state sub-runs so each lorry covers one state.
                def _octant_from_bkey(bk: str) -> str:
                    parts = bk.split("||")
                    if len(parts) >= 2:
                        sub = parts[1].split("|")
                        if len(sub) >= 2:
                            return sub[-1]
                    return ""
                def _state_from_bkey(bk: str) -> str:
                    parts = bk.split("||")
                    if len(parts) >= 2:
                        return parts[1].split("|")[0].strip().upper()
                    return ""
                _OPPOSITE = {"N": "S", "S": "N", "NE": "SW", "SW": "NE",
                             "E": "W", "W": "E", "SE": "NW", "NW": "SE"}
                _oct_base = _octant_from_bkey(base_bkey)
                _oct_cand = _octant_from_bkey(cand_bkey)
                _geo_ok = (
                    not _oct_base or not _oct_cand
                    or _oct_base == "LOCAL" or _oct_cand == "LOCAL"
                    or _oct_cand != _OPPOSITE.get(_oct_base, "")
                )
                # State boundary check — different state = different lorry.
                # Exception: same route prefix (e.g. both KV01A) always merges —
                # the route already accounts for cross-state travel.
                _st_base = _state_from_bkey(base_bkey)
                _st_cand = _state_from_bkey(cand_bkey)
                _same_state = (
                    not _st_base or not _st_cand
                    or _st_base == _st_cand
                    or base_route == cand_route  # same route code → allow cross-state merge
                )

                # City-proximity check for outstation routes: only merge city
                # clusters within the same state if their GPS centroids are
                # within _MAX_CITY_MERGE_KM_OUTSTATION of each other.
                # Urban routes (KL/Selangor) are already city-bucketed; skip check.
                _base_dest_g = _classify_dest_group(base_route,
                                                     _st_base or "")
                _cand_dest_g = _classify_dest_group(cand_route,
                                                     _st_cand or "")
                _city_dist_ok = True
                if (_base_dest_g in {"LARGE_LONG", "MEDIUM_LONG"}
                        or _cand_dest_g in {"LARGE_LONG", "MEDIUM_LONG"}):
                    # a) Cross-validate cities against the Malaysia States & Cities
                    #    reference: if both cities are known and map to different
                    #    states, hard-block the merge (border GPS can fool distance).
                    _base_cities = {
                        it.get("CITY", "").strip().upper() for it in base_bucket
                    }
                    _cand_cities = {
                        it.get("CITY", "").strip().upper() for it in cand_bucket
                    }
                    for _bc_city in _base_cities:
                        _bc_st = _CITY_TO_STATE.get(_bc_city, "")
                        for _cc_city in _cand_cities:
                            _cc_st = _CITY_TO_STATE.get(_cc_city, "")
                            if _bc_st and _cc_st and _bc_st != _cc_st:
                                _city_dist_ok = False
                                break
                        if not _city_dist_ok:
                            break
                    # b) GPS distance between centroids
                    if _city_dist_ok:
                        _bc = _live_centroids.get(base_bkey) or _live_centroids.get(base_route)
                        _cc = _live_centroids.get(cand_bkey) or _live_centroids.get(cand_route)
                        if _bc and _cc:
                            _city_dist_ok = (
                                _haversine_km(_bc[0], _bc[1], _cc[0], _cc[1])
                                <= _MAX_CITY_MERGE_KM_OUTSTATION
                            )

                _corridor_merge = _same_corridor_group(
                    base_route, cand_route, _st_base, _st_cand,
                    base_bucket[0].get("CUSTOMER NAME", ""),
                    cand_bucket[0].get("CUSTOMER NAME", ""))
                _urban_prox_merge = _geo_close

                # ── Priority-ordered merge rules (per user request) ───────────
                # Group items onto the same lorry in this priority order:
                #   1. SAME ROUTE  — identical route prefix (e.g. KV05A == KV05A)
                #   2. SAME CITY   — same state AND same city
                #   3. SAME STATE + NEAREST LONGITUDE — same state and centroids
                #      close together (longitude/GPS within the merge radius)
                _pfx_base = _extract_route_prefix(base_route)
                _pfx_cand = _extract_route_prefix(cand_route)
                _same_route_pfx = bool(_pfx_base) and _pfx_base == _pfx_cand

                _base_cities2 = {it.get("CITY", "").strip().upper()
                                 for it in base_bucket if it.get("CITY")}
                _cand_cities2 = {it.get("CITY", "").strip().upper()
                                 for it in cand_bucket if it.get("CITY")}
                _same_city = (
                    bool(_base_cities2 & _cand_cities2)
                    and (not _st_base or not _st_cand or _st_base == _st_cand)
                )

                # Same-state + nearest-longitude: both in the same state and their
                # GPS centroids are within the city-merge radius of each other.
                _lon_close = False
                if (_st_base and _st_cand and _st_base == _st_cand
                        and _bc_pref and _cc_pref):
                    _lon_close = (
                        _haversine_km(_bc_pref[0], _bc_pref[1],
                                      _cc_pref[0], _cc_pref[1])
                        <= _MAX_CITY_MERGE_KM_OUTSTATION
                    )

                _priority_merge = _same_route_pfx or _same_city or _lon_close

                if (
                    combined_w <= max_lorry_cap
                    and n_distinct <= _MAX_STOPS
                    and (_routes_on_same_way(base_route, cand_route) or _corridor_merge
                         or _urban_prox_merge or _priority_merge)
                    and (_pref_overlap or _same_route_pfx or _same_city)
                    and _geo_ok
                    and _same_state
                    and (_city_dist_ok or _corridor_merge or _urban_prox_merge
                         or _same_route_pfx or _same_city or _lon_close)
                ):
                    merged_items += list(cand_bucket)
                    in_group[j]   = True

            super_groups.append(merged_items)

        # Step 3 — if a super-group is heavier than max lorry cap, bin-pack it
        # into capacity-sized sub-groups (heaviest route bucket first).
        sorted_groups: list[list] = []
        for sg in super_groups:
            total_w = sum(it["WEIGHT"] for it in sg)
            if total_w <= max_lorry_cap:
                sorted_groups.append(sg)
                continue

            # Over-capacity super-group — split into sub-groups
            sub_buckets = defaultdict(list)
            for it in sg:
                sub_buckets[it["ROUTE"]].append(it)
            sub_list = sorted(
                sub_buckets.values(),
                key=lambda b: sum(i["WEIGHT"] for i in b),
                reverse=True,
            )

            current_sub: list = []
            current_w = 0.0
            for sub_b in sub_list:
                w = sum(it["WEIGHT"] for it in sub_b)
                if current_sub and current_w + w > max_lorry_cap:
                    sorted_groups.append(current_sub)
                    current_sub = list(sub_b)
                    current_w   = w
                else:
                    current_sub += list(sub_b)
                    current_w   += w
            if current_sub:
                sorted_groups.append(current_sub)

        # ── Step 4: geographic cross-cluster merge (Nominatim/OSM, free) ─────────
        # Same-cluster corridor merging (Step 2) only joins routes within the
        # same region.  Here we try to join groups from DIFFERENT clusters when
        # Nominatim confirms their destinations are within 300 km straight-line.
        #
        # Example: PH01-03 (Pahang/Bentong, 0.275T) + TR02 (Terengganu, 6T)
        # both use the KL→East highway and their centroids are ≈160 km apart.
        # East Malaysia (Sabah/Sarawak) is ≈1 000 km from KL → always rejected.
        def _group_centroid(items):
            """Average lat/lng centroid of all items in a group.
            Uses per-item GPS_LAT/GPS_LON when available (live LONGITUD data),
            falling back to history-based route centroid per distinct route."""
            lats, lons = [], []
            for it in items:
                _la = it.get("GPS_LAT")
                _lo = it.get("GPS_LON")
                if _la is not None and _lo is not None:
                    lats.append(_la); lons.append(_lo)
            if lats:
                return (sum(lats) / len(lats), sum(lons) / len(lons))
            # Fallback: history centroids per distinct route
            seen = set()
            for it in items:
                r = it["ROUTE"]
                if r in seen:
                    continue
                seen.add(r)
                c = _route_centroid(r)
                if c:
                    lats.append(c[0]); lons.append(c[1])
            if not lats:
                return None
            return (sum(lats) / len(lats), sum(lons) / len(lons))

        _cross_merged = [False] * len(sorted_groups)
        _new_groups: list[list] = []

        for i, base_sg in enumerate(sorted_groups):
            if _cross_merged[i]:
                continue
            merged      = list(base_sg)
            merged_cent = _group_centroid(merged)   # updated as we absorb groups

            for j in range(i + 1, len(sorted_groups)):
                if _cross_merged[j]:
                    continue
                cand_sg    = sorted_groups[j]
                cand_route = cand_sg[0]["ROUTE"]

                # Skip if all routes in cand share the same cluster as base
                # (same-cluster merging already handled by corridor merge)
                base_clusters = {_extract_route_intelligence(it["ROUTE"])["cluster"]
                                 for it in merged}
                cand_cluster  = _extract_route_intelligence(cand_route)["cluster"]
                if base_clusters == {cand_cluster}:
                    continue

                # Urban routes (KL_VALLEY / KL_CITY) never bundle with outstation.
                _LOCAL_CLUSTERS = {"KL_VALLEY", "KL_CITY"}
                if (base_clusters & _LOCAL_CLUSTERS) or cand_cluster in _LOCAL_CLUSTERS:
                    continue

                # Outstation↔urban dest-group mixing: if one side is urban and
                # the other is outstation (LARGE_LONG / MEDIUM_LONG), block merge.
                _base_dest_grps = {
                    _classify_dest_group(it["ROUTE"], it.get("STATE", ""))
                    for it in merged
                }
                _cand_dest_grp = _classify_dest_group(cand_sg[0]["ROUTE"],
                                                       cand_sg[0].get("STATE", ""))
                _base_is_outstation = any(
                    g in {"LARGE_LONG", "MEDIUM_LONG"} for g in _base_dest_grps
                )
                _cand_is_outstation = _cand_dest_grp in {"LARGE_LONG", "MEDIUM_LONG"}
                if _base_is_outstation != _cand_is_outstation:
                    continue

                combined_w = sum(it["WEIGHT"] for it in merged) + \
                             sum(it["WEIGHT"] for it in cand_sg)
                n_routes   = len({it["ROUTE"] for it in merged}) + \
                             len({it["ROUTE"] for it in cand_sg})

                if combined_w > max_lorry_cap:
                    continue
                if n_routes > _MAX_STOPS:
                    continue

                # Geographic check: candidate centroid vs merged group centroid.
                # Use tighter threshold for outstation routes so only nearby
                # cities within the same state share a lorry.
                cand_cent = _best_centroid(cand_route)
                if merged_cent is None or cand_cent is None:
                    continue
                dist_km = _haversine_km(merged_cent[0], merged_cent[1],
                                        cand_cent[0],   cand_cent[1])
                _cg_base = _classify_dest_group(
                    merged[0]["ROUTE"], merged[0].get("STATE", ""))
                _cg_cand = _classify_dest_group(
                    cand_sg[0]["ROUTE"], cand_sg[0].get("STATE", ""))
                _max_dist = (
                    _MAX_CITY_MERGE_KM_OUTSTATION
                    if (_cg_base in {"LARGE_LONG", "MEDIUM_LONG"}
                        or _cg_cand in {"LARGE_LONG", "MEDIUM_LONG"})
                    else 180.0   # urban routes: generous radius
                )
                if dist_km > _max_dist:
                    continue

                # State boundary: NEVER merge groups from different destination states.
                # Read STATE directly from item dicts — the _route_state lookup was
                # unreliable because cand_route is a full route text ("PH04-->Benta")
                # while _route_state keys are route prefixes ("PH04").
                cand_state  = cand_sg[0].get("STATE", "").strip().upper() if cand_sg else ""
                base_states = {it.get("STATE", "").strip().upper() for it in merged}
                base_states.discard("")
                base_states.discard("")
                if cand_state and base_states and cand_state not in base_states:
                    continue   # hard block — different states never share a lorry

                # Bearing check — ALL PAIRS: candidate must be directionally
                # compatible with every existing regional route in the group.
                # Using the rolling centroid alone can drift as groups merge,
                # allowing e.g. K.Selangor (304°) + Terengganu (34°) = 90°
                # to slip through after KV03A+PK02 centroid shifts to 316°.
                b_cand = _bearing_deg(_DEPOT[0], _DEPOT[1],
                                      cand_cent[0], cand_cent[1])
                depot_to_cand = _haversine_km(_DEPOT[0], _DEPOT[1],
                                              cand_cent[0], cand_cent[1])
                if depot_to_cand < 50.0:   # local route — skip bearing check
                    continue
                bearing_ok = True
                bearing_checked = False  # must have ≥1 regional route in group to merge
                for ex_it in merged:
                    ex_route = ex_it["ROUTE"]
                    ec = _best_centroid(ex_route)
                    if ec is None:
                        continue
                    d_ex = _haversine_km(_DEPOT[0], _DEPOT[1], ec[0], ec[1])
                    if d_ex < 50.0:        # local route — skip bearing check
                        continue
                    bearing_checked = True
                    b_ex = _bearing_deg(_DEPOT[0], _DEPOT[1], ec[0], ec[1])
                    diff = _bearing_diff(b_ex, b_cand)
                    # States are already equal at this point (hard block above);
                    # limit is always _CROSS_BEARING_LIMIT for same-state routes.
                    limit = _CROSS_BEARING_LIMIT
                    if diff > limit:
                        bearing_ok = False
                        break
                # Reject if merged group has no regional routes — can't validate direction
                if not bearing_checked or not bearing_ok:
                    continue

                merged += list(cand_sg)
                _cross_merged[j] = True
                # Recompute centroid to keep it accurate for subsequent candidates
                merged_cent = _group_centroid(merged)

            _new_groups.append(merged)

        sorted_groups = _new_groups

        def _parse_date_sortkey(date_str: str) -> str:
            """Convert any date string to ISO 'YYYY-MM-DD' for correct ordering.
            Returns '9999-12-31' on failure so undated groups sort last.
            Handles formats: 'd/m/yy', 'dd/mm/yy', 'YYYY-MM-DD', Excel serials.
            """
            s = (date_str or "").strip()
            if not s or s.lower() in ("nan", "none", ""):
                return "9999-12-31"
            for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d", "%-d/%-m/%y"):
                try:
                    return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
                except ValueError:
                    pass
            try:
                ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
                if pd.notna(ts):
                    return ts.strftime("%Y-%m-%d")
            except Exception:
                pass
            return "9999-12-31"

        from datetime import timedelta as _td_etd
        _ETD_TOMORROW_STR = (_today_date + _td_etd(days=1)).strftime("%Y-%m-%d")

        def _group_has_urgent_etd(g) -> bool:
            """True if any item's ETD is a real (non-NULL-sentinel) date that's
            today or tomorrow — the lorry needs to leave imminently, so this
            group should be prioritised ahead of otherwise-equal groups even
            if its delivery DATE alone wouldn't rank it first."""
            for it in g:
                etd_key = _parse_date_sortkey(it.get("ETD", ""))
                # "9999-12-31" = unparseable/blank; "1753-xx-xx" = the DB's
                # NULL-ETD sentinel (ISNULL(ZETD_0, '1753-01-01')) — neither
                # is a real ETD, so skip both rather than treat as urgent.
                if etd_key in ("9999-12-31",) or etd_key < "2000-01-01":
                    continue
                if etd_key <= _ETD_TOMORROW_STR:
                    return True
            return False

        def _group_sort_key(g):
            """Sort order:
            0. Destination priority — LARGE_LONG(0) > MEDIUM_LONG(1) > KL_SELANGOR(2)
               so long-distance groups claim large/medium lorries first.
            1. Urgent ETD (today/tomorrow) sorts ahead of non-urgent, within
               the same destination tier — it needs to leave imminently.
            2. Earliest delivery date (ascending).
            3. Total weight (descending) so heavier groups within same tier/date
               claim best-fit lorries before lighter ones.
            """
            dest_pri = _DEST_SORT_PRI.get(
                _classify_dest_group(g[0]["ROUTE"], g[0].get("STATE", "")), 2
            )
            etd_urgent = 0 if _group_has_urgent_etd(g) else 1
            dates = [_parse_date_sortkey(it.get("DATE", "")) for it in g]
            earliest = min(dates) if dates else "9999-12-31"
            return (dest_pri, etd_urgent, earliest, -sum(it["WEIGHT"] for it in g))

        # Destination-priority-first, then earliest date, then heaviest.
        # Within each group, sort items by date so earliest DOs get the first
        # lorry when capacity forces a split.
        sorted_groups.sort(key=_group_sort_key)
        for _grp in sorted_groups:
            _grp.sort(key=lambda it: _parse_date_sortkey(it.get("DATE", "")))

        # ── Split groups by MAX_TON cap ───────────────────────────────────────
        # A single VAN-only DO inside a large group (e.g. KV20A 22T) would cap
        # the whole group at 2T and cause everything to fail.  Instead, split
        # each group into per-cap sub-groups so:
        #   • VAN DOs (MAX_TON=2T) from the same route group together and fill
        #     one van sorted by nearest longitude.
        #   • Uncapped DOs keep their original group and get a normal lorry.
        _split_groups: list[list] = []
        for _sg in sorted_groups:
            _cap_buckets: dict = {}
            for _it in _sg:
                _ck = _it.get("MAX_TON")   # None means no cap
                _cap_buckets.setdefault(_ck, []).append(_it)
            if len(_cap_buckets) == 1:
                _split_groups.append(_sg)   # uniform cap — no split needed
            else:
                # Mixed caps: split into per-cap sub-groups.
                # Uncapped items are the "pool" that capped sub-groups can draw from.
                _uncapped_pool = list(_cap_buckets.get(None, []))

                # Emit capped sub-groups first (tightest cap first) so small-lorry
                # DOs claim the right vehicle before the main group is assigned.
                for _ck, _citems in sorted(
                    {k: v for k, v in _cap_buckets.items() if k is not None}.items()
                ):
                    # Sort capped (e.g. VAN) items by longitude (parsed GPS_LON
                    # float — LONGITUD is a raw "lat lon" string not in the item).
                    _citems.sort(key=lambda x: x.get("GPS_LON") or 999)

                    # Find the largest eligible lorry that satisfies this cap so we
                    # know how many nearby uncapped items can share the same vehicle.
                    _van_lorry_cap = max(
                        (float(r["TON"]) for _, r in engine.eligible_lorries.iterrows()
                         if float(r["TON"]) <= _ck),
                        default=_ck,
                    )

                    # Pull in nearest uncapped same-route items up to van capacity.
                    # Use parsed GPS_LON (float) — LONGITUD is a raw "lat lon" string.
                    _van_w = sum(x["WEIGHT"] for x in _citems)
                    _van_lons = [x["GPS_LON"] for x in _citems if x.get("GPS_LON") is not None]
                    _van_clon = (sum(_van_lons) / len(_van_lons)) if _van_lons else 0

                    _near = sorted(
                        [u for u in _uncapped_pool if u.get("GPS_LON") is not None],
                        key=lambda x: abs(x["GPS_LON"] - _van_clon),
                    )
                    _absorbed: list = []
                    for _ui in _near:
                        if _van_w + _ui["WEIGHT"] <= _van_lorry_cap:
                            _van_w += _ui["WEIGHT"]
                            _absorbed.append(_ui)
                            _citems.append(_ui)
                    for _a in _absorbed:
                        _uncapped_pool.remove(_a)

                    # Re-sort by longitude so the trip manifest is geographically ordered
                    _citems.sort(key=lambda x: x.get("GPS_LON") or 999)
                    _split_groups.append(_citems)

                # Remaining uncapped items form the normal sub-group
                if _uncapped_pool:
                    _split_groups.append(_uncapped_pool)
        sorted_groups = _split_groups

        # ── Per-direction lorry reservation ───────────────────────────────────
        # Reserve one >5T lorry for EACH distinct OUTSTATION direction that has
        # DOs, so the biggest direction (e.g. Pahang/Kuantan) cannot grab every
        # lorry and leave a smaller outstation run (e.g. KV01A/KV02A north) with
        # nothing. Reservation only happens when a direction actually has DOs;
        # urban routes need no reservation. Smallest direction reserves first
        # (it is the most easily starved) and takes the TIGHTEST-fitting >5T lorry.
        _reserved_lorry: dict[str, str] = {}   # plate → direction key
        _dir_weight: dict[str, float] = defaultdict(float)
        # Count this user's own OUTSTATION DOs — both those being assigned now and
        # those off-schedule (NOT_TODAY) that the user may choose to assign — so a
        # lorry is held for their direction. Other users' rows are excluded.
        for _it in items:
            _l = _it.get("LORRY")
            if _l not in (None, "NOT_TODAY"):
                continue                            # settled / other-user / skip
            _dg = _classify_dest_group(_it.get("ROUTE", ""), _it.get("STATE", ""))
            if _dg in _DEST_URBAN_GROUPS:
                continue                            # urban → no reservation
            _dir_weight[_direction_key(_it.get("ROUTE", ""), _it.get("STATE", ""), _it.get("CUSTOMER NAME", ""))] += _it["WEIGHT"]
        if len(_dir_weight) > 1:                     # only when directions compete
            _res_excl = sess["unavailable"] | get_assigned_today()
            _res_avail = sorted(
                (float(r["TON"]), str(r["LORRY"]).strip().upper())
                for _, r in engine.eligible_lorries.iterrows()
                if float(r["TON"]) > _OUTSTATION_MIN_TON
                and str(r["LORRY"]).strip().upper() not in _res_excl
            )
            for _dk in sorted(_dir_weight, key=lambda k: _dir_weight[k]):
                # A tiny NS/Seremban direction may ride a small lorry, so don't
                # reserve a big lorry for it (that would waste the big lorry).
                if (_dk[:2] in _TINY_OUTSTATION_PREFIXES
                        and _dir_weight[_dk] <= _TINY_OUTSTATION_MAX_TON):
                    continue
                _need = _dir_weight[_dk]
                _free = [(t, p) for t, p in _res_avail if p not in _reserved_lorry]
                if not _free:
                    break
                _fit = [(t, p) for t, p in _free if t * NAIK_FACTOR >= _need]
                _pick = _fit[0] if _fit else _free[-1]   # tightest that fits, else largest
                _reserved_lorry[_pick[1]] = _dk

        # Session-level capacity tracker so groups can share a lorry when combined
        # weight still fits (e.g. two 0.4T groups sharing VEA2818's 1.07T).
        # Seed session loads from pre-filled items (Case B re-upload)
        _session_loads: dict[str, float]  = dict(_prefill_loads)
        _session_routes: dict[str, str]   = dict(_prefill_routes)
        # Track which destination states each lorry has been assigned to THIS
        # session. Seeded from pre-filled items so re-uploads respect prior state.
        _session_lorry_states: dict[str, set] = {
            p: set(sts) for p, sts in _prefill_states.items()
        }
        _lorry_cap_map = {row["LORRY"]: float(row["TON"])
                          for _, row in engine.eligible_lorries.iterrows()}

        def _record_lorry_state(plate: str, state: str) -> None:
            """Update _session_lorry_states whenever a lorry is assigned."""
            if plate and state:
                _session_lorry_states.setdefault(plate, set()).add(state.strip().upper())

        # Track reason each DO number ended up unassigned (NO_LORRY)
        _unassigned_reasons: dict[str, str] = {}   # DO_NUMBER → reason code

        def _mark_no_lorry(items_list, reason: str) -> None:
            for _it in items_list:
                _it["LORRY"] = "NO_LORRY"
                _unassigned_reasons[_it["DO NUMBER"]] = reason

        def _assign_group(group_items):
            """Assign ONE lorry (or split) to cover ALL items in the group.
            All items in the group share the same route (one route = one lorry).
            """
            # Pre-filter: items that exceed every single lorry's capacity cannot
            # ride alone.  Instead of immediately marking them NO_LORRY, try to
            # split the single heavy DO across multiple lorries (e.g. a 27T DO
            # split across two 14T lorries).  Only if no combination of available
            # lorries can cover the total weight does the DO become NO_LORRY.
            _max_cap = (float(engine.eligible_lorries["TON"].max())
                        if not engine.eligible_lorries.empty else 0.0)
            _excl_pre = sess["unavailable"] | get_assigned_today()
            # NOTE: strict-route exclusions can't be computed here because we
            # don't yet know the route for every item that might be oversized.
            # We compute a conservative (inclusive) fleet cap; the strict check
            # is applied per-item when the multi-lorry split actually runs.
            _avail_lorries_pre = [
                (str(r["LORRY"]).strip().upper(), float(r["TON"]))
                for _, r in engine.eligible_lorries.iterrows()
                if str(r["LORRY"]).strip().upper() not in _excl_pre
            ]
            _fleet_total_cap = sum(c for _, c in _avail_lorries_pre)
            _all_group = list(group_items)
            if _max_cap > 0:
                _still_oversized = []
                for it in list(_all_group):
                    if it.get("LORRY") is None and it["WEIGHT"] > _max_cap:
                        # Heavy item — can the fleet cover it across multiple lorries?
                        # Compute usable capacity excluding lorries strictly
                        # forbidden from this item's route (e.g. BQU3875 can't
                        # serve KV routes even for a multi-lorry split).
                        _item_strict = _strict_route_excl(it.get("ROUTE", ""))
                        _usable_cap = sum(
                            c for p, c in _avail_lorries_pre
                            if p not in _item_strict
                        )
                        if it["WEIGHT"] <= _usable_cap:
                            # Fleet can cover it across multiple lorries — mark
                            # for multi-lorry split below.
                            it["_ALLOW_MULTI_LORRY"] = True
                        else:
                            # Truly impossible — more than usable fleet capacity.
                            it["LORRY"] = "NO_LORRY"
                            _unassigned_reasons[it["DO NUMBER"]] = "LOAD_EXCEEDS_ALL_LORRIES"
                group_items = [it for it in _all_group if it.get("LORRY") != "NO_LORRY"]
            else:
                _all_group = list(group_items)
            if not group_items:
                for it in _all_group:
                    sess["assigned"][it["DO NUMBER"]] = it["LORRY"]
                return

            # Rule 6b — per-lorry delivery-stop limit.
            # When a group carries too many individual DOs, split it across two
            # lorries so drivers aren't overloaded and idle ABI lorries get work.
            # MAX_STOPS_PER_LORRY (=8) was designed for route-count merging;
            # here we use a separate threshold for DO count.
            _MAX_DOS_PER_LORRY = MAX_DOS_PER_LORRY
            if len(group_items) > _MAX_DOS_PER_LORRY:
                # Only split when the combined weight truly exceeds every
                # available lorry's capacity.  If the full group fits on a
                # single lorry (even a large 14T), keep it together so that
                # lorry reaches high utilisation instead of creating two
                # under-filled lorries.
                _pre_total_w = sum(it["WEIGHT"] for it in group_items)
                _excl_check = sess["unavailable"] | get_assigned_today()
                _avail_caps = sorted(
                    float(row["TON"]) for _, row in engine.eligible_lorries.iterrows()
                    if row["LORRY"] not in _excl_check
                )
                if not any(c * NAIK_FACTOR >= _pre_total_w for c in _avail_caps):
                    # Too heavy for any single lorry — split by DATE (earliest
                    # DOs get the first lorry so older orders always ship first).
                    # group_items is already date-sorted from the pre-sort above.
                    _cap1 = max(_avail_caps) if _avail_caps else 0
                    # Split into ATOMIC components first so a split never breaks
                    # criterion 1 (same longitude) or 2 (same route + same
                    # customer CODE). Components sharing a route code are kept
                    # adjacent so same-route DOs (criterion 3) stay together when
                    # they fit, and each half is filled by weight (criterion 4).
                    _hs_comps = _atomic_components(
                        group_items,
                        _route_code_of,
                        lambda _it: (_it.get("GPS_LAT"), _it.get("GPS_LON")),
                        lambda _it: _it.get("CODE", ""))
                    # Keep each ROUTE CODE whole across the split (operator rule:
                    # same route code → same lorry) — but ONLY when the whole code
                    # fits a lorry. A route code heavier than one lorry is split
                    # back into its atomic components so the fill always makes
                    # progress (otherwise recursing on an oversized whole chunk
                    # would never terminate).
                    _cap_lim = _cap1 * NAIK_FACTOR
                    _by_code: dict = {}
                    for _cl in _hs_comps:
                        _by_code.setdefault(_route_code_of(_cl[0]), []).append(_cl)
                    _units: list = []               # each unit = list of items
                    for _code, _comps in _by_code.items():
                        _flat = [x for c in _comps for x in c]
                        if sum(x["WEIGHT"] for x in _flat) <= _cap_lim:
                            _units.append(_flat)    # whole route code, kept together
                        else:
                            _units.extend(_comps)   # too big — split into components
                    _units.sort(key=lambda u: -sum(x["WEIGHT"] for x in u))
                    half_a, half_b = [], []
                    _fill_w = 0.0
                    for _u in _units:
                        _uw = sum(x["WEIGHT"] for x in _u)
                        if _fill_w + _uw <= _cap_lim or not half_a:
                            half_a.extend(_u)
                            _fill_w += _uw
                        else:
                            half_b.extend(_u)
                    # Only recurse when BOTH halves are non-empty (a real split);
                    # otherwise the split made no progress — fall through to the
                    # normal single-lorry / multi-lorry path to avoid infinite
                    # recursion on an indivisible oversized unit.
                    if half_a and half_b:
                        _assign_group(half_a)
                        _assign_group(half_b)
                        for it in _all_group:
                            if it.get("LORRY") == "NO_LORRY":
                                sess["assigned"][it["DO NUMBER"]] = "NO_LORRY"
                        return
                # else: weight fits one lorry — fall through to normal single-lorry path

            total_w  = sum(it["WEIGHT"] for it in group_items)
            route    = group_items[0]["ROUTE"]
            customer = group_items[0]["CUSTOMER NAME"]

            # ── Multi-lorry split for single heavy items ──────────────────────
            # A single DO whose weight exceeds the largest available lorry but
            # fits across 2+ lorries in the fleet is split here.  We greedily
            # fill the biggest available lorry, then recurse on the remainder.
            # The sentinel _ALLOW_MULTI_LORRY is set upstream by the pre-filter.
            if (len(group_items) == 1
                    and group_items[0].get("_ALLOW_MULTI_LORRY")
                    and total_w > _max_cap):
                _excl_ml = sess["unavailable"] | get_assigned_today() | get_broken_lorries().keys()
                _ml_strict = _strict_route_excl(route)   # e.g. BQU3875 excluded for KV routes
                _ml_lorries = sorted(
                    [
                        (str(r["LORRY"]).strip().upper(), float(r["TON"]))
                        for _, r in engine.eligible_lorries.iterrows()
                        if str(r["LORRY"]).strip().upper() not in _excl_ml
                        and str(r["LORRY"]).strip().upper() not in _ml_strict
                    ],
                    key=lambda x: -x[1],   # largest first
                )
                _remaining_w = total_w
                _orig_item   = group_items[0]
                _part_num    = 0
                _first_plate = None
                for _ml_plate, _ml_cap in _ml_lorries:
                    if _remaining_w <= 0:
                        break
                    _portion = min(_ml_cap, _remaining_w)
                    _part_num += 1
                    # Clone the item for this portion, giving it a unique internal key
                    # so the summary can show it under each lorry separately.
                    _part_item = dict(_orig_item)
                    _part_item["WEIGHT"]            = round(_portion, 4)
                    _part_item["_ALLOW_MULTI_LORRY"] = False
                    _part_item["LORRY"]             = _ml_plate
                    _part_item["SPLIT_PART"]        = _part_num   # 1, 2, 3…
                    _part_item["DO NUMBER"]         = f"{_orig_item['DO NUMBER']}(p{_part_num})"
                    # Register in _all_group so the export loop can see all parts
                    _all_group.append(_part_item)
                    _session_loads[_ml_plate] = float(_session_loads.get(_ml_plate, 0)) + _portion
                    if _ml_plate not in _session_routes:
                        _session_routes[_ml_plate] = route
                    sess["assigned"][_part_item["DO NUMBER"]] = _ml_plate
                    _remaining_w = round(_remaining_w - _portion, 4)
                    if _first_plate is None:
                        _first_plate = _ml_plate
                # Mark original item with first lorry (the main record); it will be
                # superseded by the cloned parts in the export.
                if _remaining_w > 0:
                    _orig_item["LORRY"] = "NO_LORRY"
                    _unassigned_reasons[_orig_item["DO NUMBER"]] = "LOAD_EXCEEDS_ALL_LORRIES"
                else:
                    _orig_item["LORRY"] = _first_plate or "NO_LORRY"
                for it in _all_group:
                    if it.get("LORRY") == "NO_LORRY":
                        sess["assigned"][it["DO NUMBER"]] = "NO_LORRY"
                return

            broken_map = get_broken_lorries()
            sess["unavailable"].update(broken_map.keys())

            # ── Effective-capacity helper ─────────────────────────────────────
            # Per user preference: assign by SINGLE-trip physical capacity only —
            # always pick the smallest lorry that fits in one trip. No 2-trip
            # (morning+afternoon) capacity crediting, which previously let a small
            # van "fit" a group it could not physically carry in one load.
            def _eff_cap_for(plate: str, grp_dest: str) -> float:
                return float(_lorry_cap_map.get(plate, 0.0))

            # ── Destination group — must be defined BEFORE _session_full ─────────
            # LARGE_LONG  (Pahang/Kuantan/Terengganu/…): must use ≥14T lorry
            # MEDIUM_LONG (Seremban/NS/Rawang/T.Malim/…): must use ≥11T lorry
            # KL / SELANGOR / KL_SELANGOR: must use <11T lorry (no large/medium)
            _item_state = group_items[0].get("STATE", "")
            _dest_grp   = _classify_dest_group(route, _item_state)

            # Also exclude lorries already full or incompatible with this route
            _session_full = {
                p for p in _session_loads
                if _eff_cap_for(p, _dest_grp) - _session_loads.get(p, 0.0) < total_w
            }
            _session_incompatible = {
                p for p in _session_loads
                if _session_routes.get(p)
                and not _routes_on_same_way(route, _session_routes.get(p, ""))
                and _lorry_cap_map.get(p, 0) - _session_loads.get(p, 0) >= total_w
            }
            excluded = sess["unavailable"] | _session_full

            # ── Strict lorry-route reservations ───────────────────────────────
            # e.g. BQY7823 → Rawang only; BQU3875 → Pahang only.
            # Check each unique route code individually — a joined text would
            # falsely pass if the group happened to contain an allowed route code
            # alongside a disallowed one (e.g. "KV02A PH09 ..." lets BQY7823 slip).
            _unique_routes = {it.get("ROUTE", "") for it in group_items}
            _strict_excl: set[str] = set()
            for _ur in _unique_routes:
                _strict_excl |= _strict_route_excl(_ur)
            excluded = excluded | _strict_excl

            # ── Cross-direction incompatibility for outstation lorries ──────────
            # A lorry already serving an OUTSTATION route (LARGE_LONG or
            # MEDIUM_LONG) must not be mixed with a DIFFERENT direction — whether
            # the current group is outstation OR urban.
            #
            # Without this, a lorry heading Pahang (outstation) also picks up
            # urban KL/Selangor items that it literally cannot visit on the same
            # trip. The previous guard had `_dest_grp not in _DEST_URBAN_GROUPS`
            # which disabled the check entirely for urban groups — removed below.
            _session_incompatible_lm = {
                p for p in _session_loads
                if _session_routes.get(p)
                and _classify_dest_group(
                    _session_routes.get(p, "")) not in _DEST_URBAN_GROUPS
                and not _routes_on_same_way(route, _session_routes.get(p, ""))
            }
            excluded = excluded | _session_incompatible_lm

            # ── State-boundary exclusion ──────────────────────────────────────
            # A lorry committed to state X this session must not serve state Y.
            # Uses _session_lorry_states (running tracker, updated after each
            # assignment) so it is never confused by pre-assigned rows from
            # other days that happen to share the same items list.
            # A merged route bucket may legitimately span several COMPATIBLE
            # states (e.g. KV11A covers SELANGOR + WP KUALA LUMPUR — adjacent
            # urban states). Collect ALL states in the group, and exclude a
            # lorry only if NONE of its committed states are compatible with
            # ANY of the group's states. Exact-match exclusion here was wrongly
            # stranding same-route items that cross an urban state line.
            _grp_state = (group_items[0].get("STATE", "").strip().upper()
                          if group_items else "")
            _grp_states = {
                it.get("STATE", "").strip().upper()
                for it in group_items if it.get("STATE")
            }
            _state_excl: set[str] = set()
            if _grp_states:
                _state_excl = {
                    p for p, sts in _session_lorry_states.items()
                    if sts and not any(
                        _states_compatible(_gs, _ls)
                        for _gs in _grp_states for _ls in sts
                    )
                }
            excluded = excluded | _state_excl

            # ── Forbidden-plate exclusion (REMARKS "lorry 3875 tak boleh masuk") ─
            # Any plate a member DO forbids cannot serve this group (all members
            # ride one lorry, so the union of forbidden plates is excluded).
            _grp_forbid: set = set()
            for _it in group_items:
                _fp = _it.get("FORBID_PLATES")
                if _fp:
                    _grp_forbid |= _fp
            if _grp_forbid:
                excluded = excluded | _grp_forbid

            # ── Per-direction reservation exclusion ───────────────────────────
            # A lorry reserved for another outstation direction is off-limits to
            # this group (keeps e.g. BMN3682 free for the KV_NORTH run instead of
            # being consumed by Pahang). A group may still use a lorry reserved
            # for ITS OWN direction, or any unreserved lorry.
            _reserve_excl: set = set()
            if _reserved_lorry:
                _grp_dir = _direction_key(route,
                    group_items[0].get("STATE", "") if group_items else "",
                    group_items[0].get("CUSTOMER NAME", "") if group_items else "")
                _reserve_excl = {p for p, dk in _reserved_lorry.items()
                                 if dk != _grp_dir}
                if _reserve_excl:
                    excluded = excluded | _reserve_excl

            # ── Preferred lorry enforcement (runs BEFORE size exclusions) ─────
            # Preferred lorries bypass the standard size-minimum rules because
            # they are operationally designated for that corridor (e.g. BMN3682
            # at 8.66T handles NS04/NS05 even though MEDIUM_LONG min is 10.5T).
            #
            # Use the DOMINANT route (most items) to select the preferred lorry,
            # not just the first item — a merged group may contain multiple routes.
            _route_counts: dict[str, int] = {}
            for _it in group_items:
                _route_counts[_it.get("ROUTE", "")] = _route_counts.get(_it.get("ROUTE", ""), 0) + 1
            _dominant_route = max(_route_counts, key=lambda r: _route_counts[r])
            _preferred = _preferred_lorries_for_route(_dominant_route, engine)
            _base_excluded = excluded   # save pre-size excluded set for fallback
            # REMARKS size cap (FIELD 3) — smallest cap among this group's DOs.
            _grp_caps_pre = [it["MAX_TON"] for it in group_items
                             if it.get("MAX_TON") is not None]
            _grp_cap_pre = min(_grp_caps_pre) if _grp_caps_pre else None
            if _preferred:
                # Preferred lorries are a hint — weight fit wins.
                # Hard-exclude preferred lorries if: truly unavailable, full,
                # strictly forbidden for ANY route in this group, OR already
                # committed to a different destination state.
                _hard_excl = (sess["unavailable"] | get_assigned_today()
                               | _strict_excl | _state_excl)
                # Also build destination-group rules here so preferred check is consistent
                # with open assignment: urban routes (KL/Selangor) must use <11T lorries.
                _pref_dest_min = _DEST_MIN_TON.get(_dest_grp, 0.0)
                _pref_avail = [
                    p for p in _preferred
                    if p in _lorry_cap_map
                    and p not in _hard_excl
                    and _eff_cap_for(p, _dest_grp) - float(_session_loads.get(p, 0)) >= total_w
                    and float(_lorry_cap_map.get(p, 0)) >= _pref_dest_min
                    and (_grp_cap_pre is None or float(_lorry_cap_map.get(p, 0)) <= _grp_cap_pre)
                ]
                if _pref_avail:
                    # OWNER-FIRST: use the tightest-fitting preferred lorry
                    # unconditionally. Fall back to open fleet only when ALL
                    # preferred lorries are full, unavailable, or size-capped
                    # (i.e. _pref_avail is empty — handled below).
                    _pref_avail.sort(key=lambda p: _eff_cap_for(p, _dest_grp) - float(_session_loads.get(p, 0)))
                    if _is_kuantan(_dominant_route, customer):
                        # Kuantan priority pool (explicit request): try
                        # VJN9910/BQY7823/BQU3875 first — whichever of these
                        # three is available fits tightest wins (_pref_avail
                        # is already tightest-fit sorted, so filtering
                        # preserves that order within the pool). Only fall
                        # through to the rest of the FIT IN LORRY list
                        # (BPE9788, BQX9983, WA6899M, VER2872) when none of
                        # the three have room that day.
                        _kuantan_pool = {"VJN9910", "BQY7823", "BQU3875"}
                        _kp = [p for p in _pref_avail if p in _kuantan_pool]
                        if _kp:
                            _pref_avail = _kp + [p for p in _pref_avail if p not in _kuantan_pool]
                    else:
                        # BQY7823 priority (explicit request, non-Kuantan
                        # routes — Kuantan has its own pool rule above):
                        # claim it ahead of tightest-fit ordering whenever
                        # it's a valid candidate for this route (i.e. the
                        # route is one it's listed for in LORRY DAILY
                        # PLANNING.xlsx's FIT IN LORRY sheet — that's what
                        # populated _preferred/_pref_avail here). This lets
                        # it serve those outstation routes first; it only
                        # falls through to urban work on days those routes
                        # don't need it.
                        if "BQY7823" in _pref_avail:
                            _pref_avail.remove("BQY7823")
                            _pref_avail.insert(0, "BQY7823")
                    _chosen = _pref_avail[0]
                    for it in group_items:
                        it["LORRY"] = _chosen
                    _session_loads[_chosen] = float(_session_loads.get(_chosen, 0)) + total_w
                    if _chosen not in _session_routes:
                        _session_routes[_chosen] = _dominant_route
                    _record_lorry_state(_chosen, _grp_state)
                    for it in _all_group:
                        sess["assigned"][it["DO NUMBER"]] = it.get("LORRY", "NO_LORRY")
                    return
                # All preferred lorries full / unavailable → open weight-based assignment

            # ── Destination-based lorry size enforcement ───────────────────────
            # Tiny-NS relaxation: a very small Seremban/NS load may use a small
            # lorry (far outstation stays strict).
            _dest_min_t = _eff_dest_min_ton(route, _dest_grp, total_w)
            # Exclude undersized lorries for long-distance destinations
            if _dest_min_t > 0:
                excluded = excluded | {
                    str(r["LORRY"]).strip().upper()
                    for _, r in engine.eligible_lorries.iterrows()
                    if float(r["TON"]) < _dest_min_t
                }
            # ── Tiny-item route guard ─────────────────────────────────────────
            # Routes with very small average DO weight (e.g. KV11A ~46 kg each)
            # must use a van or small lorry — a 14T truck cannot park in those
            # narrow streets and the turns-per-km economics don't work.
            _n_items = len(group_items)
            _avg_w = total_w / _n_items if _n_items > 0 else total_w
            if _avg_w <= _TINY_ITEM_AVG_WEIGHT_T:
                # Exclude lorries too large for tiny-item routes (narrow shophouse streets)
                excluded = excluded | {
                    str(r["LORRY"]).strip().upper()
                    for _, r in engine.eligible_lorries.iterrows()
                    if float(r["TON"]) >= LORRY_TINY_EXCL_TON
                }
            # ── REMARKS size requirement (FIELD 3) ────────────────────────────
            # A DO whose REMARKS demand a lorry size (e.g. "VAN", "LORRY KECIL",
            # "BELOW 10 TON") caps the lorries this group may use.  The binding
            # cap is the SMALLEST cap among the group's items (DOs without a
            # remark impose no cap).
            if _grp_cap_pre is not None:
                excluded = excluded | {
                    str(r["LORRY"]).strip().upper()
                    for _, r in engine.eligible_lorries.iterrows()
                    if float(r["TON"]) > _grp_cap_pre
                }

            # ── Within-session lorry sharing ──────────────────────────────────
            _share_pool = [
                (_eff_cap_for(p, _dest_grp) - float(_session_loads.get(p, 0)), p)
                for p in _session_loads
                if p in _lorry_cap_map
                and _eff_cap_for(p, _dest_grp) - float(_session_loads.get(p, 0)) >= total_w
                and p not in excluded
                and float(_lorry_cap_map.get(p, 0)) >= _dest_min_t
            ]
            if _share_pool:
                _compat = sorted(
                    [(r, p) for r, p in _share_pool
                     if _routes_on_same_way(route, _session_routes.get(p, ""))]
                )
                if _compat:
                    _shared = _compat[0][1]
                    for it in group_items:
                        it["LORRY"] = _shared
                    _session_loads[_shared] = float(_session_loads.get(_shared, 0)) + total_w
                    _record_lorry_state(_shared, _grp_state)
                    for it in _all_group:
                        sess["assigned"][it["DO NUMBER"]] = it.get("LORRY", "NO_LORRY")
                    return

            # Try single lorry for the whole group
            suggestions = engine.suggest(
                route=route,
                total_ton=total_w,
                unavailable=excluded,
                top_n=1,
                customer_name=customer,
                today_date_str=_today(),
                # Tiny NS/Seremban load whose outstation min was waived → let the
                # engine consider small lorries too (don't waste a big one).
                allow_small=(_dest_min_t == 0),
            )

            if suggestions:
                single_cap  = suggestions[0]["TON_CAPACITY"]
                single_util = total_w / single_cap if single_cap > 0 else 0

                # Rule 8: tightest-fit lorry would still be <10% loaded → leave blank.
                # Exception: never apply to urban routes (KL/Selangor) or tiny-item
                # routes — those items must always ship regardless of truck fill %.
                # Rule 8 targets outstation trucks (don't waste a 14T for 100 kg on
                # a 200 km run), not inner-city vans delivering AEON shophouse goods.
                _rule8_applies = (
                    _dest_grp not in _DEST_URBAN_GROUPS
                    and _avg_w > _TINY_ITEM_AVG_WEIGHT_T
                )
                if single_util < _MIN_UTIL and _rule8_applies:
                    _mark_no_lorry(group_items, "LOAD_BELOW_MIN_UTIL")
                else:
                    split_option = None
                    # Only split if utilization is very low (<30%) — weight-based
                    # selection already picks the smallest fitting lorry, so splitting
                    # at 60% would waste an extra truck unnecessarily.
                    if single_util < 0.30:
                        split_option = engine.suggest_split(
                            route=route,
                            total_ton=total_w,
                            unavailable=excluded,
                            single_util_threshold=0.60,
                        )
                    if split_option is not None:
                        # Each lorry in the split carries a portion of the items.
                        # Build bins with their allotted weight (PORTION).
                        bins = [
                            {"lorry": s["LORRY"],
                             "rows":  [],
                             "remain": s["PORTION"]}   # how much this bin can take
                            for s in split_option
                        ]
                        for s in split_option:
                            sess["unavailable"].add(s["LORRY"])
                        # Assign each item to exactly ONE bin (greedy, heaviest first)
                        item_bin: dict[str, str] = {}
                        for it in sorted(group_items, key=lambda x: x["WEIGHT"], reverse=True):
                            placed = False
                            for bin_ in bins:
                                if bin_["remain"] >= it["WEIGHT"] - 0.001:
                                    bin_["rows"].append({"DO": it["DO NUMBER"], "W": it["WEIGHT"]})
                                    bin_["remain"] -= it["WEIGHT"]
                                    item_bin[it["DO NUMBER"]] = bin_["lorry"]
                                    placed = True
                                    break
                            if not placed:
                                bins[0]["rows"].append({"DO": it["DO NUMBER"], "W": it["WEIGHT"]})
                                item_bin[it["DO NUMBER"]] = bins[0]["lorry"]
                        # Each item gets ONE lorry plate — no more "VEA2818, W3618U" for all
                        for it in group_items:
                            it["LORRY"] = item_bin.get(it["DO NUMBER"], bins[0]["lorry"])
                            it.pop("SPLIT_LORRIES", None)
                    else:
                        lorry = suggestions[0]["LORRY"]
                        for it in group_items:
                            it["LORRY"] = lorry
            else:
                # No single lorry fits — bin-pack across multiple lorries.
                # Build bins using tightest-fit first: ask suggest() for the
                # smallest lorry that can carry the remaining weight.  Fall
                # back to the largest available only when no single lorry
                # can carry the full remainder (partial-load pass).
                remain = total_w
                bins   = []
                # Honour REMARKS size cap (VAN/BELOW 5 TON etc.) throughout the
                # bin-pack loop — the same cap applied in the single-lorry path
                # must carry through here, otherwise a VAN group with no ≤2T
                # lorry available in the single pass falls into bin-pack and
                # picks a 13T lorry, ignoring the cap entirely.
                _bp_cap_excl: set = set()
                if _grp_cap_pre is not None:
                    _bp_cap_excl = {
                        str(r["LORRY"]).strip().upper()
                        for _, r in engine.eligible_lorries.iterrows()
                        if float(r["TON"]) > _grp_cap_pre
                    }
                # Honour the outstation minimum tonnage in the bin-pack loop too:
                # the single-lorry path excludes ≤5T lorries for outstation, but
                # the partial-load fallback (suggest with total_ton≈0) would
                # otherwise let a 4.2T lorry pick up an outstation tail item.
                _bp_dest_excl: set = set()
                if _dest_min_t > 0:
                    _bp_dest_excl = {
                        str(r["LORRY"]).strip().upper()
                        for _, r in engine.eligible_lorries.iterrows()
                        if float(r["TON"]) < _dest_min_t
                    }
                # Lorries picked as bins in THIS group's bin-pack. They must not be
                # re-picked within the same group, but are only committed to
                # sess["unavailable"] AFTER item distribution — a bin that ends up
                # empty leaves its lorry free for later groups (NS, PH03, …)
                # instead of being stranded idle-but-unavailable.
                _bp_picked: set = set()
                for _ in range(10):
                    if remain <= 0:
                        break
                    # Exclude lorries that already carry too much from this
                    # session to accept 'remain' more tons — otherwise the
                    # bin-pack re-picks a lorry loaded in an earlier group
                    # (e.g. BQU3875 filled to 20T getting a second 18T run).
                    _excl_session_full = {
                        p for p in _session_loads
                        if float(_lorry_cap_map.get(p, 0))
                           - float(_session_loads.get(p, 0)) < remain
                    }
                    excl = sess["unavailable"] | get_assigned_today() | _excl_session_full | _state_excl | _bp_cap_excl | _bp_dest_excl | _bp_picked | _grp_forbid | _reserve_excl | _strict_excl
                    # Tightest-fit pass: find smallest lorry that handles remain
                    sug = engine.suggest(route=route, total_ton=remain,
                                         unavailable=excl, top_n=20,
                                         customer_name=customer)
                    if not sug:
                        # No lorry can carry full remain — grab largest available
                        # for a partial load, then continue with what's left
                        sug = engine.suggest(route=route, total_ton=0.001,
                                             unavailable=excl, top_n=20)
                        if not sug:
                            break
                        sug.sort(key=lambda x: x["TON_CAPACITY"], reverse=True)
                    # Promote preferred lorries to the front of the bin-pack
                    # selection so config priority is respected even when the
                    # engine's route-history scoring would pick a different lorry.
                    _bp_pref = _preferred_lorries_for_route(route, engine)
                    if _bp_pref:
                        _bp_pref_idx = {p: i for i, p in enumerate(_bp_pref)}
                        _sug_pref = sorted(
                            [s for s in sug if s["LORRY"] in _bp_pref_idx
                             and s["TON_CAPACITY"] >= remain],
                            key=lambda s: _bp_pref_idx[s["LORRY"]],
                        )
                        if _sug_pref:
                            _sug_rest = [s for s in sug if s["LORRY"] not in _bp_pref_idx]
                            sug = _sug_pref + _sug_rest
                    lorry   = sug[0]["LORRY"]
                    cap     = sug[0]["TON_CAPACITY"]
                    portion = min(cap, remain)
                    # Use full lorry capacity for the bin, not just the arithmetic
                    # portion. Items overflow BQU3875's last few scraps and must
                    # fit in the next bin — which needs its full capacity available,
                    # not just the remaining-weight arithmetic.
                    bins.append({"lorry": lorry, "rows": [], "remain": cap})
                    # Reserve this lorry for the current group only (local set);
                    # do NOT commit to sess["unavailable"]/_session_loads yet — that
                    # happens after distribution, for bins that actually get items.
                    _bp_picked.add(lorry)
                    remain = round(remain - cap, 6)

                if remain <= 0 and bins:
                    # Distribute items into bins. First place whole SAME-DESTINATION
                    # clusters (same route+state+city+GPS) into one bin so a single
                    # drop is never split across lorries unless it exceeds one lorry
                    # (same-route clusters may use the ×NAIK overage). Then place any
                    # remaining items individually, heaviest-first.
                    item_bin2: dict[str, str] = {}
                    max_bin_cap = max(b["remain"] + sum(
                        x["W"] for x in b["rows"]) for b in bins) if bins else 0

                    def _bp_dest_key(_it):
                        _la, _lo = _it.get("GPS_LAT"), _it.get("GPS_LON")
                        return (str(_it.get("ROUTE", "")).strip().upper(),
                                str(_it.get("STATE", "")).strip().upper(),
                                str(_it.get("CITY", "")).strip().upper(),
                                round(_la, 4) if _la is not None else None,
                                round(_lo, 4) if _lo is not None else None)

                    _bp_clusters: dict = {}
                    for it in group_items:
                        _bp_clusters.setdefault(_bp_dest_key(it), []).append(it)
                    _placed_ids: set = set()
                    # Heaviest cluster first; try to seat the whole cluster in one bin.
                    for _cl in sorted(_bp_clusters.values(),
                                      key=lambda c: -sum(x["WEIGHT"] for x in c)):
                        if len(_cl) < 2:
                            continue                 # singletons handled below
                        _cw = sum(x["WEIGHT"] for x in _cl)
                        _cmax = min((x["MAX_TON"] for x in _cl if x.get("MAX_TON") is not None),
                                    default=None)
                        for bin_ in bins:
                            _blc = float(_lorry_cap_map.get(bin_["lorry"], 0))
                            if _cmax is not None and _blc > _cmax:
                                continue
                            _bused = sum(x["W"] for x in bin_["rows"])
                            if _bused + _cw <= _blc * NAIK_FACTOR + 0.001:
                                for it in _cl:
                                    bin_["rows"].append({"DO": it["DO NUMBER"], "W": it["WEIGHT"]})
                                    bin_["remain"] -= it["WEIGHT"]
                                    item_bin2[it["DO NUMBER"]] = bin_["lorry"]
                                    _placed_ids.add(it["DO NUMBER"])
                                break
                    for it in sorted(group_items, key=lambda x: x["WEIGHT"], reverse=True):
                        if it["DO NUMBER"] in _placed_ids:
                            continue
                        placed = False
                        for bin_ in bins:
                            # Skip bins whose lorry exceeds this item's REMARKS size cap
                            if (it.get("MAX_TON") is not None
                                    and float(_lorry_cap_map.get(bin_["lorry"], 0)) > it["MAX_TON"]):
                                continue
                            if bin_["remain"] >= it["WEIGHT"] - 0.001:
                                bin_["rows"].append({"DO": it["DO NUMBER"], "W": it["WEIGHT"]})
                                bin_["remain"] -= it["WEIGHT"]
                                item_bin2[it["DO NUMBER"]] = bin_["lorry"]
                                placed = True
                                break
                        if not placed:
                            # Bins are full — try to grab one more lorry rather
                            # than giving up (greedy fill can leave tiny tail items
                            # stranded even when arithmetic says they should fit).
                            _excl_retry_sf = {
                                p for p in _session_loads
                                if float(_lorry_cap_map.get(p, 0))
                                   - float(_session_loads.get(p, 0)) < it["WEIGHT"]
                            }
                            excl_retry = sess["unavailable"] | get_assigned_today() | _excl_retry_sf | _state_excl | _bp_cap_excl | _bp_dest_excl | _bp_picked | _grp_forbid | _reserve_excl | _strict_excl
                            extra_sug  = engine.suggest(
                                route=route, total_ton=it["WEIGHT"],
                                unavailable=excl_retry, top_n=1,
                                customer_name=customer,
                                today_date_str=_today(),
                            )
                            if extra_sug:
                                extra_lorry = extra_sug[0]["LORRY"]
                                extra_cap   = extra_sug[0]["TON_CAPACITY"]
                                new_bin = {"lorry": extra_lorry, "rows": [], "remain": extra_cap}
                                bins.append(new_bin)
                                _bp_picked.add(extra_lorry)
                                new_bin["rows"].append({"DO": it["DO NUMBER"], "W": it["WEIGHT"]})
                                new_bin["remain"] -= it["WEIGHT"]
                                item_bin2[it["DO NUMBER"]] = extra_lorry
                            else:
                                item_bin2[it["DO NUMBER"]] = "NO_LORRY"
                                _unassigned_reasons[it["DO NUMBER"]] = "CAPACITY_FULL"
                    # Commit only bins that actually received items. Empty bins'
                    # lorries stay available for later groups (NS, PH03, PH04, …)
                    # instead of being stranded idle-but-unavailable.
                    for _b in bins:
                        if not _b["rows"]:
                            continue
                        _blorry = _b["lorry"]
                        _bweight = sum(x["W"] for x in _b["rows"])
                        if _dest_grp not in _DEST_URBAN_GROUPS:
                            sess["unavailable"].add(_blorry)
                        else:
                            _session_loads[_blorry] = float(_session_loads.get(_blorry, 0)) + _bweight
                    for it in group_items:
                        it["LORRY"] = item_bin2.get(it["DO NUMBER"], "NO_LORRY")
                        if it["LORRY"] == "NO_LORRY":
                            _unassigned_reasons.setdefault(it["DO NUMBER"], "CAPACITY_FULL")
                        it.pop("SPLIT_LORRIES", None)
                else:
                    # Bin-pack failed — all lorries taken or too small.
                    # Last resort: find the tightest-fitting available lorry
                    # that is NOT overloaded (combined weight ≤ capacity).
                    # If even the largest lorry can't handle the weight → NO_LORRY.
                    _excl_lr_sf = {
                        p for p in _session_loads
                        if float(_lorry_cap_map.get(p, 0))
                           - float(_session_loads.get(p, 0)) < total_w
                    }
                    excl_final = sess["unavailable"] | get_assigned_today() | _excl_lr_sf | _state_excl | _bp_cap_excl | _bp_dest_excl | _grp_forbid | _reserve_excl | _strict_excl
                    last_resort = engine.suggest_largest_available(
                        route, excl_final, _today(), total_ton=total_w)
                    if last_resort:
                        lr_cap  = last_resort[0]["TON_CAPACITY"]
                        lr_util = total_w / lr_cap if lr_cap > 0 else 0
                        if lr_util < _MIN_UTIL:
                            # Even the smallest available lorry would be <10% loaded
                            _mark_no_lorry(group_items, "CAPACITY_FULL")
                        else:
                            lorry = last_resort[0]["LORRY"]
                            for it in group_items:
                                it["LORRY"] = lorry
                    else:
                        # No lorry can carry this weight without overloading
                        _mark_no_lorry(group_items, "NO_ELIGIBLE_LORRY")

            for it in _all_group:
                sess["assigned"][it["DO NUMBER"]] = it["LORRY"]

            # Update session load and state trackers for subsequent groups
            for _it in _all_group:
                _pl = _it.get("LORRY")
                if _pl and _pl not in {"NO_LORRY", "SPLIT", "SKIPPED", "OTHER_USER", "NOT_TODAY", "REMARKS_SKIP", "OUT_SOURCE", "PAST_DATE", "WRONG_TRIP", "", None}:
                    _session_loads[_pl] = float(_session_loads.get(_pl, 0)) + _it["WEIGHT"]
                    if _pl not in _session_routes:
                        _session_routes[_pl] = route
                    _record_lorry_state(_pl, _it.get("STATE", ""))

        # ── FIT IN LORRY contention priority ───────────────────────────────────
        # FIT IN LORRY plates are a PREFERENCE, not a restriction (RULE 9A) — any
        # Available lorry owned by this user stays eligible even if not on a
        # route's list. But two groups can both prefer the SAME plate. For each
        # group with FIT IN LORRY data, simulate which preferred plate it would
        # naturally claim alone (tightest-fitting available). Where two collide,
        # let the group that pushes that plate CLOSER TO FULL go first, so it
        # claims the plate; the loser falls through to its next preferred plate
        # or the open fleet. Only reorders sorted_groups — no bucketing/merge
        # change. (Your NS05-14T vs PH09-13T on VJN9910 example.)
        if getattr(engine, "fit_in_lorry", None) and \
                sess.get("user_id", "").strip().upper() == engine.fit_in_lorry_owner:
            def _grp_dominant_route(g):
                _rc: dict = {}
                for _it in g:
                    _rc[_it.get("ROUTE", "")] = _rc.get(_it.get("ROUTE", ""), 0) + 1
                return max(_rc, key=lambda r: _rc[r]) if _rc else ""

            _grp_routes  = [_grp_dominant_route(g) for g in sorted_groups]
            _grp_weights = [sum(it["WEIGHT"] for it in g) for g in sorted_groups]
            _dry_excl = sess["unavailable"] | get_assigned_today()

            def _natural_preferred_pick(_r, _w):
                _plates = engine.fit_in_lorry_preferred(_r)
                if not _plates:
                    return None
                _cands = [p for p in _plates
                          if p in _lorry_cap_map and p not in _dry_excl
                          and float(_lorry_cap_map[p]) >= _w]
                if not _cands:
                    return None
                return min(_cands, key=lambda p: float(_lorry_cap_map[p]))

            _grp_bestfit = [_natural_preferred_pick(_r, _grp_weights[_gi])
                            for _gi, _r in enumerate(_grp_routes)]
            _plate_claimants: dict = defaultdict(list)
            for _gi, _p in enumerate(_grp_bestfit):
                if _p:
                    _plate_claimants[_p].append(_gi)

            _fil_boost = [0.0] * len(sorted_groups)
            for _plate, _gis in _plate_claimants.items():
                if len(_gis) < 2:
                    continue
                _cap = _lorry_cap_map.get(_plate)
                if not _cap:
                    continue
                _scores = {_gi: min(_grp_weights[_gi], _cap) / _cap for _gi in _gis}
                _winner = max(_gis, key=lambda gi: (_scores[gi], -gi))
                _fil_boost[_winner] = max(_fil_boost[_winner], _scores[_winner])

            if any(_fil_boost):
                _fil_order = sorted(range(len(sorted_groups)), key=lambda i: -_fil_boost[i])
                sorted_groups = [sorted_groups[i] for i in _fil_order]

        for group in sorted_groups:
            _assign_group(group)

        # ── Consolidation pass ────────────────────────────────────────────────
        # Any item still marked NO_LORRY gets a second chance: find a lorry from
        # this session with enough remaining capacity, or any eligible lorry at all.
        # Preferred lorries are tried FIRST but if all are full/unavailable, any
        # eligible lorry is used — the DO must ship.
        _excl_consol = sess["unavailable"] | get_assigned_today()

        # Sort unassigned items by date so consolidation assigns oldest DOs first.
        # Items already assigned keep their position; stable sort preserves relative
        # order within the same date.
        items.sort(key=lambda it: (
            _parse_date_sortkey(it.get("DATE", ""))
            if it.get("LORRY") in (None, "NO_LORRY") else ""
        ))

        # Pre-build lorry→assigned-states map for state boundary checks
        _consol_lorry_states: dict[str, set] = {}
        for _cit in items:
            _cpl = _cit.get("LORRY")
            _cst = _cit.get("STATE", "").strip().upper()
            if _cpl and _cst and _cpl not in {
                "NO_LORRY", "OTHER_USER", "NOT_TODAY", "REMARKS_SKIP", "OUT_SOURCE", "PAST_DATE", "WRONG_TRIP", "SPLIT", "SKIPPED", "", None
            }:
                _consol_lorry_states.setdefault(_cpl, set()).add(_cst)

        for it in items:
            if it.get("LORRY") not in (None, "NO_LORRY"):
                continue
            w = it["WEIGHT"]
            it_dest = _classify_dest_group(it.get("ROUTE", ""), it.get("STATE", ""))
            it_state = it.get("STATE", "").strip().upper()
            _it_dest_min_t = _DEST_MIN_TON.get(it_dest, 0.0)
            _it_pref = _preferred_lorries_for_route(it.get("ROUTE", ""), engine)
            _it_strict_excl = _strict_route_excl(it.get("ROUTE", ""))

            def _consol_eligible(p: str) -> bool:
                """Core eligibility check for consolidation candidates."""
                if float(_lorry_cap_map.get(p, 0)) - float(_session_loads.get(p, 0)) < w:
                    return False
                if p in _excl_consol or p in _it_strict_excl:
                    return False
                if float(_lorry_cap_map.get(p, 0)) < _it_dest_min_t:
                    return False
                # REMARKS size cap (FIELD 3) — DO must not exceed its lorry-size limit
                if it.get("MAX_TON") is not None and float(_lorry_cap_map.get(p, 0)) > it["MAX_TON"]:
                    return False
                # Hard state boundary — never mix states regardless of tier.
                # _session_lorry_states tracks states assigned in this session;
                # _consol_lorry_states tracks assignments made during this pass.
                if it_state:
                    _lst = _consol_lorry_states.get(p, set()) | _session_lorry_states.get(p, set())
                    if _lst and it_state not in _lst:
                        return False
                # Direction guard — applies to ALL routes (urban AND outstation).
                # Urban routes: ONLY allow same corridor group (never GPS-bearing match
                # alone — e.g. KV08A and KV11A both point "east" but are incompatible zones).
                # Outstation routes: also allow same-way GPS bearing.
                _lorry_rt = _session_routes.get(p, "")
                if _lorry_rt and it.get("ROUTE", "") != _lorry_rt:
                    _it_urban_r     = _classify_dest_group(it.get("ROUTE", "")) in _DEST_URBAN_GROUPS
                    _lorry_urban_r  = _classify_dest_group(_lorry_rt) in _DEST_URBAN_GROUPS
                    if _it_urban_r or _lorry_urban_r:
                        # Urban: same corridor group only
                        if not _same_corridor_group(it.get("ROUTE", ""), _lorry_rt):
                            return False
                    else:
                        # Outstation: also allow same-way bearing
                        if (not _same_corridor_group(it.get("ROUTE", ""), _lorry_rt)
                                and not _routes_on_same_way(it.get("ROUTE", ""), _lorry_rt)):
                            return False
                # GPS city-proximity guard for outstation: if this item's GPS
                # is known, only allow lorries whose assigned items' centroid
                # is within _MAX_CITY_MERGE_KM_OUTSTATION of this item.
                if (it_dest not in _DEST_URBAN_GROUPS
                        and it.get("GPS_LAT") is not None
                        and it.get("GPS_LON") is not None):
                    _p_route = _session_routes.get(p, "")
                    _p_cent  = _live_centroids.get(_p_route)
                    if _p_cent:
                        _p_dist = _haversine_km(
                            it["GPS_LAT"], it["GPS_LON"],
                            _p_cent[0],   _p_cent[1]
                        )
                        if _p_dist > _MAX_CITY_MERGE_KM_OUTSTATION:
                            return False
                return True

            # Priority tiers — state boundary always enforced; prefer GPS-nearest lorry.
            # 1. Preferred lorry for this route
            # 2. Any eligible lorry sorted by GPS proximity (nearest first)
            _it_lat = it.get("GPS_LAT")
            _it_lon = it.get("GPS_LON")

            def _gps_dist_to_lorry(plate: str) -> float:
                """Distance (km) from this item to lorry's current centroid, or 0."""
                _pr  = _session_routes.get(plate, "")
                _pc  = _live_centroids.get(_pr)
                if _pc and _it_lat is not None and _it_lon is not None:
                    return _haversine_km(_it_lat, _it_lon, _pc[0], _pc[1])
                return 0.0

            _tiers = [
                [p for p in _it_pref if p in _lorry_cap_map and _consol_eligible(p)],
                sorted(
                    [p for p in _lorry_cap_map if _consol_eligible(p)],
                    key=_gps_dist_to_lorry
                ),
            ]
            _cand_plates = next((t for t in _tiers if t), [])
            if not _cand_plates:
                continue

            # Among eligible plates, prefer same-route/corridor then tightest fit.
            # For urban items use same-corridor only; outstation may also use same-way.
            def _compat_check(plate):
                _pr = _session_routes.get(plate, "")
                if not _pr or it["ROUTE"] == _pr:
                    return True
                if it_dest in _DEST_URBAN_GROUPS or _classify_dest_group(_pr) in _DEST_URBAN_GROUPS:
                    return _same_corridor_group(it["ROUTE"], _pr)
                return _routes_on_same_way(it["ROUTE"], _pr)
            _compat = sorted(
                [(float(_lorry_cap_map.get(p, 0)) - float(_session_loads.get(p, 0)), p)
                 for p in _cand_plates
                 if _compat_check(p)]
            )
            _any_fit = sorted(
                [(float(_lorry_cap_map.get(p, 0)) - float(_session_loads.get(p, 0)), p)
                 for p in _cand_plates]
            )
            _pick = (_compat or _any_fit)[0][1]
            it["LORRY"] = _pick
            _session_loads[_pick] = float(_session_loads.get(_pick, 0)) + w
            if it_state:   # guard against empty state poisoning future checks
                _consol_lorry_states.setdefault(_pick, set()).add(it_state)
            _record_lorry_state(_pick, it_state)
            if _pick not in _session_routes:
                _session_routes[_pick] = it["ROUTE"]
            sess["assigned"][it["DO NUMBER"]] = _pick

        # ── Force-assign pass ─────────────────────────────────────────────────
        # Any item still NO_LORRY after the consolidation pass MUST ship today.
        # Relax the daily-log exclusion (only the current sess["unavailable"] and
        # physical size/state constraints still apply) so lorries that appear in
        # today's log from a PREVIOUS session don't block scheduled DOs.
        for it in items:
            if it.get("LORRY") not in (None, "NO_LORRY"):
                continue
            w         = it["WEIGHT"]
            it_route  = it.get("ROUTE", "")
            it_state  = it.get("STATE", "").strip().upper()
            it_dest   = _classify_dest_group(it_route, it.get("STATE", ""))
            _it_dest_min_t = _DEST_MIN_TON.get(it_dest, 0.0)
            _it_strict = _strict_route_excl(it_route)
            _it_is_urban = it_dest in _DEST_URBAN_GROUPS

            _force_candidates = []
            for _fp in _lorry_cap_map:
                if _fp in sess.get("unavailable", set()):
                    continue
                _fp_cap  = float(_lorry_cap_map.get(_fp, 0))
                _fp_load = float(_session_loads.get(_fp, 0))
                _fp_rem  = _fp_cap - _fp_load
                if _fp_rem < w:
                    continue
                if _fp_cap < _it_dest_min_t:
                    continue
                # REMARKS size cap (FIELD 3)
                if it.get("MAX_TON") is not None and _fp_cap > it["MAX_TON"]:
                    continue
                if _fp in _it_strict:
                    continue
                # Urban↔outstation guard
                if (not _it_is_urban
                        and _session_routes.get(_fp, "")
                        and _classify_dest_group(_session_routes.get(_fp, "")) in _DEST_URBAN_GROUPS):
                    continue
                # Route direction guard — same route, same corridor, or (outstation only) same-way bearing
                _fp_rt = _session_routes.get(_fp, "")
                if _fp_rt and it_route != _fp_rt:
                    _it_urban_fa    = _it_is_urban
                    _fp_urban_fa    = _classify_dest_group(_fp_rt) in _DEST_URBAN_GROUPS
                    if _it_urban_fa or _fp_urban_fa:
                        if not _same_corridor_group(it_route, _fp_rt):
                            continue
                    else:
                        if (not _same_corridor_group(it_route, _fp_rt)
                                and not _routes_on_same_way(it_route, _fp_rt)):
                            continue
                # State boundary
                _fp_states = _consol_lorry_states.get(_fp, set()) | _session_lorry_states.get(_fp, set())
                if it_state and _fp_states and not any(_states_compatible(it_state, s) for s in _fp_states):
                    continue
                _force_candidates.append((_fp_rem, _fp))

            if not _force_candidates:
                continue

            _force_candidates.sort()    # smallest remaining capacity first → tightest fit
            _fpick = _force_candidates[0][1]
            it["LORRY"] = _fpick
            _session_loads[_fpick] = float(_session_loads.get(_fpick, 0)) + w
            if it_state:
                _consol_lorry_states.setdefault(_fpick, set()).add(it_state)
            _record_lorry_state(_fpick, it_state)
            if _fpick not in _session_routes:
                _session_routes[_fpick] = it_route
            sess["assigned"][it["DO NUMBER"]] = _fpick
            _unassigned_reasons.pop(it["DO NUMBER"], None)

        # ── Last-resort overflow / split pass ────────────────────────────────
        # Items still unassigned after force-assign:
        #   • Groups of >1 item sharing the same city+state → try a bin-pack
        #     split across two eligible lorries first; overflow only if that fails.
        #   • Single remaining item → allow ≤1T overage directly.
        # Process items one at a time so each assignment updates _session_loads.

        def _overflow_eligible_lorries(it_w, it_route, it_state, it_dest, it_max_ton=None):
            """Return [(remaining_cap, plate)] sorted by most-remaining-cap first."""
            _it_is_urban = it_dest in _DEST_URBAN_GROUPS
            _it_strict   = _strict_route_excl(it_route)
            _it_dest_min = _DEST_MIN_TON.get(it_dest, 0.0)
            _cands = []
            for _op in _lorry_cap_map:
                if _op in sess.get("unavailable", set()):
                    continue
                _oc = float(_lorry_cap_map.get(_op, 0))
                _ol = float(_session_loads.get(_op, 0))
                _or = _oc - _ol
                if _op in _it_strict:
                    continue
                # Outstation minimum tonnage: lorries ≤5T can never serve an
                # outstation route (LARGE_LONG / MEDIUM_LONG), even as overflow.
                if _oc < _it_dest_min:
                    continue
                # REMARKS size cap (FIELD 3)
                if it_max_ton is not None and _oc > it_max_ton:
                    continue
                if (not _it_is_urban
                        and _session_routes.get(_op, "")
                        and _classify_dest_group(_session_routes.get(_op, "")) in _DEST_URBAN_GROUPS):
                    continue
                # Route direction guard — urban routes only allow same corridor group
                _op_rt = _session_routes.get(_op, "")
                if _op_rt and it_route != _op_rt:
                    _op_urban = _classify_dest_group(_op_rt) in _DEST_URBAN_GROUPS
                    if _it_is_urban or _op_urban:
                        if not _same_corridor_group(it_route, _op_rt):
                            continue
                    else:
                        if (not _same_corridor_group(it_route, _op_rt)
                                and not _routes_on_same_way(it_route, _op_rt)):
                            continue
                _op_sts = _consol_lorry_states.get(_op, set()) | _session_lorry_states.get(_op, set())
                if it_state and _op_sts and not any(_states_compatible(it_state, s) for s in _op_sts):
                    continue
                if _or - it_w < -1.0:   # more than 1T short → skip
                    continue
                _cands.append((_or, _op))
            _cands.sort(reverse=True)
            return _cands

        def _do_overflow_assign(it, candidates):
            """Commit item to first candidate; return True on success."""
            if not candidates:
                return False
            _opick = candidates[0][1]
            it["LORRY"] = _opick
            _session_loads[_opick] = float(_session_loads.get(_opick, 0)) + it["WEIGHT"]
            _st = it.get("STATE", "").strip().upper()
            if _st:
                _consol_lorry_states.setdefault(_opick, set()).add(_st)
            _record_lorry_state(_opick, _st)
            if _opick not in _session_routes:
                _session_routes[_opick] = it.get("ROUTE", "")
            sess["assigned"][it["DO NUMBER"]] = _opick
            _unassigned_reasons.pop(it["DO NUMBER"], None)
            return True

        # Group still-unassigned items by (city, state) to decide split vs overflow
        _of_groups: dict[tuple, list] = {}
        for _it in items:
            if _it.get("LORRY") not in (None, "NO_LORRY"):
                continue
            _gk = (
                _it.get("CITY",  "").strip().upper(),
                _it.get("STATE", "").strip().upper(),
            )
            _of_groups.setdefault(_gk, []).append(_it)

        for _gk, _gits in _of_groups.items():
            if len(_gits) == 1:
                # Single remaining item → overflow directly
                _it = _gits[0]
                _cands = _overflow_eligible_lorries(
                    _it["WEIGHT"], _it.get("ROUTE",""),
                    _it.get("STATE","").strip().upper(),
                    _classify_dest_group(_it.get("ROUTE",""), _it.get("STATE","")),
                    _it.get("MAX_TON"),
                )
                _do_overflow_assign(_it, _cands)
            else:
                # Multiple items — try bin-pack split across two lorries first
                _g_total = sum(x["WEIGHT"] for x in _gits)
                _g_state = _gits[0].get("STATE", "").strip().upper()
                _g_route = _gits[0].get("ROUTE", "")
                _g_dest  = _classify_dest_group(_g_route, _gits[0].get("STATE",""))
                _g_urban = _g_dest in _DEST_URBAN_GROUPS
                _g_strict= _strict_route_excl(_g_route)
                _g_dest_min = _DEST_MIN_TON.get(_g_dest, 0.0)

                # Find up to 2 eligible lorries (ignoring capacity floor, state OK)
                _split_pool = []
                for _op in _lorry_cap_map:
                    if _op in sess.get("unavailable", set()):
                        continue
                    _oc = float(_lorry_cap_map.get(_op, 0))
                    _ol = float(_session_loads.get(_op, 0))
                    _or = _oc - _ol
                    if _op in _g_strict:
                        continue
                    # Outstation minimum tonnage: ≤5T lorries can't serve outstation
                    if _oc < _g_dest_min:
                        continue
                    # Route direction guard
                    _op_rt2 = _session_routes.get(_op, "")
                    if _op_rt2 and _g_route != _op_rt2:
                        if (not _same_corridor_group(_g_route, _op_rt2)
                                and not _routes_on_same_way(_g_route, _op_rt2)):
                            continue
                    _op_sts = _consol_lorry_states.get(_op, set()) | _session_lorry_states.get(_op, set())
                    if _g_state and _op_sts and not any(_states_compatible(_g_state, s) for s in _op_sts):
                        continue
                    if _or <= 0:
                        continue
                    _split_pool.append((_or, _op))
                _split_pool.sort(reverse=True)

                _split_ok = False
                if len(_split_pool) >= 2:
                    _b1_rem, _b1 = _split_pool[0]
                    _b2_rem, _b2 = _split_pool[1]
                    # Greedy bin-pack: heaviest first
                    _b1_items, _b2_items = [], []
                    _b1_load = _b2_load = 0.0
                    for _sit in sorted(_gits, key=lambda x: x["WEIGHT"], reverse=True):
                        _sit_max = _sit.get("MAX_TON")
                        _b1_ok = (_sit_max is None or float(_lorry_cap_map.get(_b1, 0)) <= _sit_max)
                        _b2_ok = (_sit_max is None or float(_lorry_cap_map.get(_b2, 0)) <= _sit_max)
                        if _b1_ok and _b1_load + _sit["WEIGHT"] <= _b1_rem + 1.0:
                            _b1_items.append(_sit)
                            _b1_load += _sit["WEIGHT"]
                        elif _b2_ok and _b2_load + _sit["WEIGHT"] <= _b2_rem + 1.0:
                            _b2_items.append(_sit)
                            _b2_load += _sit["WEIGHT"]
                        else:
                            break   # item doesn't fit either bin
                    if len(_b1_items) + len(_b2_items) == len(_gits):
                        # All items placed across two lorries
                        for _sit in _b1_items:
                            _sit["LORRY"] = _b1
                            sess["assigned"][_sit["DO NUMBER"]] = _b1
                            _unassigned_reasons.pop(_sit["DO NUMBER"], None)
                        for _sit in _b2_items:
                            _sit["LORRY"] = _b2
                            sess["assigned"][_sit["DO NUMBER"]] = _b2
                            _unassigned_reasons.pop(_sit["DO NUMBER"], None)
                        _session_loads[_b1] = float(_session_loads.get(_b1, 0)) + _b1_load
                        _session_loads[_b2] = float(_session_loads.get(_b2, 0)) + _b2_load
                        if _g_state:
                            _consol_lorry_states.setdefault(_b1, set()).add(_g_state)
                            _consol_lorry_states.setdefault(_b2, set()).add(_g_state)
                        _record_lorry_state(_b1, _g_state)
                        _record_lorry_state(_b2, _g_state)
                        if _b1 not in _session_routes:
                            _session_routes[_b1] = _g_route
                        if _b2 not in _session_routes:
                            _session_routes[_b2] = _g_route
                        _split_ok = True

                if not _split_ok:
                    # Split failed — overflow each item individually
                    for _it in _gits:
                        if _it.get("LORRY") not in (None, "NO_LORRY"):
                            continue
                        _cands = _overflow_eligible_lorries(
                            _it["WEIGHT"], _it.get("ROUTE",""),
                            _it.get("STATE","").strip().upper(),
                            _classify_dest_group(_it.get("ROUTE",""), _it.get("STATE","")),
                            _it.get("MAX_TON"),
                        )
                        _do_overflow_assign(_it, _cands)

        # ── Lorry-swap optimisation ───────────────────────────────────────────
        # Reduce waste on large lorries (≥10T) by swapping them with a smaller
        # lorry that is currently carrying a heavier load.  Conditions for a
        # valid swap: (A is large, B is small), A's load fits on B, B's load is
        # heavier than A's (so A's utilisation improves after taking B's items).
        # Pick the swap with the biggest waste reduction each round; repeat
        # until no improving swap remains.
        _LARGE_T = LORRY_LARGE_MIN_TON
        _pit: dict[str, list] = {}
        for _it in items:
            _pl = _it.get("LORRY")
            if _pl and _pl not in {"NO_LORRY", "SPLIT", "SKIPPED", "OTHER_USER", "PAST_DATE", "WRONG_TRIP", "", None}:
                _pit.setdefault(_pl, []).append(_it)

        _swap_ok = True
        while _swap_ok:
            _swap_ok = False
            _ploads = {p: sum(x["WEIGHT"] for x in its) for p, its in _pit.items()}
            _best_delta, _best_pa, _best_pb = 0.0, None, None
            for _pa, _pa_its in list(_pit.items()):
                _cap_a = _lorry_cap_map.get(_pa, 0)
                if _cap_a < _LARGE_T:
                    continue  # A must be a large lorry
                _load_a = _ploads[_pa]
                for _pb, _pb_its in list(_pit.items()):
                    if _pb == _pa:
                        continue
                    _cap_b = _lorry_cap_map.get(_pb, 0)
                    if _cap_b >= _LARGE_T:
                        continue  # B must be smaller than A
                    _load_b = _ploads[_pb]
                    if _load_b <= _load_a:
                        continue  # swap only improves if B is heavier
                    if _load_a > _cap_b:
                        continue  # A's items must physically fit on B
                    # Only swap if A's load would fill B to ≥70%.
                    # This prevents cascading swaps that move heavy, well-fitted
                    # loads off their historically preferred smaller lorries
                    # (e.g. KV19A at 92% on BMN3682 should stay there).
                    if _load_a / _cap_b < 0.70:
                        continue
                    _delta = _load_a - _load_b  # negative = waste reduction
                    if _delta < _best_delta:
                        _best_delta = _delta
                        _best_pa, _best_pb = _pa, _pb
            if _best_pa:
                # Validate swap: both lorries must be allowed for the routes
                # they'd receive after the swap (strict reservations + direction).
                _routes_pa = {it["ROUTE"] for it in _pit[_best_pa] if it.get("ROUTE")}
                _routes_pb = {it["ROUTE"] for it in _pit[_best_pb] if it.get("ROUTE")}
                # Check strict route rules for the swap
                _swap_pa_ok = not any(
                    _best_pa in _strict_route_excl(r) for r in _routes_pb
                )
                _swap_pb_ok = not any(
                    _best_pb in _strict_route_excl(r) for r in _routes_pa
                )
                # Check geographic direction: lorry A can only receive B's routes
                # if they share a direction with A's current outstation routes (if any).
                _dest_pa_routes = [_classify_dest_group(r) for r in _routes_pa]
                _dest_pb_routes = [_classify_dest_group(r) for r in _routes_pb]
                _outstation_dir_ok = True
                for _ra in _routes_pa:
                    for _rb in _routes_pb:
                        _da = _classify_dest_group(_ra)
                        _db = _classify_dest_group(_rb)
                        if (_da not in _DEST_URBAN_GROUPS
                                and _db not in _DEST_URBAN_GROUPS
                                and not _routes_on_same_way(_ra, _rb)):
                            _outstation_dir_ok = False
                            break
                    if not _outstation_dir_ok:
                        break
                # REMARKS size cap: a swap must not place a capped item onto a
                # lorry larger than its cap. A's items move to B and B's to A.
                _cap_a2 = float(_lorry_cap_map.get(_best_pa, 0))
                _cap_b2 = float(_lorry_cap_map.get(_best_pb, 0))
                _swap_cap_ok = all(
                    _it.get("MAX_TON") is None or _cap_b2 <= _it["MAX_TON"]
                    for _it in _pit[_best_pa]
                ) and all(
                    _it.get("MAX_TON") is None or _cap_a2 <= _it["MAX_TON"]
                    for _it in _pit[_best_pb]
                )
                # Outstation minimum: after the swap B receives A's routes and A
                # receives B's. Neither lorry may end up too small for an
                # outstation route it would then have to run.
                _swap_dstmin_ok = all(
                    _cap_b2 >= _DEST_MIN_TON.get(_classify_dest_group(r), 0.0)
                    for r in _routes_pa
                ) and all(
                    _cap_a2 >= _DEST_MIN_TON.get(_classify_dest_group(r), 0.0)
                    for r in _routes_pb
                )
                if (_swap_pa_ok and _swap_pb_ok and _outstation_dir_ok
                        and _swap_cap_ok and _swap_dstmin_ok):
                    for _it in _pit[_best_pa]:
                        _it["LORRY"] = _best_pb
                        sess["assigned"][_it["DO NUMBER"]] = _best_pb
                    for _it in _pit[_best_pb]:
                        _it["LORRY"] = _best_pa
                        sess["assigned"][_it["DO NUMBER"]] = _best_pa
                    _pit[_best_pa], _pit[_best_pb] = _pit[_best_pb], _pit[_best_pa]
                    _swap_ok = True

        # ── Same-route merge pass ─────────────────────────────────────────────
        # If two lorries carry compatible routes and lorry A's entire load fits
        # inside lorry B's remaining capacity, move all of A's DOs onto B and
        # free lorry A.  Picks the merge that maximises B's utilisation after
        # the move so we prefer consolidating onto the tightest-fitting lorry.
        # Example: VEA2818 (1T, PH09, 0.6T) + WLD8738 (5T, PH09, 0.538T)
        #   → 0.6T fits on WLD8738 (remaining 4.46T) → merge → WLD8738 at 22.8%
        _merge_ok = True
        while _merge_ok:
            _merge_ok = False
            _ploads = {p: sum(x["WEIGHT"] for x in its) for p, its in _pit.items()}
            _best_gain, _best_src, _best_dst = 0.0, None, None
            for _pa in list(_pit.keys()):
                _load_a = _ploads[_pa]
                _route_a = next((it["ROUTE"] for it in _pit[_pa] if it.get("ROUTE")), "") \
                           or _session_routes.get(_pa, "")
                for _pb in list(_pit.keys()):
                    if _pb == _pa:
                        continue
                    _load_b = _ploads[_pb]
                    _cap_b  = float(_lorry_cap_map.get(_pb, 0))
                    # Route compatibility: check source route against ALL routes
                    # already on the destination lorry (not just the first item).
                    # This lets PH07 merge into BQX9983 via the shared-waypoint
                    # chain  PH07↔PH03(JERANTUT)  even though PH07 and PH02
                    # (BQX9983's first item) share no waypoints directly.
                    _routes_b = {it["ROUTE"] for it in _pit[_pb] if it.get("ROUTE")}
                    if not _routes_b:
                        _sr = _session_routes.get(_pb, "")
                        if _sr:
                            _routes_b = {_sr}
                    if _route_a and _routes_b:
                        if not any(_routes_on_same_way(_route_a, _rb) for _rb in _routes_b):
                            continue
                    # Same-route items may fill up to 10 % over rated capacity
                    # so they are never split across lorries unnecessarily.
                    _same_rt = bool(_route_a and _route_a in _routes_b)
                    _eff_cap_b = _cap_b * (SAME_ROUTE_NAIK if _same_rt else 1.0)
                    if _load_a + _load_b > _eff_cap_b:
                        continue
                    # Score by utilisation GAIN on the destination lorry so that
                    # underloaded lorries are filled first.
                    _gain = _load_a / _cap_b if _cap_b else 0
                    if _gain > _best_gain:
                        _best_gain = _gain
                        _best_src, _best_dst = _pa, _pb
            if _best_src:
                # Validate: destination lorry must be allowed for all source routes,
                # and must be a preferred lorry for any route that has preferences.
                # Also block merging if source and destination items are in
                # different destination states — state boundary must not be crossed.
                _src_routes  = {it["ROUTE"] for it in _pit[_best_src] if it.get("ROUTE")}
                _src_states  = {it.get("STATE", "").strip().upper()
                                for it in _pit[_best_src] if it.get("STATE")}
                _dst_states2 = {it.get("STATE", "").strip().upper()
                                for it in _pit[_best_dst] if it.get("STATE")}
                _src_states.discard(""); _dst_states2.discard("")
                _state_merge_ok = (
                    not _src_states or not _dst_states2
                    or bool(_src_states & _dst_states2)
                )
                # REMARKS size cap (FIELD 3): never merge a capped item (e.g. a
                # VAN DO, MAX_TON=2T) onto a lorry larger than its cap — that
                # would silently override the size requirement the planner set.
                _dst_cap = float(_lorry_cap_map.get(_best_dst, 0))
                _cap_merge_ok = all(
                    _it.get("MAX_TON") is None or _dst_cap <= _it["MAX_TON"]
                    for _it in _pit[_best_src]
                )
                # Outstation minimum: never merge outstation items onto a ≤5T
                # lorry that is too small to run that route.
                _dstmin_merge_ok = all(
                    _dst_cap >= _DEST_MIN_TON.get(_classify_dest_group(r), 0.0)
                    for r in _src_routes
                )
                _merge_route_ok = _state_merge_ok and _cap_merge_ok and _dstmin_merge_ok and not any(
                    _best_dst in _strict_route_excl(r) for r in _src_routes
                ) and not any(
                    _preferred_lorries_for_route(r, engine)
                    and _best_dst not in _preferred_lorries_for_route(r, engine)
                    for r in _src_routes
                )
                if _merge_route_ok:
                    for _it in _pit[_best_src]:
                        _it["LORRY"] = _best_dst
                        sess["assigned"][_it["DO NUMBER"]] = _best_dst
                    _pit[_best_dst].extend(_pit.pop(_best_src))
                    _merge_ok = True

        # ── Partial-transfer rebalance pass ──────────────────────────────────
        # The merge pass moves entire lorry loads (all-or-nothing).  When two
        # lorries carry compatible routes but their combined weight exceeds any
        # single lorry's capacity, the merge fails and one lorry stays severely
        # underloaded.  This pass moves individual items from well-loaded lorries
        # onto underloaded ones (< 50 % util) when routes are compatible and
        # the source stays above the threshold after the transfer — preventing
        # oscillation.  Each item may only be moved once (tracked by object id).
        # Example: BPE9788 (14T, 30% Temerloh) absorbs individual Kuantan DOs
        # from BQX9983 (10.5T, 95%) since PH05/PH09 are both east-bound.
        _REBAL_THRESHOLD = REBAL_THRESHOLD
        _rebal_moved: set = set()
        _rebal_ok = True
        while _rebal_ok:
            _rebal_ok = False
            _ploads = {p: sum(x["WEIGHT"] for x in its) for p, its in _pit.items()}
            _underloaded = sorted(
                [p for p in _pit
                 if _pit[p]
                 and float(_lorry_cap_map.get(p, 0)) > 0
                 and _ploads[p] / float(_lorry_cap_map.get(p, 0)) < _REBAL_THRESHOLD],
                key=lambda p: _ploads[p] / float(_lorry_cap_map.get(p, 1))
            )
            if not _underloaded:
                break
            _best_gain, _best_item, _best_src, _best_dst = 0.0, None, None, None
            for _dst in _underloaded:
                _cap_dst  = float(_lorry_cap_map.get(_dst, 0))
                _load_dst = _ploads[_dst]
                _routes_dst = {x["ROUTE"] for x in _pit[_dst] if x.get("ROUTE")}
                if not _routes_dst:
                    _sr = _session_routes.get(_dst, "")
                    if _sr:
                        _routes_dst = {_sr}
                for _src in list(_pit.keys()):
                    if _src == _dst:
                        continue
                    _cap_src  = float(_lorry_cap_map.get(_src, 0))
                    _load_src = _ploads[_src]
                    for _it in _pit[_src]:
                        if id(_it) in _rebal_moved:
                            continue
                        if _load_dst + _it["WEIGHT"] > _cap_dst:
                            continue
                        # REMARKS size cap: never move a capped item (VAN etc.)
                        # onto a lorry larger than its cap.
                        if (_it.get("MAX_TON") is not None
                                and _cap_dst > _it["MAX_TON"]):
                            continue
                        # Don't make the source lorry itself underloaded
                        # (skip only when source has multiple items remaining)
                        if len(_pit[_src]) > 1 and _cap_src > 0:
                            if (_load_src - _it["WEIGHT"]) / _cap_src < _REBAL_THRESHOLD:
                                continue
                        _route_it = _it.get("ROUTE", "")
                        # Strict route guard: don't move item to a lorry that
                        # is not allowed to serve the item's route
                        if _route_it and _dst in _strict_route_excl(_route_it):
                            continue
                        # Outstation minimum: never move an outstation item onto
                        # a ≤5T lorry too small to run that route.
                        if _cap_dst < _DEST_MIN_TON.get(
                                _classify_dest_group(_route_it, _it.get("STATE","")), 0.0):
                            continue
                        # Preferred lorry guard: prefer designated lorries, but
                        # don't block the move entirely — rebalance is an optimisation,
                        # not an assignment gate. Only block if _dst is strictly excluded.
                        _it_pref_rb = _preferred_lorries_for_route(_route_it, engine)
                        if _it_pref_rb and _dst not in _it_pref_rb:
                            # Still allow if no preferred lorry is present in the
                            # underloaded pool (all full/unavailable → any is OK)
                            _pref_in_pool = any(
                                p in _it_pref_rb and p in _pit
                                for p in _it_pref_rb
                            )
                            if _pref_in_pool:
                                continue
                        # State boundary: don't move item to lorry already
                        # committed to a different destination state
                        _it_state = _it.get("STATE", "").strip().upper()
                        _dst_states = {x.get("STATE", "").strip().upper()
                                       for x in _pit[_dst] if x.get("STATE")}
                        if _it_state and _dst_states and _it_state not in _dst_states:
                            continue
                        if _routes_dst and _route_it:
                            if not any(_routes_on_same_way(_route_it, _rd) for _rd in _routes_dst):
                                continue
                        _gain = _it["WEIGHT"] / _cap_dst
                        if _gain > _best_gain:
                            _best_gain = _gain
                            _best_item = _it
                            _best_src  = _src
                            _best_dst  = _dst
            if _best_item:
                _best_item["LORRY"] = _best_dst
                sess["assigned"][_best_item["DO NUMBER"]] = _best_dst
                _pit[_best_src].remove(_best_item)
                _pit[_best_dst].append(_best_item)
                _rebal_moved.add(id(_best_item))
                if not _pit[_best_src]:
                    del _pit[_best_src]
                _rebal_ok = True

        # ── Fill-to-80% pass ──────────────────────────────────────────────────
        # After all rebalance passes, lorries still below 80% utilisation may
        # absorb unassigned items from the SAME route+city+state bucket sorted
        # by GPS proximity to the lorry's route centroid.  This prevents leaving
        # items unassigned when a compatible lorry still has headroom.
        _FILL_TARGET = FILL_TARGET
        _NAIK_FACTOR = NAIK_FACTOR
        # Rebuild load map from _pit
        _fill_ploads = {p: sum(x["WEIGHT"] for x in its) for p, its in _pit.items()}
        # Collect all currently unassigned items (NO_LORRY or None/blank)
        _unassigned_pool = [
            it for it in items
            if it.get("LORRY") in (None, "", "NO_LORRY")
            and it.get("WEIGHT", 0) > 0
        ]
        # Sort pool by date ascending so earlier DOs get priority
        _unassigned_pool.sort(key=lambda it: _parse_date_sortkey(it.get("DATE", "")))

        for _fp in list(_pit.keys()):
            _cap = float(_lorry_cap_map.get(_fp, 0))
            if _cap <= 0:
                continue
            _load = _fill_ploads.get(_fp, 0.0)
            if _load / _cap >= _FILL_TARGET:
                continue  # already at target
            _fp_route = _session_routes.get(_fp, "")
            _fp_dest  = _classify_dest_group(_fp_route)
            _fp_states_set = {it.get("STATE", "").strip().upper()
                              for it in _pit.get(_fp, []) if it.get("STATE")}
            # Find candidates from the pool that match this lorry's route/state
            _candidates = []
            for _cu in _unassigned_pool:
                _cu_route = _cu.get("ROUTE", "")
                _cu_state = _cu.get("STATE", "").strip().upper()
                _cu_dest  = _classify_dest_group(_cu_route)
                # Must be same destination class (urban stays urban, outstation stays outstation)
                if _cu_dest != _fp_dest:
                    continue
                # Route must match lorry's route; urban uses corridor group only
                if _fp_route and _cu_route != _fp_route:
                    if _cu_dest in _DEST_URBAN_GROUPS or _fp_dest in _DEST_URBAN_GROUPS:
                        if not _same_corridor_group(_cu_route, _fp_route):
                            continue
                    else:
                        if (not _same_corridor_group(_cu_route, _fp_route)
                                and not _routes_on_same_way(_cu_route, _fp_route)):
                            continue
                # State boundary
                if _fp_states_set and _cu_state and _cu_state not in _fp_states_set:
                    continue
                # Strict route exclusion
                if _fp in _strict_route_excl(_cu_route):
                    continue
                # Lorry size vs outstation min tonnage
                if _cap < _DEST_MIN_TON.get(_cu_dest, 0.0):
                    continue
                # REMARKS size cap (FIELD 3) — don't overfill onto too-big a lorry
                if _cu.get("MAX_TON") is not None and _cap > _cu["MAX_TON"]:
                    continue
                _candidates.append(_cu)
            if not _candidates:
                continue
            # Sort candidates by GPS proximity to lorry's route centroid
            _fp_centroid = _route_centroid(_fp_route) if _fp_route else None
            def _dist_to_lorry(it, _fc=_fp_centroid):
                c = _route_centroid(it.get("ROUTE", ""))
                if c is None or _fc is None:
                    return float("inf")
                return _haversine_km(_fc[0], _fc[1], c[0], c[1])
            _candidates.sort(key=_dist_to_lorry)
            # Fill up to NAIK capacity
            _naik_cap = _cap * _NAIK_FACTOR
            for _cu in _candidates:
                if _load + _cu["WEIGHT"] > _naik_cap:
                    continue
                # Assign
                _cu["LORRY"] = _fp
                sess["assigned"][_cu["DO NUMBER"]] = _fp
                _pit.setdefault(_fp, []).append(_cu)
                _fill_ploads[_fp] = _fill_ploads.get(_fp, 0.0) + _cu["WEIGHT"]
                _load = _fill_ploads[_fp]
                _unassigned_pool.remove(_cu)
                if not _fp_route and _cu.get("ROUTE"):
                    _fp_route = _cu["ROUTE"]
                    _session_routes[_fp] = _fp_route

        # ── Hard capacity guard ───────────────────────────────────────────────
        # After all passes, any lorry whose assigned weight still exceeds its
        # physical capacity is trimmed: keep the nearest DOs (to Shah Alam HQ)
        # up to the lorry's capacity and mark the overflow as NO_LORRY so the
        # user can handle them manually or in a second trip.
        for _pl, _pl_items in list(_pit.items()):
            _cap = float(_lorry_cap_map.get(_pl, 0))
            if _cap <= 0:
                continue
            _total = sum(x["WEIGHT"] for x in _pl_items)
            # Same-route items may run up to 10 % over rated capacity so they
            # are never split.  Multi-route lorries use the hard cap exactly.
            _pl_routes = {it.get("ROUTE", "").strip().upper() for it in _pl_items
                          if it.get("ROUTE")}
            _eff_cap = _cap * (SAME_ROUTE_NAIK if len(_pl_routes) == 1 else 1.0)
            if _total <= _eff_cap:
                continue
            # Sort by distance from depot — nearest first
            def _dist_hq_guard(it):
                c = _route_centroid(it["ROUTE"])
                if c is None:
                    return float("inf")
                return _haversine_km(_DEPOT[0], _DEPOT[1], c[0], c[1])
            _sorted_items = sorted(_pl_items, key=_dist_hq_guard)
            _kept, _fill = [], 0.0
            for _it in _sorted_items:
                if _fill + _it["WEIGHT"] <= _eff_cap:
                    _kept.append(_it)
                    _fill += _it["WEIGHT"]
                else:
                    _it["LORRY"] = "NO_LORRY"
                    sess["assigned"][_it["DO NUMBER"]] = "NO_LORRY"
            _pit[_pl] = _kept

        # ── Reassign trimmed overflow to idle/available lorries ───────────────
        # The hard capacity guard above may strand items as NO_LORRY when a small
        # preferred lorry (e.g. a van) was chosen on its 2-trip effective capacity
        # but can only physically carry one trip.  Those items must still ship if
        # ANY eligible lorry has physical room — otherwise idle lorries sit unused
        # while DOs go unassigned.  Recompute true physical loads from items, then
        # place each stranded item on the tightest-fitting compatible lorry.
        _phys_load: dict[str, float] = defaultdict(float)
        for _it in items:
            _lp = _it.get("LORRY")
            if _lp and _lp not in ("NO_LORRY", "OTHER_USER", "NOT_TODAY", "PAST_DATE", "WRONG_TRIP",
                                    "REMARKS_SKIP", "OUT_SOURCE", "SPLIT", "SKIPPED", "", None):
                _phys_load[_lp] += _it["WEIGHT"]

        for _it in items:
            if _it.get("LORRY") != "NO_LORRY":
                continue
            _w        = _it["WEIGHT"]
            _it_route = _it.get("ROUTE", "")
            _it_state = _it.get("STATE", "").strip().upper()
            _it_dest  = _classify_dest_group(_it_route, _it.get("STATE", ""))
            _it_min_t = _DEST_MIN_TON.get(_it_dest, 0.0)
            _it_strict_r = _strict_route_excl(_it_route)
            _it_urban = _it_dest in _DEST_URBAN_GROUPS

            _cands = []
            for _cp, _ccap in _lorry_cap_map.items():
                _ccap = float(_ccap)
                if _cp in sess.get("unavailable", set()):
                    continue
                _crem = _ccap - _phys_load.get(_cp, 0.0)
                if _crem < _w:
                    continue
                if _ccap < _it_min_t:
                    continue
                if _it.get("MAX_TON") is not None and _ccap > _it["MAX_TON"]:
                    continue
                if _cp in _it_strict_r:
                    continue
                # Route/direction compatibility with whatever the lorry already carries
                _cp_rt = _session_routes.get(_cp, "")
                if _cp_rt and _it_route != _cp_rt:
                    _cp_urban = _classify_dest_group(_cp_rt) in _DEST_URBAN_GROUPS
                    if _it_urban or _cp_urban:
                        if not _same_corridor_group(_it_route, _cp_rt):
                            continue
                    elif (not _same_corridor_group(_it_route, _cp_rt)
                          and not _routes_on_same_way(_it_route, _cp_rt)):
                        continue
                # State compatibility
                _cp_states = (_consol_lorry_states.get(_cp, set())
                              | _session_lorry_states.get(_cp, set()))
                if _it_state and _cp_states and not any(
                        _states_compatible(_it_state, s) for s in _cp_states):
                    continue
                _cands.append((_crem, _cp))

            if not _cands:
                continue
            _cands.sort()                      # tightest physical fit first
            _pick = _cands[0][1]
            _it["LORRY"] = _pick
            _phys_load[_pick] += _w
            _session_loads[_pick] = float(_session_loads.get(_pick, 0)) + _w
            if _it_state:
                _consol_lorry_states.setdefault(_pick, set()).add(_it_state)
            _record_lorry_state(_pick, _it_state)
            if _pick not in _session_routes:
                _session_routes[_pick] = _it_route
            _pit.setdefault(_pick, []).append(_it)
            sess["assigned"][_it["DO NUMBER"]] = _pick
            _unassigned_reasons.pop(_it["DO NUMBER"], None)

        for item in items:
            sess["assigned"][item["DO NUMBER"]] = item["LORRY"]

        # ── Geographic / state clustering enforcement ─────────────────────────
        # Route code is atomic: all stops of the SAME route code always ride one
        # lorry, however far apart (a route may legitimately span several cities,
        # e.g. PH04 = Benta + Kuala Lipis + Lipis).  A lorry may carry MULTIPLE
        # route codes only when the codes are mutually reachable ("chain") under:
        #     same state (KL/Selangor count as one urban state)  AND
        #     (share a city  OR  nearest stops within _MAX_GEO_GAP_DEG ≈ 32 km).
        # This stops e.g. a Selangor (KV) lorry also carrying a Pahang (PH04)
        # drop, while never splitting a single route code across lorries.
        _GEO_GAP = _MAX_GEO_GAP_DEG
        _GEO_VALID = {"NO_LORRY", "OTHER_USER", "NOT_TODAY", "REMARKS_SKIP", "OUT_SOURCE", "PAST_DATE", "WRONG_TRIP",
                      "SPLIT", "SKIPPED", "", None}

        def _gps_of(_it):
            _la, _lo = _it.get("GPS_LAT"), _it.get("GPS_LON")
            return (_la, _lo) if (_la is not None and _lo is not None) else None

        def _rcode(_it):
            """Canonical route code, e.g. 'PH03', 'KV12A', 'NS04'."""
            _r = str(_it.get("ROUTE", "")).strip().upper()
            _m = re.match(r"([A-Z]+\d+[A-Z]?)", _r)
            return _m.group(1) if _m else _r[:6]

        def _pt_of(_it):
            """(lat, lon, route_code, state, city) or None when GPS missing."""
            _g = _gps_of(_it)
            if not _g:
                return None
            return (_g[0], _g[1], _rcode(_it),
                    str(_it.get("STATE", "")).strip().upper(),
                    str(_it.get("CITY", "")).strip().upper())

        def _geo_deg(_a, _b):
            return ((_a[0] - _b[0]) ** 2 + (_a[1] - _b[1]) ** 2) ** 0.5

        def _edge_ok(_a, _b):
            """May these two stops sit on one lorry together?
            - SAME route code → always yes (route code is atomic: one route =
              one trip = one lorry, regardless of how far its stops are, because
              the route code is a more reliable signal than a possibly-wrong GPS).
            - DIFFERENT route codes, SAME corridor group (e.g. PH_INT =
              PH01..PH08 Pahang interior) → yes: it is one directional
              outstation trip covering several towns, so they combine to fill
              the lorry. (Still requires the same state.)
            - Otherwise different route codes → only if SAME state AND within
              _GEO_GAP straight-line degrees (so a far/foreign drop is not
              bolted on)."""
            if _a[2] == _b[2]:
                return True
            if not _states_compatible(_a[3], _b[3]):
                return False
            if _same_corridor_group(_a[2], _b[2]):
                return True
            return _geo_deg(_a[:2], _b[:2]) <= _GEO_GAP

        def _lorry_geo_ok(_pts):
            """GPS single-linkage: all stops must form ONE connected cluster where
            neighbouring stops are same-state and within _GEO_GAP. A route that
            spans several nearby cities stays whole; a mislabeled far-away stop
            (wrong GPS) is NOT connected and must move to another lorry."""
            _n = len(_pts)
            if _n <= 1:
                return True
            _adj = {_i: set() for _i in range(_n)}
            for _i in range(_n):
                for _j in range(_i + 1, _n):
                    if _edge_ok(_pts[_i], _pts[_j]):
                        _adj[_i].add(_j)
                        _adj[_j].add(_i)
            _seen = {0}
            _stack = [0]
            while _stack:
                for _nb in _adj[_stack.pop()]:
                    if _nb not in _seen:
                        _seen.add(_nb)
                        _stack.append(_nb)
            if len(_seen) != _n:
                return False
            # ── Urban anti-chaining (consider longitude, not just adjacency) ──
            # Single-linkage alone lets a chain of close stops bridge two far
            # urban zones (e.g. central KL → Rawang, or Semenyih → KL). For
            # URBAN stops of DIFFERENT route codes, also bound the overall
            # SPREAD: no two different-route urban stops may be more than
            # _URBAN_MERGE_SPREAD apart. Same route code stays atomic (any
            # distance), so a legitimate multi-town route is never broken.
            _urb = [_p for _p in _pts
                    if _p[3] in _URBAN_COMPATIBLE_STATES and _p[0] is not None]
            for _i in range(len(_urb)):
                for _j in range(_i + 1, len(_urb)):
                    if (_urb[_i][2] != _urb[_j][2]
                            and _geo_deg(_urb[_i][:2], _urb[_j][:2]) > _URBAN_MERGE_SPREAD):
                        return False
            return True

        # Split each lorry into GPS-connected clusters (single linkage under
        # _GEO_GAP + same state); keep the heaviest cluster, detach the rest.
        _geo_by_lorry: dict[str, list] = {}
        for _it in items:
            _l = _it.get("LORRY")
            if _l not in _GEO_VALID and _pt_of(_it):
                _geo_by_lorry.setdefault(_l, []).append(_it)

        _detached: list = []
        for _l, _its in _geo_by_lorry.items():
            _pts = [_pt_of(_x) for _x in _its]
            if _lorry_geo_ok(_pts):
                continue
            # Build connected components (indices) under _edge_ok.
            _n = len(_its)
            _adj = {_i: set() for _i in range(_n)}
            for _i in range(_n):
                for _j in range(_i + 1, _n):
                    if _edge_ok(_pts[_i], _pts[_j]):
                        _adj[_i].add(_j)
                        _adj[_j].add(_i)
            _comp_of = [-1] * _n
            _comps: list = []
            for _s in range(_n):
                if _comp_of[_s] != -1:
                    continue
                _cid = len(_comps)
                _comps.append([])
                _stack = [_s]
                _comp_of[_s] = _cid
                while _stack:
                    _u = _stack.pop()
                    _comps[_cid].append(_u)
                    for _nb in _adj[_u]:
                        if _comp_of[_nb] == -1:
                            _comp_of[_nb] = _cid
                            _stack.append(_nb)
            # Keep the heaviest component; detach the rest.
            _comps.sort(key=lambda c: sum(_its[i]["WEIGHT"] for i in c), reverse=True)
            for _comp in _comps[1:]:
                for _i in _comp:
                    _its[_i]["LORRY"] = "NO_LORRY"
                    _detached.append(_its[_i])
            # ── Urban complete-linkage refinement (anti-chaining) ─────────────
            # A single connected component can still bridge two far urban zones
            # through a chain of close stops (central KL → Rawang, Semenyih →
            # KL). Within the kept component, no two DIFFERENT-route urban stops
            # may exceed _URBAN_MERGE_SPREAD. While one does, detach the LIGHTER
            # of the two offending route codes (route code stays atomic) and
            # re-home it. This enforces the spread by longitude, not just by the
            # single-linkage chain.
            _kept = list(_comps[0])
            while True:
                _worst = 0.0
                _pair = None
                for _a in _kept:
                    for _b in _kept:
                        if _a < _b:
                            _pa, _pb = _pts[_a], _pts[_b]
                            if (_pa[2] != _pb[2]
                                    and _pa[3] in _URBAN_COMPATIBLE_STATES
                                    and _pb[3] in _URBAN_COMPATIBLE_STATES):
                                _dd = _geo_deg(_pa[:2], _pb[:2])
                                if _dd > _worst:
                                    _worst = _dd
                                    _pair = (_a, _b)
                if _pair is None or _worst <= _URBAN_MERGE_SPREAD:
                    break
                _ra, _rb = _pts[_pair[0]][2], _pts[_pair[1]][2]
                _wa = sum(_its[i]["WEIGHT"] for i in _kept if _pts[i][2] == _ra)
                _wb = sum(_its[i]["WEIGHT"] for i in _kept if _pts[i][2] == _rb)
                _drop = _ra if _wa <= _wb else _rb
                _new_kept = []
                for _i in _kept:
                    if _pts[_i][2] == _drop:
                        _its[_i]["LORRY"] = "NO_LORRY"
                        _detached.append(_its[_i])
                    else:
                        _new_kept.append(_i)
                _kept = _new_kept

        if _detached:
            # Recompute physical loads and each lorry's current point tuples.
            _phys2: dict[str, float] = defaultdict(float)
            _lpts: dict[str, list] = defaultdict(list)
            for _it in items:
                _l = _it.get("LORRY")
                if _l not in _GEO_VALID:
                    _phys2[_l] += _it["WEIGHT"]
                    _p = _pt_of(_it)
                    if _p:
                        _lpts[_l].append(_p)

            # Re-home detached items, same route code first so a route stays whole.
            _detached.sort(key=lambda x: (_rcode(x), x.get("GPS_LON") or 0))
            for _it in _detached:
                _w = _it["WEIGHT"]
                _pt = _pt_of(_it)
                _rc = _rcode(_it)
                _dest = _classify_dest_group(_it.get("ROUTE", ""), _it.get("STATE", ""))
                _min_t = _DEST_MIN_TON.get(_dest, 0.0)
                _strict = _strict_route_excl(_it.get("ROUTE", ""))
                _mt = _it.get("MAX_TON")
                _best = None
                _best_key = None
                for _cp, _ccap in _lorry_cap_map.items():
                    _ccap = float(_ccap)
                    if _cp in sess.get("unavailable", set()):
                        continue
                    if _ccap - _phys2.get(_cp, 0.0) < _w:
                        continue
                    if _ccap < _min_t:
                        continue
                    if _mt is not None and _ccap > _mt:
                        continue
                    if _cp in _strict:
                        continue
                    if not _lorry_geo_ok(_lpts.get(_cp, []) + [_pt]):
                        continue
                    # Prefer a lorry already carrying this route code (keeps the
                    # route whole), then any occupied compatible lorry, then idle;
                    # within a tier, the tightest physical fit.
                    _has_same = any(_p[2] == _rc for _p in _lpts.get(_cp, []))
                    _tier = 0 if _has_same else (1 if _lpts.get(_cp) else 2)
                    _key = (_tier, _ccap - _phys2.get(_cp, 0.0))
                    if _best is None or _key < _best_key:
                        _best, _best_key = _cp, _key
                if _best:
                    _it["LORRY"] = _best
                    _phys2[_best] += _w
                    _lpts[_best].append(_pt)
                    _session_loads[_best] = float(_session_loads.get(_best, 0)) + _w
                    _record_lorry_state(_best, _it.get("STATE", "").strip().upper())
                    if _best not in _session_routes:
                        _session_routes[_best] = _it.get("ROUTE", "")
                    _unassigned_reasons.pop(_it["DO NUMBER"], None)
                else:
                    _it["LORRY"] = "NO_LORRY"
                    _unassigned_reasons.setdefault(_it["DO NUMBER"], "GEO_TOO_FAR")

            for item in items:
                sess["assigned"][item["DO NUMBER"]] = item["LORRY"]

        # ── VAN-remark consolidation (highest priority) ───────────────────────
        # Every DO whose remark is VAN (cap ≤2T) should ride a van TOGETHER with
        # nearby VAN DOs, even across different route codes. Pool all VAN DOs,
        # cluster them by GPS single-linkage (same state + within 0.29°), and
        # pack each cluster onto one ≤2T lorry — preferring a van already holding
        # part of the cluster, else the tightest-fitting free van. This pulls a
        # VAN DO off a 4.2T lorry onto the van with its VAN neighbours.
        _van_its = [_it for _it in items
                    if _it.get("MAX_TON") == 2.0 and _pt_of(_it)
                    and _it.get("LORRY") not in _GEO_VALID]
        if _van_its:
            _van_plates = [p for p, c in _lorry_cap_map.items() if float(c) <= 2.0]
            if _van_plates:
                # Physical load per lorry EXCLUDING the VAN DOs (they get re-placed).
                _vphys: dict[str, float] = defaultdict(float)
                _van_ids = {id(_x) for _x in _van_its}
                for _it in items:
                    _l = _it.get("LORRY")
                    if _l not in _GEO_VALID and id(_it) not in _van_ids:
                        _vphys[_l] += _it["WEIGHT"]
                # Single-linkage clusters of the VAN DOs.
                _vpts = [_pt_of(_x) for _x in _van_its]
                _vn = len(_van_its)
                _vadj = {_i: set() for _i in range(_vn)}
                for _i in range(_vn):
                    for _j in range(_i + 1, _vn):
                        if _edge_ok(_vpts[_i], _vpts[_j]):
                            _vadj[_i].add(_j)
                            _vadj[_j].add(_i)
                _vcomp = [-1] * _vn
                _vcomps: list = []
                for _s in range(_vn):
                    if _vcomp[_s] != -1:
                        continue
                    _cid = len(_vcomps)
                    _vcomps.append([])
                    _vcomp[_s] = _cid
                    _stk = [_s]
                    while _stk:
                        _u = _stk.pop()
                        _vcomps[_cid].append(_u)
                        for _nb in _vadj[_u]:
                            if _vcomp[_nb] == -1:
                                _vcomp[_nb] = _cid
                                _stk.append(_nb)
                for _comp in _vcomps:
                    _citems = [_van_its[_i] for _i in _comp]
                    _cw = sum(_x["WEIGHT"] for _x in _citems)
                    # Prefer a van already carrying most of this cluster, then the
                    # one with the most free room; require it to fit the whole cluster.
                    _cur = defaultdict(float)
                    for _x in _citems:
                        if _x.get("LORRY") in _van_plates:
                            _cur[_x["LORRY"]] += 1
                    _pick = None
                    for _v in sorted(_van_plates,
                                     key=lambda v: (-_cur.get(v, 0),
                                                    _vphys[v] - float(_lorry_cap_map.get(v, 0)))):
                        if _vphys[_v] + _cw <= float(_lorry_cap_map.get(_v, 0)):
                            _pick = _v
                            break
                    if _pick is None:
                        continue   # cluster larger than any single van — leave as-is
                    for _x in _citems:
                        _x["LORRY"] = _pick
                        sess["assigned"][_x["DO NUMBER"]] = _pick
                        _unassigned_reasons.pop(_x["DO NUMBER"], None)
                    _vphys[_pick] += _cw

        # ── Same-destination consolidation (highest grouping priority) ────────
        # DOs at the SAME physical destination — same route code + state + city +
        # longitude/GPS — are one drop and must ride ONE lorry, never split,
        # provided they fit (same-route load may use the ×1.05 NAIK overage).
        # Applies to ALL routes (urban and outstation). Respects size cap,
        # outstation minimum and forbidden plates.
        _dd_phys: dict[str, float] = defaultdict(float)
        _dest_groups: dict = defaultdict(lambda: defaultdict(list))  # destkey→lorry→items
        for _it in items:
            _l = _it.get("LORRY")
            if _l in _GEO_VALID:
                continue
            _dd_phys[_l] += _it["WEIGHT"]
            _g = _gps_of(_it)
            _dk = (_rcode(_it),
                   str(_it.get("STATE", "")).strip().upper(),
                   str(_it.get("CITY", "")).strip().upper(),
                   round(_g[0], 4) if _g else None,
                   round(_g[1], 4) if _g else None)
            _dest_groups[_dk][_l].append(_it)

        for _dk, _lor_map in _dest_groups.items():
            if len(_lor_map) < 2:
                continue                          # already on one lorry
            _all_it = [_x for _l in _lor_map for _x in _lor_map[_l]]
            _grp_w = sum(_x["WEIGHT"] for _x in _all_it)
            _mt = min((_x["MAX_TON"] for _x in _all_it if _x.get("MAX_TON") is not None),
                      default=None)
            _dmin = max(_DEST_MIN_TON.get(
                _classify_dest_group(_x.get("ROUTE", ""), _x.get("STATE", "")), 0.0)
                for _x in _all_it)
            _forbid = set()
            for _x in _all_it:
                if _x.get("FORBID_PLATES"):
                    _forbid |= _x["FORBID_PLATES"]
            # target = the lorry already holding the most of this destination
            _pick = None
            for _tgt in sorted(_lor_map, key=lambda l: -sum(x["WEIGHT"] for x in _lor_map[l])):
                _tcap = float(_lorry_cap_map.get(_tgt, 0))
                _new = _dd_phys[_tgt] + _grp_w - sum(_x["WEIGHT"] for _x in _lor_map.get(_tgt, []))
                if _new > _tcap * NAIK_FACTOR:
                    continue                      # would overload even with naik
                if _mt is not None and _tcap > _mt:
                    continue
                if _tcap < _dmin:
                    continue
                if _tgt in _forbid:
                    continue
                _pick = _tgt
                break
            if _pick is None:
                continue                          # doesn't fit one lorry → keep split
            for _src, _sitems in list(_lor_map.items()):
                if _src == _pick:
                    continue
                _mw = sum(_x["WEIGHT"] for _x in _sitems)
                for _x in _sitems:
                    _x["LORRY"] = _pick
                    sess["assigned"][_x["DO NUMBER"]] = _pick
                    _unassigned_reasons.pop(_x["DO NUMBER"], None)
                _dd_phys[_pick] += _mw
                _dd_phys[_src] -= _mw

        # ── Same-route consolidation ──────────────────────────────────────────
        # Pull scattered stops of ONE route code together so a route is not split
        # between a full lorry and a near-empty van (e.g. two KV05A drops at the
        # same Batu Caves address — 1.883T on WLD8738 and 0.025T alone on the van).
        # For each route code on ≥2 lorries, move the lighter lorry's items onto a
        # heavier same-route lorry that has room, respecting size caps, outstation
        # minimum and the geo/state chain.  Leaves the van free for a real load.
        _cons_phys: dict[str, float] = defaultdict(float)
        _cons_routes: dict = defaultdict(lambda: defaultdict(list))  # rcode→lorry→items
        for _it in items:
            _l = _it.get("LORRY")
            if _l not in _GEO_VALID:
                _cons_phys[_l] += _it["WEIGHT"]
                _cons_routes[_rcode(_it)][_l].append(_it)

        for _rc, _lor_map in _cons_routes.items():
            if len(_lor_map) < 2:
                continue
            # Only consolidate URBAN routes (the van-scrap problem). Outstation
            # routes are split across lorries out of capacity necessity — merging
            # them would overload a truck, so leave them alone.
            _any_it = next((x for _l in _lor_map for x in _lor_map[_l]), None)
            if _any_it is None or _classify_dest_group(
                    _any_it.get("ROUTE", ""), _any_it.get("STATE", "")) not in _DEST_URBAN_GROUPS:
                continue
            # Move from the lightest-loaded lorry (for this route) to the heaviest.
            _srcs = sorted(_lor_map, key=lambda l: sum(x["WEIGHT"] for x in _lor_map[l]))
            for _src in _srcs:
                _src_items = _lor_map.get(_src) or []
                if not _src_items:
                    continue
                _src_w = sum(x["WEIGHT"] for x in _src_items)
                _tgts = sorted((l for l in _lor_map if l != _src and _lor_map[l]),
                               key=lambda l: -sum(x["WEIGHT"] for x in _lor_map[l]))
                for _tgt in _tgts:
                    _tcap = float(_lorry_cap_map.get(_tgt, 0))
                    if _cons_phys[_tgt] + _src_w > _tcap:
                        continue
                    if any(_x.get("MAX_TON") is not None and _tcap > _x["MAX_TON"]
                           for _x in _src_items):
                        continue
                    if _tcap < min(_DEST_MIN_TON.get(
                            _classify_dest_group(_x.get("ROUTE", ""), _x.get("STATE", "")), 0.0)
                            for _x in _src_items):
                        continue
                    _tgt_pts = [_pt_of(_x) for _x in items
                                if _x.get("LORRY") == _tgt and _pt_of(_x)]
                    _add_pts = [_pt_of(_x) for _x in _src_items if _pt_of(_x)]
                    if not _lorry_geo_ok(_tgt_pts + _add_pts):
                        continue
                    for _x in _src_items:
                        _x["LORRY"] = _tgt
                        sess["assigned"][_x["DO NUMBER"]] = _tgt
                    _cons_phys[_tgt] += _src_w
                    _cons_phys[_src] -= _src_w
                    _session_loads[_tgt] = float(_session_loads.get(_tgt, 0)) + _src_w
                    _lor_map[_tgt].extend(_src_items)
                    _lor_map[_src] = []
                    break

        # Lorries already assigned today (e.g. a SPARE lorry used on the ABI run
        # before this VIVIAN run) or broken are OFF-LIMITS to every consolidation
        # / de-concentration / repack pass below — they must never receive a DO.
        _blocked_today = set(get_assigned_today()) | set(get_broken_lorries())

        # ── ATOMIC-UNIT REUNIFICATION (criteria 1 & 2) ────────────────────────
        # No atomic unit may stay split across lorries: (1) same GPS longitude,
        # (2) same route + customer CODE. Any unit that ended up split (from any
        # earlier pass, e.g. when its lorries filled up) is pulled back onto ONE
        # lorry that can hold the whole unit — preferring a lorry already holding
        # part of it, then the tightest-fitting other eligible lorry with room.
        # If no single lorry can hold the whole unit (fleet genuinely full), it
        # is left as-is rather than dropping any DO.
        _ru_all = [_it for _it in items if _it.get("LORRY") in _lorry_cap_map]
        _ru_phys: dict[str, float] = defaultdict(float)
        for _it in _ru_all:
            _ru_phys[_it["LORRY"]] += _it["WEIGHT"]
        for _comp in _atomic_components(_ru_all, _rcode, _gps_of,
                                        lambda x: x.get("CODE", "")):
            _lset = {x["LORRY"] for x in _comp}
            if len(_lset) < 2:
                continue                          # already whole
            _comp_ids = {id(x) for x in _comp}
            _cw = sum(x["WEIGHT"] for x in _comp)
            _cmt = min((x["MAX_TON"] for x in _comp
                        if x.get("MAX_TON") is not None), default=None)
            _dmin = max(_DEST_MIN_TON.get(
                _classify_dest_group(x.get("ROUTE", ""), x.get("STATE", "")), 0.0)
                for x in _comp)
            _cforbid: set = set()
            for x in _comp:
                if x.get("FORBID_PLATES"):
                    _cforbid |= x["FORBID_PLATES"]
            _cstates = {x.get("STATE", "").strip().upper()
                        for x in _comp if x.get("STATE")}

            def _ru_fits(_cand):
                if _cand in _blocked_today:
                    return False
                _ccap = float(_lorry_cap_map.get(_cand, 0))
                _here = sum(x["WEIGHT"] for x in _comp if x["LORRY"] == _cand)
                if _ru_phys[_cand] - _here + _cw > _ccap * NAIK_FACTOR:
                    return False
                if _cmt is not None and _ccap > _cmt:
                    return False
                if _ccap < _dmin:
                    return False
                if _cand in _cforbid:
                    return False
                _rstates = {str(y.get("STATE", "")).strip().upper()
                            for y in items if y.get("LORRY") == _cand and y.get("STATE")}
                if (_rstates and _cstates and not any(
                        _states_compatible(_a, _b)
                        for _a in _cstates for _b in _rstates)):
                    return False
                _other = [_pt_of(x) for x in items
                          if x.get("LORRY") == _cand and id(x) not in _comp_ids
                          and _pt_of(x)]
                _add = [_pt_of(x) for x in _comp if _pt_of(x)]
                return _lorry_geo_ok(_other + _add)

            # 1st choice: a lorry already holding part (most of it first).
            _tgt = None
            for _cand in sorted(_lset,
                                key=lambda l: -sum(x["WEIGHT"] for x in _comp
                                                   if x["LORRY"] == l)):
                if _ru_fits(_cand):
                    _tgt = _cand
                    break
            # 2nd choice: any other eligible lorry, tightest fit that holds it
            # whole (prefer ≤11T for urban units so a big lorry isn't taken).
            if _tgt is None:
                _urban_unit = all(
                    _classify_dest_group(x.get("ROUTE", ""),
                                         x.get("STATE", "")) in _DEST_URBAN_GROUPS
                    for x in _comp)
                for _cand in sorted(
                        _lorry_cap_map,
                        key=lambda l: (0 if (not _urban_unit
                                             or float(_lorry_cap_map[l]) <= _URBAN_MAX_TON)
                                       else 1,
                                       float(_lorry_cap_map[l]))):
                    if _cand in _lset:
                        continue
                    if _ru_fits(_cand):
                        _tgt = _cand
                        break
            if _tgt is None:
                continue                          # fleet full — leave split
            for x in _comp:
                if x["LORRY"] != _tgt:
                    _ru_phys[x["LORRY"]] -= x["WEIGHT"]
                    _ru_phys[_tgt] += x["WEIGHT"]
                    x["LORRY"] = _tgt
                    sess["assigned"][x["DO NUMBER"]] = _tgt
                    _unassigned_reasons.pop(x["DO NUMBER"], None)
                    _session_loads[_tgt] = float(_session_loads.get(_tgt, 0)) + x["WEIGHT"]
            _record_lorry_state(_tgt, next(iter(_cstates), ""))

        # ── OUTSTATION ROUTE CONSOLIDATION ────────────────────────────────────
        # An outstation route (Kuantan/Pahang/…) should ride the FEWEST big
        # lorries and obey its minimum tonnage (LARGE_LONG needs >11T). If a
        # route is split, consolidate it onto ONE lorry that holds its whole
        # weight and meets the min — preferring a current holder, else the
        # tightest-fitting eligible lorry. Fixes an outstation stub stranded on
        # a too-small lorry and frees that lorry / raises utilisation.
        _or_phys: dict[str, float] = defaultdict(float)
        for _it in items:
            if _it.get("LORRY") in _lorry_cap_map:
                _or_phys[_it["LORRY"]] += _it["WEIGHT"]
        # Group outstation DOs by DIRECTION (corridor), not exact route code, so
        # all routes in one corridor (e.g. Perak PK01+PK02+PK06) consolidate onto
        # ONE big lorry and the run is fully utilised — mirroring the manual.
        _out_routes: dict = {}
        for _it in items:
            if (_it.get("LORRY") in _lorry_cap_map
                    and _classify_dest_group(_it.get("ROUTE", ""),
                                             _it.get("STATE", "")) not in _DEST_URBAN_GROUPS):
                _out_routes.setdefault(_direction_key(_it.get("ROUTE", ""), _it.get("STATE", ""), _it.get("CUSTOMER NAME", "")), []).append(_it)
        for _dir, _rits in _out_routes.items():
            _lset = {x["LORRY"] for x in _rits}
            if len(_lset) < 2:
                continue                          # already on one lorry
            _rits_ids = {id(x) for x in _rits}
            _rw = sum(x["WEIGHT"] for x in _rits)
            _dg = max((_classify_dest_group(x.get("ROUTE", ""), x.get("STATE", ""))
                       for x in _rits),
                      key=lambda g: _DEST_MIN_TON.get(g, 0.0))
            _dmin = _eff_dest_min_ton(_rits[0].get("ROUTE", ""), _dg, _rw)
            _rmt = min((x["MAX_TON"] for x in _rits
                        if x.get("MAX_TON") is not None), default=None)
            _rforbid: set = set()
            for x in _rits:
                if x.get("FORBID_PLATES"):
                    _rforbid |= x["FORBID_PLATES"]

            def _out_fits(_cand):
                if _cand in _blocked_today:
                    return False
                _cap = float(_lorry_cap_map.get(_cand, 0))
                if _cap < _dmin:
                    return False
                if _rmt is not None and _cap > _rmt:
                    return False
                if _cand in _rforbid:
                    return False
                _here = sum(x["WEIGHT"] for x in _rits if x["LORRY"] == _cand)
                if _or_phys[_cand] - _here + _rw > _cap * NAIK_FACTOR:
                    return False
                # Existing stops on _cand that are NOT part of this corridor group
                # must be in the SAME direction and form one geo cluster with it.
                _other_items = [y for y in items
                                if y.get("LORRY") == _cand and id(y) not in _rits_ids]
                if any(_direction_key(y.get("ROUTE", ""), y.get("STATE", ""), y.get("CUSTOMER NAME", "")) != _dir for y in _other_items):
                    return False
                _other = [_pt_of(y) for y in _other_items if _pt_of(y)]
                _add = [_pt_of(x) for x in _rits if _pt_of(x)]
                return _lorry_geo_ok(_other + _add)

            _tgt = None
            for _cand in sorted(_lset, key=lambda l: -sum(
                    x["WEIGHT"] for x in _rits if x["LORRY"] == l)):
                if _out_fits(_cand):
                    _tgt = _cand
                    break
            if _tgt is None:
                for _cand in sorted(_lorry_cap_map, key=lambda l: float(_lorry_cap_map[l])):
                    if _cand in _lset:
                        continue
                    if _out_fits(_cand):
                        _tgt = _cand
                        break
            if _tgt is not None:
                for x in _rits:
                    if x["LORRY"] != _tgt:
                        _or_phys[x["LORRY"]] -= x["WEIGHT"]
                        _or_phys[_tgt] += x["WEIGHT"]
                        x["LORRY"] = _tgt
                        sess["assigned"][x["DO NUMBER"]] = _tgt
                        _unassigned_reasons.pop(x["DO NUMBER"], None)
                        _session_loads[_tgt] = float(_session_loads.get(_tgt, 0)) + x["WEIGHT"]
                _record_lorry_state(_tgt, next(
                    (x.get("STATE", "").strip().upper() for x in _rits if x.get("STATE")), ""))
                continue

            # Whole route fits no single lorry (capacity, or a per-stop plate
            # forbid on the only big lorry). Lift the WHOLE route off its current
            # lorries and re-pack its ATOMIC components (same GPS / same route+
            # customer stay whole) across valid big lorries (cap ≥ dmin) — each
            # component onto the emptiest lorry it may use (not forbidden, and
            # empty of OTHER routes so directions never mix). This can move the
            # non-forbidding bulk onto an idle big lorry, freeing the incumbent
            # for a plate-forbidding stop. If a component finds no valid home it
            # is restored to its original lorry (never stranded worse than start).
            _valid = [p for p in _lorry_cap_map
                      if float(_lorry_cap_map[p]) >= _dmin
                      and (_rmt is None or float(_lorry_cap_map[p]) <= _rmt)]
            _comps_r = _atomic_components(_rits, _rcode, _gps_of,
                                          lambda x: x.get("CODE", ""))
            _orig = {id(x): x["LORRY"] for x in _rits}
            # set the route aside: remove its weight from the running loads
            for x in _rits:
                _or_phys[x["LORRY"]] -= x["WEIGHT"]
                x["LORRY"] = None
            _cids_all = {id(x) for x in _rits}
            _used_here: set = set()               # lorries this repack has filled

            def _may_take(_cand, _comp, _cw, _cfp):
                if _cand in _cfp or _cand in _blocked_today:
                    return False
                # empty lorry, or one already holding ONLY this corridor — never
                # mix a different direction (urban, or another outstation corridor).
                _cand_dirs = {_direction_key(y.get("ROUTE", ""), y.get("STATE", ""), y.get("CUSTOMER NAME", "")) for y in items
                              if y.get("LORRY") == _cand and id(y) not in _cids_all}
                if _cand_dirs and _cand_dirs != {_dir}:
                    return False
                if float(_lorry_cap_map[_cand]) * NAIK_FACTOR - _or_phys[_cand] < _cw:
                    return False
                _other = [_pt_of(y) for y in items
                          if y.get("LORRY") == _cand and id(y) not in _cids_all and _pt_of(y)]
                return _lorry_geo_ok(_other + [_pt_of(x) for x in _comp if _pt_of(x)])

            for _comp in sorted(_comps_r, key=lambda c: -sum(x["WEIGHT"] for x in c)):
                _cw = sum(x["WEIGHT"] for x in _comp)
                _cfp: set = set()
                for x in _comp:
                    if x.get("FORBID_PLATES"):
                        _cfp |= x["FORBID_PLATES"]
                # First-fit-decreasing that CONSOLIDATES: fill a lorry already
                # opened by this repack (tightest fit) before opening a new one;
                # when opening a new lorry pick the LARGEST valid one so it can
                # absorb the rest of the route too (fewest lorries).
                _u = [c for c in _used_here if _may_take(c, _comp, _cw, _cfp)]
                if _u:
                    _dest = min(_u, key=lambda c: float(_lorry_cap_map[c]) * NAIK_FACTOR - _or_phys[c])
                else:
                    _fresh = [c for c in _valid
                              if c not in _used_here and _may_take(c, _comp, _cw, _cfp)]
                    _dest = max(_fresh, key=lambda c: float(_lorry_cap_map[c])) if _fresh else None
                if _dest is None:                 # restore to original lorry
                    for x in _comp:
                        x["LORRY"] = _orig[id(x)]
                        _or_phys[x["LORRY"]] += x["WEIGHT"]
                    continue
                _used_here.add(_dest)
                for x in _comp:
                    x["LORRY"] = _dest
                    sess["assigned"][x["DO NUMBER"]] = _dest
                    _unassigned_reasons.pop(x["DO NUMBER"], None)
                    _or_phys[_dest] += x["WEIGHT"]
                    _session_loads[_dest] = float(_session_loads.get(_dest, 0)) + x["WEIGHT"]
                _record_lorry_state(_dest, next(
                    (x.get("STATE", "").strip().upper() for x in _comp if x.get("STATE")), ""))

        # ── URBAN >11T DE-CONCENTRATION ───────────────────────────────────────
        # Urban (KL / Selangor) routes visit many closely-spaced small drops; a
        # lorry over _URBAN_MAX_TON (11 T) cannot physically run that many stops
        # in one trip ("only lorry under 11 TON can handle that much routes").
        # When a >11 T lorry has accumulated urban stops, move as many as
        # possible onto ≤11 T lorries that have room — keeping same-destination
        # clusters whole and respecting weight, size cap, state, geo chain and
        # forbidden plates. Whatever cannot find a ≤11 T home stays put (never
        # left unassigned). Big lorries are thus freed for heavy / outstation
        # work, and the many-stop urban runs land on the small lorries.
        _big_urban: dict[str, list] = defaultdict(list)
        for _it in items:
            _l = _it.get("LORRY")
            if (_l in _lorry_cap_map and float(_lorry_cap_map[_l]) > _URBAN_MAX_TON
                    and _classify_dest_group(_it.get("ROUTE", ""),
                                             _it.get("STATE", "")) in _DEST_URBAN_GROUPS):
                _big_urban[_l].append(_it)
        if _big_urban:
            _uc_phys: dict[str, float] = defaultdict(float)
            for _it in items:
                if _it.get("LORRY") in _lorry_cap_map:
                    _uc_phys[_it["LORRY"]] += _it["WEIGHT"]
            _small_plates = [p for p, c in _lorry_cap_map.items()
                             if float(c) <= _URBAN_MAX_TON]
            # route codes each lorry currently carries (for criterion 3 —
            # keep same-route DOs together on one lorry when possible)
            _uc_routes: dict[str, set] = defaultdict(set)
            for _it in items:
                if _it.get("LORRY") in _lorry_cap_map:
                    _uc_routes[_it["LORRY"]].add(_rcode(_it))
            for _big, _bitems in _big_urban.items():
                # Split the big lorry's urban items into ATOMIC components — same
                # longitude (1) OR same route+customer CODE (2) stays together and
                # is never split. Move whole components only, heaviest first.
                for _cit in sorted(
                        _atomic_components(_bitems, _rcode, _gps_of,
                                           lambda x: x.get("CODE", "")),
                        key=lambda c: -sum(x["WEIGHT"] for x in c)):
                    _cw = sum(x["WEIGHT"] for x in _cit)
                    _croutes = {_rcode(x) for x in _cit}
                    _cforbid: set = set()
                    for x in _cit:
                        if x.get("FORBID_PLATES"):
                            _cforbid |= x["FORBID_PLATES"]
                    _cstates = {x.get("STATE", "").strip().upper()
                                for x in _cit if x.get("STATE")}
                    _cmt = min((x["MAX_TON"] for x in _cit
                                if x.get("MAX_TON") is not None), default=None)
                    # Collect every valid ≤11T receiver, then choose by priority:
                    # (3) a lorry already carrying this route code (keep route
                    # together), else the one with the MOST free room.
                    _cands = []
                    for _r in _small_plates:
                        if _r == _big or _r in _blocked_today:
                            continue
                        _rcap = float(_lorry_cap_map[_r])
                        if _uc_phys[_r] + _cw > _rcap * NAIK_FACTOR:
                            continue
                        if _cmt is not None and _rcap > _cmt:
                            continue
                        if _r in _cforbid:
                            continue
                        _rstates = {str(y.get("STATE", "")).strip().upper()
                                    for y in items if y.get("LORRY") == _r and y.get("STATE")}
                        if (_rstates and _cstates and not any(
                                _states_compatible(_a, _b)
                                for _a in _cstates for _b in _rstates)):
                            continue
                        _rpts = [_pt_of(x) for x in items
                                 if x.get("LORRY") == _r and _pt_of(x)]
                        _add = [_pt_of(x) for x in _cit if _pt_of(x)]
                        if not _lorry_geo_ok(_rpts + _add):
                            continue
                        _cands.append(_r)
                    if not _cands:
                        continue                      # no ≤11T home — keep on big lorry
                    _best = min(
                        _cands,
                        key=lambda p: (0 if _uc_routes[p] & _croutes else 1,
                                       -(float(_lorry_cap_map[p]) - _uc_phys[p])))
                    for x in _cit:
                        x["LORRY"] = _best
                        sess["assigned"][x["DO NUMBER"]] = _best
                        _unassigned_reasons.pop(x["DO NUMBER"], None)
                    _uc_phys[_best] += _cw
                    _uc_phys[_big] -= _cw
                    _uc_routes[_best] |= _croutes
                    _session_loads[_best] = float(_session_loads.get(_best, 0)) + _cw
                    _record_lorry_state(_best, next(iter(_cstates), ""))

        # ── URBAN ROUTE-FIRST CONSOLIDATION + NEAREST-ROUTE FILL ──────────────
        # For URBAN (KL / Selangor) routes only: a whole route code should ride
        # ONE lorry (priority 1). Consolidate any urban route split across
        # lorries onto a single lorry that can hold the whole route within
        # capacity, size cap, forbidden plates and the geo spread — preferring
        # the lorry already holding most of it. Then, for a lorry with leftover
        # space, pull in the geographically NEAREST other urban route that still
        # fits (priority 2). Outstation routes (Kuantan, Perak, …) are NEVER
        # touched — they split by capacity necessity. A route stays split only
        # when no single lorry can hold it.
        def _urban_rc_items():
            _m: dict = {}
            for _it in items:
                if (_it.get("LORRY") in _lorry_cap_map
                        and _classify_dest_group(_it.get("ROUTE", ""),
                                                 _it.get("STATE", "")) in _DEST_URBAN_GROUPS):
                    _m.setdefault(_rcode(_it), []).append(_it)
            return _m

        _rc_phys: dict[str, float] = defaultdict(float)
        for _it in items:
            if _it.get("LORRY") in _lorry_cap_map:
                _rc_phys[_it["LORRY"]] += _it["WEIGHT"]

        def _route_fits_on(_cand, _rits, _rc, _rw, _rmt, _rforbid):
            if _cand in _blocked_today:
                return False
            _cap = float(_lorry_cap_map.get(_cand, 0))
            _here = sum(x["WEIGHT"] for x in _rits if x["LORRY"] == _cand)
            if _rc_phys[_cand] - _here + _rw > _cap * NAIK_FACTOR:
                return False
            if _rmt is not None and _cap > _rmt:
                return False
            if _cand in _rforbid:
                return False
            _other = [_pt_of(x) for x in items
                      if x.get("LORRY") == _cand and _rcode(x) != _rc and _pt_of(x)]
            _add = [_pt_of(x) for x in _rits if _pt_of(x)]
            return _lorry_geo_ok(_other + _add)

        # Priority 1 — consolidate each split urban route onto one lorry.
        for _rc, _rits in _urban_rc_items().items():
            _lset = {x["LORRY"] for x in _rits}
            if len(_lset) < 2:
                continue
            _rw = sum(x["WEIGHT"] for x in _rits)
            _rmt = min((x["MAX_TON"] for x in _rits
                        if x.get("MAX_TON") is not None), default=None)
            _rforbid: set = set()
            for x in _rits:
                if x.get("FORBID_PLATES"):
                    _rforbid |= x["FORBID_PLATES"]
            _rstates = {x.get("STATE", "").strip().upper() for x in _rits if x.get("STATE")}
            _tgt = None
            for _cand in sorted(_lset, key=lambda l: -sum(
                    x["WEIGHT"] for x in _rits if x["LORRY"] == l)):
                if _route_fits_on(_cand, _rits, _rc, _rw, _rmt, _rforbid):
                    _tgt = _cand
                    break
            if _tgt is None:
                for _cand in sorted(_lorry_cap_map, key=lambda l: (
                        0 if float(_lorry_cap_map[l]) <= _URBAN_MAX_TON else 1,
                        float(_lorry_cap_map[l]))):
                    if _cand in _lset:
                        continue
                    if _route_fits_on(_cand, _rits, _rc, _rw, _rmt, _rforbid):
                        _tgt = _cand
                        break
            if _tgt is None:
                continue                          # whole route fits no single lorry
            for x in _rits:
                if x["LORRY"] != _tgt:
                    _rc_phys[x["LORRY"]] -= x["WEIGHT"]
                    _rc_phys[_tgt] += x["WEIGHT"]
                    x["LORRY"] = _tgt
                    sess["assigned"][x["DO NUMBER"]] = _tgt
                    _unassigned_reasons.pop(x["DO NUMBER"], None)
                    _session_loads[_tgt] = float(_session_loads.get(_tgt, 0)) + x["WEIGHT"]
            _record_lorry_state(_tgt, next(iter(_rstates), ""))

        # Priority 2 — fill leftover space with the NEAREST other urban route.
        # For each urban lorry with spare capacity, find the whole urban route
        # (currently on another lorry, and itself un-split) whose centroid is
        # closest, and move it over if it fits within capacity and geo spread.
        def _centroid(_rits):
            _pts = [(_x["GPS_LAT"], _x["GPS_LON"]) for _x in _rits
                    if _x.get("GPS_LAT") is not None]
            if not _pts:
                return None
            return (sum(p[0] for p in _pts) / len(_pts),
                    sum(p[1] for p in _pts) / len(_pts))

        for _pass in range(2):                    # a couple of top-up rounds
            _urb = _urban_rc_items()
            _moved = False
            # lorries that currently carry an urban route, most free space first
            _urb_lorries = sorted(
                {x["LORRY"] for _r in _urb.values() for x in _r},
                key=lambda l: -(float(_lorry_cap_map[l]) - _rc_phys[l]))
            for _l in _urb_lorries:
                if _l in _blocked_today:
                    continue
                _free = float(_lorry_cap_map[_l]) * NAIK_FACTOR - _rc_phys[_l]
                if _free <= 0.05:
                    continue
                _lpts = [_pt_of(x) for x in items
                         if x.get("LORRY") == _l and _pt_of(x)]
                _lcent = _centroid([x for x in items if x.get("LORRY") == _l])
                if _lcent is None:
                    continue
                # candidate whole urban routes sitting on OTHER lorries
                _cands = []
                for _rc, _rits in _urb.items():
                    _lset = {x["LORRY"] for x in _rits}
                    if _l in _lset or len(_lset) != 1:
                        continue                  # only move a route that is whole elsewhere
                    _rw = sum(x["WEIGHT"] for x in _rits)
                    if _rw > _free:
                        continue
                    _rmt = min((x["MAX_TON"] for x in _rits
                                if x.get("MAX_TON") is not None), default=None)
                    if _rmt is not None and float(_lorry_cap_map[_l]) > _rmt:
                        continue
                    _rforbid: set = set()
                    for x in _rits:
                        if x.get("FORBID_PLATES"):
                            _rforbid |= x["FORBID_PLATES"]
                    if _l in _rforbid:
                        continue
                    _rcent = _centroid(_rits)
                    if _rcent is None:
                        continue
                    _dist = ((_lcent[0] - _rcent[0]) ** 2 + (_lcent[1] - _rcent[1]) ** 2) ** 0.5
                    if not _lorry_geo_ok(_lpts + [_pt_of(x) for x in _rits if _pt_of(x)]):
                        continue
                    _cands.append((_dist, _rc, _rits, _rw))
                if not _cands:
                    continue
                _cands.sort(key=lambda c: c[0])   # nearest route first
                _d, _rc, _rits, _rw = _cands[0]
                for x in _rits:
                    _rc_phys[x["LORRY"]] -= x["WEIGHT"]
                    _rc_phys[_l] += x["WEIGHT"]
                    x["LORRY"] = _l
                    sess["assigned"][x["DO NUMBER"]] = _l
                    _unassigned_reasons.pop(x["DO NUMBER"], None)
                    _session_loads[_l] = float(_session_loads.get(_l, 0)) + x["WEIGHT"]
                _record_lorry_state(_l, next(
                    (x.get("STATE", "").strip().upper() for x in _rits if x.get("STATE")), ""))
                _moved = True
            if not _moved:
                break

        # ── Priority 3 — partial CORRIDOR top-up (fill spare space) ───────────
        # A lorry carrying a corridor route (e.g. KV20A) that still has spare
        # space is topped up with individual same-CORRIDOR drops (e.g. KV19A
        # stops from another lorry) so the two south routes ride together and
        # the lorry is fully used. Only defined corridors (KV_SOUTH, KV_EAST,
        # KV_NORTH …) merge this way; whole GPS/customer units stay together and
        # the geo spread + size cap + forbidden plates are respected.
        _corridor_names = set(_ROUTE_CORRIDOR_GROUPS)

        def _urban_corridor_of(_it):
            if _classify_dest_group(_it.get("ROUTE", ""),
                                    _it.get("STATE", "")) not in _DEST_URBAN_GROUPS:
                return None
            _dk = _direction_key(_it.get("ROUTE", ""), _it.get("STATE", ""), _it.get("CUSTOMER NAME", ""))
            return _dk if _dk in _corridor_names else None

        _tc_phys: dict[str, float] = defaultdict(float)
        for _it in items:
            if _it.get("LORRY") in _lorry_cap_map:
                _tc_phys[_it["LORRY"]] += _it["WEIGHT"]
        for _l in sorted(_lorry_cap_map,
                         key=lambda l: -(float(_lorry_cap_map[l]) - _tc_phys[l])):
            if _l in _blocked_today:
                continue
            _free = float(_lorry_cap_map[_l]) * NAIK_FACTOR - _tc_phys[_l]
            if _free <= 0.05:
                continue
            _l_items = [x for x in items if x.get("LORRY") == _l]
            _l_dirs = {_urban_corridor_of(x) for x in _l_items} - {None}
            if not _l_dirs:
                continue                              # lorry isn't on an urban corridor
            # Candidates = same-corridor DOs on OTHER lorries OR still unassigned
            # (NO_LORRY) — so a corridor lorry with room also RESCUES stranded
            # same-corridor drops, not just rebalances.
            _cands = [x for x in items
                      if (x.get("LORRY") in _lorry_cap_map or x.get("LORRY") == "NO_LORRY")
                      and x.get("LORRY") != _l
                      and _urban_corridor_of(x) in _l_dirs]
            for _unit in sorted(_atomic_components(_cands, _rcode, _gps_of,
                                                   lambda x: x.get("CODE", "")),
                                key=lambda u: sum(x["WEIGHT"] for x in u)):
                _uw = sum(x["WEIGHT"] for x in _unit)
                if _uw > _free:
                    continue
                _umt = min((x["MAX_TON"] for x in _unit
                            if x.get("MAX_TON") is not None), default=None)
                if _umt is not None and float(_lorry_cap_map[_l]) > _umt:
                    continue
                _ufp: set = set()
                for x in _unit:
                    if x.get("FORBID_PLATES"):
                        _ufp |= x["FORBID_PLATES"]
                if _l in _ufp:
                    continue
                _lpts = [_pt_of(x) for x in items if x.get("LORRY") == _l and _pt_of(x)]
                _add = [_pt_of(x) for x in _unit if _pt_of(x)]
                if not _lorry_geo_ok(_lpts + _add):
                    continue
                for x in _unit:
                    _tc_phys[x["LORRY"]] -= x["WEIGHT"]
                    _tc_phys[_l] += x["WEIGHT"]
                    x["LORRY"] = _l
                    sess["assigned"][x["DO NUMBER"]] = _l
                    _unassigned_reasons.pop(x["DO NUMBER"], None)
                    _session_loads[_l] = float(_session_loads.get(_l, 0)) + x["WEIGHT"]
                _record_lorry_state(_l, next(
                    (x.get("STATE", "").strip().upper() for x in _unit if x.get("STATE")), ""))
                _free -= _uw
                if _free <= 0.05:
                    break

        # ── Final geo-cleanup — no far urban mixing ───────────────────────────
        # After every top-up/spillover, enforce the urban spread rule one last
        # time so nothing slips a far drop onto the wrong lorry: on any lorry,
        # two DIFFERENT urban route codes may not sit more than
        # _URBAN_MERGE_SPREAD apart. Detach the lighter offending route code to
        # NO_LORRY (the overload-rescue that runs next re-homes it on a NEAR
        # lorry, or it stays unassigned rather than ride a far-away truck).
        # A single route code spanning several towns is never split (only
        # DIFFERENT codes are compared).
        def _cleanup_far_urban():
            for _l in list(_lorry_cap_map):
                for _guard in range(20):                 # bounded refinement
                    _lit = [x for x in items if x.get("LORRY") == _l]
                    _upts = [(x, _pt_of(x)) for x in _lit if _pt_of(x)]
                    _upts = [(x, p) for x, p in _upts
                             if p[3] in _URBAN_COMPATIBLE_STATES]
                    _worst, _pair = 0.0, None
                    for _a in range(len(_upts)):
                        for _b in range(_a + 1, len(_upts)):
                            _pa, _pb = _upts[_a][1], _upts[_b][1]
                            if _pa[2] != _pb[2]:
                                _dd = _geo_deg(_pa[:2], _pb[:2])
                                if _dd > _worst:
                                    _worst, _pair = _dd, (_pa[2], _pb[2])
                    if _pair is None or _worst <= _URBAN_MERGE_SPREAD:
                        break
                    _ra, _rb = _pair
                    _wa = sum(x["WEIGHT"] for x in _lit if _route_code_of(x) == _ra)
                    _wb = sum(x["WEIGHT"] for x in _lit if _route_code_of(x) == _rb)
                    _drop = _ra if _wa <= _wb else _rb
                    for x in _lit:
                        if _route_code_of(x) == _drop:
                            x["LORRY"] = "NO_LORRY"
                            sess["assigned"][x["DO NUMBER"]] = "NO_LORRY"
                            _unassigned_reasons[x["DO NUMBER"]] = "GEO_FAR_URBAN"
        try:
            _cleanup_far_urban()
        except Exception as _e:
            import logging as _rlog
            _rlog.warning("[GEO-CLEANUP] skipped: %s", _e)

        # ── RULES-COMPLIANCE GATE (ASSIGNMENT_RULES.md / DO_BOT_SKILL.md §A) ───
        # Final deterministic audit: every assigned DO must obey the HARD rules.
        # Any violation is corrected (the DO is unassigned → NO_LORRY) and logged,
        # so a rule can never be silently broken in the output. No black box —
        # each check maps 1:1 to a written rule.
        _audit_caps = {str(r["LORRY"]).strip().upper(): float(r["TON"])
                       for _, r in engine.eligible_lorries.iterrows()}
        _audit_owner = set(_audit_caps)          # owner + SPARE fleet
        # Per-lorry total load — used to apply the tiny-NS relaxation in the audit
        # (a small Seremban load on a small lorry is allowed only when the whole
        # load on that lorry is tiny).
        _audit_load: dict = defaultdict(float)
        for _it in items:
            if _it.get("LORRY") in _audit_caps:
                _audit_load[_it["LORRY"]] += _it["WEIGHT"]
        _audit_viol: list[str] = []
        for _it in items:
            _l = _it.get("LORRY")
            if _l not in _audit_caps:
                continue                          # sentinels / blanks — skip
            _lt = _audit_caps[_l]
            _rt = _it.get("ROUTE", "")
            _dg = _classify_dest_group(_rt, _it.get("STATE", ""))
            _reason = None
            # Rule A1 — owner isolation
            if _l not in _audit_owner:
                _reason = "OWNER_ISOLATION"
            # Rule A2 — outstation minimum tonnage. Far outstation (LARGE_LONG)
            # needs >11T; nearer (MEDIUM_LONG) >5T — waived for a tiny NS load.
            elif _dg not in _DEST_URBAN_GROUPS and _lt < _eff_dest_min_ton(_rt, _dg, _audit_load[_l]):
                _reason = f"OUTSTATION_NEEDS_>{_eff_dest_min_ton(_rt, _dg, _audit_load[_l]):.0f}T"
            # Rule A5 — REMARKS / SHIP_DETAIL size cap (incl. MAX 2 TON → van)
            elif _it.get("MAX_TON") is not None and _lt > _it["MAX_TON"]:
                _reason = "SIZE_CAP_EXCEEDED"
            # Rule — REMARKS forbids this specific plate ("3875 tak boleh masuk")
            elif _it.get("FORBID_PLATES") and _l in _it["FORBID_PLATES"]:
                _reason = "PLATE_FORBIDDEN"
            if _reason:
                _audit_viol.append(f"{_it.get('DO NUMBER')}:{_l}({_lt}T):{_reason}")
                _it["LORRY"] = "NO_LORRY"
                sess["assigned"][_it["DO NUMBER"]] = "NO_LORRY"
                _unassigned_reasons[_it["DO NUMBER"]] = _reason
        if _audit_viol:
            import logging as _rlog
            _rlog.warning("[RULES-AUDIT] corrected %d violation(s): %s",
                          len(_audit_viol), "; ".join(_audit_viol[:20]))
        else:
            print(f"[RULES-AUDIT] OK — all assignments comply with the hard rules.")

        # ── Slight-overload rescue: place any capacity-stranded DOs the way a
        # human planner does (slight overload of small lorries, and reserving a
        # big lorry for a big DO) rather than leaving them unassigned. ────────
        try:
            _overload_rescue(sess)
        except Exception as _e:
            import logging as _rlog
            _rlog.warning("[OVERLOAD-RESCUE] skipped: %s", _e)

        # ── Right-size trucks: move a light load off a big lorry onto a
        # smaller idle one so big trucks don't roll out half-empty. ──────────
        try:
            _downsize_lorries(sess)
        except Exception as _e:
            import logging as _rlog
            _rlog.warning("[DOWNSIZE] skipped: %s", _e)

        # ── Urban rebalance: free a big lorry stuck with a lone small urban
        # load by repacking small-urban loads onto the smaller lorries. ──────
        try:
            _urban_rebalance(sess)
        except Exception as _e:
            import logging as _rlog
            _rlog.warning("[URBAN-REBALANCE] skipped: %s", _e)

        # ── Global optimizer (OFF by default; OPTIMIZER_ENABLED=1 to try) ─────
        # When enabled it replaces the greedy result with the CP-SAT plan; on
        # any problem it returns False and the greedy result above stands.
        try:
            import optimizer_bridge as _optb
            if _optb.optimizer_enabled():
                _optb.run_optimizer(sess, _is_urban_do, _is_kuantan)
        except Exception as _e:
            import logging as _rlog
            _rlog.warning("[OPTIMIZER] skipped, keeping greedy result: %s", _e)

        # ── (legacy for-loop removed — replaced by _assign_one above) ────────
        # The block below was the old heaviest-first loop.  Keep a dummy
        # reference so diff is minimal.
        # ── Build display groups: group items by DO NUMBER, preserving order ─
        # pending_dos is used by _build_summary; rebuild from items
        seen_do = {}
        pending_dos = []
        _remarks_skip_count = sum(1 for it in items if it.get("LORRY") in ("REMARKS_SKIP", "OUT_SOURCE"))
        for item in items:
            if item.get("LORRY") in ("OTHER_USER", "NOT_TODAY", "REMARKS_SKIP", "OUT_SOURCE", "PAST_DATE", "WRONG_TRIP"):
                continue          # keep in raw_df for export blank; hide from UI
            do_num = item["DO NUMBER"]
            # Multi-lorry split parts get their own display entry (keyed by their
            # synthetic DO NUMBER like "39877(p2)") so each lorry's section shows
            # the portion it carries instead of the full 27T.
            _display_key = do_num   # default: group under original DO number
            if item.get("SPLIT_PART"):
                _display_key = do_num   # already unique ("39877(p1)", "39877(p2)")
            if _display_key not in seen_do:
                seen_do[_display_key] = len(pending_dos)
                # Strip the synthetic part suffix for the display label
                _orig_num = do_num.split("(p")[0]
                pending_dos.append({
                    "DO NUMBER":     _orig_num,
                    "ALL_DO_NUMBERS": [do_num],
                    "ROUTE":         item["ROUTE"],
                    "CODE":          item["CODE"],
                    "CUSTOMER NAME": item["CUSTOMER NAME"],
                    "DATE":          item.get("DATE", ""),
                    "ITEMS":         [],          # list of item dicts
                    "SPLIT_PART":    item.get("SPLIT_PART"),
                })
            pending_dos[seen_do[_display_key]]["ITEMS"].append(item)

        # Compute TOTAL_TON and flatten split/single for display
        for do in pending_dos:
            do["TOTAL_TON"] = round(sum(it["WEIGHT"] for it in do["ITEMS"]), 3)

        sess["pending_dos"]          = pending_dos
        sess["change_do_page"]       = 0
        sess["unassigned_reasons"]   = _unassigned_reasons

        # ── Build and return summary ──────────────────────────────────────────
        my_items    = [it for it in items if it.get("LORRY") not in ("OTHER_USER", "NOT_TODAY", "REMARKS_SKIP", "OUT_SOURCE", "PAST_DATE", "WRONG_TRIP")]
        total_items = len(my_items)
        header = f"✅ *{total_items} item(s) across {len(pending_dos)} DO(s) auto-assigned!*"
        if _other_user_count:
            header += f"\n📌 _{_other_user_count} DO(s) from another user's routes — left blank (cross-user assignment not allowed)._"
        if _past_date_count:
            header += f"\n🗓️ _{_past_date_count} DO(s) dated before today — left unassigned (AI only assigns today's DOs and onward; assign these manually if still needed)._"
        if _wrong_trip_count:
            _other_trip = "AFTERNOON" if _trip_session == "MORNING" else "MORNING"
            header += f"\n🕐 _{_wrong_trip_count} DO(s) marked {_other_trip} TRIP in REMARKS — left unassigned (you picked the {_trip_session} trip)._"
        if _remarks_skip_count:
            header += f"\n📅 _{_remarks_skip_count} DO(s) skipped — REMARKS indicate delivery not due today._"
        if _not_today_count:
            # User's own DOs not on today's schedule — ask if they want to assign anyway
            sess["not_today_pending_count"] = _not_today_count
            sess["state"] = "AWAIT_OTHER_USER_REPLY"
        if _sched_notice:
            header += "\n" + "\n".join(_sched_notice)

        # ── Low-utilisation warnings for outstation lorries ───────────────────
        # Flag any lorry heading out of the city (LARGE_LONG / MEDIUM_LONG) at
        # less than 40% capacity — these are expensive trips with low payload.
        _lorry_items: dict[str, list] = {}
        for _it in my_items:
            _pl = _it.get("LORRY")
            if _pl and _pl not in {"NO_LORRY", "SPLIT", "SKIPPED", "OTHER_USER", "NOT_TODAY", "REMARKS_SKIP", "OUT_SOURCE", "PAST_DATE", "WRONG_TRIP", "", None}:
                _lorry_items.setdefault(_pl, []).append(_it)
        _low_util_warns = []
        for _pl, _its in _lorry_items.items():
            _pl_load = sum(_i["WEIGHT"] for _i in _its)
            _pl_cap  = float(_lorry_cap_map.get(_pl, 0)) if "_lorry_cap_map" in dir() else 0.0
            if _pl_cap <= 0:
                continue
            _util = _pl_load / _pl_cap
            # Only warn for outstation lorries (not urban)
            _pl_route = _its[0].get("ROUTE", "")
            _pl_state = _its[0].get("STATE", "")
            _pl_dest  = _classify_dest_group(_pl_route, _pl_state)
            if _pl_dest not in _DEST_URBAN_GROUPS and _util < 0.40:
                _low_util_warns.append(
                    f"⚠️ *{_pl}* outstation at only {_util*100:.0f}% load "
                    f"({_pl_load:.3f}T / {_pl_cap:.1f}T) — "
                    f"consider holding until more DOs accumulate."
                )
        if _low_util_warns:
            header += "\n" + "\n".join(_low_util_warns)

        # ── Unassigned reason breakdown ───────────────────────────────────────
        _no_lorry_items = [it for it in my_items if it.get("LORRY") == "NO_LORRY"]
        if _no_lorry_items and _unassigned_reasons:
            from collections import Counter
            _reason_counts = Counter(
                _unassigned_reasons.get(it["DO NUMBER"], "NO_ELIGIBLE_LORRY")
                for it in _no_lorry_items
            )
            _reason_labels = {
                "LOAD_EXCEEDS_ALL_LORRIES": "load > max lorry cap",
                "NO_ELIGIBLE_LORRY":        "no eligible lorry found",
                "CAPACITY_FULL":            "all lorries at capacity",
                "LOAD_BELOW_MIN_UTIL":      "load too small (min util rule)",
            }
            _reason_parts = [
                f"{v}× {_reason_labels.get(k, k)}"
                for k, v in sorted(_reason_counts.items(), key=lambda x: -x[1])
            ]
            header += f"\n❌ *Unassigned reasons:* " + ", ".join(_reason_parts)

        # ── DEBUG: show why each NO_LORRY item couldn't be assigned ──────────
        _debug_lines = []
        for _dbg_it in _no_lorry_items:
            _dbg_w     = _dbg_it["WEIGHT"]
            _dbg_route = _dbg_it.get("ROUTE", "")
            _dbg_state = _dbg_it.get("STATE", "").strip().upper()
            _dbg_dest  = _classify_dest_group(_dbg_route, _dbg_it.get("STATE", ""))
            _dbg_urban = _dbg_dest in _DEST_URBAN_GROUPS
            _dbg_strict = _strict_route_excl(_dbg_route)
            _rejections = []
            for _dp, _dc in _lorry_cap_map.items():
                _dl = float(_session_loads.get(_dp, 0))
                _dr = _dc - _dl
                if _dp in sess.get("unavailable", set()):
                    _rejections.append(f"{_dp}:UNAVAIL")
                elif not _dbg_urban and _dc < _DEST_MIN_TON.get(_classify_dest_group(_dbg_route), 0.0):
                    _rejections.append(f"{_dp}:TOO_SMALL_OUTSTATION")
                elif _dp in _dbg_strict:
                    _rejections.append(f"{_dp}:STRICT")
                elif _dr < _dbg_w and (_dr + 1.0) < _dbg_w:
                    _rejections.append(f"{_dp}:CAP({_dr:.2f}T)")
                else:
                    _rejections.append(f"{_dp}:OK(rem={_dr:.2f}T)")
            _debug_lines.append(
                f"🔍 DO {_dbg_it['DO NUMBER'][-5:]} {_dbg_w}T {_dbg_dest}\n"
                + "  unavail=" + str(sorted(sess.get("unavailable", set()))) + "\n"
                + "  " + " | ".join(_rejections)
            )
        if _debug_lines:
            header += "\n\n🛠 *DEBUG*\n" + "\n".join(_debug_lines)

        # ── Idle lorry diagnostic ─────────────────────────────────────────────
        _assigned_plates = {
            it.get("LORRY") for it in my_items
            if it.get("LORRY") not in {
                "NO_LORRY", "OTHER_USER", "NOT_TODAY", "REMARKS_SKIP", "OUT_SOURCE", "PAST_DATE", "WRONG_TRIP", "SPLIT", "SKIPPED", "", None
            }
        }
        _today_log_plates = get_assigned_today()
        _idle_lorries = [
            p for p in _lorry_cap_map
            if p not in _assigned_plates
            and p not in sess.get("unavailable", set())
            and p not in _today_log_plates
        ]
        _log_blocked_lorries = [
            p for p in _lorry_cap_map
            if p not in _assigned_plates
            and p not in sess.get("unavailable", set())
            and p in _today_log_plates
        ]

        def _lorry_diag_label(p):
            _ic = _lorry_cap_map.get(p, 0)
            _ip_routes = [pfx for pfx, plates in _ROUTE_PREFERRED_LORRY.items() if p in plates]
            if _ip_routes:
                return f"{p} ({_ic:.1f}T, preferred for {'/'.join(_ip_routes[:3])})"
            return f"{p} ({_ic:.1f}T, no preferred route)"

        if _idle_lorries:
            header += "\n🅿️ *Idle lorries:* " + ", ".join(_lorry_diag_label(p) for p in sorted(_idle_lorries))
        if _log_blocked_lorries:
            header += "\n🔒 *Blocked by today's log (prev session):* " + ", ".join(
                _lorry_diag_label(p) for p in sorted(_log_blocked_lorries)
            )

        _summ = _build_summary(sess)
        result_msgs = ([header + "\n\n" + _summ[0]] + _summ[1:]
                       if isinstance(_summ, list) else [header + "\n\n" + _summ])

        # If user's own off-schedule DOs were found:
        # → ask the YES/NO question FIRST, WITHOUT the full summary.
        # The summary + confirm buttons are shown AFTER the user answers.
        if sess.get("state") == "AWAIT_OTHER_USER_REPLY":
            _nt_count = sess.get("not_today_pending_count", 0)
            return [
                header,   # just the "X items assigned" line + schedule notice
                f"⏭ *{_nt_count} of your DO(s) are not on today's route schedule* and were left blank.\n"
                f"Do you want to assign them anyway? Reply *YES* to assign, or *NO* to leave them blank.",
            ]

        return result_msgs

    except Exception as e:
        import traceback
        traceback.print_exc()
        # nssm-run services often don't have a visible console for print_exc()
        # to land on — put the failing file:line right in the message shown
        # on screen so it's diagnosable without server log access.
        _tb = traceback.extract_tb(e.__traceback__)
        _loc = f" ({_tb[-1].filename.split(chr(92))[-1].split('/')[-1]}:{_tb[-1].lineno} in {_tb[-1].name})" if _tb else ""
        return [f"❌ Failed to process the DO file: {type(e).__name__}: {e}{_loc}\n"
                "Please screenshot this and send it back."]

def _handle_other_user_reply(phone, sess, text: str) -> list[str]:
    """Handle user's YES/NO reply about assigning their own off-schedule DOs."""
    reply = text.strip().upper()
    if reply in ("YES", "YA", "Y", "OK", "OKAY"):
        items = sess.get("items", [])
        not_today_items = [it for it in items if it.get("LORRY") == "NOT_TODAY"]
        if not not_today_items:
            sess["state"] = "CONFIRMING"
            return _build_summary(sess)

        # Preferred path: re-run the FULL assignment with the schedule filter
        # OFF, so EVERY rule (per-direction reservation, geo clustering,
        # same-destination, corridor combining, size caps, forbidden plates)
        # applies uniformly to ALL off-schedule routes — not just a simplified
        # single-suggest pass for one route.
        _bytes = sess.get("_upload_bytes")
        if _bytes:
            sess["_ignore_schedule"] = True
            try:
                _msgs = _handle_excel_upload(phone, sess, _bytes)
            finally:
                sess["_ignore_schedule"] = False
            return _msgs

        # Fallback (no stored upload): assign off-schedule DOs with the simpler
        # per-route pass below.
        # Clear NOT_TODAY marker so items are treated as unassigned
        for it in not_today_items:
            it["LORRY"] = None

        # Seed session loads AND lorry-state map from already-assigned items
        engine: LorryEngine = sess["engine"]
        _lorry_cap_map = {
            str(r["LORRY"]).strip().upper(): float(r["TON"])
            for _, r in engine.eligible_lorries.iterrows()
        }
        _session_loads: dict = {}
        _session_routes: dict = {}
        _lorry_states: dict = {}   # plate → set of destination states (for boundary check)
        for _it in items:
            _pl = _it.get("LORRY")
            if _pl and _pl not in (None, "NO_LORRY", "NOT_TODAY", "PAST_DATE", "WRONG_TRIP", "REMARKS_SKIP", "OUT_SOURCE", "OTHER_USER",
                                   "SPLIT", "SKIPPED", "", "LOAD_BELOW_MIN_UTIL"):
                _session_loads[_pl] = _session_loads.get(_pl, 0.0) + _it.get("WEIGHT", 0.0)
                if _pl not in _session_routes:
                    _session_routes[_pl] = _it.get("ROUTE", "")
                _it_state = _it.get("STATE", "").strip().upper()
                if _it_state:
                    _lorry_states.setdefault(_pl, set()).add(_it_state)

        # Assign grouped by route — apply same size and state-boundary rules as main assignment
        from collections import defaultdict as _dd
        _by_route: dict = _dd(list)
        for it in not_today_items:
            _by_route[it.get("ROUTE", "")].append(it)

        for _route, _grp in _by_route.items():
            _total_w = sum(x["WEIGHT"] for x in _grp)
            _grp_state = _grp[0].get("STATE", "").strip().upper()
            _dest_grp_r = _classify_dest_group(_route, _grp_state)
            _is_urban_r = _dest_grp_r in _DEST_URBAN_GROUPS

            # State-boundary exclusion: don't assign a lorry already committed to a
            # different (incompatible) state — e.g. W3826C on KL routes can't also do Pahang
            _state_excl = {
                p for p, sts in _lorry_states.items()
                if sts and _grp_state
                and _grp_state not in sts
                and not any(_states_compatible(_grp_state, _ls) for _ls in sts)
            }

            # Size exclusion: outstation → exclude small lorries (≤5T); urban → no upper cap
            _size_excl = {
                str(r["LORRY"]).strip().upper()
                for _, r in engine.eligible_lorries.iterrows()
                if (not _is_urban_r and float(r["TON"]) < _DEST_MIN_TON.get(_dest_grp_r, 0.0))
            }

            # A lorry flagged "unavailable" but carrying nothing (0 physical
            # load) is really idle — often one reserved for this direction — so
            # it should be usable for these off-schedule DOs. Only keep the
            # unavailable flag for lorries that actually carry something.
            _unavail_loaded = {p for p in sess.get("unavailable", set())
                               if float(_session_loads.get(p, 0)) > 0}
            _excl = (_unavail_loaded
                     | _state_excl | _size_excl
                     | {p for p, cap in _lorry_cap_map.items()
                        if cap - float(_session_loads.get(p, 0)) < _total_w})

            _suggs = engine.suggest(route=_route, total_ton=_total_w,
                                    unavailable=_excl, top_n=1,
                                    today_date_str="")
            if _suggs:
                _chosen = _suggs[0]["LORRY"]
                for _it in _grp:
                    _it["LORRY"] = _chosen
                _session_loads[_chosen] = float(_session_loads.get(_chosen, 0)) + _total_w
                if _chosen not in _session_routes:
                    _session_routes[_chosen] = _route
                if _grp_state:
                    _lorry_states.setdefault(_chosen, set()).add(_grp_state)
            else:
                for _it in _grp:
                    _it["LORRY"] = "NO_LORRY"

        # Add the newly assigned items into pending_dos so _build_summary shows them
        pending_dos = sess.get("pending_dos", [])
        seen_do = {do["DO NUMBER"]: i for i, do in enumerate(pending_dos)}
        for it in not_today_items:
            do_num = it["DO NUMBER"]
            if do_num not in seen_do:
                seen_do[do_num] = len(pending_dos)
                pending_dos.append({
                    "DO NUMBER":      do_num,
                    "ALL_DO_NUMBERS": [do_num],
                    "ROUTE":          it["ROUTE"],
                    "CODE":           it["CODE"],
                    "CUSTOMER NAME":  it["CUSTOMER NAME"],
                    "DATE":           it.get("DATE", ""),
                    "ITEMS":          [],
                })
            pending_dos[seen_do[do_num]]["ITEMS"].append(it)
        # Recompute TOTAL_TON for updated DOs
        for do in pending_dos:
            do["TOTAL_TON"] = round(sum(i["WEIGHT"] for i in do["ITEMS"]), 3)
        sess["pending_dos"] = pending_dos

        sess["state"] = "CONFIRMING"
        return _build_summary(sess)

    elif reply in ("NO", "TIDAK", "SKIP", "N"):
        sess["state"] = "CONFIRMING"
        return _build_summary(sess)

    return [
        "Please reply *YES* to assign the off-schedule DOs, or *NO* to leave them blank."
    ]


def _suggest_current(sess) -> list[str]:
    idx = sess["current_do_index"]
    dos = sess["pending_dos"]

    if idx >= len(dos):
        return _finish_session(sess)

    do = dos[idx]
    engine: LorryEngine = sess["engine"]

    # Combine session unavailable + already assigned today
    excluded = sess["unavailable"] | get_assigned_today()

    suggestions = engine.suggest(
        route=do["ROUTE"],
        total_ton=do["TOTAL_TON"],
        unavailable=excluded,
        top_n=3,
    )
    sess["suggestions"] = suggestions
    sess["state"] = "REVIEWING"

    header = (
        f"📦 DO {idx + 1}/{len(dos)}\n"
        f"  *DO#* {do['DO NUMBER']}\n"
        f"  *Customer:* {do['CUSTOMER NAME']}\n"
        f"  *Route:* {do['ROUTE']}\n"
        f"  *Total weight:* {round(do['TOTAL_TON'], 3)} T\n"
    )

    if not suggestions:
        return [
            header +
            "\n⚠️ *No eligible lorry found* (all may be assigned today or over capacity).\n"
            "Reply *skip* to skip this DO or *custom [PLATE]* to assign manually."
        ]

    lines = [header + "\n🚛 *Suggested lorries:*"]
    for i, s in enumerate(suggestions, 1):
        lines.append(
            f"  *{i}.* {s['LORRY']} ({s['TON_CAPACITY']}T, {s['USER']})\n"
            f"     _{s['REASON']}_"
        )
    lines.append(
        "\nReply:\n"
        "  • *1 / 2 / 3* — to assign that lorry\n"
        "  • *block [PLATE]* — lorry unavailable all day\n  • *broken [PLATE] [REPLACEMENT]* — log breakdown & replacement\n"
        "  • *custom [PLATE]* — to assign any plate manually\n"
        "  • *skip* — skip this DO"
    )
    return ["\n".join(lines)]


def _broken_confirmed_reply(broken: str, replacement: str, sess: dict) -> list:
    """
    After a broken+replacement pair is confirmed:
    1. Log the breakdown message.
    2. Find every item currently assigned to the broken lorry.
    3. Re-assign each one to the replacement (or auto-pick if replacement="NONE").
    4. Return the breakdown summary + updated full assignment summary.
    """
    engine = sess.get("engine")
    cap_info = ""
    if engine is not None and replacement != "NONE":
        row = engine.eligible_lorries[engine.eligible_lorries["LORRY"] == replacement]
        if not row.empty:
            cap  = float(row.iloc[0]["TON"])
            user = str(row.iloc[0]["USER"])
            cap_info = f" ({cap}T, {user})"

    # ── Find and re-assign items that used the broken lorry ──────────────────
    reassigned = []   # list of (item, new_lorry)
    pending    = sess.get("pending_dos", [])

    for do in pending:
        for item in do.get("ITEMS", []):
            item_lorry = item.get("LORRY", "")

            # Check single-lorry assignment
            if item_lorry == broken:
                new_lorry = _pick_replacement(
                    broken, replacement, item, sess, engine
                )
                item["LORRY"]         = new_lorry
                item["SPLIT_LORRIES"] = None
                if new_lorry not in ("NO_LORRY", "SPLIT", None):
                    sess["unavailable"].add(new_lorry)
                reassigned.append((item, do, new_lorry))

            # Check split bins
            elif item_lorry == "SPLIT" and item.get("SPLIT_LORRIES"):
                for bin_ in (item.get("SPLIT_LORRIES") or []):
                    if bin_["lorry"] == broken:
                        new_lorry = _pick_replacement(
                            broken, replacement, item, sess, engine,
                            portion=sum(r["W"] for r in bin_["rows"])
                        )
                        bin_["lorry"] = new_lorry
                        if new_lorry not in ("NO_LORRY", None):
                            sess["unavailable"].add(new_lorry)
                        reassigned.append((item, do, new_lorry))

    # ── Build breakdown message ───────────────────────────────────────────────
    rep_display = f"*{replacement}*{cap_info}" if replacement != "NONE" else "none"
    lines = [
        f"🔧 *Breakdown logged:*",
        f"  ❌ Broken:      *{broken}*",
        f"  ✅ Replacement: {rep_display}",
        f"",
        f"*{broken}* is blocked for today.",
    ]

    if reassigned:
        lines.append(f"")
        lines.append(f"♻️ *{len(reassigned)} item(s) re-assigned:*")
        for item, do, new_lorry in reassigned:
            itmref  = item.get("ITMREF", "") or ""
            itmref_str = f" ({itmref})" if itmref and itmref.lower() not in ("nan","") else ""
            cust    = do.get("CUSTOMER NAME", "")[:22]
            w       = round(item.get("WEIGHT", 0), 2)
            lorry_s = f"*{new_lorry}*" if new_lorry not in ("NO_LORRY", None) else "❌ no lorry"
            lines.append(f"  {cust}{itmref_str} {w}T → {lorry_s}")
    else:
        lines.append("")
        lines.append("No items were assigned to this lorry.")

    msgs = ["\n".join(lines)]

    # ── Append updated full summary if we're in CONFIRMING state ─────────────
    if sess.get("state") in ("CONFIRMING", "REVIEWING") and pending:
        _sc = _build_summary(sess)
        if isinstance(_sc, list):
            msgs.extend(_sc)
        else:
            msgs.append(_sc)

    # ── Re-export if items were re-assigned and we already have a raw_df ──────
    # If the user had already confirmed (or we are in CONFIRMING), the previous
    # export is now stale. Re-run the export silently and store new bytes so
    # app.py will send the updated file automatically.
    if reassigned and sess.get("raw_df") is not None:
        try:
            _export_result(sess)   # updates sess["export_bytes"] in place
            msgs.append("📎 Updated assignment file is being sent.")
        except Exception as _e:
            msgs.append(f"⚠️ Could not regenerate export: {_e}")

    return msgs


def _pick_replacement(broken: str, replacement: str, item: dict,
                      sess: dict, engine, portion: float = None) -> str:
    """
    Pick the best lorry to replace broken for this item.
    - If replacement is specified and available → use it directly.
    - Otherwise auto-pick tightest fit from engine.
    - Falls back to NO_LORRY if nothing available.
    """
    weight = portion if portion is not None else item.get("WEIGHT", 0)

    # Try the nominated replacement first
    if replacement and replacement != "NONE":
        excluded = sess.get("unavailable", set()) | get_assigned_today()
        if replacement not in excluded:
            return replacement
        # replacement already taken — fall through to auto-pick

    # Auto-pick: tightest single lorry excluding already-used
    excluded = (sess.get("unavailable", set()) | get_assigned_today()) - {broken}
    if engine is not None:
        suggestions = engine.suggest(
            route=item.get("ROUTE", ""),
            total_ton=weight,
            unavailable=excluded,
            top_n=1,
        )
        if suggestions:
            return suggestions[0]["LORRY"]
        # Nothing fits by weight — use largest available as last resort
        last_resort = engine.suggest_largest_available(
            item.get("ROUTE", ""), excluded)
        if last_resort:
            return last_resort[0]["LORRY"]

    return "NO_LORRY"


def _handle_broken_list(sess: dict) -> list:
    """Show all broken lorries and their replacements for today."""
    broken_map = get_broken_lorries()
    if not broken_map:
        return [
            "✅ No broken lorries recorded today.",
            {"_type": "buttons",
             "body": "What would you like to do next?",
             "buttons": [{"id": "hi", "title": "👋 Hi"}]},
        ]
    engine = sess.get("engine")
    lines = ["🔧 *Broken lorries today:*\n"]
    for broken, replacement in sorted(broken_map.items()):
        rep_info = replacement
        if engine is not None and replacement != "NONE":
            row = engine.eligible_lorries[engine.eligible_lorries["LORRY"] == replacement]
            if not row.empty:
                cap = float(row.iloc[0]["TON"])
                rep_info = f"{replacement} ({cap}T)"
        status = f"→ replaced by *{rep_info}*" if replacement != "NONE" else "→ no replacement"
        lines.append(f"  ❌ *{broken}* {status}")
    lines.append("\nType *fixed [PLATE]* if lorry was in broken list.")
    lines.append("Type *release [PLATE1] [PLATE2...]* to unblock any lorry(s) from today's log.")
    return [
        "\n".join(lines),
        {"_type": "buttons",
         "body": "What would you like to do next?",
         "buttons": [{"id": "hi", "title": "👋 Hi"}]},
    ]


def _handle_broken_replacement(phone: str, sess: dict, text: str) -> list:
    """Handle the user's reply when we asked which lorry replaces the broken one."""
    broken = sess.get("pending_broken_plate", "")
    reply  = text.strip().upper()

    if reply in ("NONE", "NO", "-", "NIL"):
        # No replacement — just block the broken lorry
        record_broken_lorry(broken, "NONE")
        sess["unavailable"].add(broken)
        # Restore previous state
        sess["state"] = sess.pop("state_before_broken", "IDLE")
        sess.pop("pending_broken_plate", None)
        prev_state = sess["state"]
        msgs = [
            f"🔧 *{broken}* marked as broken with no replacement.\n"
            f"It is blocked for today."
        ]
        if prev_state == "REVIEWING":
            msgs.append("Continuing with your DO assignments...")
            msgs += _suggest_current(sess)
        else:
            msgs.append({"_type": "buttons",
                         "body": "What would you like to do next?",
                         "buttons": [{"id": "hi", "title": "👋 Hi"}]})
        return msgs

    replacement = reply
    if replacement == broken:
        return [f"⚠️ Replacement cannot be the same as the broken lorry (*{broken}*). Try again."]

    # Validate plate exists in master (warn but don't block)
    engine = sess.get("engine")
    plate_known = False
    if engine is not None:
        plate_known = replacement in engine.eligible_lorries["LORRY"].values or \
                      replacement in engine.all_lorries["LORRY"].values

    record_broken_lorry(broken, replacement)
    sess["unavailable"].add(broken)
    sess["state"] = sess.pop("state_before_broken", "IDLE")
    sess.pop("pending_broken_plate", None)
    prev_state = sess["state"]

    msgs = _broken_confirmed_reply(broken, replacement, sess)

    if not plate_known and engine is not None:
        msgs.insert(1, f"⚠️ Note: *{replacement}* is not in the master lorry list. "
                       "Double-check the plate number.")

    # _broken_confirmed_reply already appends the updated summary
    # when state is CONFIRMING/REVIEWING — no need to call _suggest_current

    return msgs


def _handle_reviewing(phone, sess, text):
    cmd = text.strip().lower()
    suggestions = sess["suggestions"]

    if cmd in ("1", "2", "3"):
        pick = int(cmd) - 1
        if pick < len(suggestions):
            chosen = suggestions[pick]["LORRY"]
            return _assign_and_next(sess, chosen)
        return ["Invalid selection. Reply 1, 2, or 3."]

    if cmd.startswith("block "):
        plate = text.split(" ", 1)[1].strip().upper()
        sess["unavailable"].add(plate)
        # Save to daily log so it stays blocked all day across all sessions
        record_assignments_today([plate], user=sess.get("user_id"))
        return [f"🚫 {plate} blocked for the entire day (won't appear again today)."] + _suggest_current(sess)

    if cmd.startswith("custom "):
        plate = text.split(" ", 1)[1].strip().upper()
        # Reject if this plate is already assigned to this DO (shouldn't normally happen,
        # but guard against re-submitting the same suggestion plate)
        do = sess["pending_dos"][sess["current_do_index"]]
        if plate == sess["assigned"].get(do["DO NUMBER"]):
            return [f"⚠️ *{plate}* is already assigned to this DO. Choose a different lorry."]
        return _assign_and_next(sess, plate)

    if cmd == "skip":
        idx = sess["current_do_index"]
        do = sess["pending_dos"][idx]
        sess["assigned"][do["DO NUMBER"]] = "SKIPPED"
        sess["current_do_index"] += 1
        return [f"⏭️ DO {do['DO NUMBER']} skipped."] + _suggest_current(sess)

    return ["Please reply with 1, 2, 3, *block [PLATE]*, *custom [PLATE]*, or *skip*."]


def _assign_and_next(sess, lorry_plate):
    idx = sess["current_do_index"]
    do = sess["pending_dos"][idx]
    sess["assigned"][do["DO NUMBER"]] = lorry_plate
    sess["current_do_index"] += 1
    # Block this lorry from appearing again in the same session
    sess["unavailable"].add(lorry_plate)
    return [f"✅ *{lorry_plate}* assigned to DO {do['DO NUMBER']}."] + _suggest_current(sess)



def _lorry_picker_buttons(sess: dict, do_num: str, page: int = 0) -> list:
    """
    Show up to 2 lorry options as tappable buttons + a Next/Prev navigation button.
    WhatsApp allows max 3 buttons per message, so:
      Button 1: Lorry option A
      Button 2: Lorry option B (if available)
      Button 3: "Next ▶" or "◀ Prev" (navigation)

    Each button tap sends "select_lorry [DO#] [PLATE]" back to the bot.
    Navigation sends "select_do [DO#] [page]".
    """
    engine: LorryEngine = sess.get("engine")
    if not engine:
        return ["❌ No engine loaded. Please restart with hi."]

    # Find the target DO/item
    target = None
    for it in sess.get("items", []):
        if it["DO NUMBER"] == do_num:
            target = it
            break
    if not target:
        return [f"❌ DO# {do_num} not found."]

    # Release current lorry from excluded so it appears as an option
    cur_lorry = target.get("LORRY", "")
    split_plates = set()
    if cur_lorry == "SPLIT" and target.get("SPLIT_LORRIES"):
        for b in target["SPLIT_LORRIES"]:
            split_plates.add(b["lorry"])
    excluded = (sess.get("unavailable", set()) | get_assigned_today()) - {cur_lorry} - split_plates

    weight = target["WEIGHT"]
    route  = target["ROUTE"]
    cust   = target.get("CUSTOMER NAME", "")

    # Get suggestions — fetch enough for pagination
    suggestions = engine.suggest(
        route=route, total_ton=weight,
        unavailable=excluded, top_n=20,
        customer_name=cust,
    )

    # Build option list: auto-pick first, then suggestions
    options = [{"plate": "__AUTO__", "label": "Auto-pick best", "desc": "Bot chooses optimal lorry"}]
    for s in suggestions:
        util = round((weight / s["TON_CAPACITY"]) * 100, 1) if s["TON_CAPACITY"] > 0 else 0
        freq = f"{s['FREQ']}trips" if s["FREQ"] > 0 else "new"
        options.append({
            "plate": s["LORRY"],
            "label": s["LORRY"],
            "desc":  f"{s['TON_CAPACITY']}T {util}% {freq}",
        })

    PER_PAGE = 2  # 2 lorry options + 1 nav button = 3 total
    total_pages = max(1, -(-len(options) // PER_PAGE))  # ceiling div
    page = max(0, min(page, total_pages - 1))

    slice_start = page * PER_PAGE
    slice_end   = slice_start + PER_PAGE
    page_opts   = options[slice_start:slice_end]

    # Build buttons
    buttons = []
    for opt in page_opts:
        label = opt["label"][:18] + ".." if len(opt["label"]) > 20 else opt["label"]
        buttons.append({
            "id":    f"select_lorry {do_num} {opt['plate']}",
            "title": label,
        })

    # Navigation button
    if total_pages > 1:
        if page < total_pages - 1:
            buttons.append({"id": f"select_do {do_num} {page+1}", "title": f"More ({page+1+1}/{total_pages})"})
        else:
            buttons.append({"id": f"select_do {do_num} 0",       "title": "From start"})

    # Truncate to max 3
    buttons = buttons[:3]

    # Body text: show current assignment + lorry details
    cur_label = ", ".join(split_plates) if split_plates else (cur_lorry if cur_lorry and cur_lorry not in ("NO_LORRY","SPLIT","") else "None")
    details = []
    for opt in page_opts:
        if opt["plate"] != "__AUTO__":
            details.append(f"  {opt['label']}: {opt['desc']}")
    detail_str = "\n".join(details)
    page_info  = f"Page {page+1}/{total_pages}" if total_pages > 1 else ""

    body = (
        f"DO: {do_num}\n"
        f"Weight: {round(weight,3)}T  Route: {route[:30]}\n"
        f"Current: {cur_label}\n"
        f"{page_info}\n"
        f"{detail_str}"
    ).strip()

    return [{"_type": "buttons", "body": body[:1024], "buttons": buttons}]


def _build_summary(sess) -> str:
    """Build a clean, mobile-friendly assignment summary."""
    pending  = sess["pending_dos"]   # list of DO groups, each with ITEMS list
    no_lorry = []
    lines    = []

    # Show broken lorry notice at top of summary if any are active
    broken_map = get_broken_lorries()
    if broken_map:
        broken_lines = ["🔧 *Active breakdowns today:*"]
        for bp, rp in sorted(broken_map.items()):
            rep = f"replaced by *{rp}*" if rp != "NONE" else "no replacement"
            broken_lines.append(f"  ❌ {bp} → {rep}")
        lines.append("\n".join(broken_lines))
        lines.append("─" * 20)

    taken_today   = get_assigned_today()
    broken_today  = set(get_broken_lorries().keys())   # lorries marked broken
    # Collect plates assigned in this session (already visible on item rows)
    session_plates = set(
        it.get("LORRY","") for do in sess.get("pending_dos",[]) for it in do.get("ITEMS",[])
        if it.get("LORRY") not in ("SPLIT","NO_LORRY",None,"")
    )
    # Also collect split bin lorries from this session
    for do in sess.get("pending_dos",[]):
        for it in do.get("ITEMS",[]):
            if it.get("LORRY") == "SPLIT":
                for b in (it.get("SPLIT_LORRIES") or []):
                    session_plates.add(b.get("lorry",""))
    # Blocked = plates in today's log that are NOT in this session AND NOT broken
    # (broken lorries already shown under Active breakdowns header above)
    extra_blocked = (taken_today - session_plates - broken_today) - {""}

    lines.append("📋 *ASSIGNMENT SUMMARY*")
    if extra_blocked:
        lines.append("⛔ Blocked: " + ", ".join(sorted(extra_blocked)))
    lines.append("━━━━━━━━━━━━━━━━━━━━━")
    lines.append("")

    engine = sess.get("engine")

    # ── Build lorry-grouped view ──────────────────────────────────────────
    # Collect all items and build per-lorry buckets (exclude other-user rows)
    all_items = [it for do in pending for it in do.get("ITEMS", [])
                 if it.get("LORRY") not in ("OTHER_USER", "NOT_TODAY", "REMARKS_SKIP", "OUT_SOURCE", "PAST_DATE", "WRONG_TRIP")]

    # Map DO NUMBER → (customer_short, route_code, date)
    do_meta: dict[str, tuple] = {}
    for do in pending:
        dn  = do["DO NUMBER"]
        cust = do["CUSTOMER NAME"][:10]
        route_code = do["ROUTE"].split(" - ")[0].strip()[:8] if " - " in do["ROUTE"] else do["ROUTE"][:8]
        dt  = do.get("DATE", "")
        if dt and dt.lower() in ("nan", "none", ""):
            dt = ""
        do_meta[dn] = (cust, route_code, dt)

    # Group items by lorry plate
    from collections import defaultdict as _dd
    lorry_items: dict[str, list] = _dd(list)  # plate → [item, ...]
    no_lorry_items: list = []
    for it in all_items:
        if it["LORRY"] in ("NO_LORRY", None):
            no_lorry_items.append(it)
        else:
            lorry_items[it["LORRY"]].append(it)

    # Lorry capacities
    cap_map: dict[str, float] = {}
    if engine is not None:
        for _, r in engine.eligible_lorries.iterrows():
            cap_map[r["LORRY"]] = float(r["TON"])

    # Sort lorries: by earliest date among their items, then by total weight desc
    def _lorry_sort(plate):
        its = lorry_items[plate]
        dates = [_parse_date_sortkey(it.get("DATE", "")) for it in its]
        return (min(dates) if dates else "9999-12-31", -sum(i["WEIGHT"] for i in its))

    # _parse_date_sortkey may not be in scope here (defined inside _handle_excel_upload).
    # Use a local re-implementation for sorting.
    def _dsort(s):
        s = (s or "").strip()
        if not s or s.lower() in ("nan", "none", ""):
            return "9999-12-31"
        for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
        try:
            ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
            if pd.notna(ts):
                return ts.strftime("%Y-%m-%d")
        except Exception:
            pass
        return "9999-12-31"

    sorted_lorries = sorted(lorry_items.keys(),
        key=lambda p: (
            min((_dsort(it.get("DATE","")) for it in lorry_items[p]), default="9999-12-31"),
            -sum(i["WEIGHT"] for i in lorry_items[p])
        ))

    for plate in sorted_lorries:
        its   = lorry_items[plate]
        total_w = round(sum(i["WEIGHT"] for i in its), 3)
        cap     = cap_map.get(plate)

        if cap and cap > 0:
            # Lorry size class label (LARGE ≥11T, MEDIUM 5–11T, SMALL 2–5T, VAN <2T)
            if cap >= 11.0:
                _size_tag = "LARGE"
            elif cap >= 5.0:
                _size_tag = "MEDIUM"
            elif cap >= 2.0:
                _size_tag = "SMALL"
            else:
                _size_tag = "VAN"

            # Trip display: LARGE (≥11T) never splits; MEDIUM/SMALL on LOCAL routes may split.
            _trip_info = ""
            _all_local_its = all(
                _classify_dest_group(i.get("ROUTE", "")) in _DEST_URBAN_GROUPS
                for i in its
            )
            if cap < 11.0 and _all_local_its and total_w > cap * 1.02:
                _t1w = round(cap, 1)
                _t2w = round(total_w - cap, 3)
                _trip_info = (f"  🌅Morning(T1):{_t1w}T"
                              f"  🌆Afternoon(T2):{_t2w}T")

            util_pct = round(min(total_w, cap) / cap * 100, 1)
            if util_pct > 100:
                util_tag = f"🔴 {util_pct}% OVER"
            elif util_pct >= 75:
                util_tag = f"✅ {util_pct}%"
            elif util_pct >= 50:
                util_tag = f"🟡 {util_pct}%"
            else:
                util_tag = f"⚠️ {util_pct}%"
            cap_str = f"{cap}T/{_size_tag}"
        else:
            util_tag  = ""
            cap_str   = "?"
            _trip_info = ""

        lines.append(f"🚛 *{plate}* ({cap_str})  {util_tag}  _{total_w}T_{_trip_info}")

        # One line per DO under this lorry: DO# first, then route→dest, customer, weight, date
        for it in sorted(its, key=lambda x: _dsort(x.get("DATE", ""))):
            dn   = it["DO NUMBER"]
            dn_short = dn[-5:] if len(dn) >= 5 else dn
            w    = round(it["WEIGHT"], 3)
            cust, rcode, dt = do_meta.get(dn, (dn, "", ""))
            dt_tag = f" [{dt}]" if dt else ""
            _dest_lbl = {
                "LARGE_LONG": "🟥", "MEDIUM_LONG": "🟧",
                "SELANGOR": "🟦", "KL": "🟩", "KL_SELANGOR": "🟩",
            }.get(_classify_dest_group(it.get("ROUTE", ""), it.get("STATE", "")), "")
            lines.append(f"  {dn_short}  {_dest_lbl}{rcode}  {cust}  {w}T{dt_tag}")

        lines.append("")   # blank line between lorries

    # ── No-lorry items ────────────────────────────────────────────────────
    _reasons = sess.get("unassigned_reasons", {})
    if no_lorry_items:
        lines.append(f"❌ *NO LORRY ({len(no_lorry_items)} item(s)):*")
        for it in no_lorry_items:
            dn   = it["DO NUMBER"]
            dn_short = dn[-5:] if len(dn) >= 5 else dn
            w    = round(it["WEIGHT"], 3)
            cust, rcode, dt = do_meta.get(dn, (dn, "", ""))
            dt_tag = f" [{dt}]" if dt else ""
            _rsn = _reasons.get(dn, "")
            _rsn_tag = f" ({_rsn})" if _rsn else ""
            lines.append(f"  {dn_short}  {rcode}  {cust}  {w}T{dt_tag}{_rsn_tag}")
        lines.append("")

    # ── Footer ────────────────────────────────────────────────────────────
    assigned_ok = len(all_items) - len(no_lorry_items)
    unassigned  = len(no_lorry_items)

    lines.append(f"✅ {assigned_ok} assigned  ❌ {unassigned} unassigned  🚛 {len(sorted_lorries)} lorry(s)")
    summary_text = "\n".join(lines)

    result = [summary_text]

    # ── Yes / No confirm buttons ──────────────────────────────────────────
    result.append({
        "_type": "buttons",
        "body":  "Confirm assignments?",
        "buttons": [
            {"id": "yes", "title": "✅ Yes, Export"},
            {"id": "no",  "title": "❌ Cancel"},
        ],
    })

    # Change Assignment (DO picker) is hidden for now.
    return result

def _handle_confirming(phone, sess, text):
    cmd = text.strip().lower()

    # Pagination for Change DO list
    if cmd.startswith("change_do_page "):
        try:
            page = int(cmd.split()[1])
            sess["change_do_page"] = page
        except (IndexError, ValueError):
            sess["change_do_page"] = 0
        return _build_summary(sess)

    # ── Propagate change to same-route DOs ───────────────────────────────────
    if cmd == "propagate yes":
        ctx = sess.pop("_propagate_ctx", None)
        if ctx:
            plate     = ctx["plate"]
            old_lorry = ctx["old_lorry"]
            dos       = ctx["dos"]
            updated   = []
            for do in sess["pending_dos"]:
                if do["DO NUMBER"] not in dos:
                    continue
                for it in do.get("ITEMS", []):
                    if it.get("LORRY") == old_lorry:
                        it["LORRY"]         = plate
                        it["SPLIT_LORRIES"] = None
                sess["assigned"][do["DO NUMBER"]] = plate
                updated.append(do["DO NUMBER"])
            msg = f"✅ Updated {len(updated)} DO(s) to *{plate}*: {', '.join(updated)}"
        else:
            msg = "✅ No propagation context found."
        return [msg] + _build_summary(sess)

    if cmd == "propagate no":
        sess.pop("_propagate_ctx", None)
        return ["✅ Change applied to selected DO only."] + _build_summary(sess)

    if cmd in ("yes", "confirm", "ok"):
        return _export_result(sess)

    if cmd in ("no", "cancel"):
        reset_session(phone)
        return ["❌ Cancelled. Send *hi* to start again."]

    # change [DO#] [PLATE1] [PLATE2] ... — reassign with 1 plate (single) or 2+ (split)
    # change [DO#]                       — auto-pick next best lorry
    if cmd.startswith("change "):
        parts = text.strip().split()
        if len(parts) < 2:
            return ["Usage: *change [DO#] [PLATE]* — single lorry\n"
                    "       *change [DO#] [PLATE1] [PLATE2]* — split across lorries\n"
                    "       *change [DO#]* — auto-pick next best"]

        do_num = parts[1].upper()

        # Find the target DO and ALL its items
        target_item = None
        target_do   = None
        for do in sess["pending_dos"]:
            if do["DO NUMBER"] == do_num:
                target_do = do
                if do.get("ITEMS"):
                    target_item = do["ITEMS"][0]   # primary ref for weight/route
                break

        if target_item is None:
            return [f"❌ DO# *{do_num}* not found. Check the number and try again."]

        engine: LorryEngine = sess["engine"]
        old_lorry = target_item["LORRY"]   # "SPLIT", "NO_LORRY", or plate string

        # ── Release old lorry(s) from unavailable pool ─────────────────────
        def _release_item_lorries(item):
            if item["LORRY"] == "SPLIT" and item.get("SPLIT_LORRIES"):
                for b in (item.get("SPLIT_LORRIES") or []):
                    sess["unavailable"].discard(b["lorry"])
            elif item["LORRY"] not in (None, "NO_LORRY", "SPLIT"):
                sess["unavailable"].discard(item["LORRY"])

        _release_item_lorries(target_item)

        plates = [p.upper() for p in parts[2:]]  # 0 = auto, 1 = single, 2+ = split

        # ── AUTO mode: no plate given ───────────────────────────────────────
        if not plates:
            # Always exclude current lorry(s) — user wants something DIFFERENT
            old_plates = set()
            if old_lorry == "SPLIT" and target_item.get("SPLIT_LORRIES"):
                old_plates = {b["lorry"] for b in (target_item.get("SPLIT_LORRIES") or [])}
            elif old_lorry not in (None, "NO_LORRY", "SPLIT"):
                old_plates = {old_lorry}

            # Merge broken lorries into unavailable — their replacements stay free
            broken_map = get_broken_lorries()
            sess["unavailable"].update(broken_map.keys())  # block broken lorries
            excluded = sess["unavailable"] | get_assigned_today() | old_plates

            # Step 1: try a different single lorry
            suggestions = engine.suggest(
                route=target_item["ROUTE"],
                total_ton=target_item["WEIGHT"],
                unavailable=excluded,
                top_n=1,
            )
            if suggestions:
                new_lorry = suggestions[0]["LORRY"]
                # Update ALL items in this DO
                for it in target_do.get("ITEMS", []):
                    it["LORRY"]         = new_lorry
                    it["SPLIT_LORRIES"] = None
                sess["assigned"][do_num]     = new_lorry
                sess["unavailable"].add(new_lorry)
                reason = suggestions[0]["REASON"]
                return [
                    f"✅ {do_num} auto-reassigned → *{new_lorry}*\n_{reason}_"
                ] + (_s2 if isinstance(_s2 := _build_summary(sess), list) else [_s2])

            # Step 2: no single lorry fits — greedy split across smaller lorries
            remain = target_item["WEIGHT"]
            bins   = []
            excl   = set(excluded)
            for _ in range(10):
                if remain <= 0:
                    break
                sug = engine.suggest(route=target_item["ROUTE"], total_ton=remain,
                                     unavailable=excl, top_n=1)
                if sug:
                    lorry_s, cap_s = sug[0]["LORRY"], sug[0]["TON_CAPACITY"]
                else:
                    all_sug = engine.suggest(route=target_item["ROUTE"], total_ton=0.01,
                                             unavailable=excl, top_n=20)
                    all_sug.sort(key=lambda x: x["TON_CAPACITY"], reverse=True)
                    if not all_sug:
                        break
                    lorry_s, cap_s = all_sug[0]["LORRY"], all_sug[0]["TON_CAPACITY"]
                portion = round(min(cap_s, remain), 6)
                bins.append({"lorry": lorry_s,
                             "rows": [{"DO": do_num, "W": portion}],
                             "cap": round(cap_s - portion, 4)})
                excl.add(lorry_s)
                remain = round(remain - cap_s, 6)

            if remain <= 0 and bins:
                for b in bins:
                    sess["unavailable"].add(b["lorry"])
                target_item["LORRY"]         = "SPLIT"
                target_item["SPLIT_LORRIES"] = bins
                sess["assigned"][do_num]     = "SPLIT"
                plate_str = " + ".join(b["lorry"] for b in bins)
                return [
                    f"✅ {do_num} → split: *{plate_str}*\n"
                    f"(no single lorry available — split across {len(bins)} lorries)"
                ] + (_s2 if isinstance(_s2 := _build_summary(sess), list) else [_s2])

            # Nothing works at all
            target_item["LORRY"]     = "NO_LORRY"
            sess["assigned"][do_num] = "NO_LORRY"
            return [f"⚠️ No alternative lorry available for DO *{do_num}*. "
                    "All eligible lorries are assigned or blocked."]

        # ── SINGLE plate ────────────────────────────────────────────────────
        if len(plates) == 1:
            plate = plates[0]
            # Check not already used elsewhere in this batch
            for do in sess["pending_dos"]:
                for it in do.get("ITEMS", []):
                    if it is target_item:
                        continue
                    if it["LORRY"] == plate:
                        return [f"❌ *{plate}* is already assigned to "
                                f"DO {do['DO NUMBER']} in this batch. Use a different lorry."]
            blocked_today = get_assigned_today()
            if plate in blocked_today:
                return [f"❌ *{plate}* is already assigned/blocked today. Use a different lorry."]

            old_lorry = target_item.get("LORRY", "")
            sess["_last_change_old_lorry"] = old_lorry

            target_item["LORRY"]         = plate
            target_item["SPLIT_LORRIES"] = None
            # Update ALL items in this DO
            for it in target_do.get("ITEMS", []):
                it["LORRY"]         = plate
                it["SPLIT_LORRIES"] = None
            sess["assigned"][do_num] = plate
            old_lorry = sess.get("_last_change_old_lorry")
            sess["unavailable"].add(plate)

            # ── Check if other DOs share the same route AND old lorry ────────
            # If so, ask user whether to propagate the change to them too
            same_route_dos = []
            changed_route  = target_do.get("ITEMS", [{}])[0].get("ROUTE", "")
            for do in sess["pending_dos"]:
                if do["DO NUMBER"] == do_num:
                    continue
                for it in do.get("ITEMS", []):
                    if (it.get("ROUTE") == changed_route and
                            it.get("LORRY") == old_lorry and
                            old_lorry not in (None, "", "NO_LORRY", "SPLIT")):
                        same_route_dos.append(do["DO NUMBER"])
                        break

            if same_route_dos:
                # Store context for propagation confirmation
                sess["_propagate_ctx"] = {
                    "plate":    plate,
                    "old_lorry": old_lorry,
                    "dos":      same_route_dos,
                }
                do_list_str = ", ".join(same_route_dos[:5])
                return [
                    f"✅ {do_num} → *{plate}*\n\n"
                    f"The following DOs share the same route with *{old_lorry}*:\n"
                    f"{do_list_str}\n\n"
                    "Apply the same change to these DOs too?",
                    {"_type": "buttons",
                     "body": "Propagate change?",
                     "buttons": [
                         {"id": "propagate yes", "title": "Yes, update all"},
                         {"id": "propagate no",  "title": "No, keep as-is"},
                     ]}
                ]

            return [
                f"✅ {do_num} → *{plate}*\n"
                "Reply *yes* to confirm or *change [DO#] [PLATE]* to adjust more."
            ] + _build_summary(sess)

        # ── SPLIT: 2 or more plates ─────────────────────────────────────────
        # Validate plates
        blocked_today = get_assigned_today()
        errors = []
        for plate in plates:
            for do in sess["pending_dos"]:
                for it in do.get("ITEMS", []):
                    if it is target_item:
                        continue
                    if it["LORRY"] == plate:
                        errors.append(f"*{plate}* already assigned to DO {do['DO NUMBER']}")
            if plate in blocked_today:
                errors.append(f"*{plate}* is blocked today")
        if errors:
            # Restore old assignment before returning error
            if old_lorry not in (None, "NO_LORRY", "SPLIT"):
                sess["unavailable"].add(old_lorry)
            elif target_item.get("SPLIT_LORRIES"):
                for b in (target_item.get("SPLIT_LORRIES") or []):
                    sess["unavailable"].add(b["lorry"])
            return ["❌ " + " | ".join(errors)]

        # Build bins: distribute weight across lorries in order given
        remain = target_item["WEIGHT"]
        bins   = []
        for plate in plates:
            # Look up this lorry's capacity from the engine
            row = engine.eligible_lorries[engine.eligible_lorries["LORRY"] == plate]
            cap = float(row.iloc[0]["TON"]) if not row.empty else remain
            portion = round(min(cap, remain), 6)
            bins.append({
                "lorry":  plate,
                "rows":   [{"DO": do_num, "W": portion}],
                "cap":    round(cap - portion, 4),
            })
            sess["unavailable"].add(plate)
            remain = round(remain - portion, 6)
            if remain <= 0:
                break

        if remain > 0:
            # Lorries given can't cover full weight
            for b in bins:
                sess["unavailable"].discard(b["lorry"])
            return [f"⚠️ The lorries given can only carry "
                    f"{round(target_item['WEIGHT'] - remain, 2)}T of "
                    f"{target_item['WEIGHT']}T. Add more lorries or choose larger ones."]

        target_item["LORRY"]         = "SPLIT"
        target_item["SPLIT_LORRIES"] = bins
        # Update ALL items in this DO
        for it in target_do.get("ITEMS", []):
            it["LORRY"]         = "SPLIT"
            it["SPLIT_LORRIES"] = bins
        sess["assigned"][do_num]     = "SPLIT"

        plate_str = " + ".join(b["lorry"] for b in bins)
        return [
            f"✅ {do_num} → *{plate_str}*\n(split load)\n"
            "Reply *yes* to confirm or *change [DO#] [PLATE]* to adjust more."
        ] + (_s2 if isinstance(_s2 := _build_summary(sess), list) else [_s2])

    # block [PLATE] — mark lorry unavailable all day and re-run auto-assign
    if cmd.startswith("block "):
        plate = text.split(" ", 1)[1].strip().upper()
        sess["unavailable"].add(plate)
        record_assignments_today([plate], user=sess.get("user_id"))
        # Re-run auto-assign for any DO currently assigned to this plate
        engine: LorryEngine = sess["engine"]
        changed = []
        for do in sess["pending_dos"]:
            do_num = do["DO NUMBER"]
            for item in do.get("ITEMS", []):
                if item["LORRY"] != plate:
                    continue
                # Merge broken lorries into unavailable — their replacements stay free
                broken_map = get_broken_lorries()
                sess["unavailable"].update(broken_map.keys())
                excluded = sess["unavailable"] | get_assigned_today()
                suggestions = engine.suggest(
                    route=item["ROUTE"],
                    total_ton=item["WEIGHT"],
                    unavailable=excluded,
                    top_n=1,
                )
                if suggestions:
                    new_lorry = suggestions[0]["LORRY"]
                    item["LORRY"] = new_lorry
                    sess["assigned"][do_num] = new_lorry
                    sess["unavailable"].add(new_lorry)
                    changed.append(f"  • {do_num} → *{new_lorry}*")
                else:
                    item["LORRY"] = "NO_LORRY"
                    sess["assigned"][do_num] = "NO_LORRY"
                    changed.append(f"  • {do_num} → *No lorry available*")
        msg = [f"🚫 *{plate}* blocked for today."]
        if changed:
            msg.append("Re-assigned affected DOs:\n" + "\n".join(changed))
        # Show full updated summary so user can review before confirming
        return ["\n".join(msg)] + (_s2 if isinstance(_s2 := _build_summary(sess), list) else [_s2])

    return ["Please reply *yes*, *change [DO#] [PLATE]*, *block [PLATE]*, or *no*."]

def _finish_session(sess) -> list[str]:
    sess["state"] = "CONFIRMING"
    assigned = sess["assigned"]
    lines = ["🎉 *All DOs reviewed!*\n\nAssignment summary:"]
    for do_num, lorry in assigned.items():
        lines.append(f"  • {do_num} → *{lorry}*")
    lines.append("\nReply *yes* to get the filled Excel file, or *no* to redo.")
    return ["\n".join(lines)]


def _export_result(sess) -> list[str]:
    try:
        return _export_result_inner(sess)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return [f"❌ Export failed: {e}\nYour assignments are saved. Try typing *yes* again or send *hi* to restart."]

def _export_result_inner(sess) -> list[str]:
    from datetime import date as _date
    today_str = _date.today().strftime("%d-%m-%Y")  # e.g. 12-05-2026
    is_new_fmt = sess.get("is_new_format", False)

    # ── Work on a copy of the raw uploaded DataFrame ─────────────────────────
    new_df: pd.DataFrame = sess["raw_df"].copy()

    # Ensure LICENSE column exists as object dtype (string-capable)
    if "LICENSE" not in new_df.columns:
        new_df["LICENSE"] = ""
    new_df["LICENSE"] = new_df["LICENSE"].astype(object)

    # Only touch DATE for old format; new format leaves DATE exactly as uploaded
    if not is_new_fmt:
        if "DATE" not in new_df.columns:
            new_df["DATE"] = ""
        new_df["DATE"] = new_df["DATE"].astype(object)

    # Blank out any stale sentinel strings in LICENSE so we only write real plates
    new_df["LICENSE"] = new_df["LICENSE"].astype(str).replace({"nan": "", "None": ""})

    SENTINELS = {"SKIPPED", "NO_LORRY", "SPLIT", "OTHER_USER", "NOT_TODAY", "REMARKS_SKIP", "OUT_SOURCE", "PAST_DATE", "WRONG_TRIP", "", None}
    confirmed_plates = []
    assigned_row_idxs = set()

    # ── Write LICENSE per original row index (item["ROW_IDX"]) ──────────────
    for do in sess.get("pending_dos", []):
        for item in do.get("ITEMS", []):
            lorry   = item.get("LORRY")
            row_idx = item.get("ROW_IDX")

            if lorry == "SPLIT" and item.get("SPLIT_LORRIES"):
                bins = (item.get("SPLIT_LORRIES") or [])
                if row_idx is not None and bins:
                    all_plates = ", ".join(b["lorry"] for b in bins
                                          if b["lorry"] not in SENTINELS)
                    new_df.loc[row_idx, "LICENSE"] = all_plates
                    if not is_new_fmt:
                        new_df.loc[row_idx, "DATE"] = today_str
                    assigned_row_idxs.add(row_idx)
                    for b in bins:
                        if b["lorry"] not in SENTINELS:
                            confirmed_plates.append(b["lorry"])
                continue

            if lorry in SENTINELS or row_idx is None:
                continue

            new_df.loc[row_idx, "LICENSE"] = lorry
            if not is_new_fmt:
                new_df.loc[row_idx, "DATE"] = today_str
            assigned_row_idxs.add(row_idx)
            if lorry not in SENTINELS:
                confirmed_plates.append(lorry)

    # ── For new format: enforce correct column order ──────────────────────────
    # Column N (index 13, 1-based col 14) must be LICENSE.
    # We rebuild the column order to match the required spec:
    # NO DATE DO-NUMBER CODE ROUTE CUSTOMER-NAME BRANCH GROSS-WEIGHT REMARKS
    # VALIDATED INVOICE-NO INV-DATE SITE LICENSE DRIVER LORRY-ASST1 LORRY-ASST2 DISTANCE
    NEW_FMT_COLS = [
        "NO", "DATE", "DO NUMBER", "CODE", "ROUTE", "CUSTOMER NAME",
        "BRANCH", "GROSS WEIGHT", "REMARKS", "VALIDATED", "INVOICE NO",
        "INV DATE", "SITE", "LICENSE", "DRIVER", "LORRY ASST1", "LORRY ASST2", "DISTANCE",
    ]
    # Drop internal helper columns added during processing (not in original spec)
    _INTERNAL_COLS = {"WEIGHT(T)"}

    # Add DEST_STATE column showing classified destination per row
    # Uses actual STATE column from uploaded file when available; falls back to route prefix.
    if "ROUTE" in new_df.columns:
        _dest_state_map = {
            "LARGE_LONG":  "OUTSTATION-LARGE",
            "MEDIUM_LONG": "OUTSTATION-MEDIUM",
            "KL":          "KUALA LUMPUR",
            "SELANGOR":    "SELANGOR",
            "KL_SELANGOR": "KL/SELANGOR",
        }
        def _row_dest_state(row):
            st = _state_from_row(row)
            grp = _classify_dest_group(str(row.get("ROUTE", "")), st)
            return _dest_state_map.get(grp, st or "")
        new_df["DEST_STATE"] = new_df.apply(_row_dest_state, axis=1)

    if is_new_fmt:
        # Reorder columns: required spec first, then any extras from the upload
        ordered = [c for c in NEW_FMT_COLS if c in new_df.columns]
        extras  = [c for c in new_df.columns if c not in NEW_FMT_COLS and c not in _INTERNAL_COLS]
        new_df  = new_df[ordered + extras]

        # Rows for history append = only the newly assigned rows
        new_rows = new_df.loc[sorted(assigned_row_idxs)].copy()
    else:
        # Old format: only export assigned rows (existing behaviour)
        new_rows = new_df.loc[sorted(assigned_row_idxs)].copy()
        if "DATE" in new_rows.columns:
            new_rows["DATE"] = today_str

    # ── Append assigned rows into master history file ─────────────────────────
    _hist_path = _resolve_history_path()
    try:
        existing_df = pd.read_excel(_hist_path)
        existing_df.columns = [c.strip().upper() for c in existing_df.columns]
        # Old format history: normalise DATE strings
        if not is_new_fmt and "DATE" in existing_df.columns:
            def _fmt_date(v):
                try:
                    ts = pd.to_datetime(v, errors="coerce")
                    return str(v) if pd.isna(ts) else ts.strftime("%d-%m-%Y")
                except Exception:
                    return str(v)
            existing_df["DATE"] = existing_df["DATE"].apply(_fmt_date)
        # Align columns for concat
        hist_rows = new_rows.drop(columns=[c for c in _INTERNAL_COLS if c in new_rows.columns],
                                  errors="ignore")
        for col in hist_rows.columns:
            if col not in existing_df.columns:
                existing_df[col] = ""
        for col in existing_df.columns:
            if col not in hist_rows.columns:
                hist_rows[col] = ""
        hist_rows = hist_rows[existing_df.columns]
        combined  = pd.concat([existing_df, hist_rows], ignore_index=True)
    except Exception:
        combined = new_rows.drop(columns=[c for c in _INTERNAL_COLS if c in new_rows.columns],
                                 errors="ignore")

    if "DATE" in combined.columns:
        combined["DATE"] = combined["DATE"].astype(str)
    _tmp_path = _hist_path + "._tmp.xlsx"
    combined.to_excel(_tmp_path, index=False, engine="openpyxl")
    os.replace(_tmp_path, _hist_path)

    # ── Build export bytes ────────────────────────────────────────────────────
    # New format: send back the FULL uploaded file (all rows) with LICENSE filled
    # Old format: send only the newly assigned rows (existing behaviour)
    if is_new_fmt:
        out_df = new_df.drop(columns=[c for c in _INTERNAL_COLS if c in new_df.columns],
                             errors="ignore").copy()
        # DATE is already formatted as "d/m/yy" string at upload time (see _handle_excel_upload)
    else:
        out_df = new_rows.copy()
        if "DATE" in out_df.columns:
            out_df["DATE"] = out_df["DATE"].astype(str)

    # ── TRIP column: mark morning (1) vs afternoon (2) for lorries used twice ──
    # A lorry that appears on two different route groups in the same session did
    # two trips.  The first group it served = Trip 1 (morning); second = Trip 2.
    if "LICENSE" in out_df.columns and "ROUTE" in out_df.columns:
        _KV_RE = re.compile(r'^KV\d', re.IGNORECASE)
        # Build plate → ordered list of row indices (preserving output order)
        _plate_rows: dict[str, list] = {}
        for _ri, _r in out_df.iterrows():
            _pl = str(_r.get("LICENSE", "")).strip().upper()
            if _pl and _pl.lower() not in ("nan", "none", "", "no_lorry",
                                           "split", "skipped", "other_user"):
                _plate_rows.setdefault(_pl, []).append(_ri)

        # Track trip number per lorry per row.
        # Trip rules (based on lorry size AND route type):
        #   LARGE  (≥14T): always TRIP 1 — one outstation run per day.
        #   MEDIUM (11–14T) on OUTSTATION routes: TRIP 1 only.
        #   MEDIUM (11–14T) on LOCAL (KV/KL/Selangor) routes: TRIP 1 morning,
        #     TRIP 2 afternoon when cumulative weight exceeds capacity.
        #     Morning: 8:00–11:59am  |  Afternoon: 1:00–5:30pm
        #   SMALL  (<11T) on LOCAL routes: TRIP 1 / TRIP 2 by capacity.
        _trip_vals = {}
        for _pl, _ridxs in _plate_rows.items():
            _cap_val = 0.0
            _eng = sess.get("engine")
            if _eng is not None:
                for _, _er in _eng.eligible_lorries.iterrows():
                    if str(_er["LORRY"]).strip().upper() == _pl:
                        _cap_val = float(_er["TON"])
                        break

            # Determine if ALL rows for this lorry are on LOCAL routes
            _lorry_routes = [str(out_df.at[_ri, "ROUTE"]).strip() for _ri in _ridxs]
            _lorry_all_local = all(
                _classify_dest_group(r) in _DEST_URBAN_GROUPS
                for r in _lorry_routes
            )

            _trip_num = 1
            _cum_w    = 0.0
            for _ri in _ridxs:
                _wt_raw = pd.to_numeric(
                    out_df.at[_ri, "WEIGHT(T)"] if "WEIGHT(T)" in out_df.columns
                    else out_df.at[_ri, "GROSS WEIGHT"],
                    errors="coerce"
                )
                _wt = (float(_wt_raw) / 1000.0
                       if "GROSS WEIGHT" in out_df.columns and "WEIGHT(T)" not in out_df.columns
                       else float(_wt_raw or 0))

                # LARGE lorries (≥11T): always TRIP 1 regardless of route
                # MEDIUM/SMALL on LOCAL routes: allow TRIP 2 when capacity overflows
                _can_trip2 = (
                    _cap_val > 0
                    and _cap_val < 11.0       # not a large lorry
                    and _lorry_all_local      # only local (KV) routes
                    and _cum_w + _wt > _cap_val * 1.02
                )
                if _can_trip2:
                    _trip_num = 2
                    _cum_w    = _wt
                else:
                    _cum_w += _wt
                _trip_vals[_ri] = _trip_num

        if _trip_vals:
            _trip_col = [
                str(_trip_vals[i]) if i in _trip_vals else ""
                for i in range(len(out_df))
            ]
            # Only insert TRIP column if any lorry does 2 trips (and it isn't
            # already present from a re-uploaded prior output).
            if any(v == 2 for v in _trip_vals.values()) and "TRIP" not in out_df.columns:
                _lic_loc = out_df.columns.get_loc("LICENSE")
                out_df.insert(_lic_loc, "TRIP", _trip_col)

    # ── Preserve the uploaded file's exact column layout ──────────────────────
    # Return the same columns, in the same order, as the file the user uploaded.
    # Only the assignment columns (LICENSE, etc.) are filled in — no added
    # TRIP / DEST_STATE / WEIGHT(T), and INVOICE DATE keeps its original spot.
    _orig_cols = sess.get("_orig_cols")
    if _orig_cols:
        for _c in _orig_cols:
            if _c not in out_df.columns:
                out_df[_c] = ""
        out_df = out_df[[_c for _c in _orig_cols if _c in out_df.columns]]

    buf = io.BytesIO()
    out_df.to_excel(buf, index=False, engine="openpyxl")

    # For new format: force DATE column cells to Text format so Excel never
    # re-interprets "11/5/26" as a date serial and shows the timestamp again.
    if is_new_fmt and "DATE" in out_df.columns:
        from openpyxl import load_workbook as _load_wb
        buf.seek(0)
        _wb = _load_wb(buf)
        _ws = _wb.active
        _date_col = out_df.columns.get_loc("DATE") + 1  # 1-based
        for _row in _ws.iter_rows(min_row=2, min_col=_date_col, max_col=_date_col):
            for _cell in _row:
                _cell.number_format = "@"  # @ = Text — prevents Excel date re-parsing
        buf = io.BytesIO()
        _wb.save(buf)

    buf.seek(0)
    sess["export_bytes"] = buf.read()

    # Generate trip manifest for drivers (second file sent alongside the export)
    try:
        sess["trip_manifest_bytes"] = _generate_trip_manifest(sess)
    except Exception as _tm_err:
        print(f"⚠️ Trip manifest generation failed: {_tm_err}")
        sess["trip_manifest_bytes"] = None

    # Persist confirmed plates to daily log (attributed to this user)
    record_assignments_today(list(set(confirmed_plates)), user=sess.get("user_id"))

    sess["state"] = "DONE"
    row_count = len(new_rows)
    total_count = len(new_df)
    if is_new_fmt:
        summary = (f"✅ *{row_count}/{total_count} rows* assigned and appended to history.\n"
                   f"📎 Sending you the complete file with LICENSE filled.")
    else:
        summary = (f"✅ *{row_count} rows* appended to the master trip route file.\n"
                   f"📎 Sending you a copy of the newly assigned rows.")
    return [
        summary,
        {
            "_type": "buttons",
            "body": "Tap below to start a new session, or type *hi* anytime.",
            "buttons": [{"id": "hi", "title": "👋 Hi"}],
        }
    ]

def _generate_trip_manifest(sess) -> bytes:
    """
    Build a driver-friendly trip manifest Excel workbook.
    Stops are sorted geographically using a greedy nearest-neighbour algorithm
    starting from the depot, so the driver follows the most logical road sequence.
    Within a given date, stops are chained by proximity; dates appear in order.
    Columns: # | DATE | DO# | Customer | Route/Area | WT(T) | Dist | Remarks
    """
    import math
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import date as _date, datetime as _dt
    from collections import defaultdict as _dd
    from lorry_engine import _haversine_km, _DEPOT, road_matrix_km, _osrm_table_km

    wb = Workbook()
    wb.remove(wb.active)

    generated_str = _date.today().strftime("%d-%m-%Y")
    raw_df = sess.get("raw_df")

    _DEPOT = (3.0340, 101.5563)   # Eng Sheng HQ, Shah Alam

    # ── Geo helpers ───────────────────────────────────────────────────────────
    def _haversine(lat1, lon1, lat2, lon2) -> float:
        R = 6371.0
        dlat = math.radians(lat2 - lat1)
        dlon = math.radians(lon2 - lon1)
        a = (math.sin(dlat / 2) ** 2
             + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2))
             * math.sin(dlon / 2) ** 2)
        return R * 2 * math.asin(math.sqrt(min(1.0, a)))

    def _parse_latlon(row_idx) -> tuple[float, float] | None:
        """Parse 'lat lon' or 'lat, lon' from the LONGITUD column."""
        if raw_df is None or row_idx is None:
            return None
        try:
            v = str(raw_df.loc[row_idx, "LONGITUD"]).strip()
            if not v or v.lower() in ("nan", "none", ""):
                return None
            v = v.replace(",", " ")
            parts = v.split()
            if len(parts) >= 2:
                return (float(parts[0]), float(parts[1]))
        except Exception:
            pass
        return None

    def _nn_sort(pairs: list) -> list:
        """
        Order stops using a greedy nearest-neighbour chain starting from the
        depot, so the driver travels the shortest practical path without
        zig-zagging across the route.

        Distances are REAL driving distances via OSRM (free OpenStreetMap
        routing) when reachable, falling back to straight-line haversine when
        OSRM is unavailable.

        Stops with no GPS coordinates are appended at the end, sorted by route
        then customer name.
        """
        with_coords = [(do, it, _parse_latlon(it.get("ROW_IDX"))) for do, it in pairs]
        has_coords  = [(do, it, ll) for do, it, ll in with_coords if ll is not None]
        no_coords   = [(do, it)     for do, it, ll in with_coords if ll is None]

        if not has_coords:
            no_coords.sort(key=lambda x: (x[0]["ROUTE"], x[0]["CUSTOMER NAME"]))
            return no_coords

        # Build coordinate list: index 0 is the depot, 1..N are the stops.
        coords = [(_DEPOT[0], _DEPOT[1])] + [ll for _, _, ll in has_coords]
        dist = road_matrix_km(coords)     # real road km (OSRM) or haversine

        # Greedy nearest-neighbour starting from the depot (index 0).
        n = len(has_coords)
        visited = [False] * n
        ordered = []
        cur = 0                           # current matrix index (0 = depot)
        for _ in range(n):
            nearest = min(
                (i for i in range(n) if not visited[i]),
                key=lambda i: dist[cur][i + 1],
            )
            visited[nearest] = True
            do, it, _ll = has_coords[nearest]
            ordered.append((do, it))
            cur = nearest + 1             # +1 because depot occupies index 0

        # Append stops with no coordinates sorted by route then customer
        no_coords.sort(key=lambda x: (x[0]["ROUTE"], x[0]["CUSTOMER NAME"]))
        ordered.extend(no_coords)
        return ordered

    # ── Date helpers ──────────────────────────────────────────────────────────
    def _fmt_date(s) -> str:
        s = str(s).strip()
        if not s or s.lower() in ("nan", "none", "nat", ""):
            return ""
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y"):
            try:
                return _dt.strptime(s, fmt).strftime("%d-%m-%Y")
            except ValueError:
                pass
        try:
            ts = pd.to_datetime(s, format="mixed", dayfirst=True, errors="coerce")
            if pd.notna(ts):
                return ts.strftime("%d-%m-%Y")
        except Exception:
            pass
        return s

    def _date_sortkey(s) -> str:
        """Return YYYY-MM-DD for chronological sort, '9999-12-31' on failure."""
        s = str(s or "").strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y"):
            try:
                return _dt.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
        try:
            ts = pd.to_datetime(s, format="mixed", dayfirst=True, errors="coerce")
            if pd.notna(ts):
                return ts.strftime("%Y-%m-%d")
        except Exception:
            pass
        return "9999-12-31"

    # ── Gather items per lorry ────────────────────────────────────────────────
    lorry_pairs: dict[str, list] = _dd(list)
    no_lorry_pairs: list         = []

    for do in sess.get("pending_dos", []):
        for it in do.get("ITEMS", []):
            lorry = it.get("LORRY") or "NO_LORRY"
            if lorry in ("NO_LORRY", None, ""):
                no_lorry_pairs.append((do, it))
            elif lorry != "SPLIT":
                lorry_pairs[lorry].append((do, it))

    engine  = sess.get("engine")
    cap_map: dict[str, float] = {}
    if engine is not None:
        for _, r in engine.eligible_lorries.iterrows():
            cap_map[r["LORRY"]] = float(r["TON"])

    sorted_lorries = sorted(
        lorry_pairs.keys(),
        key=lambda p: sum(it["WEIGHT"] for _, it in lorry_pairs[p]),
        reverse=True,
    )

    # ── Shared styles ─────────────────────────────────────────────────────────
    _thin       = Side(style="thin", color="BBBBBB")
    _brd        = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _TITLE_FILL = PatternFill("solid", fgColor="1F4E79")
    _TITLE_FONT = Font(color="FFFFFF", bold=True, size=11)
    _HDR_FILL   = PatternFill("solid", fgColor="2E75B6")
    _HDR_FONT   = Font(color="FFFFFF", bold=True, size=9)
    _ALT_FILL   = PatternFill("solid", fgColor="DEEBF7")
    _DATE_FILL  = PatternFill("solid", fgColor="E2EFDA")  # green tint = first row of new date
    _FOOT_FILL  = PatternFill("solid", fgColor="FCE4D6")
    _FOOT_FONT  = Font(bold=True, size=9)
    _NL_FILL    = PatternFill("solid", fgColor="C00000")

    HEADERS    = ["#", "DATE", "DO #", "CUSTOMER", "ROUTE / AREA", "WT (T)", "DIST", "REMARKS / NOTES"]
    COL_WIDTHS = [4,   11,     9,      22,          32,             8,        8,      42]

    def _apply_headers(ws, row):
        for ci, (hdr, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            c = ws.cell(row, ci, hdr)
            c.font = _HDR_FONT; c.fill = _HDR_FILL; c.border = _brd
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[row].height = 16

    def _route_display(route_str: str) -> str:
        if " - " in route_str:
            parts = route_str.split(" - ")
            return f"{parts[0]}: {' - '.join(parts[1:3])}"[:32]
        return route_str[:32]

    def _raw_val(row_idx, col: str) -> str:
        if raw_df is None or row_idx is None:
            return ""
        try:
            v = str(raw_df.loc[row_idx, col]).strip()
            if v in ("nan", "None", "NaN", ""):
                return ""
            # Ignore GPS-coordinate strings accidentally placed in DISTANCE column
            if col == "DISTANCE" and re.match(r"^-?\d+\.\d+\s+-?\d+\.\d+", v):
                return ""
            return v
        except Exception:
            return ""

    # ── Trip-time estimation constants ───────────────────────────────────────
    _DEPART_HOUR      = 7       # driver leaves HQ at 07:00
    _DEADLINE_HOUR    = 13      # must return by 13:00
    _AVG_SPEED_KMH    = 60.0   # average road speed
    _ROAD_FACTOR      = 1.4    # haversine × factor ≈ road distance
    _STOP_MIN         = 20     # average unloading time per stop (minutes)
    _OUTSTATION_KM    = 150    # last-stop distance from HQ > this → overnight

    def _est_return(sorted_pairs_: list) -> tuple[str, str, bool]:
        """
        Estimate return time to HQ and 1pm margin.
        Returns (return_time_str, margin_str, is_outstation).
        """
        from datetime import timedelta as _td, datetime as _dtt
        if not sorted_pairs_:
            return ("—", "—", False)

        # Last stop GPS
        last_ll = None
        for _do, _it in reversed(sorted_pairs_):
            ll = _parse_latlon(_it.get("ROW_IDX"))
            if ll:
                last_ll = ll
                break

        # Outstation check
        if last_ll:
            dist_to_hq = _haversine(_DEPOT[0], _DEPOT[1], last_ll[0], last_ll[1]) * _ROAD_FACTOR
            if dist_to_hq > _OUTSTATION_KM:
                return ("OUTSTATION", f"{dist_to_hq:.0f} km back", True)

        # Estimate travel + stop time
        n_stops    = len(sorted_pairs_)
        stop_min   = n_stops * _STOP_MIN

        # Sum segment distances: HQ→stop1→stop2→…→last→HQ
        coords = [_DEPOT]
        for _do, _it in sorted_pairs_:
            ll = _parse_latlon(_it.get("ROW_IDX"))
            if ll:
                coords.append(ll)
        if last_ll:
            coords.append(_DEPOT)

        # Prefer real OSRM driving distance; fall back to haversine×road-factor.
        total_km = 0.0
        _osrm = _osrm_table_km(coords) if len(coords) >= 2 else None
        if _osrm is not None:
            for i in range(len(coords) - 1):
                total_km += _osrm[i][i + 1]
        else:
            for i in range(len(coords) - 1):
                total_km += _haversine(coords[i][0], coords[i][1],
                                       coords[i+1][0], coords[i+1][1]) * _ROAD_FACTOR

        drive_min  = (total_km / _AVG_SPEED_KMH) * 60
        total_min  = stop_min + drive_min

        depart = _dtt.today().replace(hour=_DEPART_HOUR, minute=0, second=0, microsecond=0)
        arrive = depart + _td(minutes=total_min)
        deadline = depart.replace(hour=_DEADLINE_HOUR)

        margin_min = int((deadline - arrive).total_seconds() / 60)
        margin_str = (f"+{margin_min}min early" if margin_min >= 0
                      else f"⚠️ {abs(margin_min)}min LATE")
        return_km  = (_haversine(_DEPOT[0], _DEPOT[1], last_ll[0], last_ll[1]) * _ROAD_FACTOR
                      if last_ll else 0)
        return (arrive.strftime("%H:%M"), f"{margin_str}  |  {return_km:.0f}km back", False)

    _RTN_FILL = PatternFill("solid", fgColor="D9D9D9")  # light grey for return row
    _RTN_FONT = Font(bold=True, italic=True, size=9, color="404040")
    _LATE_FILL = PatternFill("solid", fgColor="FF0000")
    _LATE_FONT = Font(bold=True, size=9, color="FFFFFF")

    # ── One sheet per lorry ───────────────────────────────────────────────────
    last_col = get_column_letter(len(HEADERS))
    for plate in sorted_lorries:
        pairs    = lorry_pairs[plate]
        cap      = cap_map.get(plate, 0)
        total_w  = round(sum(it["WEIGHT"] for _, it in pairs), 3)
        util_pct = round(total_w / cap * 100, 1) if cap > 0 else 0

        ws = wb.create_sheet(title=plate[:31].replace("/", "-"))
        ws.freeze_panes = "A3"

        sorted_pairs: list = _nn_sort(pairs)
        ret_time, ret_note, is_outstation = _est_return(sorted_pairs)

        # Title row (row 1) — includes return time summary
        ws.merge_cells(f"A1:{last_col}1")
        util_icon = "✅" if util_pct >= 75 else ("🟡" if util_pct >= 50 else "⚠️")
        if is_outstation:
            ret_summary = "🌙 OUTSTATION — overnight, no 1pm return"
        else:
            late_flag = "⚠️" if "LATE" in ret_note else "🕐"
            ret_summary = f"{late_flag} Est. return HQ: {ret_time}  ({ret_note})"
        title_txt = (f"TRIP MANIFEST — {plate}   |   {cap}T cap   "
                     f"|   {total_w}T ({util_pct}%) {util_icon}   "
                     f"|   {ret_summary}   |   {generated_str}")
        t = ws.cell(1, 1, title_txt)
        t.font = _TITLE_FONT; t.fill = _TITLE_FILL
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        _apply_headers(ws, 2)

        prev_date = None
        for seq, (do, it) in enumerate(sorted_pairs, 1):
            dr      = seq + 2
            row_idx = it.get("ROW_IDX")
            dist_val    = _raw_val(row_idx, "DISTANCE")
            remarks_val = _raw_val(row_idx, "REMARKS")

            dn_short  = do["DO NUMBER"][-5:] if len(do["DO NUMBER"]) >= 5 else do["DO NUMBER"]
            date_disp = _fmt_date(do.get("DATE", ""))

            date_changed = (date_disp != prev_date)
            prev_date    = date_disp
            fill = _DATE_FILL if date_changed else (_ALT_FILL if seq % 2 == 0 else None)

            row_data = [seq, date_disp, dn_short, do["CUSTOMER NAME"][:22],
                        _route_display(do["ROUTE"]), round(it["WEIGHT"], 3),
                        dist_val, remarks_val]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(dr, ci, val)
                c.border    = _brd
                c.alignment = Alignment(vertical="top", wrap_text=(ci == len(HEADERS)))
                if fill:
                    c.fill = fill
            if remarks_val:
                ws.row_dimensions[dr].height = min(60, max(15, len(remarks_val) // 5 * 8))

        # ↩ RETURN TO HQ row
        rtn_row = len(sorted_pairs) + 3
        is_late = not is_outstation and "LATE" in ret_note
        rtn_fill = _LATE_FILL if is_late else _RTN_FILL
        rtn_font = _LATE_FONT if is_late else _RTN_FONT
        if is_outstation:
            rtn_label = "🌙  OUTSTATION — driver stays overnight, no same-day return"
        else:
            rtn_label = f"↩  RETURN TO HQ — Eng Sheng Shah Alam"
        ws.merge_cells(f"A{rtn_row}:E{rtn_row}")
        c = ws.cell(rtn_row, 1, rtn_label)
        c.font = rtn_font; c.fill = rtn_fill; c.border = _brd
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[rtn_row].height = 16

        # Est return time in col 6-7, margin in col 7-last
        c = ws.cell(rtn_row, 6, "—")
        c.font = rtn_font; c.fill = rtn_fill; c.border = _brd
        c.alignment = Alignment(horizontal="center")

        ws.merge_cells(f"G{rtn_row}:{last_col}{rtn_row}")
        if is_outstation:
            rtn_detail = "1 day 1 trip — overnight stay"
        else:
            rtn_detail = f"Est. arrive HQ: {ret_time}   {ret_note}   (depart 07:00, 60km/h avg)"
        c = ws.cell(rtn_row, 7, rtn_detail)
        c.font = rtn_font; c.fill = rtn_fill; c.border = _brd
        c.alignment = Alignment(horizontal="left", vertical="center")

        # Footer (total / utilisation)
        fr = rtn_row + 1
        ws.merge_cells(f"A{fr}:E{fr}")
        c = ws.cell(fr, 1, f"TOTAL — {len(sorted_pairs)} stop(s)")
        c.font = _FOOT_FONT; c.fill = _FOOT_FILL; c.border = _brd
        c.alignment = Alignment(horizontal="right")

        c = ws.cell(fr, 6, total_w)
        c.font = _FOOT_FONT; c.fill = _FOOT_FILL; c.border = _brd
        c.alignment = Alignment(horizontal="center")

        ws.merge_cells(f"G{fr}:{last_col}{fr}")
        c = ws.cell(fr, 7, f"{util_icon} {util_pct}% utilisation  ({total_w}T / {cap}T)")
        c.font = _FOOT_FONT; c.fill = _FOOT_FILL; c.border = _brd
        c.alignment = Alignment(horizontal="center")

    # ── NO LORRY sheet ────────────────────────────────────────────────────────
    if no_lorry_pairs:
        ws = wb.create_sheet(title="NO LORRY")
        ws.merge_cells(f"A1:{last_col}1")
        t = ws.cell(1, 1,
            f"UNASSIGNED — {len(no_lorry_pairs)} item(s)  |  Generated: {generated_str}  — Needs manual assignment")
        t.font = Font(color="FFFFFF", bold=True, size=11)
        t.fill = _NL_FILL
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        _apply_headers(ws, 2)
        nl_sorted = _nn_sort(no_lorry_pairs)
        for seq, (do, it) in enumerate(nl_sorted, 1):
            dr = seq + 2
            row_idx     = it.get("ROW_IDX")
            remarks_val = _raw_val(row_idx, "REMARKS")
            dn_short    = do["DO NUMBER"][-5:] if len(do["DO NUMBER"]) >= 5 else do["DO NUMBER"]
            date_disp   = _fmt_date(do.get("DATE", ""))
            for ci, val in enumerate(
                [seq, date_disp, dn_short, do["CUSTOMER NAME"][:22],
                 _route_display(do["ROUTE"]), round(it["WEIGHT"], 3), "",
                 remarks_val or "⚠️ No lorry assigned"],
                1,
            ):
                c = ws.cell(dr, ci, val)
                c.border = _brd
                c.alignment = Alignment(vertical="top", wrap_text=(ci == len(HEADERS)))
        for ci, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

    generated_str = _date.today().strftime("%d-%m-%Y")
    raw_df = sess.get("raw_df")

    # ── Normalise a date string to DD-MM-YYYY for display ────────────────────
    def _fmt_date(s) -> str:
        s = str(s).strip()
        if not s or s.lower() in ("nan", "none", "nat", ""):
            return ""
        # Pandas Timestamp str e.g. "2026-05-18 00:00:00"
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y"):
            try:
                return _dt.strptime(s, fmt).strftime("%d-%m-%Y")
            except ValueError:
                pass
        try:
            ts = pd.to_datetime(s, format="mixed", dayfirst=True, errors="coerce")
            if pd.notna(ts):
                return ts.strftime("%d-%m-%Y")
        except Exception:
            pass
        return s

    # ── Sort key: date then route ─────────────────────────────────────────────
    def _sort_key(pair):
        do, _ = pair
        s = str(do.get("DATE", "") or "").strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y"):
            try:
                return (_dt.strptime(s, fmt).strftime("%Y-%m-%d"), do["ROUTE"], do["CUSTOMER NAME"])
            except ValueError:
                pass
        try:
            ts = pd.to_datetime(s, format="mixed", dayfirst=True, errors="coerce")
            if pd.notna(ts):
                return (ts.strftime("%Y-%m-%d"), do["ROUTE"], do["CUSTOMER NAME"])
        except Exception:
            pass
        return ("9999-12-31", do["ROUTE"], do["CUSTOMER NAME"])

    # ── Gather items per lorry ────────────────────────────────────────────────
    lorry_pairs: dict[str, list] = _dd(list)
    no_lorry_pairs: list         = []

    for do in sess.get("pending_dos", []):
        for it in do.get("ITEMS", []):
            lorry = it.get("LORRY") or "NO_LORRY"
            if lorry in ("NO_LORRY", None, ""):
                no_lorry_pairs.append((do, it))
            elif lorry != "SPLIT":
                lorry_pairs[lorry].append((do, it))

    engine  = sess.get("engine")
    cap_map: dict[str, float] = {}
    if engine is not None:
        for _, r in engine.eligible_lorries.iterrows():
            cap_map[r["LORRY"]] = float(r["TON"])

    sorted_lorries = sorted(
        lorry_pairs.keys(),
        key=lambda p: sum(it["WEIGHT"] for _, it in lorry_pairs[p]),
        reverse=True,
    )

    # ── Shared styles ─────────────────────────────────────────────────────────
    _thin       = Side(style="thin", color="BBBBBB")
    _brd        = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _TITLE_FILL = PatternFill("solid", fgColor="1F4E79")
    _TITLE_FONT = Font(color="FFFFFF", bold=True, size=11)
    _HDR_FILL   = PatternFill("solid", fgColor="2E75B6")
    _HDR_FONT   = Font(color="FFFFFF", bold=True, size=9)
    _ALT_FILL   = PatternFill("solid", fgColor="DEEBF7")
    _DATE_FILL  = PatternFill("solid", fgColor="E2EFDA")   # green tint for date change rows
    _FOOT_FILL  = PatternFill("solid", fgColor="FCE4D6")
    _FOOT_FONT  = Font(bold=True, size=9)
    _NL_FILL    = PatternFill("solid", fgColor="C00000")

    # DATE column added as column 2 (after #)
    HEADERS    = ["#", "DATE", "DO #", "CUSTOMER", "ROUTE / AREA", "WT (T)", "DIST", "REMARKS / NOTES"]
    COL_WIDTHS = [4,   11,     9,      22,          32,             8,        8,      42]

    def _apply_headers(ws, row):
        for ci, (hdr, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            c = ws.cell(row, ci, hdr)
            c.font = _HDR_FONT; c.fill = _HDR_FILL; c.border = _brd
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[row].height = 16

    def _route_display(route_str: str) -> str:
        if " - " in route_str:
            parts = route_str.split(" - ")
            return f"{parts[0]}: {' - '.join(parts[1:3])}"[:32]
        return route_str[:32]

    def _raw_val(row_idx, col: str) -> str:
        if raw_df is None or row_idx is None:
            return ""
        try:
            v = str(raw_df.loc[row_idx, col]).strip()
            if v in ("nan", "None", "NaN", ""):
                return ""
            # Ignore GPS-coordinate strings accidentally placed in DISTANCE column
            if col == "DISTANCE" and re.match(r"^-?\d+\.\d+\s+-?\d+\.\d+", v):
                return ""
            return v
        except Exception:
            return ""

    # ── One sheet per lorry ───────────────────────────────────────────────────
    last_col = get_column_letter(len(HEADERS))
    for plate in sorted_lorries:
        pairs    = lorry_pairs[plate]
        cap      = cap_map.get(plate, 0)
        total_w  = round(sum(it["WEIGHT"] for _, it in pairs), 3)
        util_pct = round(total_w / cap * 100, 1) if cap > 0 else 0

        ws = wb.create_sheet(title=plate[:31].replace("/", "-"))
        ws.freeze_panes = "A3"

        # Title — generated date (not DO date)
        ws.merge_cells(f"A1:{last_col}1")
        util_icon = "✅" if util_pct >= 75 else ("🟡" if util_pct >= 50 else "⚠️")
        title_txt = (f"TRIP MANIFEST — {plate}   |   {cap}T capacity   "
                     f"|   {total_w}T loaded ({util_pct}%) {util_icon}   "
                     f"|   Generated: {generated_str}")
        t = ws.cell(1, 1, title_txt)
        t.font = _TITLE_FONT; t.fill = _TITLE_FILL
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        _apply_headers(ws, 2)

        # Sort: date first so same-day deliveries are grouped, then route, then customer
        sorted_pairs = sorted(pairs, key=_sort_key)

        prev_date = None
        for seq, (do, it) in enumerate(sorted_pairs, 1):
            dr      = seq + 2
            row_idx = it.get("ROW_IDX")
            dist_val    = _raw_val(row_idx, "DISTANCE")
            remarks_val = _raw_val(row_idx, "REMARKS")

            dn_short  = do["DO NUMBER"][-5:] if len(do["DO NUMBER"]) >= 5 else do["DO NUMBER"]
            date_disp = _fmt_date(do.get("DATE", ""))

            # Shade first row of each new date group in green tint so dates are visually separated
            date_changed = (date_disp != prev_date)
            prev_date    = date_disp
            fill = _DATE_FILL if date_changed else (_ALT_FILL if seq % 2 == 0 else None)

            row_data = [
                seq,
                date_disp,
                dn_short,
                do["CUSTOMER NAME"][:22],
                _route_display(do["ROUTE"]),
                round(it["WEIGHT"], 3),
                dist_val,
                remarks_val,
            ]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(dr, ci, val)
                c.border    = _brd
                c.alignment = Alignment(vertical="top", wrap_text=(ci == len(HEADERS)))
                if fill:
                    c.fill = fill
            if remarks_val:
                ws.row_dimensions[dr].height = min(60, max(15, len(remarks_val) // 5 * 8))

        # Footer
        fr = len(sorted_pairs) + 3
        ws.merge_cells(f"A{fr}:E{fr}")
        c = ws.cell(fr, 1, f"TOTAL — {len(sorted_pairs)} stop(s)")
        c.font = _FOOT_FONT; c.fill = _FOOT_FILL; c.border = _brd
        c.alignment = Alignment(horizontal="right")

        c = ws.cell(fr, 6, total_w)
        c.font = _FOOT_FONT; c.fill = _FOOT_FILL; c.border = _brd
        c.alignment = Alignment(horizontal="center")

        ws.merge_cells(f"G{fr}:{last_col}{fr}")
        c = ws.cell(fr, 7, f"{util_icon} {util_pct}% utilisation  ({total_w}T / {cap}T)")
        c.font = _FOOT_FONT; c.fill = _FOOT_FILL; c.border = _brd
        c.alignment = Alignment(horizontal="center")

    # ── NO LORRY sheet ────────────────────────────────────────────────────────
    if no_lorry_pairs:
        ws = wb.create_sheet(title="NO LORRY")
        ws.merge_cells(f"A1:{last_col}1")
        t = ws.cell(1, 1,
            f"UNASSIGNED — {len(no_lorry_pairs)} item(s)  |  Generated: {generated_str}  — Needs manual assignment")
        t.font = Font(color="FFFFFF", bold=True, size=11)
        t.fill = _NL_FILL
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        _apply_headers(ws, 2)
        for seq, (do, it) in enumerate(sorted(no_lorry_pairs, key=_sort_key), 1):
            dr = seq + 2
            row_idx     = it.get("ROW_IDX")
            remarks_val = _raw_val(row_idx, "REMARKS")
            dn_short    = do["DO NUMBER"][-5:] if len(do["DO NUMBER"]) >= 5 else do["DO NUMBER"]
            date_disp   = _fmt_date(do.get("DATE", ""))
            for ci, val in enumerate(
                [seq, date_disp, dn_short, do["CUSTOMER NAME"][:22],
                 _route_display(do["ROUTE"]), round(it["WEIGHT"], 3), "",
                 remarks_val or "⚠️ No lorry assigned"],
                1,
            ):
                c = ws.cell(dr, ci, val)
                c.border = _brd
                c.alignment = Alignment(vertical="top", wrap_text=(ci == len(HEADERS)))
        for ci, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def get_trip_manifest_bytes(phone: str) -> bytes | None:
    """Return trip manifest bytes if available (generated at Yes confirm), then clear."""
    sess = sessions.get(phone, {})
    if sess.get("state") in ("DONE", "CONFIRMING"):
        data = sess.get("trip_manifest_bytes")
        if data:
            sess["trip_manifest_bytes"] = None
        return data
    return None


def get_export_bytes(phone: str) -> bytes | None:
    """Return export bytes if available (DONE or re-exported after post-yes block), then clear."""
    sess = sessions.get(phone, {})
    if sess.get("state") in ("DONE", "CONFIRMING"):
        data = sess.get("export_bytes")
        if data:
            sess["export_bytes"] = None  # clear after first retrieval
        return data
    return None