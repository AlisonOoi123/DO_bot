"""
Web front-end for the DO lorry-assignment engine.

This is a thin, responsive web UI over the SAME assignment engine that powers
the WhatsApp bot (bot.py). It reuses the exact same handlers and rules —
nothing about the assignment logic changes; only the interface does.

Flow (mirrors the WhatsApp flow):
    1. Login (pick your name — ABI / VIVIAN / …)
    2. (ABI / VIVIAN only) upload today's master lorry file
    3. Choose Today / Tomorrow
    4. Upload the DO Excel file
    5. See the assignment result on screen and download the filled Excel

Run locally (e.g. on your office PC), then open it from any phone or PC on the
same network:

    pip install flask
    python web_app.py
    → open http://<this-pc-ip>:8000  (e.g. http://10.0.0.229:8000)

The page is responsive: it fits a desktop portal and a phone screen.
"""
from __future__ import annotations

import os
import re
import secrets
import io
from datetime import datetime

import pandas as pd
from flask import (
    Flask, request, jsonify, send_file, make_response, Response, redirect,
)

import bot   # the existing engine — reused as-is
import do_source   # live DO fetch from the ERP DB — see do_source.py

app = Flask(__name__)

# Each browser gets a random session token stored in a cookie. We map that token
# to a bot session (bot.get_session(token)) so the engine's per-user state
# machine works exactly as it does for a WhatsApp phone number.
_COOKIE = "do_sid"

# ─────────────────────────────────────────────────────────────────────────────
# Credential-based login — validate against data/credentials.xlsx
# (columns: email, password, ip, device_name). Keep this file OUT of git; it
# holds plaintext staff passwords (see .gitignore).
# ─────────────────────────────────────────────────────────────────────────────
_CRED_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          "data", "credentials.xlsx")


def _load_credentials() -> dict:
    """Return {email_lower: {'password':..,'ip':..,'device_name':..}}.

    Read fresh each login so edits to the file take effect without a restart.
    """
    try:
        df = pd.read_excel(_CRED_PATH)
    except Exception:
        return {}
    df.columns = [str(c).strip().lower() for c in df.columns]

    def _cell(v):
        # Blank Excel cells read as NaN; normalise to "" instead of "nan".
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        s = str(v).strip()
        return "" if s.lower() in ("nan", "none") else s

    out = {}
    for _, r in df.iterrows():
        email = _cell(r.get("email", "")).lower()
        if not email:
            continue
        out[email] = {
            "password":    _cell(r.get("password", "")),
            "ip":          _cell(r.get("ip", "")),
            "device_name": _cell(r.get("device_name", "")),
        }
    return out


def _check_credentials(email: str, password: str, device_name: str = ""):
    """Return (ok, error_message). Password must match; device name, if the
    user typed one, must match the record too."""
    creds = _load_credentials()
    if not creds:
        return False, "No credentials configured. Add data/credentials.xlsx."
    rec = creds.get((email or "").strip().lower())
    # Reject unknown email, an account with no password set (incomplete row),
    # or a wrong password.
    if not rec or not rec["password"] or (password or "").strip() != rec["password"]:
        return False, "Invalid email or password."
    # Device name is optional. Only enforce it when BOTH the account has one and
    # the user typed one. Compare loosely: ignore case and all whitespace
    # (incl. stray spaces Excel sometimes leaves in a cell).
    def _norm(v):
        return re.sub(r"\s+", "", str(v or "")).upper()
    dev, want = _norm(device_name), _norm(rec["device_name"])
    if dev and want and dev != want:
        return False, "Device name does not match this account."
    return True, None


def _is_authed(sess) -> bool:
    return bool(sess.get("_authed"))


def _sid() -> str:
    """Return the caller's BASE (browser-cookie) session id, creating one if
    absent. This identifies the browser/login only, not a specific
    planner — see _active_user_sid() for the per-planner id that all real
    engine/DO operations use."""
    sid = request.cookies.get(_COOKIE)
    if not sid:
        sid = "web_" + secrets.token_hex(8)
    return sid


def _active_user_sid() -> str:
    """Per-planner session id for real engine/DO state: base::<user>. Every
    endpoint that touches the engine, fetched DOs, the board, or lorry
    toggles resolves its session through this — not _sid() directly — so
    clicking the OTHER planner's tab and coming back doesn't wipe this
    planner's in-progress work. The two planners share one browser cookie
    (one login) but get fully independent engine state.

    Falls back to the bare base id before any planner tab has been picked
    yet (matches the previous single-session behaviour for that window).
    _with_cookie always recovers the base id by splitting on '::', so the
    browser's cookie itself never becomes planner-specific."""
    base = _sid()
    outer = bot.get_session(base)
    user = outer.get("_active_user")
    return f"{base}::{user}" if user else base


def _with_cookie(payload, sid, status=200):
    resp = make_response(jsonify(payload), status)
    base = sid.split("::", 1)[0]
    resp.set_cookie(_COOKIE, base, samesite="Lax", max_age=60 * 60 * 8)
    return resp


def _file_bytes():
    """Read the uploaded file from the request (field name 'file')."""
    f = request.files.get("file")
    if not f:
        return None
    return f.read()


# ─────────────────────────────────────────────────────────────────────────────
# Result rendering — turn the engine's session state into structured JSON so the
# browser can draw a clean table instead of parsing chat text.
# ─────────────────────────────────────────────────────────────────────────────
_SENTINELS = {"NO_LORRY", "NO_ELIGIBLE_LORRY", "SPLIT", "SKIPPED",
              "OTHER_USER", "NOT_TODAY", "REMARKS_SKIP", "OUT_SOURCE",
              "WRONG_TRIP", "PAST_DATE", "NOT_PICKED", "", None}

# Not actionable by the logged-in planner at all today — not "their" route
# (OTHER_USER), outsourced (OUT_SOURCE), explicitly skipped (REMARKS_SKIP),
# the wrong half-day (WRONG_TRIP), more than 30 days stale (PAST_DATE —
# bot.py's own comment says these should be "left alone entirely", not just
# hidden from a lorry lane), or unchecked by the user on the Picking List
# (NOT_PICKED). _result_json (table view) has always kept these out of the
# actionable "unassigned" list, showing only a count via dos_other/
# skipped_other — _board_json must apply the same exclusion or a planner's
# board pool ends up showing the OTHER planner's routes (or DOs they
# deliberately excluded, or DOs long past their date) as if they were theirs
# to work on today.
#
# NOT_TODAY is deliberately NOT in this set (by explicit request) — a DO off
# today's SCHD schedule still needs to be visible on the board so the
# planner can see and manually place it if they choose; bot.py's own
# assignment engine already refuses to auto-assign a real plate to anything
# classified NOT_TODAY (see the LORRY priority chain in _handle_excel_upload),
# so AI Assign still leaves it alone regardless of board visibility here.
_NOT_MINE_TODAY = {"OTHER_USER", "OUT_SOURCE", "REMARKS_SKIP", "WRONG_TRIP", "PAST_DATE", "NOT_PICKED"}


def _result_json(sess) -> dict:
    """Group the assigned items by lorry and list what couldn't be assigned."""
    items = sess.get("items", []) or []
    engine = sess.get("engine")
    # NaN (e.g. a lorry with an unresolvable/blank capacity slipping into the
    # fleet somewhere upstream) is not valid JSON — Python's json.dumps emits
    # a bare `NaN` token that the browser's JSON.parse() rejects outright,
    # breaking the whole response. Skip any capacity that isn't a real finite
    # number rather than letting it corrupt the fleet dicts below.
    def _finite_ton(r):
        try:
            v = float(r["TON"])
        except (TypeError, ValueError):
            return None
        return v if pd.notna(v) and v not in (float("inf"), float("-inf")) else None

    caps = {}
    if engine is not None and getattr(engine, "eligible_lorries", None) is not None:
        try:
            for _, r in engine.eligible_lorries.iterrows():
                v = _finite_ton(r)
                if v is not None:
                    caps[str(r["LORRY"]).strip().upper()] = v
        except Exception:
            caps = {}
    # Full cross-owner fleet (for the manual "type a plate" box only — the
    # checkbox grid above it stays scoped to this user's own fleet via `caps`).
    full_caps = {}
    if engine is not None and getattr(engine, "all_lorries", None) is not None:
        try:
            for _, r in engine.all_lorries.iterrows():
                v = _finite_ton(r)
                if v is not None:
                    full_caps[str(r["LORRY"]).strip().upper()] = v
        except Exception:
            full_caps = dict(caps)
    else:
        full_caps = dict(caps)

    lorries: dict[str, dict] = {}
    unassigned: list[dict] = []
    skipped_other = 0

    for it in items:
        lorry = it.get("LORRY")
        do = {
            "do": str(it.get("DO NUMBER", "")),
            "route": str(it.get("ROUTE", "")),
            "customer": str(it.get("CUSTOMER NAME", "")),
            "weight": round(_w, 3) if pd.notna(_w := (float(it.get("WEIGHT", 0) or 0))) else 0.0,
            "state": str(it.get("STATE", "")),
            "date": str(it.get("DATE", "")),
        }
        if lorry in _NOT_MINE_TODAY:
            skipped_other += 1
            continue
        if lorry in _SENTINELS:
            _reasons = sess.get("unassigned_reasons")
            _reason = _reasons.get(do["do"]) if isinstance(_reasons, dict) else None
            # Fall back to the sentinel itself (e.g. "PAST_DATE") when there's
            # no explicit recorded reason — a real label beats the generic
            # "NO_LORRY" for a DO the AI deliberately skipped, not one it
            # just couldn't fit anywhere.
            if not _reason:
                _reason = lorry if lorry not in ("NO_LORRY", "", None) else "NO_LORRY"
            do["reason"] = _reason
            unassigned.append(do)
            continue
        grp = lorries.setdefault(lorry, {
            "lorry": lorry,
            # full_caps (not caps) so a cross-owner plate assigned via the
            # manual box (assign_specific_dos) still shows its true capacity/
            # utilisation instead of "—" just because it's not this user's own.
            "capacity": full_caps.get(str(lorry).strip().upper()),
            "load": 0.0,
            "dos": [],
        })
        grp["load"] = round(grp["load"] + do["weight"], 3)
        grp["dos"].append(do)

    lorry_list = sorted(lorries.values(), key=lambda g: g["lorry"])
    for g in lorry_list:
        cap = g["capacity"]
        g["util"] = round(100.0 * g["load"] / cap, 1) if cap else None

    total_dos = sum(len(g["dos"]) for g in lorry_list)
    used_plates = {g["lorry"] for g in lorry_list}
    # Full fleet for the reassign picker (plate, capacity, whether already used).
    fleet = [{"plate": p, "capacity": round(c, 3), "used": p in used_plates}
             for p, c in sorted(caps.items())]
    # Cross-owner fleet for the manual "type a plate" box (§ assign_specific_dos
    # in bot.py accepts any owner's plate — this just lets the client know a
    # typed plate is real before it submits, without a false "unknown plate").
    full_fleet = [{"plate": p, "capacity": round(c, 3)} for p, c in sorted(full_caps.items())]
    return {
        "lorries": lorry_list,
        "unassigned": unassigned,
        "unassigned_weight": round(sum(d["weight"] for d in unassigned), 3),
        "fleet": fleet,
        "full_fleet": full_fleet,
        "summary": {
            "lorries_used": len(lorry_list),
            "dos_assigned": total_dos,
            "dos_unassigned": len(unassigned),
            "dos_other": skipped_other,
        },
    }


def _finite_ton(r):
    try:
        v = float(r["TON"])
    except (TypeError, ValueError):
        return None
    return v if pd.notna(v) and v not in (float("inf"), float("-inf")) else None


def _caps_from_pairs(pairs):
    out = {}
    for p, t in pairs or []:
        try:
            v = float(t)
        except (TypeError, ValueError):
            continue
        if pd.notna(v) and v not in (float("inf"), float("-inf")):
            out[str(p).strip().upper()] = v
    return out


def _fleet_lists(sess) -> dict:
    """This planner's current holder-based fleet — own plates plus any
    SPARE/staged one actually claimed for today, and the shared staging
    pool — with today's on/off state. This is the single source of truth
    for "who currently holds this plate": once a plate's holder is the
    OTHER planner (not staging), it must not appear here at all, or a
    stale claim would leak a plate back onto the wrong screen.

    Refreshes _my_zone_fleet/_staging_fleet/eligible_lorries fresh on every
    call, not just after this session's own toggle/claim actions — the
    OTHER planner may have released or claimed a plate since this
    session's fields were last computed (they're only written on a local
    write, not kept live), and a stale read here would show a plate as
    still staged (or still available) when it's actually just been taken.
    """
    if sess.get("engine") is not None and sess.get("user_id"):
        try:
            bot.refresh_eligible_from_toggle(sess)
        except Exception:
            pass
    engine = sess.get("engine")
    caps = _caps_from_pairs(sess.get("_my_zone_fleet"))
    if not caps and sess.get("_my_zone_fleet") is None:
        # Fallback only if the holder-aware fleet was never computed for
        # this session (shouldn't normally happen post-login).
        if engine is not None and getattr(engine, "eligible_lorries", None) is not None:
            try:
                for _, r in engine.eligible_lorries.iterrows():
                    v = _finite_ton(r)
                    if v is not None:
                        caps[str(r["LORRY"]).strip().upper()] = v
            except Exception:
                caps = {}
    staging_caps = _caps_from_pairs(sess.get("_staging_fleet"))
    _off_plates = set()
    try:
        _off_plates = bot.get_unavailable_plates_for(sess.get("user_id") or "")
    except Exception:
        pass
    return {
        "lorries": [{"plate": p, "capacity": c, "on": p not in _off_plates}
                    for p, c in sorted(caps.items())],
        "staging": [{"plate": p, "capacity": c, "broken": bot.is_staging_plate_broken(p)}
                    for p, c in sorted(staging_caps.items())],
    }


def _board_json(sess) -> dict:
    """Drag-and-drop board view: every DO (assigned or not) plus every lorry
    with its current capacity, shaped for the board UI. Unassigned DOs are
    also grouped by route for the pool panel."""
    # _fleet_lists() already refreshes holder/toggle state fresh — reuse its
    # caps/staging_caps rather than recomputing (also runs the refresh once).
    _fl = _fleet_lists(sess)
    caps = {l["plate"]: l["capacity"] for l in _fl["lorries"]}
    staging_caps = {l["plate"]: l["capacity"] for l in _fl["staging"]}
    _off_plates = {l["plate"] for l in _fl["lorries"] if not l["on"]}
    items = sess.get("items", []) or []
    engine = sess.get("engine")

    # Board lanes = the logged-in planner's OWN CURRENTLY-HELD fleet, PLUS —
    # unlike the setup screen's toggle grid (_fleet_lists alone) — a plate
    # that already has a DO on it from elsewhere (e.g. a manual table-view
    # assignment onto a BIG/SELAYANG plate) still gets a lane here so that
    # assignment stays visible, even though it's not itself a drop target.
    # This enrichment is intentionally NOT part of _fleet_lists: applying it
    # to the toggle grid too let a plate the OTHER planner had since claimed
    # reappear on this planner's setup screen just because an old DO in
    # sess["items"] still referenced it from before the claim.
    _all_caps = {}
    if engine is not None and getattr(engine, "all_lorries", None) is not None:
        try:
            for _, r in engine.all_lorries.iterrows():
                v = _finite_ton(r)
                if v is not None:
                    _all_caps[str(r["LORRY"]).strip().upper()] = v
        except Exception:
            _all_caps = {}

    orders = []
    routes: dict[str, dict] = {}
    manual_only: dict[str, dict] = {}
    # Dropping points per lane: distinct (CODE, DROPPOINT) pairs among that
    # plate's assigned DOs — BPDLVCUST.ZDROPPOINT_0 per customer (BPCNUM_0),
    # by explicit request. Two DOs for the SAME customer at the SAME drop
    # point count as 1; the SAME customer at two DIFFERENT drop points counts
    # as 2 (see do_source.py's fetch — ZDROPPOINT_0 is now pulled in as the
    # "DROPPOINT" column). Blank DROPPOINT (manual uploads, or no BPDLVCUST
    # match) falls back to one shared point per customer code.
    _drop_points: dict[str, set] = {}
    _effective_mo_codes = _effective_manual_only_codes(sess)
    for it in items:
        lorry = it.get("LORRY")
        if lorry in _NOT_MINE_TODAY:
            continue
        route = str(it.get("ROUTE", ""))
        code = str(it.get("CODE", ""))
        assigned_plate = lorry if lorry and lorry not in _SENTINELS else None
        if assigned_plate:
            _drop_points.setdefault(assigned_plate, set()).add(
                (code.strip().upper(), str(it.get("DROPPOINT", "")).strip().upper()))
        weight = float(it.get("WEIGHT", 0) or 0)
        weight = round(weight, 3) if pd.notna(weight) else 0.0
        # A real reason (e.g. PAST_DATE) beats no explanation at
        # all for a pool card the AI deliberately skipped rather than one it
        # just couldn't fit anywhere (NO_LORRY etc. show no reason tag).
        _reason = lorry if lorry in _SENTINELS and lorry not in ("NO_LORRY", "", None) else ""
        _is_manual_only = code.strip().upper() in _effective_mo_codes
        orders.append({
            "do": str(it.get("DO NUMBER", "")),
            "code": code,
            "route": route,
            "customer": str(it.get("CUSTOMER NAME", "")),
            "weight": weight,
            "distance": str(it.get("DISTANCE", "")),
            "date": str(it.get("DATE", "")),
            "remarks": str(_rm) if pd.notna(_rm := it.get("REMARKS")) else "",
            "products": str(it.get("PRODUCTS", "")),
            "lorry": assigned_plate,
            "reason": _reason,
            "manualOnly": _is_manual_only,
        })
        if assigned_plate is None and _is_manual_only:
            _mo_code = code.strip().upper()
            mo = manual_only.setdefault(_mo_code, {
                "code": code,
                "customer": str(it.get("CUSTOMER NAME", "")),
                "dos": 0, "weight": 0.0,
            })
            mo["dos"] += 1
            mo["weight"] = round(mo["weight"] + weight, 3)
        elif assigned_plate is None:
            rt = routes.setdefault(route, {"route": route, "dos": 0, "weight": 0.0})
            rt["dos"] += 1
            rt["weight"] = round(rt["weight"] + weight, 3)
        elif assigned_plate not in caps and assigned_plate in _all_caps:
            caps[assigned_plate] = _all_caps[assigned_plate]

    # Aging breakdown for the unassigned pool: how many unassigned DOs are
    # more than N days past their own DATE — cumulative thresholds (a DO
    # 10 days old counts toward over3, over5, AND over7), so each number
    # answers "how many are at least this stale" on its own.
    _aging = {"over3": 0, "over5": 0, "over7": 0}
    _today_ts = pd.Timestamp(datetime.now().date())
    for _o in orders:
        if _o["lorry"]:
            continue
        _dt = pd.to_datetime(_o["date"], dayfirst=True, errors="coerce")
        if pd.isna(_dt):
            continue
        _age = (_today_ts - _dt).days
        if _age > 7: _aging["over7"] += 1
        if _age > 5: _aging["over5"] += 1
        if _age > 3: _aging["over3"] += 1

    return {
        "orders": orders,
        "aging": _aging,
        "routes": sorted(routes.values(), key=lambda r: r["route"]),
        "manual_only": sorted(manual_only.values(), key=lambda m: m["code"]),
        "lorries": [{"plate": p, "capacity": c, "on": p not in _off_plates,
                     "droppingPoints": len(_drop_points.get(p, ()))}
                    for p, c in sorted(caps.items())],
        # Shared staging pool: plates parked unclaimed (the 4 SPARE plates
        # by default, or anything either planner has voluntarily released)
        # that this planner could drag into their own lane list above.
        # Never used for assignment until claimed — see claim_or_release_plate.
        "staging": [{"plate": p, "capacity": c, "broken": bot.is_staging_plate_broken(p)}
                    for p, c in sorted(staging_caps.items())],
    }


def _snapshot_rows(sess) -> list[dict]:
    """Build the 'Snapshot' export rows — one per lorry in this planner's
    current fleet, matching the external tracking-sheet format (NO / LORRY /
    DRIVER / MUATAN(KG) / JUMLAH (KEDAI) / ISI (KG) / LAGI BOLEH MASUK (KG) /
    DONE / AREA / PERCENTAGE OF DELIVERY / REMARKS).

    MUATAN(KG) is sourced from the same holder-based fleet capacity
    (_fleet_lists) the board itself uses for utilisation %, so it always
    matches "the LORRY DAILY PLANNING we use" rather than any raw sheet
    column. DRIVER and REMARKS are left blank — no reliable source for
    either exists in this system yet.
    """
    _fl = _fleet_lists(sess)
    caps = {l["plate"]: l["capacity"] for l in _fl["lorries"]}
    items = sess.get("items", []) or []

    per_lorry: dict[str, dict] = {p: {"isi": 0.0, "codes": set(), "routes": set()} for p in caps}
    for it in items:
        lorry = it.get("LORRY")
        if lorry in _SENTINELS or lorry in _NOT_MINE_TODAY:
            continue
        g = per_lorry.setdefault(lorry, {"isi": 0.0, "codes": set(), "routes": set()})
        g["isi"] += float(it.get("WEIGHT", 0) or 0)
        code = str(it.get("CODE", "")).strip()
        if code:
            g["codes"].add(code)
        route = str(it.get("ROUTE", "")).strip()
        if route:
            g["routes"].add(route)

    rows = []
    for i, plate in enumerate(sorted(per_lorry.keys()), start=1):
        g = per_lorry[plate]
        muatan_kg = round((caps.get(plate) or 0) * 1000, 1)
        isi_kg = round(g["isi"] * 1000, 1)
        rows.append({
            "NO": i,
            "LORRY": plate,
            "DRIVER": "",
            "MUATAN(KG)": muatan_kg,
            "JUMLAH (KEDAI)": len(g["codes"]),
            "ISI (KG)": isi_kg,
            "LAGI BOLEH MASUK (KG)": round(muatan_kg - isi_kg, 1),
            "DONE": "✓" if g["codes"] else "",
            "AREA": ", ".join(sorted(g["routes"])),
            "PERCENTAGE OF DELIVERY": round(100.0 * isi_kg / muatan_kg, 1) if muatan_kg else 0.0,
            "REMARKS": "",
        })
    return rows


def _board_save_rows(sess) -> list[dict]:
    """Every DO in this planner's own scope (assigned AND unassigned),
    shaped for do_source.save_board_to_erp. Includes unassigned rows too —
    save_board_to_erp only uses them for the ZLORRY totals, never writes
    them to SDELIVERY. Items outside this planner's scope (another user's
    route, off-schedule, outsourced, wrong trip) are excluded entirely —
    their SDELIVERY rows are never touched, since they're not this
    planner's DOs to manage.

    ai_assigned mirrors _ai_plate_snapshot (see _run_dos_upload): True only
    for a DO still sitting on the exact plate the last AI Assign run put it
    on — any manual drag since then naturally falls out of this check."""
    ai_snapshot = sess.get("_ai_plate_snapshot") or {}
    rows = []
    for it in sess.get("items", []) or []:
        lorry = it.get("LORRY")
        if lorry in _NOT_MINE_TODAY:
            continue
        plate = lorry if lorry and lorry not in _SENTINELS else None
        do_number = str(it.get("DO NUMBER", "")).strip()
        rows.append({
            "do_number": do_number,
            "plate": plate,
            "weight_kg": round(float(it.get("WEIGHT", 0) or 0) * 1000, 3),
            "ai_assigned": bool(plate) and ai_snapshot.get(do_number) == plate,
        })
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# API endpoints — each drives the existing engine handler for the current step.
# ─────────────────────────────────────────────────────────────────────────────
@app.before_request
def _guard():
    """All /api/* endpoints require a logged-in session."""
    if request.path.startswith("/api/"):
        sess = bot.get_session(_sid())
        if not _is_authed(sess):
            return _with_cookie({"error": "auth_required"}, _sid(), 401)


@app.errorhandler(Exception)
def _api_error_handler(e):
    """An unhandled exception in any /api/* endpoint used to fall through to
    Flask's default HTML error page — the frontend's `.json()` call on that
    HTML then threw its own confusing "Unexpected token '<'" error, hiding
    the real problem completely (e.g. the /api/master-grid crash this was
    added for). Every /api/* endpoint now always gets back real JSON with
    the failing file:line, matching _handle_excel_upload's on-screen error
    convention, instead of a blank/broken screen."""
    if not request.path.startswith("/api/"):
        return e
    import traceback
    traceback.print_exc()
    code = getattr(e, "code", 500) if hasattr(e, "code") else 500
    _tb = traceback.extract_tb(e.__traceback__)
    _loc = f" ({_tb[-1].filename.split(chr(92))[-1].split('/')[-1]}:{_tb[-1].lineno} in {_tb[-1].name})" if _tb else ""
    return _with_cookie({
        "error": f"Server error: {type(e).__name__}: {e}{_loc}",
    }, _sid(), code if isinstance(code, int) and 400 <= code < 600 else 500)


@app.route("/auth", methods=["POST"])
def auth():
    sid = _sid()
    sess = bot.get_session(sid)
    email = request.form.get("email", "")
    password = request.form.get("password", "")
    device = request.form.get("device_name", "")
    ok, err = _check_credentials(email, password, device)
    if not ok:
        return Response(_login_page(err), mimetype="text/html", status=401)
    sess["_authed"] = True
    sess["_email"] = email.strip()
    sess["_device"] = device.strip()
    resp = make_response(redirect("/"))
    resp.set_cookie(_COOKIE, sid, samesite="Lax", max_age=60 * 60 * 8)
    return resp


@app.route("/logout", methods=["POST", "GET"])
def logout():
    sid = _sid()
    # Clears auth plus BOTH planners' in-progress engine state — each lives
    # in its own base::<user> session (see _active_user_sid), so the plain
    # base session alone isn't enough to wipe everything on logout.
    for _u in ("ABI", "VIVIAN"):
        bot.reset_session(f"{sid}::{_u}")
    bot.reset_session(sid)
    resp = make_response(redirect("/"))
    resp.set_cookie(_COOKIE, sid, samesite="Lax", max_age=60 * 60 * 8)
    return resp


@app.route("/api/state")
def api_state():
    """Current session snapshot — also used to restore the Result screen
    after a real browser refresh (or an explicit Refresh button click),
    since the assignment itself lives in the server session, not the page.
    state/user/result come from the currently ACTIVE planner's own session
    (see _active_user_sid) so a refresh restores whichever planner tab was
    last open, not a stale mix; email is account-level, from the base
    session."""
    base = _sid()
    sess = bot.get_session(_active_user_sid())
    return _with_cookie({
        "state": sess.get("state", "IDLE"),
        "user": sess.get("user_id"),
        "email": bot.get_session(base).get("_email"),
        "needs_picking": bool(sess.get("_picking_pending")),
        "result": _result_json(sess) if sess.get("items") else None,
    }, base)


@app.route("/api/users")
def api_users():
    sid = _sid()
    try:
        users = bot._get_valid_users()
    except Exception:
        users = ["ABI", "VIVIAN"]
    # The portal only has planner tabs for ABI and VIVIAN — BIG/SELAYANG have
    # no tab (so the AI never auto-assigns using their lorries; those plates
    # stay reachable only via manual plate-number assignment), and SPARE
    # isn't a planner at all — its lorries are folded into ABI's and
    # VIVIAN's own fleet automatically (see _parse_master_lorry).
    _allowed = {"ABI", "VIVIAN"}
    users = [u for u in users if str(u).strip().upper() in _allowed]
    return _with_cookie({"users": list(users)}, sid)


def _planner_already_active(sess: dict, user: str) -> bool:
    """True if this planner's OWN session already has real work in flight —
    either DOs already fetched (state can be CONFIRMING, or parked at
    AWAIT_OTHER_USER_REPLY on a pending off-schedule prompt re-triggered by
    a later AI Assign re-run — either way sess["items"] exists and is worth
    restoring), or day/trip already picked and just waiting on Fetch
    (AWAIT_EXCEL, no items yet). Checking sess["items"] directly instead of
    enumerating every in-progress state is what actually holds up: any new
    intermediate state added later still gets recognised as "has work"
    correctly, as long as it left items behind."""
    if sess.get("user_id") != user:
        return False
    return bool(sess.get("items")) or sess.get("state") == "AWAIT_EXCEL"


@app.route("/api/login", methods=["POST"])
def api_login():
    """Picking a planner tab activates their OWN persistent session
    (base::<user>) — if that planner already has in-progress or completed
    work (they'd set up and maybe fetched before switching to the other
    tab), this reports it as already_active instead of re-running the
    whole login/master-upload flow, which would otherwise silently wipe
    it. A genuinely first-time pick for that planner still runs the full
    flow as before."""
    base = _sid()
    outer = bot.get_session(base)
    user = str((request.json or {}).get("user", "")).strip().upper()
    outer["_active_user"] = user
    sid = f"{base}::{user}" if user else base
    sess = bot.get_session(sid)

    if _planner_already_active(sess, user):
        return _with_cookie({
            "messages": [],
            "state": sess.get("state"),
            "user": sess.get("user_id"),
            "needs_master": False,
            "already_active": True,
            "has_items": bool(sess.get("items")),
            "needs_picking": bool(sess.get("_picking_pending")),
            "trip_day": sess.get("trip_day"),
            "trip_session": sess.get("trip_session"),
        }, base)

    msgs = bot._handle_user_id(sid, sess, str(user))
    return _with_cookie({
        "messages": msgs,
        "state": sess.get("state"),
        "user": sess.get("user_id"),
        "already_active": False,
        # AWAIT_MASTER_UPLOAD → needs master file; AWAIT_TRIP_DAY → skip to day
        "needs_master": sess.get("state") == "AWAIT_MASTER_UPLOAD",
    }, base)


@app.route("/api/master", methods=["POST"])
def api_master():
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    fb = _file_bytes()
    if fb is None:
        return _with_cookie({"error": "No file uploaded."}, sid, 400)
    msgs = bot._handle_master_upload(sid, sess, fb)
    return _with_cookie({
        "messages": msgs,
        "state": sess.get("state"),
        # advanced to AWAIT_TRIP_DAY on success; stayed put on validation error
        "ok": sess.get("state") == "AWAIT_TRIP_DAY",
    }, sid)


# ─────────────────────────────────────────────────────────────────────────────
# Master lorry — inline-editable grid instead of a file upload.
# Loads the committed default (data/master_lorry.xlsx) as today's starting
# point; edits made in the portal apply to THIS session only (never written
# back to the default file — tomorrow always starts from the same baseline,
# matching the existing daily-reset behaviour). Columns: LORRY, TON,
# DESCRIPTION (kg), Ori_User (reference only, not read by the bot), USER,
# Status — only USER and Status are meant to be edited.
# ─────────────────────────────────────────────────────────────────────────────
_MASTER_LORRY_DEFAULT_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "data", "master_lorry.xlsx")


@app.route("/api/master-default")
def api_master_default():
    """Return today's starting master-lorry grid as JSON rows for the
    portal's editable table.

    Sourced from the logged-in session's own engine.all_lorries — the SAME
    live LORRY DAILY PLANNING.xlsx (MUATAN sheet) LorryEngine already parsed
    at login — not a separately committed sample file. Two parallel copies
    of "today's fleet" (one live, one a static snapshot) can only drift
    apart; this was confirmed live as the cause of ABI-owned plates
    (BQU3875, BQX7228, BQX9983, BQY7823, VEA2818) showing up in VIVIAN's own
    fleet list, sourced from the stale sample file's ownership column."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    engine = sess.get("engine")
    if engine is not None and getattr(engine, "all_lorries", None) is not None and not engine.all_lorries.empty:
        rows = []
        for _, r in engine.all_lorries.iterrows():
            plate = str(r.get("LORRY", "")).strip().upper()
            if not plate:
                continue
            try:
                ton = float(r.get("TON"))
            except (TypeError, ValueError):
                ton = None
            user = str(r.get("USER", "")).strip().upper()
            rows.append({
                "lorry": plate,
                "ton": ton,
                "ori_user": user,
                "user": user,
                "status": str(r.get("Status", "")).strip().upper() or "AVAILABLE",
            })
        return _with_cookie({"rows": rows}, sid)

    # Fall back to the committed sample only if no session engine exists yet
    # (shouldn't normally happen — this endpoint is always called after login).
    try:
        df = pd.read_excel(_MASTER_LORRY_DEFAULT_PATH)
    except Exception as e:
        return _with_cookie({"error": f"Could not read the default master lorry file: {e}"}, sid, 500)
    df.columns = [str(c).strip().upper() for c in df.columns]
    rows = []
    for _, r in df.iterrows():
        plate = str(r.get("LORRY", "")).strip().upper()
        if not plate or plate in ("NAN", "NONE"):
            continue
        try:
            ton = round(float(r.get("DESCRIPTION")) / 1000.0, 4)
        except Exception:
            try:
                ton = float(r.get("TON"))
            except Exception:
                ton = None
        rows.append({
            "lorry": plate,
            "ton": ton,
            "ori_user": str(r.get("ORI_USER", "")).strip().upper(),
            "user": str(r.get("USER", "")).strip().upper(),
            "status": str(r.get("STATUS", "")).strip().upper() or "AVAILABLE",
        })
    return _with_cookie({"rows": rows}, sid)


@app.route("/api/master-grid", methods=["POST"])
def api_master_grid():
    """Accept the (possibly edited) master-lorry grid from the portal and
    feed it through the exact same validation/parsing as a file upload —
    built in-memory, never written to disk, so this stays session-only."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    rows = (request.json or {}).get("rows", []) or []
    if not rows:
        return _with_cookie({"error": "No rows submitted."}, sid, 400)
    df = pd.DataFrame([{
        "LORRY": str(r.get("lorry", "")).strip().upper(),
        "TON": r.get("ton"),
        "DESCRIPTION": (float(r["ton"]) * 1000.0) if r.get("ton") is not None else None,
        "USER": str(r.get("user", "")).strip().upper(),
        "Status": str(r.get("status", "")).strip().upper(),
    } for r in rows])
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    msgs = bot._handle_master_upload(sid, sess, buf.getvalue())
    return _with_cookie({
        "messages": msgs,
        "state": sess.get("state"),
        "ok": sess.get("state") == "AWAIT_TRIP_DAY",
    }, sid)


@app.route("/api/day", methods=["POST"])
def api_day():
    """day: "today"/"tomorrow", or an ISO date string ("YYYY-MM-DD") from the
    web app's calendar picker — which day's DOs to plan for. This is purely
    which SCHD-sheet weekday to filter routes against; it never filters DOs
    by their own DATE column."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    day = (request.json or {}).get("day", "today")
    # Keep the SCHD schedule filter ON so off-schedule DOs are surfaced and the
    # user is asked whether to assign them anyway (see /api/dos + /api/offschedule).
    sess.pop("_ignore_schedule", None)
    msgs = bot._handle_trip_day(sid, sess, str(day))
    is_error = bool(msgs) and str(msgs[0]).strip().startswith("❌")
    resp = {
        "messages": msgs,
        "state": sess.get("state"),
        "ok": not is_error and sess.get("state") in ("AWAIT_EXCEL", "AWAIT_TRIP_DAY"),
    }
    if is_error:
        resp["error"] = msgs[0]
    return _with_cookie(resp, sid)


def _reset_items_to_unassigned(sess: dict) -> None:
    """Send every already-assigned item back to the pool.

    _handle_excel_upload always runs its full two-pass bin-packing optimiser
    internally — that's unchanged and untouched. The user wants Fetch to only
    fetch (everything lands in Unassigned), with the actual fitting deferred
    until they explicitly click AI Assign. Rather than risk surgery on that
    ~4000-line assignment routine, this undoes its result at the web layer:
    it's safe because _handle_excel_upload never writes to the persistent
    daily-assignment log itself (record_assignments_today only fires later,
    from the confirm/export step) — so nothing outside sess["items"] needs
    undoing. Clicking AI Assign re-runs _handle_excel_upload from scratch
    (see /api/board/ai-assign), which naturally produces a full, correct
    assignment regardless of this reset."""
    for it in sess.get("items", []) or []:
        lorry = it.get("LORRY")
        if lorry and lorry not in _SENTINELS:
            it["LORRY"] = "NO_LORRY"


def _base_manual_only_codes(sess: dict) -> set[str]:
    """Customer codes whose live ROUTE resolved to "NA" — data-driven off
    the ERP's DRN_0 via do_source.ROUTE_MAP (see the Warehouse Route &
    Remarks Update portal), not a hardcoded list. A customer with no route
    assigned can never be usefully auto-assigned, so it's always excluded
    from AI Assign by default."""
    codes = set()
    for it in sess.get("items", []) or []:
        if str(it.get("ROUTE", "")).strip().upper() == "NA":
            code = str(it.get("CODE", "")).strip().upper()
            if code:
                codes.add(code)
    return codes


def _effective_manual_only_codes(sess: dict) -> set[str]:
    """The base NA-route set, adjusted by this planner session's own
    drag-and-drop overrides: dragging a manual-only group's header out to
    the Unassigned pool releases it (False) so AI Assign may fit it again;
    dragging any single DO onto the Manual Assign Only zone flags its whole
    customer code (True) even if its route isn't NA. Overrides are
    per-session only."""
    codes = _base_manual_only_codes(sess)
    for code, forced in (sess.get("_manual_only_override") or {}).items():
        if forced:
            codes.add(code)
        else:
            codes.discard(code)
    return codes


def _reassert_na_manual_only(sess: dict, do_numbers) -> None:
    """Call right before unassigning any DO (single cancel or a lane-level
    bulk drop): if its live ROUTE is still "NA", re-flag its customer code
    as manual-only so it lands back in MANUAL ASSIGN ONLY, not plain
    Unassigned — even if this code's group was released earlier in the
    session (e.g. dragged out via its header). The route is still NA, so
    it still needs a human to place it."""
    do_numbers = set(do_numbers)
    for it in sess.get("items", []) or []:
        if str(it.get("DO NUMBER", "")).strip() not in do_numbers:
            continue
        if str(it.get("ROUTE", "")).strip().upper() != "NA":
            continue
        code = str(it.get("CODE", "")).strip().upper()
        if code:
            sess.setdefault("_manual_only_override", {})[code] = True


def _force_manual_only_unassigned(sess: dict) -> None:
    """DOs for the effective manual-only customer codes must never carry a
    real plate — even a fresh AI Assign run has to leave them for the
    planner to place by hand. Same "undo at the web layer" approach as
    _reset_items_to_unassigned, applied unconditionally (not just when
    deferring AI assignment)."""
    codes = _effective_manual_only_codes(sess)
    for it in sess.get("items", []) or []:
        if str(it.get("CODE", "")).strip().upper() not in codes:
            continue
        lorry = it.get("LORRY")
        if lorry and lorry not in _SENTINELS:
            it["LORRY"] = "NO_LORRY"


def _run_dos_upload(sid: str, sess: dict, fb: bytes, assign_now: bool = False) -> dict:
    """Shared by /api/dos (manual upload), /api/dos-fetch/use (direct from the
    live DB fetch), and /api/board/ai-assign (explicit re-assign). Fetch/
    upload land every DO in Unassigned (assign_now=False) — the AI only
    actually fits lorries when the user clicks AI Assign (assign_now=True)."""
    msgs = bot._handle_excel_upload(sid, sess, fb)
    # If some DOs aren't on the chosen day's route schedule, the engine parks
    # in AWAIT_OTHER_USER_REPLY wanting a YES/NO ("assign the off-schedule
    # ones too?"). Every fresh re-parse (fetch, Picking List's Next, AI
    # Assign) independently re-derives this from scratch, so it can come up
    # again on any of them. There's no UI surfaced for it any more (that was
    # the now-hidden Table view's job), so auto-answer NO — off-schedule DOs
    # default to excluded, same as the original fetch flow's own explicit
    # choice — and continue, rather than parking the state machine and
    # returning early. Returning early used to skip _reset_items_to_unassigned
    # below entirely, so whatever _handle_excel_upload's internal auto-assign
    # pass had already computed silently stayed on the board.
    if sess.get("state") == "AWAIT_OTHER_USER_REPLY":
        msgs = bot._handle_other_user_reply(sid, sess, "NO")
    _force_manual_only_unassigned(sess)
    if not assign_now:
        _reset_items_to_unassigned(sess)
        # Marks "fetched, waiting on the Picking List" — cleared by
        # /api/picking-list/confirm once the user actually confirms their
        # picks. Distinguishes this from "picks confirmed, board ready" so
        # switching planner tabs and back (see _planner_already_active)
        # restores to the right screen instead of skipping the picking step.
        sess["_picking_pending"] = True
        sess["_ai_plate_snapshot"] = {}
    else:
        # Snapshot DO NUMBER -> plate for everything the engine itself just
        # placed (never a manual drag) — the Excel watermark later compares
        # each row's CURRENT plate against this snapshot, so any manual move
        # after this point (which changes the DO's LORRY away from here)
        # naturally falls out of the watermark instead of needing every
        # manual-edit endpoint to separately clear a flag.
        sess["_ai_plate_snapshot"] = {
            str(it.get("DO NUMBER", "")).strip(): it.get("LORRY")
            for it in sess.get("items", []) or []
            if it.get("LORRY") and it.get("LORRY") not in _SENTINELS
        }
    result = _result_json(sess) if sess.get("items") else None
    resp = {"messages": msgs, "state": sess.get("state"), "result": result}
    # bot.py's error messages consistently lead with "❌" — surface that as a
    # real error instead of silently falling through to a generic "no DOs
    # loaded" board message once /api/board is queried afterwards.
    if not sess.get("items") and msgs and str(msgs[0]).strip().startswith("❌"):
        resp["error"] = msgs[0]
    if sess.get("_last_parse_diagnostics"):
        resp["parse_diagnostics"] = sess["_last_parse_diagnostics"]
    return resp


@app.route("/api/dos", methods=["POST"])
def api_dos():
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    fb = _file_bytes()
    if fb is None:
        return _with_cookie({"error": "No file uploaded."}, sid, 400)
    # A fresh upload is a fresh picking decision — previous Picking List
    # unchecks shouldn't silently carry over onto different data.
    sess.pop("_unpicked_dos", None)
    return _with_cookie(_run_dos_upload(sid, sess, fb), sid)


# ─────────────────────────────────────────────────────────────────────────────
# Live DO fetch — pull today's Delivery Report straight from the ERP DB
# (do_source.py) instead of requiring a manual file upload. The fetched
# bytes are held in the session only (never written to disk) so the user
# can download them to review/edit, and either re-upload that file through
# the normal dropzone, or click "Use this" to proceed directly.
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/api/dos-fetch", methods=["POST"])
def api_dos_fetch():
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    _etd_days = (request.json or {}).get("etd_days")
    try:
        _etd_days = int(_etd_days) if _etd_days not in (None, "") else None
        if _etd_days is not None and _etd_days < 0:
            raise ValueError
    except (TypeError, ValueError):
        return _with_cookie({"error": "ETD window must be a whole number of days (0 or more)."}, sid, 400)
    # 0 or blank both mean "no ETD filter — fetch everything", not "only
    # today's ETD" — a window of exactly zero days isn't a useful choice.
    if _etd_days == 0:
        _etd_days = None
    # Remembered so _handle_excel_upload can widen its past-date cutoff to
    # match — a DO within the window the user explicitly asked for shouldn't
    # get excluded as "too old" just because its own DATE trails its ETD.
    sess["etd_days"] = _etd_days
    try:
        report_df = do_source.fetch_delivery_report(etd_days=_etd_days)
    except Exception as e:
        return _with_cookie({"error": f"Could not fetch DOs from the system: {e}"}, sid, 500)
    xbytes = do_source.report_to_xlsx_bytes(report_df)
    sess["_fetched_do_bytes"] = xbytes
    resp = {
        "count": int(len(report_df)),
        "weight": round(float(pd.to_numeric(report_df["GROSS WEIGHT"], errors="coerce").fillna(0).sum()) / 1000.0, 3),
        # Always attached (not just on a 0-count fetch) — a suspiciously LOW
        # but non-zero count is just as hard to diagnose blind, and this is
        # the only way to see which SQL-side filter stage (not-LOAN, site,
        # not-yet-validated, known route, 30-day floor) is actually
        # responsible without direct DB access.
        "diagnostics": dict(do_source.LAST_FETCH_DIAGNOSTICS),
    }
    return _with_cookie(resp, sid)


@app.route("/api/dos-fetch/download")
def api_dos_fetch_download():
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    data = sess.get("_fetched_do_bytes")
    if not data:
        return _with_cookie({"error": "Nothing fetched yet — click Fetch DOs first."}, sid, 400)
    return send_file(
        io.BytesIO(data),
        as_attachment=True,
        download_name="Delivery_Report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/dos-fetch/use", methods=["POST"])
def api_dos_fetch_use():
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    data = sess.get("_fetched_do_bytes")
    if not data:
        return _with_cookie({"error": "Nothing fetched yet — click Fetch DOs first."}, sid, 400)
    # A fresh fetch is a fresh picking decision — previous Picking List
    # unchecks shouldn't silently carry over onto today's newly-pulled DOs.
    sess.pop("_unpicked_dos", None)
    return _with_cookie(_run_dos_upload(sid, sess, data), sid)


@app.route("/api/picking-list")
def api_picking_list():
    """The current planner's own actionable DOs (same scope as the board
    pool — route-owned, on-schedule, right trip), shaped for the Picking
    List checklist: every DO the engine would otherwise consider, each
    already-unpicked DO (from a previous confirm this session) reported as
    unchecked, everything else checked by default."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    items = sess.get("items", []) or []
    unpicked = sess.get("_unpicked_dos") or set()
    _effective_mo_codes = _effective_manual_only_codes(sess)
    rows = []
    for it in items:
        lorry = it.get("LORRY")
        # A DO already excluded for a DIFFERENT reason (other user's route,
        # off-schedule, outsourced, wrong trip) isn't this planner's to pick
        # either way — only list what NOT_PICKED itself could still affect.
        if lorry in _NOT_MINE_TODAY and lorry != "NOT_PICKED":
            continue
        do_num = str(it.get("DO NUMBER", ""))
        code = str(it.get("CODE", ""))
        rows.append({
            "do": do_num,
            "code": code,
            "customer": str(it.get("CUSTOMER NAME", "")),
            "route": str(it.get("ROUTE", "")),
            "weight": round(float(it.get("WEIGHT", 0) or 0), 3),
            "date": str(it.get("DATE", "")),
            "checked": do_num not in unpicked,
            "special": code.strip().upper() in _effective_mo_codes,
        })
    # Date first (earliest-due first), then DO number, then customer code,
    # then route code — a stable, predictable reading order for the
    # checklist rather than whatever order the engine happened to build
    # items in. Date is parsed for the sort key only (dayfirst, matching
    # every other date parse in this app) so e.g. "5-1-2026" sorts before
    # "12-1-2026"; an unparseable date sorts last rather than breaking the
    # whole sort. NA-route (manual-only) rows always sort to the very end,
    # in their own section, regardless of date.
    def _picking_sort_key(r):
        _dt = pd.to_datetime(r["date"], dayfirst=True, errors="coerce")
        _date_key = _dt if pd.notna(_dt) else pd.Timestamp.max
        return (r["special"], _date_key, r["do"], r["code"], r["route"])
    rows.sort(key=_picking_sort_key)
    return _with_cookie({"rows": rows}, sid)


@app.route("/api/picking-list/confirm", methods=["POST"])
def api_picking_list_confirm():
    """Save which DOs the user unchecked and re-run the fetch/parse — bot.py
    drops those rows before building sess["items"] at all (see
    _handle_excel_upload), so the exclusion survives every later re-run (AI
    Assign, Reassign) too, not just this one response. That's why it's
    stored on the session rather than just filtered client-side."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    file_bytes = sess.get("_upload_bytes")
    if not file_bytes:
        return _with_cookie({"error": "No DO file on record for this session — fetch or upload one first."}, sid, 400)
    unpicked = (request.json or {}).get("unpicked", []) or []
    sess["_unpicked_dos"] = {str(d).strip() for d in unpicked}
    resp = _run_dos_upload(sid, sess, file_bytes, assign_now=False)
    sess["_picking_pending"] = False   # confirmed — board is now the destination
    return _with_cookie(resp, sid)


@app.route("/api/offschedule", methods=["POST"])
def api_offschedule():
    """User's answer to 'assign the off-schedule DOs too?' — YES re-runs the
    assignment with the schedule filter off; NO leaves them unassigned."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    assign = bool((request.json or {}).get("assign"))
    msgs = bot._handle_other_user_reply(sid, sess, "YES" if assign else "NO")
    # Same "land in Unassigned, don't auto-fit" rule as the initial fetch —
    # this resolves what's IN SCOPE, not whether the AI has fitted it yet.
    _reset_items_to_unassigned(sess)
    sess["_picking_pending"] = True   # next stop is the Picking List, not the board
    result = _result_json(sess) if sess.get("items") else None
    return _with_cookie({
        "messages": msgs,
        "state": sess.get("state"),
        "result": result,
    }, sid)


@app.route("/api/reassign", methods=["POST"])
def api_reassign():
    """Assign the still-unassigned DOs onto the lorries the user says are now
    available — no re-upload. Returns the refreshed result."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    plates = (request.json or {}).get("lorries", []) or []
    if isinstance(plates, str):
        plates = [p for p in re.split(r"[,\s]+", plates) if p]
    outcome = bot.reassign_unassigned(sess, plates)
    result = _result_json(sess) if sess.get("items") else None
    return _with_cookie({"outcome": outcome, "result": result}, sid)


@app.route("/api/assign-specific", methods=["POST"])
def api_assign_specific():
    """Manually assign a hand-picked list of unassigned DOs onto ONE named
    lorry — the counterpart to /api/reassign (which auto-bin-packs across
    multiple ticked lorries). All-or-nothing: rejects an unrecognised plate,
    a DO that's no longer unassigned, or a selection over capacity."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    body = request.json or {}
    plate = str(body.get("plate", ""))
    dos = body.get("dos", []) or []
    outcome = bot.assign_specific_dos(sess, plate, dos)
    if not outcome.get("ok"):
        return _with_cookie(outcome, sid, 400)
    outcome["result"] = _result_json(sess) if sess.get("items") else None
    return _with_cookie(outcome, sid)


# ─────────────────────────────────────────────────────────────────────────────
# Drag-and-drop board — a visual alternative to the static result table.
# Same engine, same session state (sess["items"]) as the rest of the wizard;
# these just expose it in a shape suited to a board UI instead of a report.
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/api/trip-session", methods=["POST"])
def api_trip_session():
    """Which trip of the day to assign for — always exactly one of
    "1"/"2"/"3"/"4" (no "Any" option, by explicit request): excludes DOs
    whose REMARKS explicitly say a different trip (a REMARKS with no
    trip-timing note is unaffected)."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    session = str((request.json or {}).get("session", "1")).strip().lower()
    if session not in ("1", "2", "3", "4"):
        return _with_cookie({"ok": False, "error": "Invalid session."}, sid, 400)
    # Same 2pm/4pm cutoffs as /api/day, re-checked here too since the date
    # and trip are set independently — a stale client clock or a direct API
    # call must not bypass them.
    _target_date = bot._resolve_trip_day_date(sess.get("trip_day"))
    _now = datetime.now()
    if _target_date == _now.date():
        if _now.hour >= 16:
            return _with_cookie({"ok": False, "error": "It's after 4pm — today is closed for planning. Pick tomorrow's date first."}, sid, 400)
        if _now.hour >= 14 and session in ("1", "2"):
            return _with_cookie({"ok": False, "error": "It's after 2pm — Trip 1/2 are closed for today. Pick Trip 3/4, or plan for tomorrow."}, sid, 400)
    sess["trip_session"] = session
    return _with_cookie({"ok": True, "session": session}, sid)


@app.route("/api/lorry-toggles")
def api_lorry_toggles():
    """This planner's own fleet + the shared staging pool, with today's
    on/off and claim state — used by the setup screen so lorries can be
    turned off/claimed BEFORE fetching/assigning, not just from the board
    afterwards. Works even before any DOs are loaded."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    if not sess.get("engine"):
        return _with_cookie({"lorries": [], "staging": []}, sid)
    return _with_cookie(_fleet_lists(sess), sid)


@app.route("/api/board")
def api_board():
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    if not sess.get("items"):
        return _with_cookie({"error": "No DOs loaded yet — upload or fetch a DO file first."}, sid, 400)
    return _with_cookie(_board_json(sess), sid)


@app.route("/api/board/move", methods=["POST"])
def api_board_move():
    """Drag-and-drop one DO onto a lorry (or back to the pool if lorry is
    empty/omitted). Never rejected for a rule violation — bot.board_move()
    always applies the move and returns warnings for the card to flag."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    body = request.json or {}
    do = str(body.get("do", "")).strip()
    plate = body.get("lorry")
    if not do:
        return _with_cookie({"error": "Missing 'do'."}, sid, 400)
    if not plate:
        _reassert_na_manual_only(sess, [do])
    outcome = bot.board_move(sess, do, plate)
    if not outcome.get("ok"):
        return _with_cookie(outcome, sid, 400)
    outcome["board"] = _board_json(sess)
    return _with_cookie(outcome, sid)


@app.route("/api/board/move-route", methods=["POST"])
def api_board_move_route():
    """Drag an entire route group (all its currently-unassigned DOs) onto
    one lorry in a single drop, instead of one card at a time."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    body = request.json or {}
    route = str(body.get("route", "")).strip()
    plate = body.get("lorry")
    if not route:
        return _with_cookie({"error": "Missing 'route'."}, sid, 400)
    outcome = bot.board_move_route(sess, route, plate)
    if not outcome.get("ok"):
        return _with_cookie(outcome, sid, 400)
    outcome["board"] = _board_json(sess)
    return _with_cookie(outcome, sid)


@app.route("/api/board/mark-manual-only", methods=["POST"])
def api_board_mark_manual_only():
    """Drag a single DO card onto the 'MANUAL ASSIGN ONLY' zone: flags its
    whole customer code as manual-only for this session — every DO sharing
    that code (even ones already on a lorry) moves into the section
    together, and AI Assign won't touch any of them until the group is
    dragged back out to Unassigned."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    do_num = str((request.json or {}).get("do", "")).strip()
    if not do_num:
        return _with_cookie({"error": "Missing 'do'."}, sid, 400)
    code = None
    for it in sess.get("items", []) or []:
        if str(it.get("DO NUMBER", "")).strip() == do_num:
            code = str(it.get("CODE", "")).strip().upper()
            break
    if not code:
        return _with_cookie({"ok": False, "error": "DO not found on this board."}, sid, 404)
    overrides = sess.setdefault("_manual_only_override", {})
    overrides[code] = True
    for it in sess.get("items", []) or []:
        if str(it.get("CODE", "")).strip().upper() != code:
            continue
        lorry = it.get("LORRY")
        if lorry and lorry not in _SENTINELS:
            it["LORRY"] = "NO_LORRY"
    sess.pop("export_bytes", None)
    return _with_cookie({"ok": True, "code": code, "board": _board_json(sess)}, sid)


@app.route("/api/board/move-code", methods=["POST"])
def api_board_move_code():
    """Drag a whole 'MANUAL ASSIGN ONLY' customer group by its header:
    dropped on the Unassigned pool it releases the code (AI Assign may fit
    it normally again); dropped on a lorry lane it manually assigns every
    one of that code's DOs to that plate (a manual edit like any other —
    a later AI Assign can still overwrite it). Either way the code is no
    longer force-excluded, since the planner just placed it deliberately."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    body = request.json or {}
    code = str(body.get("code", "")).strip().upper()
    plate = body.get("lorry")
    if not code:
        return _with_cookie({"error": "Missing 'code'."}, sid, 400)
    overrides = sess.setdefault("_manual_only_override", {})
    overrides[code] = False
    moved = 0
    for it in sess.get("items", []) or []:
        if str(it.get("CODE", "")).strip().upper() != code:
            continue
        it["LORRY"] = plate if plate else "NO_LORRY"
        sess.setdefault("assigned", {})[str(it.get("DO NUMBER", ""))] = it["LORRY"]
        moved += 1
    sess.pop("export_bytes", None)
    return _with_cookie({"ok": True, "moved": moved, "board": _board_json(sess)}, sid)


@app.route("/api/board/drop-lane", methods=["POST"])
def api_board_drop_lane():
    """Drop every DO currently assigned to one lorry back into Unassigned in
    one click, instead of removing cards one at a time."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    plate = str((request.json or {}).get("lorry", "")).strip()
    if not plate:
        return _with_cookie({"error": "Missing 'lorry'."}, sid, 400)
    lane_do_numbers = [
        str(it.get("DO NUMBER", "")) for it in sess.get("items", []) or []
        if it.get("LORRY") == plate
    ]
    _reassert_na_manual_only(sess, lane_do_numbers)
    dropped = 0
    for it in sess.get("items", []) or []:
        if it.get("LORRY") != plate:
            continue
        it["LORRY"] = "NO_LORRY"
        sess.setdefault("assigned", {})[str(it.get("DO NUMBER", ""))] = "NO_LORRY"
        dropped += 1
    sess.pop("export_bytes", None)
    return _with_cookie({"ok": True, "dropped": dropped, "board": _board_json(sess)}, sid)


@app.route("/api/board/toggle-lorry", methods=["POST"])
def api_board_toggle_lorry():
    """Flip a lorry's on/off availability for today. Off means it's not
    available for AI assignment (and any DOs on it get unassigned); a SPARE
    plate turned on for one planner is forced off for the other."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    body = request.json or {}
    plate = str(body.get("plate", "")).strip()
    on = bool(body.get("on"))
    if not plate:
        return _with_cookie({"error": "Missing 'plate'."}, sid, 400)
    outcome = bot.toggle_lorry_availability(sess, plate, on)
    if not outcome.get("ok"):
        return _with_cookie(outcome, sid, 400)
    outcome["board"] = _board_json(sess)
    return _with_cookie(outcome, sid)


@app.route("/api/board/claim-plate", methods=["POST"])
def api_board_claim_plate():
    """Staging-station drag-and-drop: pull an unclaimed plate from the
    shared staging pool into this planner's own lane list ('claim'), or
    park one of this planner's own plates into staging for the other
    planner to pick up ('release')."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    body = request.json or {}
    plate = str(body.get("plate", "")).strip()
    action = body.get("action")
    if not plate or action not in ("claim", "release"):
        return _with_cookie({"error": "Missing 'plate' or invalid 'action'."}, sid, 400)
    outcome = bot.claim_or_release_plate(sess, plate, action)
    if not outcome.get("ok"):
        return _with_cookie(outcome, sid, 400)
    outcome["board"] = _board_json(sess)
    return _with_cookie(outcome, sid)


@app.route("/api/board/staging-broken", methods=["POST"])
def api_board_staging_broken():
    """Mark a staging-pool plate broken/fixed — visible to both planners at
    once, since a plate parked unclaimed in staging isn't "mine" to switch
    off just for myself."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    body = request.json or {}
    plate = str(body.get("plate", "")).strip()
    broken = bool(body.get("broken"))
    if not plate:
        return _with_cookie({"error": "Missing 'plate'."}, sid, 400)
    outcome = bot.toggle_staging_broken(sess, plate, broken)
    if not outcome.get("ok"):
        return _with_cookie(outcome, sid, 400)
    outcome["board"] = _board_json(sess)
    return _with_cookie(outcome, sid)


@app.route("/api/board/reset-holders", methods=["POST"])
def api_board_reset_holders():
    """Escape hatch for the staging station: put every plate back to its
    master-file default owner, undoing all claim/release moves made today
    (e.g. leftover from testing the drag-and-drop, or a plate parked
    somewhere by mistake) — without touching on/off (broken) toggles."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    outcome = bot.reset_plate_holders(sess)
    outcome["board"] = _board_json(sess)
    return _with_cookie(outcome, sid)


@app.route("/api/board/ai-assign", methods=["POST"])
def api_board_ai_assign():
    """Re-run the real auto-assignment engine from scratch over the originally
    uploaded/fetched DO file — same code path as the initial upload, just
    re-triggered on demand. Overwrites any manual board edits made since."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    file_bytes = sess.get("_upload_bytes")
    if not file_bytes:
        return _with_cookie({"error": "No DO file on record for this session — upload or fetch one first."}, sid, 400)
    outcome = _run_dos_upload(sid, sess, file_bytes, assign_now=True)
    outcome["board"] = _board_json(sess) if sess.get("items") else None
    return _with_cookie(outcome, sid)


@app.route("/api/board/refetch", methods=["POST"])
def api_board_refetch():
    """Pull fresh DOs from the live DB and merge in ONLY the ones not already
    on this board — unlike AI Assign, this must never disturb anything the
    planner has already assigned or left in the unassigned pool. DOs update
    continuously in the source system, so a planner mid-way through a board
    needs a way to pick up newly-appeared DOs without losing their progress.

    Approach: snapshot every item's current real-plate assignment by DO
    NUMBER, append only the genuinely new rows to the original raw sheet,
    re-parse the combined file the normal deferred way (assign_now=False,
    same as a fresh fetch — new rows land in Unassigned, nothing auto-fits),
    then restore the snapshot onto the DO numbers that already existed. Old
    rows re-classify identically to before (same day, same data) since
    nothing about them changed, so restoring only the ones that were
    actually assigned a real plate is sufficient to make the refetch a
    no-op for everything except the new DOs."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    old_raw = sess.get("raw_df")
    if old_raw is None or not sess.get("items"):
        return _with_cookie({"error": "No board to refetch into — fetch or upload DOs first."}, sid, 400)
    try:
        fresh_df = do_source.fetch_delivery_report(etd_days=sess.get("etd_days"))
    except Exception as e:
        return _with_cookie({"error": f"Could not fetch DOs from the system: {e}"}, sid, 500)
    _diagnostics = dict(do_source.LAST_FETCH_DIAGNOSTICS)

    _known = set(old_raw["DO NUMBER"].astype(str).str.strip())
    fresh_df = fresh_df.copy()
    fresh_df["DO NUMBER"] = fresh_df["DO NUMBER"].astype(str).str.strip()
    new_rows = fresh_df[~fresh_df["DO NUMBER"].isin(_known)]
    if new_rows.empty:
        # No new DOs, so no re-parse happens here — but the classification
        # breakdown from whenever this board was first built (Fetch DOs ->
        # Use, or the last refetch that DID bring in new rows) is still
        # sitting in the session and still explains the board's current
        # total, so surface it rather than only the SQL-side numbers.
        return _with_cookie({
            "new_count": 0,
            "board": _board_json(sess),
            "diagnostics": _diagnostics,
            "parse_diagnostics": sess.get("_last_parse_diagnostics"),
        }, sid)

    # Snapshot: DO NUMBER -> its current real lorry plate, for every item
    # that's actually assigned right now (drag-and-drop, a prior AI Assign,
    # or a manual pick) — this is exactly what must survive the refetch.
    _prev_assigned = {
        str(it.get("DO NUMBER", "")).strip(): it["LORRY"]
        for it in sess["items"]
        if it.get("LORRY") and it["LORRY"] not in _SENTINELS
    }
    # Preserved separately from _run_dos_upload's reset below (assign_now=
    # False wipes _ai_plate_snapshot since a normal fetch lands everything
    # unassigned) — restored alongside each DO's plate so a refetch doesn't
    # silently strip the AI watermark off assignments that didn't change.
    _prev_ai_snapshot = dict(sess.get("_ai_plate_snapshot") or {})

    combined = pd.concat(
        [old_raw, new_rows.reindex(columns=old_raw.columns, fill_value="")],
        ignore_index=True,
    )
    combined_bytes = do_source.report_to_xlsx_bytes(combined)

    outcome = _run_dos_upload(sid, sess, combined_bytes, assign_now=False)
    for it in sess.get("items", []) or []:
        do_num = str(it.get("DO NUMBER", "")).strip()
        if do_num in _prev_assigned and it.get("LORRY") == "NO_LORRY":
            it["LORRY"] = _prev_assigned[do_num]
            if do_num in _prev_ai_snapshot:
                sess.setdefault("_ai_plate_snapshot", {})[do_num] = _prev_ai_snapshot[do_num]

    outcome["new_count"] = int(len(new_rows))
    outcome["board"] = _board_json(sess) if sess.get("items") else None
    outcome["diagnostics"] = _diagnostics
    return _with_cookie(outcome, sid)


@app.route("/api/download")
def api_download():
    """Same workbook as the Print button: a LORRY DAILY STATUS summary sheet
    plus one sheet per assigned plate — download it and print straight from
    Excel, no separate Print button needed."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    xbytes = _print_excel_bytes(sess)
    if xbytes is None:
        return _with_cookie({"error": "Nothing to download yet."}, sid, 400)
    return send_file(
        io.BytesIO(xbytes),
        as_attachment=True,
        download_name="DO_Assigned.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/board/download-unassigned")
def api_download_unassigned():
    """TEMPORARY debug tool — every DO NOT currently on a real lorry plate,
    whatever the reason (still unassigned, or excluded as NOT_TODAY/
    OTHER_USER/PAST_DATE/WRONG_TRIP/REMARKS_SKIP/OUT_SOURCE), one row each,
    with the reason plainly labelled — for cross-checking the fetch/
    classification pipeline against an external reference report. Remove
    once that's no longer needed."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    items = sess.get("items", []) or []
    if not items:
        return _with_cookie({"error": "Nothing fetched yet."}, sid, 400)

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "UNASSIGNED"
    headers = ["DO NUMBER", "CODE", "ROUTE", "CUSTOMER NAME", "DATE", "WEIGHT", "REASON", "REMARKS"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    rows = 0
    for it in items:
        lorry = it.get("LORRY")
        if lorry and lorry not in _SENTINELS:
            continue   # has a real plate — not what this tool is for
        reason = lorry if lorry else "NO_LORRY"
        ws.append([
            it.get("DO NUMBER", ""), it.get("CODE", ""), it.get("ROUTE", ""),
            it.get("CUSTOMER NAME", ""), it.get("DATE", ""),
            round(float(it.get("WEIGHT", 0) or 0), 3), reason, it.get("REMARKS", ""),
        ])
        rows += 1
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, min(40, len(h) + 2))

    buf = io.BytesIO()
    wb.save(buf)
    return send_file(
        io.BytesIO(buf.getvalue()),
        as_attachment=True,
        download_name="Unassigned_DOs.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _print_excel_bytes(sess) -> bytes | None:
    """Excel driver hand-off workbook: sheet 1 is the lorry-status summary
    (same shape/formulas as the Snapshot export sheet — DRIVER left blank,
    assigned lorries only), then one sheet per lorry that actually has DOs
    assigned, each with the SAME full column set as the main Download
    export (NO/DATE/ADD/TRIP/DO NUMBER/.../SALES REP) filtered to that
    plate's own rows — the export bytes are the single source of truth for
    both, so the two never drift apart. Idle/unused lorries are skipped
    entirely. Returns None if there's nothing to export yet."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font

    data = sess.get("export_bytes")
    if not data:
        try:
            bot._export_result(sess)
            data = sess.get("export_bytes")
        except Exception:
            data = None
    if not data:
        return None

    export_df = pd.read_excel(io.BytesIO(data), sheet_name=0)
    export_df.columns = [str(c).strip() for c in export_df.columns]
    if "DRIVER" in export_df.columns:
        export_df["DRIVER"] = ""   # always blank on the print sheet
    columns = list(export_df.columns)

    by_plate: dict[str, "pd.DataFrame"] = {}
    if "LICENSE" in export_df.columns:
        _blank = {"", "nan", "none", "no_lorry", "split", "skipped"}
        plate_col = export_df["LICENSE"].astype(str).str.strip()
        mask = ~plate_col.str.lower().isin(_blank)
        for plate, g in export_df[mask].groupby(plate_col[mask]):
            by_plate[str(plate)] = g

    snap_rows = [r for r in _snapshot_rows(sess) if r["DONE"] == "✓"]

    wb = Workbook()
    ws = wb.active
    ws.title = "LORRY DAILY STATUS"
    if snap_rows:
        headers = list(snap_rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in snap_rows:
            ws.append([r[h] for h in headers])
        for col_idx, h in enumerate(headers, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, len(h) + 2)

    _used_names = {"LORRY DAILY STATUS"}
    for plate in sorted(by_plate.keys()):
        name = re.sub(r'[\\/*?:\[\]]', '_', plate)[:31] or "PLATE"
        base_name, i = name, 1
        while name in _used_names:
            i += 1
            name = f"{base_name[:28]}_{i}"
        _used_names.add(name)
        pws = wb.create_sheet(name)
        pws.append(columns)
        for cell in pws[1]:
            cell.font = Font(bold=True)
        for _, row in by_plate[plate].iterrows():
            pws.append([None if pd.isna(row[c]) else row[c] for c in columns])
        for col_idx, h in enumerate(columns, start=1):
            pws.column_dimensions[pws.cell(row=1, column=col_idx).column_letter].width = max(10, min(30, len(str(h)) + 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.route("/api/board/save-preview", methods=["POST"])
def api_board_save_preview():
    """Read-only — computes exactly what /api/board/save would write,
    without touching SQL, so the Save button can show a confirm dialog
    first (writing straight to production ERP tables, by explicit
    request)."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    if not sess.get("items"):
        return _with_cookie({"error": "No board to save yet."}, sid, 400)
    rows = _board_save_rows(sess)
    if not rows:
        return _with_cookie({"error": "Nothing in scope to save."}, sid, 400)
    add_dt = bot._resolve_trip_day_date(sess.get("trip_day"))
    trip = str(sess.get("trip_session") or "1")
    plate_totals: dict[str, float] = {}
    for r in rows:
        if r["plate"]:
            plate_totals[r["plate"]] = plate_totals.get(r["plate"], 0.0) + r["weight_kg"]
    return _with_cookie({
        "add_date": add_dt.strftime("%d/%m/%Y"),
        "trip": trip,
        "dos_total": len(rows),
        "dos_assigned": sum(1 for r in rows if r["plate"]),
        "dos_unassigned": sum(1 for r in rows if not r["plate"]),
        "lorries": [{"plate": p, "ton_kg": round(t, 1)} for p, t in sorted(plate_totals.items())],
    }, sid)


@app.route("/api/board/save", methods=["POST"])
def api_board_save():
    """Writes the board to SQL — see do_source.save_board_to_erp for exactly
    what gets written and why. Always re-derives from the current session
    state (never trusts anything the client sent) and runs as one
    transaction on the ERP side."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    if not sess.get("items"):
        return _with_cookie({"error": "No board to save yet."}, sid, 400)
    rows = _board_save_rows(sess)
    if not rows:
        return _with_cookie({"error": "Nothing in scope to save."}, sid, 400)
    add_dt = bot._resolve_trip_day_date(sess.get("trip_day"))
    trip = str(sess.get("trip_session") or "1")
    try:
        result = do_source.save_board_to_erp(rows, add_dt, trip)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return _with_cookie({"ok": False, "error": f"Save failed: {e}"}, sid, 500)
    return _with_cookie({"ok": True, **result}, sid)


@app.route("/api/reset", methods=["POST"])
def api_reset():
    sid = _active_user_sid()
    bot.reset_session(sid)
    return _with_cookie({"ok": True}, sid)


@app.route("/api/reset-engine", methods=["POST"])
def api_reset_engine():
    """Clear this planner's assignment state machine (for start-over)
    WITHOUT logging the browser out — auth lives on the separate base
    session (see _active_user_sid), so there's nothing to preserve here."""
    sid = _active_user_sid()
    bot.reset_session(sid)
    return _with_cookie({"ok": True}, sid)


# Which engine state each wizard step corresponds to (for the Back button).
_STEP_STATE = {
    "master": "AWAIT_MASTER_UPLOAD",
    "day":    "AWAIT_TRIP_DAY",
    "dos":    "AWAIT_EXCEL",
}


@app.route("/api/back", methods=["POST"])
def api_back():
    """Step the wizard backward. Rewinding to an earlier step just moves the
    engine's state pointer back — the loaded master/engine is kept, so there's
    no slow Excel re-parse. Going all the way back to the user picker clears
    this planner's own session (auth lives separately — see
    _active_user_sid — so the browser stays logged in)."""
    sid = _active_user_sid()
    sess = bot.get_session(sid)
    target = (request.json or {}).get("target")
    if target == "login":
        bot.reset_session(sid)
        sess = bot.get_session(sid)
    elif target in _STEP_STATE and sess.get("user_id"):
        sess["state"] = _STEP_STATE[target]
    return _with_cookie({"ok": True, "state": sess.get("state")}, sid)


@app.route("/")
def index():
    sess = bot.get_session(_sid())
    if not _is_authed(sess):
        return _login_response()
    return _with_cookie_html(_PAGE)


@app.route("/login")
def login_page():
    """Explicit login URL (http://<host>:8000/login). If already logged in,
    go straight to the app."""
    sess = bot.get_session(_sid())
    if _is_authed(sess):
        return make_response(redirect("/"))
    return _login_response()


def _no_cache(resp):
    # Never let the browser cache the app HTML/JS, so a git pull + restart is
    # picked up on the next load without needing a hard refresh.
    resp.headers["Cache-Control"] = "no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    return resp


def _login_response():
    resp = make_response(Response(_login_page(), mimetype="text/html"))
    resp.set_cookie(_COOKIE, _sid(), samesite="Lax", max_age=60 * 60 * 8)
    return _no_cache(resp)


def _with_cookie_html(html):
    resp = make_response(Response(html, mimetype="text/html"))
    resp.set_cookie(_COOKIE, _sid(), samesite="Lax", max_age=60 * 60 * 8)
    return _no_cache(resp)


def _login_page(error: str = "") -> str:
    err_html = (f'<div class="alert alert-danger">{error}</div>'
                if error else "")
    return _LOGIN_PAGE.replace("<!--ERROR-->", err_html)


# Login page — styled to match the sample "Aging Reports Portal" login.
_LOGIN_PAGE = r"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1, maximum-scale=1">
    <title>Login - Lorry Assignment Portal</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        body { background-color: #f5f7fa; }
        .login-card { max-width: 400px; margin: 80px auto; border-radius: 12px; }
        .device-hint { font-size: 12px; color: #6c757d; }
        @media (max-width: 480px) { .login-card { margin: 32px auto; } }
    </style>
</head>
<body>
<div class="container">
    <div class="card shadow-sm p-4 login-card">
        <h3 class="mb-3 text-center">🚚 Lorry Assignment Portal</h3>
        <!--ERROR-->
        <form method="post" action="/auth">
            <div class="mb-3">
                <label class="form-label">Company Email</label>
                <input type="email" id="companyEmail" name="email" class="form-control" required placeholder="you@engsheng.com">
            </div>
            <div class="mb-3">
                <label class="form-label">Password</label>
                <input type="password" name="password" class="form-control" required>
            </div>
            <div class="mb-3">
                <label class="form-label">Device Name <small class="text-muted">(optional)</small></label>
                <input type="text" id="deviceName" name="device_name" class="form-control" placeholder="e.g. HQ-INV-XXXX">
                <div class="device-hint mt-1">
                    💡 If set on your account, it must match to log in.
                </div>
            </div>
            <button class="btn btn-primary w-100">Login</button>
        </form>
    </div>
</div>
<script>
    (function () {
        var EMAIL_KEY = 'esreports_email';
        var emailInput = document.getElementById('companyEmail');
        var deviceInput = document.getElementById('deviceName');
        var form = deviceInput.closest('form');
        // Remember the email per-browser, and the device name PER EMAIL, so a
        // previous user's device name never carries over and blocks a new user.
        function devKey(email) { return 'esreports_device::' + (email || '').trim().toLowerCase(); }
        function loadDevice() {
            try {
                var d = localStorage.getItem(devKey(emailInput.value));
                deviceInput.value = d || '';
            } catch (e) {}
        }
        try {
            var e = localStorage.getItem(EMAIL_KEY);
            if (e) emailInput.value = e;
        } catch (e) {}
        loadDevice();
        emailInput.addEventListener('input', loadDevice);   // switch device when email changes
        form.addEventListener('submit', function () {
            try {
                var em = emailInput.value.trim();
                if (em) localStorage.setItem(EMAIL_KEY, em);
                if (deviceInput.value.trim()) localStorage.setItem(devKey(em), deviceInput.value.trim());
                else localStorage.removeItem(devKey(em));
            } catch (e) {}
        });
    })();
</script>
</body>
</html>"""


# ─────────────────────────────────────────────────────────────────────────────
# The single-page responsive UI (inline so this stays one self-contained file).
# ─────────────────────────────────────────────────────────────────────────────
_PAGE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<title>Lorry Assignment</title>
<style>
  :root{
    --bg:#0f172a; --card:#1e293b; --card2:#273449; --line:#334155;
    --ink:#e2e8f0; --muted:#94a3b8; --brand:#38bdf8; --brand2:#0ea5e9;
    --ok:#22c55e; --warn:#f59e0b; --bad:#ef4444; --stage:#c084fc; --pink:#F6D4DC; --radius:14px;
  }
  @media (prefers-color-scheme: light){
    :root{ --bg:#f1f5f9; --card:#ffffff; --card2:#f8fafc; --line:#e2e8f0;
      --ink:#0f172a; --muted:#64748b; --brand:#0284c7; --brand2:#0369a1; --stage:#9333ea; --pink:#be185d; }
  }
  *{box-sizing:border-box}
  html,body{margin:0;padding:0}
  body{background:var(--bg);color:var(--ink);
    font-family:system-ui,-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;
    line-height:1.45;-webkit-text-size-adjust:100%}
  .wrap{max-width:960px;margin:0 auto;padding:16px}
  header{display:flex;align-items:center;gap:10px;padding:8px 0 16px}
  header .logo{font-size:26px}
  header h1{font-size:19px;margin:0;font-weight:700}
  header .who{margin-left:auto;font-size:13px;color:var(--muted)}
  .logout-btn{display:inline-flex;align-items:center;gap:6px;margin-left:14px;
    text-decoration:none;font-size:14px;font-weight:600;color:#fff;background:var(--bad);
    padding:8px 16px;border-radius:10px;white-space:nowrap;box-shadow:0 1px 2px rgba(0,0,0,.15)}
  .logout-btn:hover{filter:brightness(1.08)}
  .logout-btn:active{transform:translateY(1px)}
  .card{background:var(--card);border:1px solid var(--line);border-radius:var(--radius);
    padding:18px;margin-bottom:16px;box-shadow:0 1px 2px rgba(0,0,0,.08)}
  .step-title{font-size:13px;text-transform:uppercase;letter-spacing:.05em;
    color:var(--muted);margin:0 0 12px;font-weight:700}
  .hidden{display:none!important}
  .btn{appearance:none;border:0;border-radius:10px;padding:12px 16px;font-size:15px;
    font-weight:600;cursor:pointer;background:var(--brand);color:#04222f;width:100%}
  .btn:active{transform:translateY(1px)}
  .btn.secondary{background:var(--card2);color:var(--ink);border:1px solid var(--line)}
  .btn.back{background:transparent;color:var(--muted);border:1px solid var(--line);
    width:auto;margin-top:14px;padding:9px 16px;font-size:14px}
  .btn.back:hover{color:var(--ink);border-color:var(--brand)}
  .btn:disabled{opacity:.5;cursor:not-allowed}
  .row{display:flex;gap:10px;flex-wrap:wrap}
  .row>*{flex:1 1 140px}
  .grid-users{display:flex;flex-wrap:wrap;justify-content:center;gap:10px}
  .userbtn{background:var(--card2);border:1px solid var(--line);border-radius:12px;
    padding:16px 28px;min-width:120px;font-size:16px;font-weight:700;cursor:pointer;
    color:var(--ink);text-align:center;flex:0 0 auto}
  .userbtn:active{border-color:var(--brand)}
  .tp-badge{display:inline-block;font-size:11px;letter-spacing:.25em;font-weight:700;
    color:#241a00;background:var(--warn);padding:3px 10px;border-radius:3px}
  .tp-h1{margin:10px 0 4px;font-size:24px}
  .tp-sub{color:var(--muted);font-size:13.5px;margin:0 0 16px}
  .tp-tabs{display:flex;flex-wrap:wrap;gap:8px}
  .tp-tab{padding:8px 20px;border-radius:8px;border:1px solid var(--line);
    background:var(--card2);font-weight:800;cursor:pointer;font-size:14px;
    color:var(--muted)}
  .tp-tab.active{background:var(--warn);color:#241a00;border-color:var(--warn)}
  .tp-tab:disabled{opacity:.6;cursor:wait}
  .tp-etd-row{display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin-top:14px;
    padding-top:14px;border-top:1px solid var(--line)}
  .tp-etd-row label{font-size:13.5px;color:var(--muted);display:flex;align-items:center;gap:6px}
  .tp-etd-row input[type=number]{width:64px;padding:7px 8px;border-radius:8px;
    border:1px solid var(--line);background:var(--card2);color:var(--ink);font-size:14px}
  .tp-section-label{font-size:13.5px;color:var(--muted);margin:0 0 8px}
  .tp-toggle-section{margin-top:14px;padding-top:14px;border-top:1px solid var(--line)}
  .tp-toggle-grid{display:flex;flex-wrap:wrap;gap:8px;min-height:40px;border-radius:10px;
    transition:box-shadow .12s}
  .tp-toggle-grid.tpzone-active{box-shadow:0 0 0 2px rgba(56,189,248,.45)}
  .tp-staging-station{background:color-mix(in srgb, var(--stage) 14%, var(--card2));
    border:1px dashed var(--stage);border-radius:12px;
    padding:12px;margin-bottom:14px;transition:box-shadow .12s,border-color .12s}
  .tp-staging-station.tpzone-active{border-color:var(--brand);box-shadow:0 0 0 2px rgba(56,189,248,.35)}
  .tp-staging-label{font-size:12px;font-weight:700;letter-spacing:.04em;color:var(--stage);
    margin-bottom:8px;display:flex;align-items:baseline;gap:8px;flex-wrap:wrap}
  .tp-staging-hint{font-size:11px;font-weight:400;color:var(--muted);opacity:.75}
  .tp-staging-grid{display:flex;flex-wrap:wrap;gap:8px;min-height:36px}
  .tp-staging-grid:empty::after{content:'Nothing parked here right now';color:var(--muted);
    font-size:12.5px;font-style:italic;opacity:.6}
  .tp-toggle-chip{cursor:grab;touch-action:none}
  .tp-toggle-chip.tp-dragging{opacity:.35}
  .tp-toggle-chip.tp-broken{border-color:var(--bad)}
  .tp-toggle-chip .tp-broken-tag{color:var(--bad);font-size:10px;font-weight:700;margin-left:2px}
  .tp-chip-ghost{position:fixed;transform:translate(-50%,-120%);display:flex;gap:6px;
    align-items:center;background:var(--card2);border:1px solid var(--brand);border-radius:8px;
    padding:7px 12px;pointer-events:none;box-shadow:0 8px 24px rgba(0,0,0,.4);
    font-size:12.5px;font-family:ui-monospace,Menlo,monospace;z-index:999}
  .tp-toggle-chip{display:flex;align-items:center;gap:7px;padding:7px 10px;border-radius:8px;
    border:1px solid var(--line);background:var(--card2);font-size:12.5px;
    font-family:ui-monospace,Menlo,monospace}
  .tp-toggle-chip[data-tphome="STAGING"]{border-color:var(--stage);
    background:color-mix(in srgb, var(--stage) 16%, var(--card2))}
  .tp-toggle-chip[data-tphome="STAGING"].tp-broken{border-color:var(--bad)}
  .tp-trip-btn{opacity:.55}
  .tp-trip-btn.active{opacity:1}
  .tp-fetch-row{margin-top:14px;padding-top:14px;border-top:1px solid var(--line)}
  .drop{display:flex;flex-direction:column;align-items:center;justify-content:center;
    box-sizing:border-box;width:100%;min-height:150px;gap:6px;
    border:2px dashed var(--line);border-radius:14px;padding:28px 16px;text-align:center;
    color:var(--muted);cursor:pointer;background:var(--card2);transition:border-color .15s,background .15s}
  .drop:hover{border-color:var(--brand)}
  .drop.hot{border-color:var(--brand);color:var(--ink);background:transparent}
  .drop input{display:none}
  .msg{font-size:14px;white-space:pre-wrap;background:var(--card2);border:1px solid var(--line);
    border-radius:10px;padding:12px;margin-top:12px;color:var(--muted)}
  .msg.err{color:var(--bad);border-color:var(--bad)}
  .stat{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px}
  .pill{background:var(--card2);border:1px solid var(--line);border-radius:999px;
    padding:6px 12px;font-size:13px;font-weight:600}
  .pill b{color:var(--brand)}
  .pill.bad b{color:var(--bad)}
  .lorry{border:1px solid var(--line);border-radius:12px;margin-bottom:12px;overflow:hidden}
  .lorry h3{margin:0;padding:12px 14px;background:var(--card2);font-size:15px;
    display:flex;align-items:center;gap:10px;flex-wrap:wrap}
  .lorry h3 .cap{margin-left:auto;font-size:13px;color:var(--muted);font-weight:600}
  .bar{height:6px;background:var(--line);border-radius:6px;overflow:hidden;margin:0 14px 10px}
  .bar>i{display:block;height:100%;background:var(--ok)}
  .bar>i.hi{background:var(--warn)}
  table{width:100%;border-collapse:collapse;font-size:13px}
  th,td{text-align:left;padding:8px 14px;border-top:1px solid var(--line);vertical-align:top}
  th{color:var(--muted);font-weight:600;font-size:11px;text-transform:uppercase;letter-spacing:.04em}
  td.w{text-align:right;white-space:nowrap;font-variant-numeric:tabular-nums}
  .scroll{overflow-x:auto}
  .picking-section-divider td{font-size:11px;letter-spacing:.06em;color:#f59e0b;
    font-weight:700;text-transform:uppercase;padding:14px 8px 8px;
    border-top:1px dashed var(--line)}
  .lorry.collapsed .scroll{display:none}
  .lorry h3.collapsible{cursor:pointer}
  .unassigned h3{color:var(--bad)}
  .foot{color:var(--muted);font-size:12px;text-align:center;padding:8px 0 24px}
  .spin{display:inline-block;width:16px;height:16px;border:2px solid var(--line);
    border-top-color:var(--brand);border-radius:50%;animation:sp .7s linear infinite;vertical-align:-3px}
  @keyframes sp{to{transform:rotate(360deg)}}
  a.dl{display:inline-block;text-decoration:none}

  /* ---- Drag-and-drop board ---- */
  body.board-active .wrap{max-width:1400px}
  .board-top{display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-bottom:14px}
  .board-top .btn{width:auto}
  .board-top .btn-ai{background:linear-gradient(135deg,var(--brand),#a855f7);color:#fff}
  .board-top .btn-save{background:#059669;color:#fff}
  .board-top .btn-save:hover{background:#047857}
  .board-top-actions{display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-left:auto}
  .board-download-dropdown{position:relative}
  .board-download-menu{position:absolute;top:calc(100% + 6px);right:0;z-index:20;
    background:var(--card);border:1px solid var(--line);border-radius:10px;
    box-shadow:0 8px 24px rgba(0,0,0,.35);padding:6px;min-width:230px;display:flex;flex-direction:column;gap:2px}
  .board-download-menu a{display:block;padding:8px 10px;border-radius:7px;font-size:13px;
    font-weight:600;color:var(--ink);text-decoration:none;white-space:nowrap}
  .board-download-menu a:hover{background:var(--card2)}
  .board-plan-row{display:flex;align-items:center;gap:10px;flex-wrap:wrap}
  .board-plan-field{display:flex;align-items:center;gap:7px;font-size:12.5px;color:var(--muted);
    font-weight:600;background:var(--card2);border:1px solid var(--line);border-radius:20px;
    padding:4px 12px}
  .board-plan-field input[type=date]{padding:4px 6px;border-radius:6px;border:1px solid var(--line);
    background:var(--card);color:var(--ink);font-size:12.5px;font-family:inherit}
  .board-plan-trip-btns{display:flex;gap:5px}
  .board-plan-trip-btns .tp-trip-btn{padding:5px 10px;font-size:12px}
  .board-plan-warn{font-size:12px;font-weight:700;color:var(--warn);white-space:nowrap}
  .board-assigned-stat{font-size:12.5px;color:var(--muted);background:var(--card2);
    border:1px solid var(--line);border-radius:20px;padding:4px 12px;font-weight:600;
    font-family:ui-monospace,Menlo,monospace}
  .board-assigned-stat b{color:var(--ok)}
  .board-grid{display:grid;grid-template-columns:minmax(280px,380px) 1px 1fr;gap:16px}
  .board-divider{background:var(--line);border-radius:1px}
  @media (max-width:860px){ .board-grid{grid-template-columns:1fr} .board-divider{display:none} }
  .board-pool,.board-lanes-wrap{min-width:0}
  .board-pool-label{font-size:11px;letter-spacing:.08em;color:var(--muted);
    font-weight:700;text-transform:uppercase;margin:2px 2px 10px}
  .board-pool-section-hdr{font-size:11px;letter-spacing:.08em;color:#f59e0b;
    font-weight:700;text-transform:uppercase;margin:14px 2px 10px;
    border-top:1px dashed var(--line);padding-top:12px}
  .board-route{border:1px solid var(--line);border-radius:10px;margin-bottom:8px;
    overflow:hidden;background:var(--card2)}
  .board-route-head{display:flex;align-items:center;gap:8px;padding:9px 11px;cursor:pointer}
  .board-route-head:hover{background:var(--line)}
  .board-route-chev{font-size:10px;color:var(--muted);transition:transform .15s}
  .board-route.open .board-route-chev{transform:rotate(90deg)}
  .board-route-dot{width:9px;height:9px;border-radius:50%;flex-shrink:0}
  .board-route-code{font-weight:700;font-size:12.5px;font-family:ui-monospace,Menlo,monospace}
  .board-route-name{font-size:11.5px;color:var(--muted);white-space:nowrap;
    overflow:hidden;text-overflow:ellipsis;min-width:0}
  .board-route-meta{margin-left:auto;font-size:11px;color:var(--muted);flex-shrink:0;
    font-family:ui-monospace,Menlo,monospace}
  .board-route-body{display:none;flex-direction:column;gap:6px;padding:2px 8px 8px}
  .board-route.open .board-route-body{display:flex}
  .board-lanes-tools{display:flex;gap:8px;margin-bottom:10px}
  .board-pool-aging{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:10px}
  .board-pool-aging .aging-pill{font-size:11px;font-weight:700;color:var(--bad);
    background:color-mix(in srgb, var(--bad) 14%, var(--card2));border:1px solid var(--bad);
    border-radius:20px;padding:3px 10px}
  .mini-btn{background:var(--card2);border:1px solid var(--line);color:var(--muted);
    border-radius:8px;padding:6px 12px;font-size:12px;cursor:pointer}
  .mini-btn:hover{color:var(--ink);border-color:var(--brand)}
  .board-lanes{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:12px}
  .board-lane{background:var(--card2);border:1px solid var(--line);border-radius:12px;
    padding:12px;min-height:120px;transition:box-shadow .12s,border-color .12s}
  .board-lane.collapsed{min-height:0;padding:9px 12px}
  .board-lane.collapsed .board-cap-track{margin-bottom:0}
  .board-lane.collapsed .board-lane-body,.board-lane.collapsed .board-lane-naik{display:none}
  .board-lane.zone-active{border-color:var(--brand);box-shadow:0 0 0 2px rgba(56,189,248,.35)}
  .board-mo-zone{border-radius:10px;transition:box-shadow .15s}
  .board-mo-zone.zone-active{box-shadow:0 0 0 2px #f59e0b}
  .board-mo-zone .board-empty{border:1px dashed var(--line);border-radius:10px;
    padding:14px;text-align:center;color:var(--muted);font-size:12.5px}
  .board-lane-head{display:flex;align-items:baseline;gap:7px;margin-bottom:2px;
    cursor:pointer;flex-wrap:wrap}
  .board-lane-chev{font-size:9px;color:var(--muted);transition:transform .15s;align-self:center}
  .board-lane.collapsed .board-lane-chev{transform:rotate(-90deg)}
  .board-lane-plate{font-weight:700;font-size:14px;font-family:ui-monospace,Menlo,monospace}
  .board-lane-count{font-size:11px;color:var(--muted);font-family:ui-monospace,Menlo,monospace;
    background:var(--bg);border-radius:10px;padding:1px 7px}
  .board-lane-drops{font-size:11px;color:var(--muted);font-family:ui-monospace,Menlo,monospace;
    background:var(--bg);border-radius:10px;padding:1px 7px}
  .board-lane-load{margin-left:auto;font-size:11px;font-weight:700;
    font-family:ui-monospace,Menlo,monospace}
  .board-lane.lane-off{opacity:.5}
  .board-lane.lane-off .board-lane-body{pointer-events:none}
  .lane-toggle{position:relative;width:30px;height:17px;border-radius:9px;background:var(--ok);
    flex-shrink:0;cursor:pointer;transition:background .15s;border:none;padding:0}
  .lane-toggle.off{background:var(--line)}
  .lane-toggle::after{content:'';position:absolute;top:2px;left:2px;width:13px;height:13px;
    border-radius:50%;background:#fff;transition:left .15s}
  .lane-toggle.off::after{left:15px}
  .board-lane-naik{font-size:10px;color:var(--muted);width:100%;text-align:right;
    font-family:ui-monospace,Menlo,monospace;margin-bottom:6px}
  .board-cap-track{height:5px;background:var(--bg);border-radius:3px;overflow:hidden;margin-bottom:9px}
  .board-cap-fill{height:100%;transition:width .2s;background:var(--ok)}
  .board-cap-fill.hi{background:var(--warn)}
  .board-cap-fill.over{background:var(--bad)}
  .lane-max-btn{background:none;border:1px solid var(--line);border-radius:6px;color:var(--muted);
    width:22px;height:22px;flex-shrink:0;cursor:pointer;font-size:12px;line-height:1;
    display:inline-flex;align-items:center;justify-content:center}
  .lane-max-btn:hover{color:var(--ink);border-color:var(--brand)}
  .lane-drop-btn{background:none;border:1px solid var(--bad);border-radius:6px;color:var(--bad);
    width:22px;height:22px;flex-shrink:0;cursor:pointer;font-size:13px;line-height:1;font-weight:700;
    display:inline-flex;align-items:center;justify-content:center}
  .lane-drop-btn:hover:not(:disabled){background:var(--bad);color:#fff}
  .lane-drop-btn:disabled{opacity:.3;cursor:not-allowed}
  .board-lanes.has-maximized{grid-template-columns:1fr}
  .board-lane-maximized{padding:20px;min-height:60vh}
  .board-lane-maximized .board-lane-plate{font-size:20px}
  .board-lane-maximized .board-lane-body{gap:10px}
  .board-lane-maximized .board-card{padding:10px 14px;font-size:13.5px}
  .board-lane-body{display:flex;flex-direction:column;gap:6px;min-height:36px}
  .board-empty{color:var(--muted);font-size:12.5px;font-style:italic;padding:6px 2px;opacity:.7}
  .board-card{display:flex;gap:8px;background:var(--card);border:1px solid var(--line);
    border-radius:8px;padding:7px 9px;cursor:grab;font-size:12.5px;touch-action:none;position:relative}
  .board-card.dragging{opacity:.25}
  .board-card.warned{border-color:var(--warn)}
  .board-card .b-stripe{width:4px;align-self:stretch;border-radius:2px;flex-shrink:0}
  .board-card .b-body{min-width:0;flex:1}
  .board-card .b-top{display:flex;align-items:center;gap:7px}
  .board-card .b-id{font-family:ui-monospace,Menlo,monospace;font-weight:700;
    color:var(--brand);font-size:12px}
  .board-card .b-code{font-family:ui-monospace,Menlo,monospace;font-weight:700;
    color:var(--ink);font-size:12px}
  .board-card .b-dist{font-family:ui-monospace,Menlo,monospace;font-weight:600;
    color:var(--muted);font-size:11px}
  .board-card .b-kg{margin-left:auto;font-family:ui-monospace,Menlo,monospace;
    font-weight:700;font-size:12px;color:var(--ok);flex-shrink:0}
  .board-card .b-delete{display:inline-flex;align-items:center;justify-content:center;
    width:18px;height:18px;border-radius:4px;border:none;background:rgba(239,68,68,.15);
    color:var(--bad);font-size:14px;font-weight:bold;cursor:pointer;flex-shrink:0;
    line-height:1;padding:0;transition:background .15s,color .15s}
  .board-card .b-delete:hover{background:var(--bad);color:#fff}
  .board-card .b-cust{font-size:12px;font-weight:600;margin-top:2px;white-space:nowrap;
    overflow:hidden;text-overflow:ellipsis}
  .board-card .b-meta{color:var(--muted);font-size:10.5px;margin-top:1px;
    font-family:ui-monospace,Menlo,monospace}
  .board-card .b-products-toggle{cursor:pointer;user-select:none;font-weight:700}
  .board-card .b-products-toggle:hover{text-decoration:underline}
  .board-card .b-products-list{margin-top:2px}
  .board-card .b-warn{color:var(--warn);font-size:10.5px;margin-top:3px}
  .board-ghost{position:fixed;transform:translate(-50%,-120%);display:flex;gap:8px;
    align-items:center;background:var(--card2);border:1px solid var(--brand);border-radius:8px;
    padding:8px 12px;pointer-events:none;box-shadow:0 8px 24px rgba(0,0,0,.4);
    font-size:12.5px;z-index:999;max-width:80vw}
  .board-toast{position:fixed;left:50%;bottom:20px;transform:translateX(-50%);
    background:var(--warn);color:#1a1300;padding:10px 18px;border-radius:10px;
    font-size:13px;font-weight:600;z-index:1000;box-shadow:0 6px 20px rgba(0,0,0,.35);
    max-width:90vw}
</style>
</head>
<body>
<div class="wrap">
  <header>
    <span class="logo">🚚</span>
    <h1>Lorry Assignment</h1>
    <span class="who" id="who"></span>
    <a href="/logout" class="logout-btn">⎋ Logout</a>
  </header>

  <!-- Planner bar — always visible. Picking a planner logs in and loads
       today's lorries automatically; fetching DOs waits for the ETD window
       below so the user chooses what to pull before anything is assigned. -->
  <div class="card" id="card-login">
    <span class="tp-badge">TRUCK PLANNING</span>
    <h1 class="tp-h1">Loading Board</h1>
    <p class="tp-sub">Pick a planner, set the ETD window, then Fetch DOs — everything
      lands in Unassigned first. Drag a card onto a lorry yourself, or click AI Assign
      to let the algorithm fit them. Bar shows MUATAN capacity — amber means inside the
      +10% naik allowance, red means over the limit.</p>
    <div class="tp-tabs" id="users"><span class="spin"></span></div>
    <div class="msg hidden" id="login-msg"></div>
    <div class="tp-etd-row hidden" id="tp-etd-row">
      <label for="etd-days-input">1&#41; ETD window: &plusmn;
        <input type="number" id="etd-days-input" min="0" step="1" value="0" placeholder="ALL"> day(s) &mdash; 0 or blank = all
      </label>
    </div>
    <div class="tp-toggle-section hidden" id="tp-toggle-section">
      <p class="tp-section-label" style="display:flex;align-items:baseline;gap:10px;flex-wrap:wrap">
        <span>2&#41; Lorries available today &mdash; drag a plate into Staging to share it,
          drag one out to claim it, or turn off any not running:</span>
        <button class="mini-btn" id="tp-reset-holders" type="button" style="margin-left:auto">&#8635; Reset ALL lorries below to their default owner</button>
      </p>
      <div class="tp-staging-station" data-tpzone="STAGING" id="tp-staging-station">
        <div class="tp-staging-label">🅿️ STAGING STATION
          <span class="tp-staging-hint">shared &mdash; drag a plate here to lend it, drag one out to claim it</span>
        </div>
        <div class="tp-staging-grid" id="tp-staging-grid"></div>
      </div>
      <div class="tp-toggle-grid" data-tpzone="MINE" id="tp-toggle-grid"></div>
    </div>
    <div class="tp-fetch-row hidden" id="tp-fetch-row">
      <button class="btn" id="btn-fetch-assign" style="width:auto">📥 Fetch DOs</button>
    </div>
  </div>

  <!-- Step 2: master lorry grid (today's defaults, editable in place) -->
  <div class="card hidden" id="card-master">
    <p class="step-title">Step 2 · Today's lorries — adjust if needed</p>
    <p style="font-size:13px;color:var(--muted);margin:0 0 12px">
      Starts from the default fleet list. Change <b>User</b> or <b>Status</b> for
      today only — nothing here is saved back to the default for tomorrow.
    </p>
    <div class="msg hidden" id="master-msg"></div>
    <div class="scroll">
      <table>
        <thead><tr><th>Lorry</th><th class="w">Ton</th><th>Ori. User</th><th>User</th><th>Status</th></tr></thead>
        <tbody id="master-grid-rows"></tbody>
      </table>
    </div>

    <div style="margin-top:14px;padding-top:12px;border-top:1px solid var(--line)">
      <p style="font-size:13px;color:var(--muted);margin:0 0 8px">+ Add a lorry not on this list (today only):</p>
      <div class="row">
        <input id="new-lorry-plate" type="text" placeholder="Plate (e.g. ABC1234)"
               style="flex:1 1 130px;padding:10px 12px;border-radius:10px;border:1px solid var(--line);background:var(--card2);color:var(--ink)">
        <input id="new-lorry-ton" type="number" step="0.001" placeholder="Ton"
               style="flex:1 1 90px;padding:10px 12px;border-radius:10px;border:1px solid var(--line);background:var(--card2);color:var(--ink)">
        <input id="new-lorry-oriuser" type="text" placeholder="Ori. User"
               style="flex:1 1 110px;padding:10px 12px;border-radius:10px;border:1px solid var(--line);background:var(--card2);color:var(--ink)">
        <select id="new-lorry-user" style="flex:1 1 110px;padding:10px 12px;border-radius:10px;border:1px solid var(--line);background:var(--card2);color:var(--ink)"></select>
        <select id="new-lorry-status" style="flex:1 1 110px;padding:10px 12px;border-radius:10px;border:1px solid var(--line);background:var(--card2);color:var(--ink)">
          <option value="AVAILABLE">AVAILABLE</option><option value="BLOCK">BLOCK</option>
        </select>
        <button class="btn secondary" id="btn-add-lorry" style="flex:0 0 auto;width:auto">+ Add</button>
      </div>
      <div class="msg hidden" id="add-lorry-msg"></div>
    </div>

    <div class="row" style="margin-top:14px">
      <button class="btn back" data-back="login">← Back</button>
      <button class="btn" id="btn-master-next" style="flex:0 0 auto;width:auto">Next →</button>
    </div>

    <p style="font-size:13px;color:var(--muted);text-align:center;margin:14px 0 10px">— or —</p>

    <label class="drop" id="drop-master">
      <div>📄 Upload a master lorry .xlsx instead<br><small>or drag &amp; drop</small></div>
      <input type="file" id="file-master" accept=".xlsx">
    </label>
  </div>

  <!-- Step 3: day -->
  <div class="card hidden" id="card-day">
    <p class="step-title">Step 3 · Which day?</p>
    <div class="row">
      <button class="btn" data-day="today">Today</button>
      <button class="btn secondary" data-day="tomorrow">Tomorrow</button>
    </div>
    <div class="msg hidden" id="day-msg"></div>
    <button class="btn back" data-back="afterlogin">← Back</button>
  </div>

  <!-- Step 4: DO file -->
  <div class="card hidden" id="card-dos">
    <p class="step-title">Step 4 · Get today's DOs</p>

    <div style="margin-bottom:14px">
      <button class="btn" id="btn-dos-fetch">📥 Fetch DOs from system</button>
      <div class="msg hidden" id="dos-fetch-msg"></div>
      <div id="dos-fetch-result" class="hidden" style="margin-top:10px;font-size:14px">
        <span id="dos-fetch-summary"></span>
        <span id="dos-fetch-diag-toggle" style="cursor:pointer;color:var(--muted);text-decoration:underline;margin-left:8px;font-size:12px">ℹ️ fetch details</span>
        <div id="dos-fetch-diag" class="hidden" style="margin-top:6px;font-size:12px;color:var(--muted);font-family:ui-monospace,Menlo,monospace;white-space:pre-wrap"></div>
        <div class="row" style="margin-top:8px">
          <a class="dl" href="/api/dos-fetch/download"><button class="btn secondary" style="width:auto">⬇️ Download to review</button></a>
          <button class="btn" id="btn-dos-fetch-use" style="width:auto">Use this directly →</button>
        </div>
      </div>
    </div>

    <p style="font-size:13px;color:var(--muted);text-align:center;margin:4px 0 10px">— or —</p>

    <label class="drop" id="drop-dos">
      <div>📎 Upload a DO .xlsx (e.g. after editing the fetched file)<br><small>or drag &amp; drop</small></div>
      <input type="file" id="file-dos" accept=".xlsx">
    </label>
    <div class="msg hidden" id="dos-msg"></div>
    <button class="btn back" data-back="day">← Back</button>
  </div>

  <!-- Step 4b: off-schedule question -->
  <div class="card hidden" id="card-offsched">
    <p class="step-title">One moment · Off-schedule DOs</p>
    <div class="msg" id="offsched-msg"></div>
    <div class="row" style="margin-top:14px">
      <button class="btn" id="offsched-yes">Yes, assign them too</button>
      <button class="btn secondary" id="offsched-no">No, leave them blank</button>
    </div>
    <button class="btn back" data-back="dos">← Back (re-upload DO file)</button>
  </div>

  <!-- Step 4c: picking list — review the fetched DOs, uncheck any not needed today -->
  <div class="card hidden" id="card-picking">
    <p class="step-title" style="margin-bottom:2px">Picking List</p>
    <p style="font-size:13px;color:var(--muted);margin:0 0 14px">
      All of today's DOs are checked by default. Uncheck any you don't need to assign today, then continue.
    </p>
    <div class="row" style="margin-bottom:10px;align-items:center">
      <button class="mini-btn" id="picking-select-all">Select all</button>
      <button class="mini-btn" id="picking-deselect-all">Deselect all</button>
      <span id="picking-count" style="margin-left:auto;font-size:12.5px;color:var(--muted);font-weight:600"></span>
    </div>
    <div class="msg hidden" id="picking-msg"></div>
    <span id="picking-fetch-diag-toggle" class="hidden" style="cursor:pointer;color:var(--muted);text-decoration:underline;font-size:12px">ℹ️ fetch details</span>
    <div id="picking-fetch-diag" class="hidden" style="margin:6px 0;font-size:12px;color:var(--muted);font-family:ui-monospace,Menlo,monospace;white-space:pre-wrap"></div>
    <div class="scroll" style="max-height:60vh">
      <table>
        <thead><tr><th style="width:32px"></th><th>Date</th><th>DO</th><th>Code</th><th>Customer</th><th>Route</th><th class="w">Weight</th></tr></thead>
        <tbody id="picking-rows"></tbody>
      </table>
    </div>
    <div class="row" style="margin-top:14px">
      <button class="btn" id="picking-next" style="width:auto">Next →</button>
    </div>
  </div>

  <!-- Step 5: drag-and-drop board -->
  <div class="card hidden" id="card-board">
    <div class="board-top">
      <span class="board-assigned-stat" id="board-assigned-stat"></span>
      <div class="board-plan-row" id="board-plan-row">
        <label class="board-plan-field">Assign for
          <input type="date" id="tp-day-date">
        </label>
        <span class="board-plan-warn hidden" id="tp-day-late-warn" title="It's after 2pm — deliveries planned for today may be tight. Consider planning for tomorrow instead.">⚠️ After 2pm</span>
        <span class="board-plan-field" style="border:none;padding:0;background:none">Trip
          <span class="board-plan-trip-btns">
            <button class="btn tp-trip-btn active" id="tp-trip-1" style="width:auto" data-tptrip="1">1</button>
            <button class="btn secondary tp-trip-btn" id="tp-trip-2" style="width:auto" data-tptrip="2">2</button>
            <button class="btn secondary tp-trip-btn" id="tp-trip-3" style="width:auto" data-tptrip="3">3</button>
            <button class="btn secondary tp-trip-btn" id="tp-trip-4" style="width:auto" data-tptrip="4">4</button>
          </span>
        </span>
      </div>
      <div class="board-top-actions">
        <button class="btn btn-save" id="btn-board-save">💾 Save to SQL</button>
        <div class="board-download-dropdown" id="board-download-dropdown">
          <button class="btn secondary" id="board-download-toggle">⬇️ Download ▾</button>
          <div class="board-download-menu hidden" id="board-download-menu">
            <a class="dl" href="/api/download">⬇️ Download &amp; Print (assigned)</a>
            <a class="dl" href="/api/board/download-unassigned" title="Temporary — every DO not on a real plate, with the reason it's not, for checking">🔎 Download unassigned (temp)</a>
          </div>
        </div>
        <button class="btn btn-ai" id="btn-board-ai">🤖 AI Assign</button>
        <button class="btn secondary hidden" id="btn-board-table">📋 Table view</button>
      </div>
    </div>
    <div class="msg hidden" id="board-msg"></div>
    <div style="margin:-4px 0 8px">
      <span id="board-fetch-diag-toggle" class="hidden" style="cursor:pointer;color:var(--muted);text-decoration:underline;font-size:12px">ℹ️ fetch details</span>
      <div id="board-fetch-diag" class="hidden" style="margin-top:6px;font-size:12px;color:var(--muted);font-family:ui-monospace,Menlo,monospace;white-space:pre-wrap"></div>
    </div>
    <div class="board-grid">
      <section class="board-pool" data-zone="">
        <div class="board-pool-label" id="board-pool-label">UNASSIGNED</div>
        <div class="board-lanes-tools">
          <button class="mini-btn" id="board-pool-collapse-all">Collapse all routes</button>
          <button class="mini-btn" id="board-pool-expand-all">Expand all</button>
        </div>
        <div class="board-pool-aging" id="board-pool-aging"></div>
        <div id="board-routes"></div>
      </section>
      <div class="board-divider"></div>
      <div class="board-lanes-wrap">
        <div class="board-lanes-tools">
          <button class="mini-btn" id="board-collapse-all">Collapse all lorries</button>
          <button class="mini-btn" id="board-expand-all">Expand all</button>
        </div>
        <div class="board-lanes" id="board-lanes"></div>
      </div>
    </div>
    <div class="row" style="margin-top:14px">
      <button class="btn secondary" id="btn-board-back">🔄 Refetch DOs</button>
      <button class="btn secondary" id="btn-board-restart">Start over</button>
    </div>
  </div>

  <!-- Step 5b: result (table view / export / manual-assign tools) -->
  <div class="card hidden" id="card-result">
    <p class="step-title">Result</p>
    <div class="row" style="margin-bottom:14px">
      <button class="btn secondary" id="btn-result-board" style="width:auto">🧩 Board view</button>
    </div>
    <div class="stat" id="result-stat"></div>
    <div style="margin-bottom:14px">
      <a class="dl" href="/api/download"><button class="btn">⬇️ Download filled Excel</button></a>
    </div>
    <div id="result-body"></div>

    <!-- Reassign leftover DOs onto now-available lorries (no re-upload) -->
    <div id="reassign-box" class="hidden" style="margin-top:16px;border-top:1px solid var(--line);padding-top:14px">
      <p class="step-title">Leftover DOs — assign to available lorries</p>
      <div class="msg" id="reassign-msg"></div>
      <p style="font-size:14px;color:var(--muted);margin:4px 0 10px">
        <b id="reassign-count">0</b> DO(s), <b id="reassign-weight">0</b> T unassigned.
        Tick the lorries that are now free and assign them — no need to re-upload.
      </p>
      <div id="reassign-fleet" class="grid-users" style="justify-content:flex-start;margin-bottom:12px"></div>
      <button class="btn" id="btn-reassign" style="width:auto">Assign leftover DOs to ticked lorries</button>
    </div>

    <!-- Manually assign hand-picked DOs onto ONE named lorry (no re-upload) -->
    <div id="manual-assign-box" class="hidden" style="margin-top:16px;border-top:1px solid var(--line);padding-top:14px">
      <p class="step-title">Or assign specific DOs to one lorry</p>
      <div class="row">
        <input id="manual-plate" type="text" placeholder="Lorry plate (e.g. VEA2818)"
               style="flex:1 1 200px;padding:12px 14px;border-radius:10px;border:1px solid var(--line);
                      background:var(--card2);color:var(--ink);font-size:15px">
        <button class="btn secondary" id="btn-manual-lookup" style="flex:0 0 auto;width:auto;padding:12px 18px">Find DOs for this plate</button>
      </div>
      <div class="msg hidden" id="manual-msg"></div>
      <div id="manual-picker" class="hidden" style="margin-top:12px">
        <div class="scroll"><table><thead><tr><th></th><th>DO</th><th>Route</th><th>Customer</th><th class="w">Weight</th></tr></thead>
          <tbody id="manual-do-rows"></tbody>
        </table></div>
        <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;margin-top:10px;font-size:14px">
          <span>Selected: <b id="manual-selected-total">0.000</b>T / <span id="manual-plate-cap">0</span>T
            <span id="manual-cap-warn" class="hidden" style="color:var(--bad);font-weight:700"> — over capacity!</span></span>
          <button class="btn" id="btn-manual-assign" style="width:auto" disabled>Assign selected to this lorry</button>
        </div>
      </div>
    </div>

    <div class="row" style="margin-top:14px">
      <button class="btn secondary" id="btn-refresh">🔄 Refresh</button>
      <button class="btn secondary" id="btn-restart">Start over</button>
    </div>
  </div>

  <div class="foot">Same assignment engine as the WhatsApp bot · works on phone &amp; desktop</div>
</div>

<script>
const $ = s => document.querySelector(s);
const show = (id,on=true)=>{ $(id).classList.toggle('hidden',!on); };
const setMsg = (id,txt,err=false)=>{ const e=$(id); if(!txt){e.classList.add('hidden');return;}
  e.textContent=Array.isArray(txt)?txt.join('\n'):txt; e.classList.toggle('err',err); e.classList.remove('hidden'); };

// If the session was lost (server returns 401), send the user back to the
// login page to re-authenticate rather than showing a confusing error.
function _checkAuth(r){ if(r.status===401){ window.location.href='/login'; throw new Error('auth'); } return r; }
async function jpost(url,body){ const r=_checkAuth(await fetch(url,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body||{})})); return r.json(); }
async function fpost(url,file){ const fd=new FormData(); fd.append('file',file); const r=_checkAuth(await fetch(url,{method:'POST',body:fd})); return r.json(); }

// ---- Wizard state (so Back can rewind the engine and replay prior steps) ----
let selUser=null, masterFile=null, selDay=null, needsMaster=false;
// #card-login is now the always-visible planner-tabs bar, not a wizard step —
// it's deliberately left out of ALL_CARDS so hideAll()/goTo() never hide it.
const ALL_CARDS=['#card-master','#card-day','#card-dos','#card-offsched','#card-picking','#card-board','#card-result'];
function hideAll(){ ALL_CARDS.forEach(id=>show(id,false)); }
function clearMsgs(){ ['#login-msg','#master-msg','#day-msg','#dos-msg','#offsched-msg','#board-msg'].forEach(id=>setMsg(id,null)); }

// Rewind to an earlier step (used by the master-grid card's own "← Back",
// a rare fallback path — see autoLoadPlanner). Only moves the engine's state
// pointer back (server keeps the loaded master), so it's instant.
async function goTo(step){
  clearMsgs();
  document.body.classList.remove('board-active');
  await jpost('/api/back',{target:step});
  const inpD=document.getElementById('file-dos');
  const inpM=document.getElementById('file-master');
  if(step==='login'){ selUser=null; masterFile=null; selDay=null; if(inpD)inpD.value=''; if(inpM)inpM.value=''; }
  if(step==='master'){ masterFile=null; if(inpM)inpM.value=''; loadMasterGrid(); }
  if(step==='day'){ selDay=null; }
  if(step==='dos'){ if(inpD)inpD.value=''; }
  hideAll();
  if(step!=='login'){ show('#card-'+step, true); }
}
function wireBackButtons(){
  document.querySelectorAll('[data-back]').forEach(b=>b.onclick=()=>{
    let t=b.dataset.back;
    if(t==='afterlogin') t = needsMaster ? 'master' : 'login';
    goTo(t);
  });
}

// ---- Planner tabs (always visible). Picking one runs the whole pipeline —
// login, default lorries, today, live DO fetch from the system, auto-assign —
// and lands straight on the board. No manual wizard steps in the normal path;
// the master/day/dos cards only resurface as a fallback if something needs
// a human's attention (e.g. a lorry double-booked between two planners). ----
let validUsers=[];
async function loadUsers(){
  const d = await (await fetch('/api/users')).json();
  validUsers = d.users||[];
  renderPlannerTabs();
}
function renderPlannerTabs(){
  const box=$('#users'); box.innerHTML='';
  validUsers.forEach(u=>{
    const b=document.createElement('button');
    b.className='tp-tab'+(u===selUser?' active':'');
    b.textContent=u;
    b.onclick=()=>autoLoadPlanner(u);
    box.appendChild(b);
  });
}
function setActiveTab(user){ selUser=user; renderPlannerTabs(); }

function clearBoardPanels(){
  BOARD=null;
  const routes=$('#board-routes'), lanes=$('#board-lanes');
  if(routes) routes.innerHTML=''; if(lanes) lanes.innerHTML='';
}
function showBoardWithError(msg){
  hideAll(); show('#card-board',true);
  document.body.classList.add('board-active');
  clearBoardPanels();
  setMsg('#board-msg', msg, true);
}

function _resetEtdDays(){
  // Always defaults to 0 (= all) on every login/restore — by explicit
  // request it must never remember a previously-typed ETD window.
  $('#etd-days-input').value = '0';
}

// Login + today's lorries + trip day happen automatically — fast, no
// judgement call involved. Fetching DOs does NOT: the user picks the ETD
// window first (below) and clicks Fetch & Assign themselves, rather than
// the AI silently deciding what to pull. autoFetch=true (used by
// Refetch/Start-over, an explicit click in its own right) skips re-asking
// and reuses whatever's currently in the ETD-window field.
async function autoLoadPlanner(user, autoFetch){
  clearMsgs();
  setActiveTab(user);
  document.querySelectorAll('.tp-tab').forEach(b=>b.disabled=true);
  hideAll();
  show('#tp-etd-row', false);
  show('#tp-toggle-section', false);
  show('#tp-fetch-row', false);
  try{
    setMsg('#login-msg', 'Loading '+user+"'s lorries… ", false);
    const dl = await jpost('/api/login',{user});
    if(!dl.user){ setMsg('#login-msg', dl.messages||'Could not load '+user, true); return; }
    $('#who').textContent = 'Logged in as '+dl.user;

    if(dl.already_active){
      // This planner already has in-progress or completed work from
      // before switching to the other tab — restore it directly instead
      // of re-running the whole setup wizard, which would wipe it.
      setMsg('#login-msg', null);
      setDayDate(dl.trip_day || _todayISO());
      setTripActive(dl.trip_session || '1');
      await updateTimeRestrictions();
      _resetEtdDays();
      show('#tp-etd-row', true);
      show('#tp-toggle-section', true);
      show('#tp-fetch-row', true);
      await loadLorryToggles();
      if(dl.needs_picking){ await showPickingList(); }
      else if(dl.has_items){ showBoard(); }
      return;
    }

    needsMaster = !!dl.needs_master;

    if(needsMaster){
      const dm = await (await fetch('/api/master-default')).json();
      if(dm.error){ setMsg('#login-msg', dm.error, true); return; }
      const dg = await jpost('/api/master-grid',{rows: dm.rows||[]});
      if(dg.error){
        // A real server error (crash) — surface it right here, in the same
        // screen every user sees, instead of silently switching to the old
        // editable-grid layout (that's for a genuine data conflict below,
        // not a bug — switching layouts on a crash just hides the error).
        setMsg('#login-msg', dg.error, true);
        return;
      }
      if(!dg.ok){
        // Genuine data issue: e.g. the same plate marked Available for two
        // planners today, or no Available lorries found at all. Needs a
        // human to fix — drop into the editable grid so they can, instead
        // of guessing which planner should keep a conflicting plate.
        setMsg('#login-msg', null);
        show('#card-master', true); loadMasterGrid();
        setMsg('#master-msg', dg.messages||'Please review and continue.', true);
        return;
      }
    }

    // Default to the earliest still-plannable day/trip (today unless the
    // 2pm/4pm cutoffs have already closed it) — the board's own
    // Assign-for/Trip controls let the user switch before re-fetching.
    const _defaultDay = _earliestAllowedDate();
    setDayDate(_defaultDay);
    const dd = await jpost('/api/day',{day:_defaultDay});
    if(!dd.ok){ setMsg('#login-msg', dd.messages||'Could not set trip day', true); return; }
    const _defaultTrip = (_defaultDay===_todayISO() && new Date().getHours()>=14) ? '3' : '1';
    setTripActive(_defaultTrip);
    await jpost('/api/trip-session',{session:_defaultTrip});
    await updateTimeRestrictions();

    setMsg('#login-msg', null);
    _resetEtdDays();
    show('#tp-etd-row', true);
    show('#tp-toggle-section', true);
    show('#tp-fetch-row', true);
    await loadLorryToggles();
    if(autoFetch) await fetchAndAssign();
  } finally {
    document.querySelectorAll('.tp-tab').forEach(b=>b.disabled=false);
  }
}

let LAST_TOGGLES = null;

function tpChipHtml(l, zone){
  if(zone==='STAGING'){
    const broken = !!l.broken;
    return `<span class="tp-toggle-chip${broken?' tp-broken':''}" data-plate="${esc(l.plate)}" data-tphome="STAGING">`+
      `<button class="lane-toggle${broken?' off':''}" title="${broken?'Marked broken — click to mark fixed':'Working — click to mark broken'}"></button>`+
      `${esc(l.plate)}${broken?'<span class="tp-broken-tag">BROKEN</span>':''}</span>`;
  }
  const on = l.on !== false;
  return `<span class="tp-toggle-chip" data-plate="${esc(l.plate)}" data-tphome="MINE">`+
    `<button class="lane-toggle${on?'':' off'}" title="${on?'Available — click to turn off':'Not available — click to turn on'}"></button>`+
    `${esc(l.plate)}</span>`;
}

async function loadLorryToggles(){
  const d = await (await fetch('/api/lorry-toggles', {credentials:'same-origin'})).json();
  LAST_TOGGLES = d;
  renderLorryToggles();
}

function renderLorryToggles(){
  const d = LAST_TOGGLES;
  const grid = $('#tp-toggle-grid');
  const stagingGrid = $('#tp-staging-grid');
  if(!d || (!d.lorries?.length && !d.staging?.length)){
    grid.innerHTML = '<span style="color:var(--muted);font-size:12.5px">No lorries loaded yet.</span>';
    stagingGrid.innerHTML = '';
    return;
  }
  grid.innerHTML = (d.lorries||[]).map(l=>tpChipHtml(l,'MINE')).join('');
  stagingGrid.innerHTML = (d.staging||[]).map(l=>tpChipHtml(l,'STAGING')).join('');

  document.querySelectorAll('#tp-toggle-grid .tp-toggle-chip, #tp-staging-grid .tp-toggle-chip').forEach(chip=>{
    const home = chip.dataset.tphome;
    chip.querySelector('.lane-toggle').onclick = async(e)=>{
      e.stopPropagation();
      const btn = chip.querySelector('.lane-toggle');
      const turningOn = btn.classList.contains('off');
      if(home==='STAGING'){
        await jpost('/api/board/staging-broken', {plate: chip.dataset.plate, broken: !turningOn});
        // Refresh whichever views are visible — loadBoard() alone doesn't
        // touch this toggle grid, so without also reloading it the chip's
        // own broken/fixed state looked stuck even though the click worked.
        if(BOARD) await loadBoard();
        await loadLorryToggles();
      } else {
        // Shared with the board's own lane switches — so whichever one the
        // user clicks, both this grid AND the board (if already showing)
        // reflect the change immediately, not just on next reload.
        await toggleLorry(chip.dataset.plate, turningOn);
      }
    };
    chip.addEventListener('pointerdown', e=>{
      if(e.target.closest('.lane-toggle')) return;
      startChipDrag(e, chip.dataset.plate, home);
    });
  });
}

// ---- Staging-station drag-and-drop (own pointer-based mechanism, shared
// zones marked with data-tpzone="MINE"/"STAGING" on the two containers) ----
let chipDrag=null, chipGhost=null;
function startChipDrag(e, plate, fromZone){
  e.preventDefault();
  chipDrag = {plate, fromZone};
  chipGhost = document.createElement('div');
  chipGhost.className = 'tp-chip-ghost';
  chipGhost.textContent = plate;
  document.body.appendChild(chipGhost);
  moveChipGhost(e.clientX, e.clientY);
  document.querySelectorAll(`#tp-toggle-grid .tp-toggle-chip[data-plate="${plate}"], #tp-staging-grid .tp-toggle-chip[data-plate="${plate}"]`)
    .forEach(c=>c.classList.add('tp-dragging'));
}
function moveChipGhost(x,y){ if(chipGhost){ chipGhost.style.left=x+'px'; chipGhost.style.top=y+'px'; } }
document.addEventListener('pointermove', e=>{
  if(!chipDrag) return;
  moveChipGhost(e.clientX, e.clientY);
  document.querySelectorAll('.tpzone-active').forEach(el=>el.classList.remove('tpzone-active'));
  const z = document.elementsFromPoint(e.clientX, e.clientY).find(el=>el.closest?.('[data-tpzone]'));
  const zoneEl = z ? z.closest('[data-tpzone]') : null;
  if(zoneEl) zoneEl.classList.add('tpzone-active');
});
$('#tp-reset-holders').onclick=async()=>{
  if(!confirm('Put EVERY lorry plate back to its default owner today — ABI\'s plates to ABI, VIVIAN\'s to VIVIAN, SPARE plates to Staging — whether it\'s currently sitting in Staging or in someone\'s own lorries list below? This undoes all claim/release moves made today.')) return;
  const d = await jpost('/api/board/reset-holders', {});
  if(!d.ok){ boardToast(d.message||d.error||'Reset failed'); return; }
  await loadLorryToggles();
  if(BOARD) await loadBoard();
  boardToast('All lorries reset to today\'s default owner.');
};
document.addEventListener('pointerup', async e=>{
  if(!chipDrag) return;
  const {plate, fromZone} = chipDrag;
  chipDrag = null;
  if(chipGhost){ chipGhost.remove(); chipGhost=null; }
  document.querySelectorAll('.tpzone-active').forEach(el=>el.classList.remove('tpzone-active'));
  document.querySelectorAll('.tp-dragging').forEach(el=>el.classList.remove('tp-dragging'));
  const z = document.elementsFromPoint(e.clientX, e.clientY).find(el=>el.closest?.('[data-tpzone]'));
  const zoneEl = z ? z.closest('[data-tpzone]') : null;
  const toZone = zoneEl ? zoneEl.dataset.tpzone : null;
  if(!toZone || toZone===fromZone) return;   // dropped nowhere, or back where it came from
  const action = toZone==='MINE' ? 'claim' : 'release';
  const d = await jpost('/api/board/claim-plate', {plate, action});
  if(!d.ok){ boardToast(d.message||d.error||'Move failed'); await loadLorryToggles(); return; }
  await loadLorryToggles();
  if(BOARD) await loadBoard();
});

// ---- Assign-for date (calendar) — which SCHD weekday to plan against.
// Deliberately NOT a filter on the DOs' own DATE column (see /api/day).
// Cutoff rules (by explicit request): after 4pm, today is closed for
// planning entirely — the earliest pickable date becomes tomorrow. Between
// 2pm and 4pm today is still pickable but Trip 1/2 close first (only
// Trip 3/4 may be used for today in that window).
function _todayISO(){
  const d = new Date();
  return new Date(d.getTime() - d.getTimezoneOffset()*60000).toISOString().slice(0,10);
}
function _tomorrowISO(){
  const d = new Date(Date.now() + 24*3600*1000);
  return new Date(d.getTime() - d.getTimezoneOffset()*60000).toISOString().slice(0,10);
}
function _earliestAllowedDate(){
  return new Date().getHours() >= 16 ? _tomorrowISO() : _todayISO();
}
function updateLateWarning(){
  const inp = $('#tp-day-date'), warn = $('#tp-day-late-warn');
  if(!inp || !warn) return;
  const isToday = inp.value === _todayISO();
  const isLate = new Date().getHours() >= 14;
  warn.classList.toggle('hidden', !(isToday && isLate));
}
function setDayDate(dateStr){
  const inp = $('#tp-day-date');
  if(inp) inp.value = dateStr;
  updateLateWarning();
}
function setTripActive(session){
  ['1','2','3','4'].forEach(s=>{
    const btn = $('#tp-trip-'+s);
    if(!btn) return;
    const isActive = s===session;
    btn.classList.toggle('active', isActive);
    btn.classList.toggle('secondary', !isActive);
  });
}
// Re-checked on load, on date change, and every minute (so a page left
// open across the 2pm/4pm cutoffs updates itself without a refresh).
async function updateTimeRestrictions(){
  const inp = $('#tp-day-date');
  if(!inp) return;
  const earliest = _earliestAllowedDate();
  inp.min = earliest;
  if(inp.value && inp.value < earliest){
    setDayDate(earliest);
    await jpost('/api/day', {day: earliest});
    boardToast("It's after 4pm — today is closed for planning. Moved to tomorrow.");
  }
  const isToday = inp.value === _todayISO();
  const block12 = isToday && new Date().getHours() >= 14;
  ['1','2'].forEach(s=>{
    const btn = $('#tp-trip-'+s);
    if(btn) btn.disabled = block12;
  });
  if(block12){
    const activeTrip = document.querySelector('.tp-trip-btn.active')?.dataset.tptrip;
    if(activeTrip==='1' || activeTrip==='2'){
      setTripActive('3');
      await jpost('/api/trip-session', {session: '3'});
      boardToast("It's after 2pm — Trip 1/2 are closed for today. Switched to Trip 3.");
    }
  }
  updateLateWarning();
}
const _dayDateInput = $('#tp-day-date');
if(_dayDateInput){
  _dayDateInput.min = _earliestAllowedDate();
  _dayDateInput.addEventListener('change', async()=>{
    const val = _dayDateInput.value;
    if(!val) return;
    const dd = await jpost('/api/day', {day: val});
    if(!dd.ok){
      boardToast(dd.error || (dd.messages && dd.messages[0]) || 'Could not set that date');
      const fallback = _earliestAllowedDate();
      setDayDate(fallback);
      await jpost('/api/day', {day: fallback});
    }
    await updateTimeRestrictions();
  });
  setInterval(updateTimeRestrictions, 60000);
}

document.querySelectorAll('.tp-trip-btn').forEach(b=>{
  b.onclick = async()=>{
    if(b.disabled) return;
    const prev = document.querySelector('.tp-trip-btn.active')?.dataset.tptrip;
    setTripActive(b.dataset.tptrip);
    const d = await jpost('/api/trip-session', {session: b.dataset.tptrip});
    if(!d.ok){
      boardToast(d.error || 'Could not set that trip.');
      if(prev) setTripActive(prev);
      await updateTimeRestrictions();
    }
  };
});

async function fetchAndAssign(){
  const btn = $('#btn-fetch-assign');
  const raw = $('#etd-days-input').value;
  const etdDays = raw==='' ? null : parseInt(raw, 10);
  if(etdDays!==null && (isNaN(etdDays) || etdDays<0)){
    setMsg('#login-msg', 'ETD window must be a whole number of days (0 or more).', true);
    return;
  }
  btn.disabled = true;
  try{
    setMsg('#login-msg', 'Fetching DOs from system… ', false);
    const df = await jpost('/api/dos-fetch', {etd_days: etdDays});
    if(df.error){ showBoardWithError(df.error); return; }
    if(!df.count){
      let _diagMsg = '';
      if(df.diagnostics){
        const d = df.diagnostics;
        _diagMsg = ` [diagnostics: raw SQL rows=${d.raw_rows_from_sql}, `+
          `after not-LOAN=${d.after_not_loan}, after site 1SA=${d.after_site_1SA}, `+
          `after not-yet-validated=${d.after_not_validated}, after known route=${d.after_known_route}, `+
          `after 30-day floor=${d.after_30day_floor}`+
          (d.after_etd_window!=null?`, after ETD window=${d.after_etd_window}`:'')+
          `. Site codes seen: ${(d.distinct_stofcy_seen||[]).join(', ')||'none'}. `+
          `Types seen: ${(d.distinct_sdhtyp_seen||[]).join(', ')||'none'}.]`;
      }
      showBoardWithError(
        `The system returned 0 DOs for ${selUser}` +
        (etdDays!=null?` with the ETD window set to ±${etdDays} day(s)`:'') +
        `. Double-check the ETD window (0 or blank = all), the Today/Tomorrow ` +
        `pick, and that today's DOs are actually in the system for ${selUser}'s routes.` +
        _diagMsg
      );
      return;
    }

    setMsg('#login-msg', 'Assigning lorries… ', false);
    const du = await jpost('/api/dos-fetch/use',{});
    if(du.offschedule){
      // Which routes run today/tomorrow comes from the LORRY DAILY PLANNING
      // file's SCHD sheet for the day the user picked — a route not on that
      // day's list stays unassigned rather than being auto-included.
      await jpost('/api/offschedule',{assign:false});
    } else if(du.error){
      showBoardWithError(du.messages||du.error||'Could not assign lorries.');
      return;
    }
    setMsg('#login-msg', null);
    await showPickingList();
  } finally {
    btn.disabled = false;
  }
}
$('#btn-fetch-assign').onclick = fetchAndAssign;

async function restartPlanner(){
  document.body.classList.remove('board-active');
  await jpost('/api/back',{target:'login'});
  hideAll();
  const u = selUser || validUsers[0];
  if(u) autoLoadPlanner(u, true);
}

// ---- file drop helper ----
function wireDrop(dropId,inputId,onFile){
  const drop=$(dropId), input=$(inputId);
  input.addEventListener('change',()=>{ if(input.files[0]) onFile(input.files[0]); });
  ['dragover','dragenter'].forEach(ev=>drop.addEventListener(ev,e=>{e.preventDefault();drop.classList.add('hot');}));
  ['dragleave','drop'].forEach(ev=>drop.addEventListener(ev,e=>{e.preventDefault();drop.classList.remove('hot');}));
  drop.addEventListener('drop',e=>{ const f=e.dataTransfer.files[0]; if(f) onFile(f); });
}

// ---- Step 2: master lorry grid (editable, no upload) ----
const STATUS_OPTS=['AVAILABLE','BLOCK'];
function masterUserOpts(){
  // The portal's planner tabs are ABI/VIVIAN only, but a row in this grid
  // can still legitimately be a shared SPARE lorry — offer that as a third
  // option here even though it's not a tab.
  const base = validUsers.length ? validUsers : [selUser];
  return base.includes('SPARE') ? base : [...base, 'SPARE'];
}
function masterRowHtml(r){
  const userOpts=masterUserOpts();
  const userSelect = userOpts.map(u=>`<option value="${esc(u)}" ${u===r.user?'selected':''}>${esc(u)}</option>`).join('');
  const statusSelect = STATUS_OPTS.map(s=>`<option value="${esc(s)}" ${s===r.status?'selected':''}>${esc(s)}</option>`).join('');
  return `<tr data-lorry="${esc(r.lorry)}" data-ton="${r.ton!=null?r.ton:''}">`+
         `<td>${esc(r.lorry)}</td>`+
         `<td class="w">${r.ton!=null?Number(r.ton).toFixed(2):'—'}</td>`+
         `<td>${esc(r.ori_user||'')}</td>`+
         `<td><select class="mg-user">${userSelect}</select></td>`+
         `<td><select class="mg-status">${statusSelect}</select></td>`+
         `</tr>`;
}
async function loadMasterGrid(){
  setMsg('#master-msg','Loading today\'s lorries… ',false);
  const d=await (await fetch('/api/master-default')).json();
  if(d.error){ setMsg('#master-msg', d.error, true); return; }
  setMsg('#master-msg', null);
  $('#master-grid-rows').innerHTML=(d.rows||[]).map(masterRowHtml).join('');
  const userOpts=masterUserOpts();
  $('#new-lorry-user').innerHTML=userOpts.map(u=>`<option value="${esc(u)}">${esc(u)}</option>`).join('');
}

function addLorryRow(){
  const plate=($('#new-lorry-plate').value||'').trim().toUpperCase();
  const tonRaw=$('#new-lorry-ton').value;
  const ton=tonRaw!=='' ? parseFloat(tonRaw) : NaN;
  if(!plate){ setMsg('#add-lorry-msg','Enter a plate number.',true); return; }
  if(isNaN(ton) || ton<=0){ setMsg('#add-lorry-msg','Enter a valid Ton (> 0).',true); return; }
  const existing=[...document.querySelectorAll('#master-grid-rows tr')].map(tr=>tr.dataset.lorry);
  if(existing.includes(plate)){ setMsg('#add-lorry-msg',`${plate} is already in the list above.`,true); return; }
  setMsg('#add-lorry-msg', null);
  $('#master-grid-rows').insertAdjacentHTML('beforeend', masterRowHtml({
    lorry: plate, ton: ton,
    ori_user: ($('#new-lorry-oriuser').value||'').trim().toUpperCase(),
    user: $('#new-lorry-user').value,
    status: $('#new-lorry-status').value,
  }));
  $('#new-lorry-plate').value=''; $('#new-lorry-ton').value=''; $('#new-lorry-oriuser').value='';
}
$('#btn-add-lorry').onclick=addLorryRow;

async function submitMasterGrid(){
  const rows=[...document.querySelectorAll('#master-grid-rows tr')].map(tr=>({
    lorry: tr.dataset.lorry,
    ton: tr.dataset.ton!=='' ? parseFloat(tr.dataset.ton) : null,
    user: tr.querySelector('.mg-user').value,
    status: tr.querySelector('.mg-status').value,
  }));
  const btn=$('#btn-master-next'); btn.disabled=true;
  setMsg('#master-msg','Saving… ',false);
  try{
    const d=await jpost('/api/master-grid',{rows});
    setMsg('#master-msg', d.messages||d.error, !d.ok);
    if(d.ok){ show('#card-master',false); show('#card-day',true); }
  } finally { btn.disabled=false; }
}
$('#btn-master-next').onclick=submitMasterGrid;

// ---- Step 2b: master lorry file upload (alternative to the grid) ----
wireDrop('#drop-master','#file-master',async(f)=>{
  setMsg('#master-msg','Reading master file… ',false);
  const d=await fpost('/api/master',f);
  setMsg('#master-msg',d.messages||d.error,!d.ok);
  if(d.ok){ masterFile=f; show('#card-master',false); show('#card-day',true); }
});

// ---- Step 3: day ----
document.querySelectorAll('[data-day]').forEach(b=>b.onclick=async()=>{
  setMsg('#day-msg','Setting day… ',false);
  selDay=b.dataset.day;
  const d=await jpost('/api/day',{day:b.dataset.day});
  setMsg('#day-msg',d.messages,false);
  show('#card-day',false); show('#card-dos',true);
});

// ---- Step 4: DOs ----
// Shared by manual upload and "use fetched DOs directly" — same response shape.
let LAST_PARSE_DIAG = null;
function handleDosResponse(d){
  LAST_PARSE_DIAG = d.parse_diagnostics || null;
  if(d.offschedule){
    const os=d.offschedule;
    const dayTxt = os.day==='tomorrow' ? 'tomorrow' : 'today';
    setMsg('#offsched-msg',
      `⏭ ${os.count} DO(s) are NOT on ${dayTxt}'s route schedule.\n\n`+
      `Do you still want to assign lorries to them?`);
    show('#card-dos',false); show('#card-offsched',true);
    return true;
  }
  if(d.result){ show('#card-dos',false); showPickingList(); return true; }
  return false;
}

wireDrop('#drop-dos','#file-dos',async(f)=>{
  setMsg('#dos-msg','Assigning lorries… this can take a moment ',false);
  const d=await fpost('/api/dos',f);
  if(handleDosResponse(d)){ setMsg('#dos-msg',null); }
  else { setMsg('#dos-msg',d.messages||d.error||'Could not process file',true); }
});

// ---- Step 4: fetch DOs from system (do_source.py, live DB) ----
async function doFetchDos(){
  const btn=$('#btn-dos-fetch'); btn.disabled=true;
  setMsg('#dos-fetch-msg','Fetching from system… ',false);
  show('#dos-fetch-result', false);
  try{
    const d=await jpost('/api/dos-fetch',{});
    if(d.error){ setMsg('#dos-fetch-msg', d.error, true); return; }
    setMsg('#dos-fetch-msg', null);
    $('#dos-fetch-summary').textContent = `Found ${d.count} DO(s), ${d.weight}T total.`;
    $('#dos-fetch-diag').textContent = buildDiagText(d) || '(no diagnostics returned)';
    show('#dos-fetch-diag', false);
    show('#dos-fetch-result', true);
  } finally { btn.disabled=false; }
}
$('#dos-fetch-diag-toggle').onclick=()=>{
  show('#dos-fetch-diag', $('#dos-fetch-diag').classList.contains('hidden'));
};
$('#btn-dos-fetch').onclick=doFetchDos;

async function useFetchedDos(){
  const btn=$('#btn-dos-fetch-use'); btn.disabled=true;
  setMsg('#dos-fetch-msg','Assigning lorries… this can take a moment ',false);
  try{
    const d=await jpost('/api/dos-fetch/use',{});
    if(handleDosResponse(d)){ setMsg('#dos-fetch-msg', null); }
    else { setMsg('#dos-fetch-msg', d.messages||d.error||'Could not process', true); }
  } finally { btn.disabled=false; }
}
$('#btn-dos-fetch-use').onclick=useFetchedDos;

async function answerOffsched(assign){
  setMsg('#offsched-msg', assign?'Assigning off-schedule DOs… ':'Finalising… ');
  const d=await jpost('/api/offschedule',{assign});
  if(d.result){ show('#card-offsched',false); showPickingList(); }
  else { setMsg('#offsched-msg', d.messages||d.error||'Something went wrong',true); }
}
document.getElementById('offsched-yes').onclick=()=>answerOffsched(true);
document.getElementById('offsched-no').onclick=()=>answerOffsched(false);

// ---- Step 5: render ----
function esc(s){ return String(s).replace(/[&<>]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[c])); }
let lastFleet=[], lastFullFleet=[], lastUnassignedList=[];
function renderResult(r){
  const s=r.summary;
  $('#result-stat').innerHTML =
    `<span class="pill">Lorries used <b>${s.lorries_used}</b></span>`+
    `<span class="pill">DOs assigned <b>${s.dos_assigned}</b></span>`+
    (s.dos_unassigned?`<span class="pill bad">Unassigned <b>${s.dos_unassigned}</b></span>`:'')+
    (s.dos_other?`<span class="pill">Other user / skipped <b>${s.dos_other}</b></span>`:'');
  let html='';
  r.lorries.forEach(g=>{
    const cap=g.capacity!=null?g.capacity.toFixed(2)+'T':'—';
    const util=g.util!=null?g.util:0;
    html+=`<div class="lorry"><h3>🚚 ${esc(g.lorry)} <span class="cap">${g.load.toFixed(2)}T / ${cap}${g.util!=null?' · '+g.util+'%':''}</span></h3>`;
    html+=`<div class="bar"><i class="${util>100?'hi':''}" style="width:${Math.min(util,100)}%"></i></div>`;
    html+=`<div class="scroll"><table><thead><tr><th>DO</th><th>Route</th><th>Customer</th><th class="w">Weight</th></tr></thead><tbody>`;
    g.dos.forEach(d=>{ html+=`<tr><td>${esc(d.do)}</td><td>${esc(d.route)}</td><td>${esc(d.customer)}</td><td class="w">${d.weight.toFixed(3)}T</td></tr>`; });
    html+=`</tbody></table></div></div>`;
  });
  if(r.unassigned.length){
    // PAST_DATE (dated before today, not auto-assigned) and any other
    // reasoned skip live in this same list — the Date column is what makes
    // "why is this here" legible instead of needing a separate section.
    html+=`<div class="lorry unassigned"><h3 class="collapsible">⚠️ Unassigned (${r.unassigned.length})</h3>`;
    html+=`<div class="scroll"><table><thead><tr><th>DO</th><th>Route</th><th>Customer</th><th class="w">Weight</th><th>Date</th><th>Reason</th></tr></thead><tbody>`;
    r.unassigned.forEach(d=>{ html+=`<tr><td>${esc(d.do)}</td><td>${esc(d.route)}</td><td>${esc(d.customer)}</td><td class="w">${d.weight.toFixed(3)}T</td><td>${esc(d.date||'')}</td><td>${esc(d.reason||'')}</td></tr>`; });
    html+=`</tbody></table></div></div>`;
  }
  $('#result-body').innerHTML=html;
  $('#result-body').querySelectorAll('.lorry h3.collapsible').forEach(h=>{
    h.onclick=()=>h.closest('.lorry').classList.toggle('collapsed');
  });

  // ---- Reassign leftover DOs onto now-available lorries ----
  lastFleet=r.fleet||[];
  lastFullFleet=r.full_fleet||r.fleet||[];
  lastUnassignedList=r.unassigned||[];
  const box=$('#reassign-box');
  if(r.unassigned.length && r.fleet){
    $('#reassign-count').textContent=r.unassigned.length;
    $('#reassign-weight').textContent=(r.unassigned_weight||0).toFixed(3);
    let f='';
    r.fleet.forEach(l=>{
      f+=`<label class="userbtn" style="cursor:pointer;display:flex;align-items:center;gap:6px;padding:10px 14px">`+
         `<input type="checkbox" class="reassign-cb" value="${esc(l.plate)}"> ${esc(l.plate)} `+
         `<small style="color:var(--muted)">${l.capacity.toFixed(1)}T${l.used?' · in use':''}</small></label>`;
    });
    $('#reassign-fleet').innerHTML=f;
    setMsg('#reassign-msg',null);
    box.classList.remove('hidden');
  } else {
    box.classList.add('hidden');
  }

  // ---- Manually assign hand-picked DOs onto one named lorry ----
  show('#manual-assign-box', r.unassigned.length>0);
  $('#manual-picker').classList.add('hidden');
  setMsg('#manual-msg', null);
}

// ==================== Drag-and-drop board ====================
let BOARD=null;
let boardOpenRoutes=new Set(), boardCollapsedLanes=new Set();
let boardMaximizedPlate=null;
let boardDrag=null, boardGhost=null;
let routeDragCandidate=null, routeDrag=null, routeGhost=null, routeDragJustHappened=false;

// ---- Picking List — review the fetched/assigned-scope DOs, uncheck any
// not needed today, before landing on the board. ----
let PICKING_ROWS = [];
async function showPickingList(){
  hideAll();
  document.body.classList.remove('board-active');
  show('#card-picking', true);
  setMsg('#picking-msg', 'Loading… ', false);
  const d = await (await fetch('/api/picking-list')).json();
  if(d.error){ setMsg('#picking-msg', d.error, true); return; }
  setMsg('#picking-msg', null);
  renderPickingList(d.rows || []);
  const diagText = buildDiagText({parse_diagnostics: LAST_PARSE_DIAG});
  $('#picking-fetch-diag').textContent = diagText;
  show('#picking-fetch-diag', false);
  show('#picking-fetch-diag-toggle', !!diagText);
}
$('#picking-fetch-diag-toggle').onclick=()=>{
  show('#picking-fetch-diag', $('#picking-fetch-diag').classList.contains('hidden'));
};
function renderPickingList(rows){
  PICKING_ROWS = rows;
  const tbody = $('#picking-rows');
  let html='', sawSpecial=false;
  rows.forEach(r=>{
    if(r.special && !sawSpecial){
      sawSpecial=true;
      html+=`<tr class="picking-section-divider"><td colspan="7">MANUAL ASSIGN ONLY — not auto-assigned by AI, place these by hand</td></tr>`;
    }
    html+=`
    <tr>
      <td><input type="checkbox" class="picking-cb" data-do="${esc(r.do)}" ${r.checked?'checked':''}></td>
      <td>${esc(r.date)}</td>
      <td>${esc(r.do)}</td>
      <td>${esc(r.code)}</td>
      <td>${esc(r.customer)}</td>
      <td>${esc(r.route)}</td>
      <td class="w">${r.weight.toFixed(3)}T</td>
    </tr>`;
  });
  tbody.innerHTML = html;
  tbody.querySelectorAll('.picking-cb').forEach(cb=>{ cb.onchange = updatePickingCount; });
  updatePickingCount();
}
function updatePickingCount(){
  const total = PICKING_ROWS.length;
  const checked = $('#picking-rows').querySelectorAll('.picking-cb:checked').length;
  $('#picking-count').textContent = `${checked} of ${total} selected`;
}
$('#picking-select-all').onclick = ()=>{
  $('#picking-rows').querySelectorAll('.picking-cb').forEach(cb=>cb.checked=true);
  updatePickingCount();
};
$('#picking-deselect-all').onclick = ()=>{
  $('#picking-rows').querySelectorAll('.picking-cb').forEach(cb=>cb.checked=false);
  updatePickingCount();
};
$('#picking-next').onclick = async()=>{
  const btn = $('#picking-next');
  btn.disabled = true;
  try{
    const unpicked = [...$('#picking-rows').querySelectorAll('.picking-cb')]
      .filter(cb=>!cb.checked).map(cb=>cb.dataset.do);
    setMsg('#picking-msg', 'Assigning… ', false);
    const d = await jpost('/api/picking-list/confirm', {unpicked});
    if(d.error){ setMsg('#picking-msg', d.error, true); return; }
    setMsg('#picking-msg', null);
    showBoard();
  } finally { btn.disabled = false; }
};

function showBoard(){
  hideAll(); show('#card-board',true);
  document.body.classList.add('board-active');
  loadBoard();
}
function showResultTable(result){
  hideAll();
  document.body.classList.remove('board-active');
  if(result) renderResult(result);
  show('#card-result',true);
}
$('#btn-board-table').onclick=async()=>{
  const d=await (await fetch('/api/state')).json();
  showResultTable(d && d.result ? d.result : null);
};
$('#btn-result-board').onclick=showBoard;
$('#btn-board-back').onclick=refetchDOs;
$('#btn-board-restart').onclick=restartPlanner;

function boardToast(msg){
  const t=document.createElement('div');
  t.className='board-toast'; t.textContent=msg;
  document.body.appendChild(t);
  setTimeout(()=>t.remove(),4200);
}

async function loadBoard(){
  setMsg('#board-msg','Loading board… ',false);
  const d=await (await fetch('/api/board')).json();
  if(d.error){ setMsg('#board-msg', d.error, true); return; }
  setMsg('#board-msg', null);
  BOARD=d;
  if(!boardOpenRoutes.size && d.routes.length) boardOpenRoutes.add(d.routes[0].route);
  renderBoard();
}

function boardOrdersInPool(route){ return BOARD.orders.filter(o=>o.route===route && !o.lorry && !o.manualOnly); }
function boardOrdersManualOnly(code){ return BOARD.orders.filter(o=>o.code===code && !o.lorry && o.manualOnly); }
function boardOrdersOnLorry(plate){ return BOARD.orders.filter(o=>o.lorry===plate); }
function boardSumKg(list){ return list.reduce((s,o)=>s+o.weight,0); }
function fmtT(w){ return w.toFixed(3)+'T'; }

const BOARD_ROUTE_COLORS=['#5ab0ff','#c58bff','#ffd166','#7ee8b2','#ff9e7d',
  '#8fd3ff','#f2a6d8','#b6e37a','#ffc98a','#9fb8ff'];
// Only these item codes (SDELIVERYD.ITMREF_0) render bold in the product
// list — everything else stays normal weight, by explicit request.
const BOARD_BOLD_ITMREFS = new Set(['SS007','SS005','SS006','BL002','BL004',
  'AG001','KM101-1','KM101-2','CK102','CK107','CK004','CD001','KP065']);
function boardRouteColor(route){
  const idx=(BOARD.routes||[]).findIndex(r=>r.route===route);
  return BOARD_ROUTE_COLORS[(idx<0?0:idx)%BOARD_ROUTE_COLORS.length];
}

function boardCardEl(o){
  const el=document.createElement('div');
  el.className='board-card'+(o._warned?' warned':'');
  el.dataset.do=o.do;
  const color=boardRouteColor(o.route);
  const deleteBtn = o.lorry
    ? `<button class="b-delete" title="Remove from ${esc(o.lorry)} and return to unassigned">&times;</button>`
    : '';
  // Each raw item is "ITMREF_0|description xqty" (see do_source.py) — the
  // ref decides bold vs. normal weight per line, then gets stripped before
  // display. A part with no '|' (stale cached board data from before this
  // format existed) falls back to plain, unbolded display.
  const productItems = o.products ? o.products.split(';').map(p=>p.trim()).filter(Boolean)
    .map(p=>{
      const bar = p.indexOf('|');
      return bar === -1
        ? {label: p, bold: false}
        : {label: p.slice(bar+1), bold: BOARD_BOLD_ITMREFS.has(p.slice(0, bar).trim())};
    }) : [];
  const productsHtml = productItems.length
    ? `<div class="b-meta b-products" style="color:var(--pink)">
         <div class="b-products-toggle">&#9656; ${productItems.length} product${productItems.length===1?'':'s'}</div>
         <div class="b-products-list hidden">${productItems.map(p=>`<div style="font-weight:${p.bold?700:400}">&bull; ${esc(p.label)}</div>`).join('')}</div>
       </div>`
    : '';
  el.innerHTML=`
    <span class="b-stripe" style="background:${color}"></span>
    <div class="b-body">
      <div class="b-top"><span class="b-id">${esc(o.do)}</span>${o.code?`<span class="b-code">${esc(o.code)}</span>`:''}${o.distance?`<span class="b-dist">${esc(o.distance)}</span>`:''}<span class="b-kg">${fmtT(o.weight)}</span>${deleteBtn}</div>
      <div class="b-cust">${esc(o.customer)}</div>
      <div class="b-meta">${esc(o.route)} &middot; ${esc(o.date)}</div>
      ${o.remarks?`<div class="b-meta" style="color:var(--warn);font-weight:700">${esc(o.remarks)}</div>`:''}
      ${productsHtml}
      ${o.reason?`<div class="b-meta" style="color:var(--bad)">${esc(o.reason)}</div>`:''}
      ${o._warned?`<div class="b-warn">⚠️ ${esc(o._warned)}</div>`:''}
    </div>`;
  if(o.lorry){
    el.querySelector('.b-delete').addEventListener('click', e=>{
      e.stopPropagation();
      moveBoardCard(o.do, null);
    });
  }
  const _prodToggle = el.querySelector('.b-products-toggle');
  if(_prodToggle){
    _prodToggle.addEventListener('click', e=>{
      e.stopPropagation();
      const list = el.querySelector('.b-products-list');
      const expanding = list.classList.contains('hidden');
      list.classList.toggle('hidden', !expanding);
      _prodToggle.innerHTML = `${expanding?'&#9662;':'&#9656;'} ${productItems.length} product${productItems.length===1?'':'s'}`;
    });
  }
  el.addEventListener('pointerdown', e=>{
    if(e.target.classList.contains('b-delete') || e.target.closest('.b-products-toggle')) return;
    startBoardDrag(e,o.do);
  });
  if(boardDrag && boardDrag.doId===o.do) el.classList.add('dragging');
  return el;
}

function renderBoard(){
  if(!BOARD) return;
  const wrap=$('#board-routes'); wrap.innerHTML='';
  let totalUn=0;
  BOARD.routes.forEach(rt=>{
    const list=boardOrdersInPool(rt.route);
    if(!list.length) return;
    totalUn+=list.length;
    const div=document.createElement('div');
    div.className='board-route'+(boardOpenRoutes.has(rt.route)?' open':'');
    div.innerHTML=`
      <div class="board-route-head">
        <span class="board-route-chev">&#9654;</span>
        <span class="board-route-dot" style="background:${boardRouteColor(rt.route)}"></span>
        <span class="board-route-code">${esc(rt.route)}</span>
        <span class="board-route-meta">${list.length} DO &middot; ${fmtT(rt.weight)}</span>
      </div>
      <div class="board-route-body"></div>`;
    const _head=div.querySelector('.board-route-head');
    _head.onclick=()=>{
      if(routeDragJustHappened){ routeDragJustHappened=false; return; }
      boardOpenRoutes.has(rt.route)?boardOpenRoutes.delete(rt.route):boardOpenRoutes.add(rt.route);
      renderBoard();
    };
    _head.addEventListener('pointerdown', e=>{
      if(e.target.closest('.board-route-chev')) return;
      routeDragCandidate={kind:'route', key:rt.route, startX:e.clientX, startY:e.clientY, count:list.length, weight:rt.weight};
    });
    const body=div.querySelector('.board-route-body');
    list.forEach(o=>body.appendChild(boardCardEl(o)));
    wrap.appendChild(div);
  });
  // Manual-assign-only customers (live ROUTE == "NA", plus any this
  // session has dragged in/out) — AI Assign never touches these; shown as
  // their own drop zone below the normal route groups so the planner
  // remembers to place them by hand. Drag a single DO onto this zone to
  // flag its whole customer code; drag a group's header back onto the
  // Unassigned pool to release it for AI Assign again.
  const moGroups=BOARD.manual_only||[];
  const moWrap=document.createElement('div');
  moWrap.dataset.zone='MO';
  moWrap.className='board-mo-zone';
  const hdr=document.createElement('div');
  hdr.className='board-pool-section-hdr';
  hdr.textContent='MANUAL ASSIGN ONLY';
  moWrap.appendChild(hdr);
  let moCount=0;
  moGroups.forEach(mo=>{
    const list=boardOrdersManualOnly(mo.code);
    if(!list.length) return;
    moCount+=list.length;
    totalUn+=list.length;
    const key='MO:'+mo.code;
    const div=document.createElement('div');
    div.className='board-route'+(boardOpenRoutes.has(key)?' open':'');
    div.innerHTML=`
      <div class="board-route-head">
        <span class="board-route-chev">&#9654;</span>
        <span class="board-route-dot" style="background:#f59e0b"></span>
        <span class="board-route-code">${esc(mo.code)} - ${esc(mo.customer)}</span>
        <span class="board-route-meta">${list.length} DO &middot; ${fmtT(mo.weight)}</span>
      </div>
      <div class="board-route-body"></div>`;
    const _moHead=div.querySelector('.board-route-head');
    _moHead.onclick=()=>{
      if(routeDragJustHappened){ routeDragJustHappened=false; return; }
      boardOpenRoutes.has(key)?boardOpenRoutes.delete(key):boardOpenRoutes.add(key);
      renderBoard();
    };
    _moHead.addEventListener('pointerdown', e=>{
      if(e.target.closest('.board-route-chev')) return;
      routeDragCandidate={kind:'mo', key:mo.code, startX:e.clientX, startY:e.clientY, count:list.length, weight:mo.weight};
    });
    const body=div.querySelector('.board-route-body');
    list.forEach(o=>body.appendChild(boardCardEl(o)));
    moWrap.appendChild(div);
  });
  if(!moCount){
    const empty=document.createElement('div');
    empty.className='board-empty';
    empty.textContent='Drag a DO here to exclude it from AI Assign';
    moWrap.appendChild(empty);
  }
  wrap.appendChild(moWrap);
  $('#board-pool-label').textContent=`UNASSIGNED · ${totalUn} DO${totalUn===1?'':'s'}`;
  const _aging=BOARD.aging||{};
  const _agingParts=[
    _aging.over7?`<span class="aging-pill">More than 7 days: ${_aging.over7} DO${_aging.over7===1?'':'s'}</span>`:'',
    _aging.over5?`<span class="aging-pill">More than 5 days: ${_aging.over5} DO${_aging.over5===1?'':'s'}</span>`:'',
    _aging.over3?`<span class="aging-pill">More than 3 days: ${_aging.over3} DO${_aging.over3===1?'':'s'}</span>`:'',
  ].filter(Boolean);
  $('#board-pool-aging').innerHTML=_agingParts.join('');
  const totalAssigned = BOARD.orders.filter(o=>o.lorry).length;
  $('#board-assigned-stat').innerHTML = `Assigned <b>${totalAssigned}</b> / ${BOARD.orders.length} DO${BOARD.orders.length===1?'':'s'}`;

  const lanes=$('#board-lanes');
  lanes.className='board-lanes'+(boardMaximizedPlate?' has-maximized':'');
  lanes.innerHTML='';
  BOARD.lorries.forEach(t=>{
    if(boardMaximizedPlate && t.plate!==boardMaximizedPlate) return;   // hidden while another lane is focused
    const maximized = t.plate===boardMaximizedPlate;
    const list=boardOrdersOnLorry(t.plate);
    const load=boardSumKg(list);
    const pctReal=t.capacity?Math.round(load/t.capacity*100):0;
    const pct=Math.min(100,pctReal);   // bar fill stays capped at 100%; text below shows the real number
    const over=t.capacity && load>t.capacity;
    const fillClass= over?'over': pct>=85?'hi':'';
    const collapsed=!maximized && boardCollapsedLanes.has(t.plate);
    const isOn=t.on!==false;
    const lane=document.createElement('section');
    lane.className='board-lane'+(collapsed?' collapsed':'')+(isOn?'':' lane-off')+(maximized?' board-lane-maximized':'');
    if(isOn) lane.dataset.zone=t.plate;
    lane.innerHTML=`
      <div class="board-lane-head">
        <span class="board-lane-chev">&#9660;</span>
        <span class="board-lane-plate">${esc(t.plate)}</span>
        <button class="lane-toggle${isOn?'':' off'}" title="${isOn?'Available today — click to turn off':'Not available today — click to turn on'}"></button>
        <button class="lane-max-btn" title="${maximized?'Minimize':'Maximize — focus assigning on this lorry'}">${maximized?'&#10529;':'&#10530;'}</button>
        <button class="lane-drop-btn" title="Drop all of this lorry's DOs back to Unassigned" ${list.length?'':'disabled'}>&#128465;</button>
        <span class="board-lane-count">${list.length} DO${list.length===1?'':'s'}</span>
        <span class="board-lane-drops" title="Dropping points — distinct customer + drop point combinations on this lorry">${t.droppingPoints||0} drop${(t.droppingPoints||0)===1?'':'s'}</span>
        <span class="board-lane-load" style="color:${over?'var(--bad)':'var(--ink)'}">${fmtT(load)} / ${t.capacity!=null?t.capacity.toFixed(2)+'T':'—'}${t.capacity!=null?' &middot; '+pctReal+'%':''}</span>
      </div>
      <div class="board-lane-naik">naik limit ${t.capacity!=null?t.capacity.toFixed(2)+'T':'—'}</div>
      <div class="board-cap-track"><div class="board-cap-fill ${fillClass}" style="width:${pct}%"></div></div>
      <div class="board-lane-body"></div>`;
    lane.querySelector('.board-lane-head').onclick=(e)=>{
      if(e.target.closest('.lane-toggle') || e.target.closest('.lane-max-btn') || e.target.closest('.lane-drop-btn')) return;
      collapsed?boardCollapsedLanes.delete(t.plate):boardCollapsedLanes.add(t.plate);
      renderBoard();
    };
    lane.querySelector('.lane-toggle').onclick=async(e)=>{
      e.stopPropagation();
      await toggleLorry(t.plate, !isOn);
    };
    lane.querySelector('.lane-max-btn').onclick=(e)=>{
      e.stopPropagation();
      boardMaximizedPlate = maximized ? null : t.plate;
      renderBoard();
    };
    lane.querySelector('.lane-drop-btn').onclick=async(e)=>{
      e.stopPropagation();
      if(!list.length) return;
      if(!confirm(`Drop all ${list.length} DO(s) from ${t.plate} back to Unassigned?`)) return;
      await dropLane(t.plate);
    };
    const body=lane.querySelector('.board-lane-body');
    if(!list.length) body.innerHTML='<div class="board-empty">Drop DOs here</div>';
    list.forEach(o=>body.appendChild(boardCardEl(o)));
    lanes.appendChild(lane);
  });
}
$('#board-collapse-all').onclick=()=>{
  if(BOARD) BOARD.lorries.forEach(t=>boardCollapsedLanes.add(t.plate));
  renderBoard();
};
$('#board-expand-all').onclick=()=>{ boardCollapsedLanes.clear(); renderBoard(); };
$('#board-pool-collapse-all').onclick=()=>{ boardOpenRoutes.clear(); renderBoard(); };
$('#board-pool-expand-all').onclick=()=>{
  if(BOARD){
    BOARD.routes.forEach(rt=>boardOpenRoutes.add(rt.route));
    (BOARD.manual_only||[]).forEach(mo=>boardOpenRoutes.add('MO:'+mo.code));
  }
  renderBoard();
};

function startBoardDrag(e,doId){
  e.preventDefault(); e.stopPropagation();
  boardDrag={doId};
  const o=BOARD.orders.find(x=>x.do===doId);
  boardGhost=document.createElement('div');
  boardGhost.className='board-ghost';
  boardGhost.innerHTML=`<span style="color:var(--brand);font-weight:700;font-family:ui-monospace,Menlo,monospace">${esc(o.do)}</span>`+
    `<span style="overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${esc(o.customer)}</span>`+
    `<span style="color:var(--muted);flex-shrink:0">${fmtT(o.weight)}</span>`;
  document.body.appendChild(boardGhost);
  moveBoardGhost(e.clientX,e.clientY);
  renderBoard();
}
function moveBoardGhost(x,y,ghost){ ghost=ghost||boardGhost; if(ghost){ ghost.style.left=x+'px'; ghost.style.top=y+'px'; } }
function boardZoneAt(x,y){
  for(const el of document.elementsFromPoint(x,y)){
    const z=el.closest?.('[data-zone]');
    if(z) return z;
  }
  return null;
}
document.addEventListener('pointermove', e=>{
  if(routeDragCandidate && !routeDrag){
    const dx=e.clientX-routeDragCandidate.startX, dy=e.clientY-routeDragCandidate.startY;
    if(Math.hypot(dx,dy) > 6){
      routeDrag={kind:routeDragCandidate.kind, key:routeDragCandidate.key};
      routeGhost=document.createElement('div');
      routeGhost.className='board-ghost';
      routeGhost.innerHTML=`<span style="color:var(--brand);font-weight:700;font-family:ui-monospace,Menlo,monospace">${esc(routeDrag.key)}</span>`+
        `<span style="overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${routeDrag.kind==='mo'?'whole customer':'whole route'}</span>`+
        `<span style="color:var(--muted);flex-shrink:0">${routeDragCandidate.count} DO &middot; ${fmtT(routeDragCandidate.weight)}</span>`;
      document.body.appendChild(routeGhost);
    }
  }
  if(routeDrag){
    moveBoardGhost(e.clientX,e.clientY,routeGhost);
    document.querySelectorAll('.zone-active').forEach(el=>el.classList.remove('zone-active'));
    const z=boardZoneAt(e.clientX,e.clientY);
    // Whole-route drags only land on a lorry lane; whole-customer (manual-
    // only) drags may also land on the Unassigned pool zone ('') to release it.
    if(z && (z.dataset.zone || routeDrag.kind==='mo')) z.classList.add('zone-active');
    return;
  }
  if(!boardDrag) return;
  moveBoardGhost(e.clientX,e.clientY,boardGhost);
  document.querySelectorAll('.zone-active').forEach(el=>el.classList.remove('zone-active'));
  const z=boardZoneAt(e.clientX,e.clientY);
  if(z) z.classList.add('zone-active');
});
document.addEventListener('pointerup', async e=>{
  if(routeDragCandidate || routeDrag){
    const wasDragging=!!routeDrag;
    const z=wasDragging?boardZoneAt(e.clientX,e.clientY):null;
    const dragKind=routeDrag?routeDrag.kind:null;
    const dragKey=routeDrag?routeDrag.key:null;
    routeDragCandidate=null; routeDrag=null;
    if(routeGhost){ routeGhost.remove(); routeGhost=null; }
    document.querySelectorAll('.zone-active').forEach(el=>el.classList.remove('zone-active'));
    if(wasDragging){
      routeDragJustHappened=true;
      if(dragKind==='mo'){
        // Dropped back on its own zone ('MO') is a no-op; the pool ('') or
        // any lorry plate both release the code (see /api/board/move-code).
        if(z && z.dataset.zone!==undefined && z.dataset.zone!=='MO'){
          await moveBoardCode(dragKey, z.dataset.zone || null);
        } else {
          renderBoard();
        }
      } else if(z && z.dataset.zone){
        await moveBoardRoute(dragKey, z.dataset.zone);
      } else {
        renderBoard();
      }
    }
    return;
  }
  if(!boardDrag) return;
  const z=boardZoneAt(e.clientX,e.clientY);
  const doId=boardDrag.doId;
  boardDrag=null;
  if(boardGhost){ boardGhost.remove(); boardGhost=null; }
  document.querySelectorAll('.zone-active').forEach(el=>el.classList.remove('zone-active'));
  if(!z){ renderBoard(); return; }
  const zone=z.dataset.zone;  // '' = pool, 'MO' = manual-assign-only, else a plate
  if(zone==='MO'){ await markManualOnly(doId); return; }
  await moveBoardCard(doId, zone || null);
});

async function moveBoardRoute(route, plate){
  const d=await jpost('/api/board/move-route',{route, lorry:plate});
  if(!d.ok){ boardToast(d.message||d.error||'Move failed'); await loadBoard(); return; }
  if(d.warnings && d.warnings.length){
    boardToast(`${route}: ${d.moved}/${d.total} moved · ${d.warnings.slice(0,2).join(' · ')}`);
  } else {
    boardToast(`${route}: ${d.moved} DO(s) moved to ${plate}.`);
  }
  if(d.board){ BOARD=d.board; }
  renderBoard();
}

async function moveBoardCode(code, plate){
  const d=await jpost('/api/board/move-code',{code, lorry:plate});
  if(!d.ok){ boardToast(d.message||d.error||'Move failed'); await loadBoard(); return; }
  boardToast(plate ? `${code}: ${d.moved} DO(s) moved to ${plate}.` : `${code}: released — AI Assign can place it now.`);
  if(d.board){ BOARD=d.board; }
  renderBoard();
}

async function markManualOnly(doId){
  const d=await jpost('/api/board/mark-manual-only',{do:doId});
  if(!d.ok){ boardToast(d.message||d.error||'Move failed'); await loadBoard(); return; }
  boardToast(`${d.code}: moved to Manual Assign Only.`);
  if(d.board){ BOARD=d.board; }
  renderBoard();
}

async function moveBoardCard(doId, plate){
  const d=await jpost('/api/board/move',{do:doId, lorry:plate});
  if(!d.ok){ boardToast(d.message||d.error||'Move failed'); await loadBoard(); return; }
  if(d.warnings && d.warnings.length){ boardToast(`${doId}: ${d.warnings.join(' · ')}`); }
  if(d.board){ BOARD=d.board;
    if(plate){
      const o=BOARD.orders.find(x=>x.do===doId);
      if(o) o._warned = (d.warnings&&d.warnings.length) ? d.warnings[0] : null;
    }
  }
  renderBoard();
}

async function dropLane(plate){
  const d=await jpost('/api/board/drop-lane',{lorry:plate});
  if(!d.ok){ boardToast(d.message||d.error||'Drop failed'); await loadBoard(); return; }
  boardToast(`${plate}: ${d.dropped} DO(s) dropped back to Unassigned.`);
  if(d.board){ BOARD=d.board; }
  renderBoard();
}

async function toggleLorry(plate, on){
  const d=await jpost('/api/board/toggle-lorry',{plate, on});
  if(!d.ok){ boardToast(d.message||d.error||'Toggle failed'); await loadBoard(); return; }
  if(d.board){ BOARD=d.board; }
  if(!on && d.unassigned_count){
    boardToast(`${plate} turned off · ${d.unassigned_count} DO(s) sent back to unassigned.`);
  } else if(on && d.refilled_count){
    boardToast(`${plate} turned on · ${d.refilled_count} DO(s) it was carrying put back on.`);
  }
  // Both toggle points (the setup grid above and the board's own lane
  // switches) call this — keep whichever one isn't the one just clicked
  // in sync too, since they can be visible at the same time.
  if(BOARD) renderBoard();
  const _toggleSection = $('#tp-toggle-section');
  if(_toggleSection && !_toggleSection.classList.contains('hidden')) await loadLorryToggles();
}

async function aiAssign(){
  const btn=$('#btn-board-ai'); btn.disabled=true;
  setMsg('#board-msg','AI is assigning… this can take a moment ',false);
  try{
    const d=await jpost('/api/board/ai-assign',{});
    if(d.error){ setMsg('#board-msg', d.error, true); return; }
    setMsg('#board-msg', null);
    if(d.board){ BOARD=d.board; renderBoard(); }
  } finally { btn.disabled=false; }
}
$('#btn-board-ai').onclick=aiAssign;

$('#board-download-toggle').onclick=(e)=>{
  e.stopPropagation();
  show('#board-download-menu', $('#board-download-menu').classList.contains('hidden'));
};
document.addEventListener('click', (e)=>{
  if(!e.target.closest('#board-download-dropdown')) show('#board-download-menu', false);
});

async function saveBoardToSql(){
  const btn=$('#btn-board-save'); btn.disabled=true;
  try{
    const p=await jpost('/api/board/save-preview',{});
    if(p.error){ boardToast(p.error); return; }
    const lines=[
      `Save to SQL — Assign for ${p.add_date}, Trip ${p.trip}`,
      '',
      `${p.dos_assigned} DO(s) assigned across ${p.lorries.length} lorry(s):`,
      ...p.lorries.map(l=>`  ${l.plate}: ${l.ton_kg.toFixed(1)} KG`),
    ];
    if(p.dos_unassigned){
      lines.push('', `${p.dos_unassigned} unassigned DO(s) in scope — left untouched, not written.`);
    }
    lines.push('', 'This writes directly to the live ERP database (SDELIVERY + ZLORRY). Continue?');
    if(!confirm(lines.join('\n'))) return;
    setMsg('#board-msg','Saving to SQL… ',false);
    const d=await jpost('/api/board/save',{});
    if(!d.ok){ setMsg('#board-msg', d.error||'Save failed', true); return; }
    setMsg('#board-msg', null);
    boardToast(`Saved: ${d.dos_written} DO(s), ${d.lorries_written} lorry(s) written to SQL. Downloading Excel…`);
    window.location.href='/api/download';
  } finally { btn.disabled=false; }
}
$('#btn-board-save').onclick=saveBoardToSql;

function buildDiagText(d){
  const parts=[];
  if(d && d.diagnostics && Object.keys(d.diagnostics).length){
    const dg=d.diagnostics;
    parts.push(`SQL FETCH\n`+
      `raw SQL rows: ${dg.raw_rows_from_sql}\n`+
      `after not-LOAN: ${dg.after_not_loan}\n`+
      `after site 1SA: ${dg.after_site_1SA}\n`+
      `after not-yet-validated: ${dg.after_not_validated}\n`+
      `after known route: ${dg.after_known_route}\n`+
      `after 30-day floor: ${dg.after_30day_floor}`+
      (dg.after_etd_window!=null?`\nafter ETD window: ${dg.after_etd_window}`:'')+
      `\nsite codes seen: ${(dg.distinct_stofcy_seen||[]).join(', ')||'none'}`+
      `\ntypes seen: ${(dg.distinct_sdhtyp_seen||[]).join(', ')||'none'}`);
  }
  if(d && d.parse_diagnostics){
    const pg=d.parse_diagnostics;
    parts.push(`AFTER-FETCH CLASSIFICATION\n`+
      `rows parsed: ${pg.raw_rows_parsed}\n`+
      `excluded — other user's route: ${pg.other_user}\n`+
      `excluded — not on today's schedule: ${pg.not_today}\n`+
      `excluded — older than 30-day cutoff: ${pg.past_date}\n`+
      `excluded — wrong trip: ${pg.wrong_trip}\n`+
      `excluded — REMARKS skip: ${pg.remarks_skip}\n`+
      `= actionable (this planner's pool): ${pg.actionable}\n`+
      (pg.today_used_by_app_server?`\napp server's "today": ${pg.today_used_by_app_server}`:'')+
      (pg.past_date_cutoff_used?`\npast-date cutoff used: ${pg.past_date_cutoff_used}`:'')+
      (pg.past_date_samples&&pg.past_date_samples.length?`\nmost-recent excluded past-date DOs: `+
        pg.past_date_samples.map(s=>`${s.do} (${s.date})`).join(', '):'')+
      (pg.other_user_route_prefixes&&pg.other_user_route_prefixes.length?`\nexcluded other-user route prefixes: `+
        pg.other_user_route_prefixes.join(', '):''));
  }
  return parts.join('\n\n');
}

function renderFetchDiag(d, toggleSel, diagSel){
  const text = buildDiagText(d);
  show(diagSel, false);
  if(!text){ show(toggleSel, false); return; }
  $(diagSel).textContent = text;
  show(toggleSel, true);
}
$('#board-fetch-diag-toggle').onclick=()=>{
  show('#board-fetch-diag', $('#board-fetch-diag').classList.contains('hidden'));
};

async function refetchDOs(){
  const btn=$('#btn-board-back'); btn.disabled=true;
  setMsg('#board-msg','Checking for new DOs… ',false);
  show('#board-fetch-diag-toggle', false); show('#board-fetch-diag', false);
  try{
    const d=await jpost('/api/board/refetch',{});
    renderFetchDiag(d, '#board-fetch-diag-toggle', '#board-fetch-diag');
    if(d.error){ setMsg('#board-msg', d.error, true); return; }
    if(d.board){ BOARD=d.board; renderBoard(); }
    const n=d.new_count||0;
    setMsg('#board-msg', n ? `✅ Found ${n} new DO(s) — added to Unassigned.` : 'No new DOs found — everything is already on the board.', false);
  } finally { btn.disabled=false; }
}

async function doReassign(){
  const plates=[...document.querySelectorAll('.reassign-cb:checked')].map(c=>c.value);
  if(!plates.length){ setMsg('#reassign-msg','Tick at least one lorry first.',true); return; }
  setMsg('#reassign-msg','Assigning leftover DOs… ');
  const d=await jpost('/api/reassign',{lorries:plates});
  const o=d.outcome||{};
  if(d.result){
    renderResult(d.result);
    let m=`✅ Assigned ${o.assigned||0} of ${o.total||0} leftover DO(s)`+(o.used&&o.used.length?` to ${o.used.join(', ')}`:'')+'.';
    if(o.still&&o.still.length) m+=`  ⚠️ ${o.still.length} still don't fit (size/route/state rules) — tick another lorry.`;
    if(o.skipped&&o.skipped.length) m+=`  (Ignored ${o.skipped.join(', ')} — already in use today.)`;
    setMsg('#reassign-msg', m, (o.assigned||0)===0);
  } else {
    setMsg('#reassign-msg', d.error||'Reassign failed', true);
  }
}
document.getElementById('btn-reassign').onclick=doReassign;

// ---- Manually assign hand-picked DOs onto one named lorry ----
function manualLookup(){
  const plate=($('#manual-plate').value||'').trim().toUpperCase();
  if(!plate) return;
  const found=lastFullFleet.find(l=>l.plate===plate);
  if(!found){
    setMsg('#manual-msg', `Unknown plate "${esc(plate)}". Check the spelling, or update the master lorry file first.`, true);
    $('#manual-picker').classList.add('hidden');
    return;
  }
  setMsg('#manual-msg', null);
  $('#manual-plate-cap').textContent=found.capacity.toFixed(2);
  let rows='';
  lastUnassignedList.forEach(d=>{
    rows+=`<tr><td><input type="checkbox" class="manual-do-cb" data-w="${d.weight}" value="${esc(d.do)}"></td>`+
          `<td>${esc(d.do)}</td><td>${esc(d.route)}</td><td>${esc(d.customer)}</td><td class="w">${d.weight.toFixed(3)}T</td></tr>`;
  });
  $('#manual-do-rows').innerHTML=rows;
  document.querySelectorAll('.manual-do-cb').forEach(cb=>cb.addEventListener('change',()=>updateManualTotal(found.capacity)));
  updateManualTotal(found.capacity);
  $('#manual-picker').classList.remove('hidden');
}

function updateManualTotal(cap){
  const checked=[...document.querySelectorAll('.manual-do-cb:checked')];
  const total=checked.reduce((s,c)=>s+parseFloat(c.dataset.w),0);
  $('#manual-selected-total').textContent=total.toFixed(3);
  $('#manual-cap-warn').classList.toggle('hidden', total<=cap+1e-6);
  $('#btn-manual-assign').disabled = checked.length===0;
}

async function manualAssign(){
  const plate=($('#manual-plate').value||'').trim().toUpperCase();
  const checked=[...document.querySelectorAll('.manual-do-cb:checked')].map(c=>c.value);
  if(!plate || !checked.length) return;
  const btn=$('#btn-manual-assign'); btn.disabled=true;
  setMsg('#manual-msg','Assigning… ',false);
  try{
    const d=await jpost('/api/assign-specific',{plate:plate, dos:checked});
    if(d.ok){
      renderResult(d.result);
      setMsg('#manual-msg', `✅ Assigned ${d.assigned} DO(s) to ${plate} (${d.weight}T).`, false);
      show('#manual-msg', true);
    } else {
      setMsg('#manual-msg', d.message || d.error || 'Could not assign.', true);
      btn.disabled=false;
    }
  } catch(e){ setMsg('#manual-msg','Something went wrong.',true); btn.disabled=false; }
}
$('#btn-manual-lookup').onclick=manualLookup;
$('#btn-manual-assign').onclick=manualAssign;
$('#manual-plate').addEventListener('keydown', e=>{ if(e.key==='Enter'){ e.preventDefault(); manualLookup(); } });

$('#btn-restart').onclick=restartPlanner;

async function doRefresh(){
  const d=await (await fetch('/api/state')).json();
  if(d && d.result){ renderResult(d.result); }
}
$('#btn-refresh').onclick=doRefresh;

// On load: render the planner tabs, then either restore an already-active
// session (e.g. after a real browser refresh) straight onto its board, or
// log the first planner in and load today's lorries automatically — DOs
// still wait for the user to set an ETD window and click Fetch & Assign.
async function boot(){
  await loadUsers();
  const d = await (await fetch('/api/state')).json();
  if(d && d.email){ $('#who').textContent = d.email; }
  if(d && d.user && d.result){
    setActiveTab(d.user);
    // A real browser refresh lands here directly (skipping autoLoadPlanner
    // entirely), so the 1)/2)/3) setup rows — hidden by default in the HTML
    // — need to be shown here too, or they just stay gone after a refresh.
    _resetEtdDays();
    show('#tp-etd-row', true);
    show('#tp-toggle-section', true);
    show('#tp-fetch-row', true);
    setDayDate(_earliestAllowedDate());
    await updateTimeRestrictions();
    await loadLorryToggles();
    if(d.needs_picking){ await showPickingList(); } else { showBoard(); }
    return;
  }
  if(validUsers.length){ autoLoadPlanner(validUsers[0]); }
}
wireBackButtons();
boot();
</script>
</body>
</html>"""


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8000"))
    host = os.environ.get("HOST", "0.0.0.0")
    print(f"\n  DO Lorry-Assignment web app running.")
    print(f"  Open on this PC:      http://127.0.0.1:{port}/login")
    print(f"  Open from a phone:    http://<this-pc-ip>:{port}/login  "
          f"(e.g. http://10.0.0.229:{port}/login)\n")
    # Prefer waitress (a production WSGI server) for stable 24/7 running as a
    # Windows service; fall back to Flask's dev server if it isn't installed.
    try:
        from waitress import serve
        print("  Serving with waitress (production).\n")
        serve(app, host=host, port=port, threads=8)
    except ImportError:
        print("  waitress not installed — using Flask dev server. For 24/7, run"
              " 'pip install waitress'.\n")
        app.run(host=host, port=port, debug=False)
